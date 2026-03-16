from __future__ import annotations

import json
import os
import tempfile
from datetime import datetime, timedelta
from typing import Any

try:
    import numpy as np
except Exception:  # pragma: no cover
    np = None

try:
    import pandas as pd
except Exception:  # pragma: no cover
    pd = None

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo
except Exception:  # pragma: no cover
    Workbook = None
    Font = None
    PatternFill = None
    Table = None
    TableStyleInfo = None

from fastapi_modulo.db import SessionLocal
from fastapi_modulo.modulos.web.modelos.db_models import WebLoginAttempt, WebSecurityEvent


def _hours_since(hours: int) -> datetime:
    return datetime.utcnow() - timedelta(hours=max(1, int(hours)))


def _normalize_event_metadata(row: WebSecurityEvent) -> dict[str, Any]:
    payload = row.metadata_json or {}
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError:
            payload = {}
    return payload if isinstance(payload, dict) else {}


def load_access_history(hours: int = 24) -> list[dict[str, Any]]:
    db = SessionLocal()
    try:
        rows = (
            db.query(WebLoginAttempt)
            .filter(WebLoginAttempt.created_at >= _hours_since(hours))
            .order_by(WebLoginAttempt.created_at.desc())
            .all()
        )
        return [
            {
                "created_at": row.created_at.isoformat() if row.created_at else "",
                "tenant_id": row.tenant_id,
                "username": row.username,
                "ip": row.ip,
                "user_agent": row.user_agent,
                "success": bool(row.success),
            }
            for row in rows
        ]
    finally:
        db.close()


def load_screen_usage(hours: int = 24) -> list[dict[str, Any]]:
    db = SessionLocal()
    try:
        rows = (
            db.query(WebSecurityEvent)
            .filter(
                WebSecurityEvent.created_at >= _hours_since(hours),
                WebSecurityEvent.event_type == "screen_view",
            )
            .order_by(WebSecurityEvent.created_at.desc())
            .all()
        )
        items: list[dict[str, Any]] = []
        for row in rows:
            metadata = _normalize_event_metadata(row)
            items.append(
                {
                    "created_at": row.created_at.isoformat() if row.created_at else "",
                    "tenant_id": row.tenant_id,
                    "user_id": row.user_id,
                    "username": row.username,
                    "role": str(metadata.get("role") or ""),
                    "path": str(metadata.get("path") or ""),
                    "module_name": str(metadata.get("module_name") or ""),
                    "screen_name": str(metadata.get("screen_name") or ""),
                }
            )
        return items
    finally:
        db.close()


def _dataframe(records: list[dict[str, Any]], columns: list[str]):
    if pd is not None:
        frame = pd.DataFrame(records)
        for column in columns:
            if column not in frame.columns:
                frame[column] = ""
        return frame[columns]
    normalized: list[dict[str, Any]] = []
    for record in records:
        normalized.append({column: record.get(column, "") for column in columns})
    return normalized


def build_access_history_dataframe(hours: int = 24):
    columns = ["created_at", "tenant_id", "username", "ip", "user_agent", "success"]
    return _dataframe(load_access_history(hours), columns)


def build_screen_usage_dataframe(hours: int = 24):
    columns = ["created_at", "tenant_id", "user_id", "username", "role", "path", "module_name", "screen_name"]
    return _dataframe(load_screen_usage(hours), columns)


def _safe_mean_std(values: list[float]) -> tuple[float, float]:
    if not values:
        return 0.0, 0.0
    if np is not None:
        array = np.array(values, dtype=float)
        return float(array.mean()), float(array.std())
    mean = sum(values) / len(values)
    variance = sum((value - mean) ** 2 for value in values) / len(values)
    return mean, variance ** 0.5


def analyze_failed_login_patterns(hours: int = 24) -> dict[str, Any]:
    records = load_access_history(hours)
    failures = [item for item in records if not item.get("success")]
    if not failures:
        return {
            "window_hours": int(hours),
            "failed_attempts": 0,
            "peak_hours": [],
            "top_ips": [],
            "top_usernames": [],
            "anomalies": [],
        }
    if pd is not None:
        frame = pd.DataFrame(failures)
        frame["created_at"] = pd.to_datetime(frame["created_at"], errors="coerce")
        frame["hour_bucket"] = frame["created_at"].dt.strftime("%Y-%m-%d %H:00")
        hour_counts = frame.groupby("hour_bucket").size().reset_index(name="attempts").sort_values("attempts", ascending=False)
        ip_counts = frame.groupby("ip").size().reset_index(name="attempts").sort_values("attempts", ascending=False)
        user_counts = frame.groupby("username").size().reset_index(name="attempts").sort_values("attempts", ascending=False)
        values = [float(value) for value in hour_counts["attempts"].tolist()]
        mean_value, std_value = _safe_mean_std(values)
        anomalies = hour_counts[hour_counts["attempts"] > (mean_value + max(std_value, 1.0))].to_dict(orient="records")
        return {
            "window_hours": int(hours),
            "failed_attempts": int(len(frame.index)),
            "peak_hours": hour_counts.head(6).to_dict(orient="records"),
            "top_ips": ip_counts.head(10).to_dict(orient="records"),
            "top_usernames": user_counts.head(10).to_dict(orient="records"),
            "anomalies": anomalies,
        }
    hour_counts: dict[str, int] = {}
    ip_counts: dict[str, int] = {}
    user_counts: dict[str, int] = {}
    for item in failures:
        hour_bucket = str(item.get("created_at") or "")[:13] + ":00"
        hour_counts[hour_bucket] = hour_counts.get(hour_bucket, 0) + 1
        ip = str(item.get("ip") or "")
        username = str(item.get("username") or "")
        ip_counts[ip] = ip_counts.get(ip, 0) + 1
        user_counts[username] = user_counts.get(username, 0) + 1
    values = [float(value) for value in hour_counts.values()]
    mean_value, std_value = _safe_mean_std(values)
    peak_hours = [{"hour_bucket": key, "attempts": value} for key, value in sorted(hour_counts.items(), key=lambda item: item[1], reverse=True)[:6]]
    anomalies = [item for item in peak_hours if item["attempts"] > (mean_value + max(std_value, 1.0))]
    return {
        "window_hours": int(hours),
        "failed_attempts": len(failures),
        "peak_hours": peak_hours,
        "top_ips": [{"ip": key, "attempts": value} for key, value in sorted(ip_counts.items(), key=lambda item: item[1], reverse=True)[:10]],
        "top_usernames": [{"username": key, "attempts": value} for key, value in sorted(user_counts.items(), key=lambda item: item[1], reverse=True)[:10]],
        "anomalies": anomalies,
    }


def analyze_module_usage(hours: int = 24) -> dict[str, Any]:
    records = load_screen_usage(hours)
    if not records:
        return {
            "window_hours": int(hours),
            "total_views": 0,
            "top_modules": [],
            "top_modules_by_role": [],
            "top_screens": [],
        }
    if pd is not None:
        frame = pd.DataFrame(records)
        module_counts = frame.groupby("module_name").size().reset_index(name="views").sort_values("views", ascending=False)
        role_module_counts = (
            frame.groupby(["role", "module_name"])
            .size()
            .reset_index(name="views")
            .sort_values("views", ascending=False)
        )
        screen_counts = frame.groupby("screen_name").size().reset_index(name="views").sort_values("views", ascending=False)
        return {
            "window_hours": int(hours),
            "total_views": int(len(frame.index)),
            "top_modules": module_counts.head(10).to_dict(orient="records"),
            "top_modules_by_role": role_module_counts.head(20).to_dict(orient="records"),
            "top_screens": screen_counts.head(20).to_dict(orient="records"),
        }
    module_counts: dict[str, int] = {}
    role_module_counts: dict[tuple[str, str], int] = {}
    screen_counts: dict[str, int] = {}
    for item in records:
        module_name = str(item.get("module_name") or "")
        role_name = str(item.get("role") or "")
        screen_name = str(item.get("screen_name") or "")
        module_counts[module_name] = module_counts.get(module_name, 0) + 1
        role_module_counts[(role_name, module_name)] = role_module_counts.get((role_name, module_name), 0) + 1
        screen_counts[screen_name] = screen_counts.get(screen_name, 0) + 1
    return {
        "window_hours": int(hours),
        "total_views": len(records),
        "top_modules": [{"module_name": key, "views": value} for key, value in sorted(module_counts.items(), key=lambda item: item[1], reverse=True)[:10]],
        "top_modules_by_role": [{"role": key[0], "module_name": key[1], "views": value} for key, value in sorted(role_module_counts.items(), key=lambda item: item[1], reverse=True)[:20]],
        "top_screens": [{"screen_name": key, "views": value} for key, value in sorted(screen_counts.items(), key=lambda item: item[1], reverse=True)[:20]],
    }


def build_backend_analytics(hours: int = 24) -> dict[str, Any]:
    access_records = load_access_history(hours)
    failed = sum(1 for item in access_records if not item.get("success"))
    successful = sum(1 for item in access_records if item.get("success"))
    total = failed + successful
    return {
        "window_hours": int(hours),
        "successful_logins": successful,
        "failed_logins": failed,
        "failure_rate": round((failed / total) * 100, 2) if total else 0.0,
        "failed_login_patterns": analyze_failed_login_patterns(hours),
        "module_usage": analyze_module_usage(hours),
    }


def export_access_history_excel(hours: int = 24, output_path: str = "") -> str:
    if Workbook is None:
        return ""
    resolved_path = output_path or os.path.join(
        tempfile.gettempdir(),
        f"web_access_history_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Historial"
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78") if PatternFill is not None else None
    header_font = Font(color="FFFFFF", bold=True) if Font is not None else None

    history = load_access_history(hours)
    headers = ["created_at", "tenant_id", "username", "ip", "user_agent", "success"]
    sheet.append(headers)
    for record in history:
        sheet.append([record.get(header, "") for header in headers])
    for cell in sheet[1]:
        if header_fill is not None:
            cell.fill = header_fill
        if header_font is not None:
            cell.font = header_font
    sheet.auto_filter.ref = sheet.dimensions
    if Table is not None and len(history) >= 1:
        table = Table(displayName="AccessHistory", ref=sheet.dimensions)
        if TableStyleInfo is not None:
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        sheet.add_table(table)

    peaks_sheet = workbook.create_sheet("Analitica")
    peaks_headers = ["Seccion", "Campo", "Valor"]
    peaks_sheet.append(peaks_headers)
    analytics = build_backend_analytics(hours)
    rows = [
        ("Resumen", "successful_logins", analytics.get("successful_logins", 0)),
        ("Resumen", "failed_logins", analytics.get("failed_logins", 0)),
        ("Resumen", "failure_rate", analytics.get("failure_rate", 0.0)),
    ]
    for item in analytics.get("failed_login_patterns", {}).get("peak_hours", []):
        rows.append(("Picos", str(item.get("hour_bucket") or ""), int(item.get("attempts") or 0)))
    for item in analytics.get("module_usage", {}).get("top_modules", []):
        rows.append(("Modulos", str(item.get("module_name") or ""), int(item.get("views") or 0)))
    for row in rows:
        peaks_sheet.append(list(row))
    for cell in peaks_sheet[1]:
        if header_fill is not None:
            cell.fill = header_fill
        if header_font is not None:
            cell.font = header_font
    peaks_sheet.auto_filter.ref = peaks_sheet.dimensions

    workbook.save(resolved_path)
    return resolved_path
