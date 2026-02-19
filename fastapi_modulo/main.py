
import json
import base64
import hashlib
import hmac
import smtplib
import csv
import sqlite3
from io import BytesIO, StringIO
from datetime import date, datetime, timedelta
from html import escape
import os
import re
import secrets
import time
import struct
import ipaddress
from textwrap import dedent
from email.message import EmailMessage
from urllib.parse import quote, urlparse
from cryptography.fernet import Fernet, InvalidToken
from cryptography.exceptions import InvalidSignature
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.primitives.asymmetric import ec, padding, rsa

from typing import Any, Dict, List, Optional, Set
from fastapi import FastAPI, Request, Body, HTTPException, UploadFile, File, Form, Depends
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, Response
import httpx
from openpyxl import Workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError, create_model
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.exceptions import HTTPException as StarletteHTTPException
from sqlalchemy import create_engine, Column, String, Integer, JSON, ForeignKey, Boolean, Date, DateTime, UniqueConstraint, func
from sqlalchemy.orm import declarative_base, sessionmaker, relationship

HIDDEN_SYSTEM_USERS = {"0konomiyaki"}
IDENTIDAD_LOGIN_CONFIG_PATH = "fastapi_modulo/identidad_login.json"
IDENTIDAD_LOGIN_IMAGE_DIR = "fastapi_modulo/templates/imagenes"
DOCUMENTS_UPLOAD_DIR = "fastapi_modulo/uploads/documentos"
DEFAULT_LOGIN_IDENTITY = {
    "favicon_filename": "icon.png",
    "logo_filename": "icon.png",
    "desktop_bg_filename": "fondo.jpg",
    "mobile_bg_filename": "movil.jpg",
    "company_short_name": "AVAN",
    "login_message": "Incrementando el nivel de eficiencia",
}
PLANTILLAS_STORE_PATH = "fastapi_modulo/plantillas_store.json"
SYSTEM_REPORT_HEADER_TEMPLATE_ID = "system-report-header"
AUTH_COOKIE_NAME = "auth_session"
DATOS_PRELIMINARES_STORE_PATH = "fastapi_modulo/datos_preliminares_store.json"
REGIONES_STORE_PATH = "fastapi_modulo/regiones_store.json"
SUCURSALES_STORE_PATH = "fastapi_modulo/sucursales_store.json"
DEFAULT_DATOS_GENERALES = {
    "responsable_general": "",
    "primer_anio_proyeccion": "",
    "anios_proyeccion": "3",
    "sociedad": "",
    "figura_juridica": "",
    "calle": "",
    "numero_exterior": "",
    "numero_interior": "",
    "colonia": "",
    "ciudad": "",
    "municipio": "",
    "estado": "",
    "cp": "",
    "pais": "",
    "ifb_activos_m3": "",
    "ifb_activos_m2": "",
    "ifb_activos_m1": "",
    "ifb_pasivos_m3": "",
    "ifb_pasivos_m2": "",
    "ifb_pasivos_m1": "",
    "ifb_capital_m3": "",
    "ifb_capital_m2": "",
    "ifb_capital_m1": "",
    "ifb_ingresos_m3": "",
    "ifb_ingresos_m2": "",
    "ifb_ingresos_m1": "",
    "ifb_egresos_m3": "",
    "ifb_egresos_m2": "",
    "ifb_egresos_m1": "",
    "ifb_resultado_m3": "",
    "ifb_resultado_m2": "",
    "ifb_resultado_m1": "",
}
SIPET_PREMIUM_UI_TEMPLATE_CSS = dedent("""
    .sipet-ui-template .table-excel {
        width: 100% !important;
        border-collapse: collapse !important;
        border-spacing: 0 !important;
        background: transparent !important;
    }
    .sipet-ui-template .table-excel thead th {
        text-align: left !important;
        font-size: 13px !important;
        letter-spacing: .08em;
        text-transform: uppercase;
        color: rgba(15,23,42,.75) !important;
        background: linear-gradient(180deg, rgba(255,255,255,.92), rgba(255,255,255,.74)) !important;
        border-bottom: 1px solid rgba(15,23,42,.10) !important;
        border-right: 1px solid rgba(15,23,42,.10) !important;
        padding: 14px 16px !important;
    }
    .sipet-ui-template .table-excel thead th:last-child {
        border-right: 0 !important;
    }
    .sipet-ui-template .table-excel tbody td {
        border-bottom: 1px solid rgba(15,23,42,.08) !important;
        border-right: 1px solid rgba(15,23,42,.10) !important;
        background: #ffffff !important;
        padding: 12px !important;
        vertical-align: middle !important;
    }
    .sipet-ui-template .table-excel tbody td:last-child {
        border-right: 0 !important;
    }
    .sipet-ui-template .table-excel tbody tr:nth-child(even) td {
        background: #ecfdf3 !important;
    }
    .sipet-ui-template .table-excel tbody tr:hover td {
        background: #dcfce7 !important;
    }
    .sipet-ui-template .table-excel tbody tr:last-child td {
        border-bottom: 0 !important;
    }
    .sipet-ui-template .table-excel--compact tbody td {
        padding: 8px !important;
    }
    .sipet-ui-template .table-excel--compact tbody td.year {
        font-size: 16px !important;
        font-weight: 750 !important;
        letter-spacing: 0 !important;
        padding: 8px 12px !important;
    }
    .sipet-ui-template .table-excel--compact .macro-input {
        min-height: 34px !important;
        padding: 8px 10px !important;
    }
    .sipet-ui-template .table-excel .table-input {
        width: 100%;
        height: 36px;
        border: 1px solid #cbd5e1;
        border-radius: 8px;
        padding: 0 10px;
        background: #ffffff;
        color: #0f172a;
    }
    .sipet-ui-template .table-excel .table-input.num {
        text-align: right;
        font-variant-numeric: tabular-nums;
    }
    .sipet-ui-template .table-excel .table-actions-cell,
    .sipet-ui-template .table-excel .table-actions-head {
        text-align: center;
    }
    .sipet-ui-template .table-excel .table-add-btn {
        width: 34px;
        height: 34px;
        border-radius: 8px;
        border: 1px solid #0f172a;
        background: #ffffff;
        color: #0f172a;
        font-size: 1.1rem;
        font-weight: 700;
        cursor: pointer;
    }
    .sipet-ui-template .table-excel .table-delete-btn {
        width: 34px;
        height: 34px;
        border-radius: 8px;
        border: 1px solid #991b1b;
        background: #ffffff;
        color: #991b1b;
        font-size: 1rem;
        font-weight: 700;
        cursor: pointer;
    }
    .sipet-ui-template .table-primary-btn {
        height: 36px;
        padding: 0 12px;
        border-radius: 8px;
        border: 1px solid #0f172a;
        background: #0f172a;
        color: #ffffff;
        font-size: 0.95rem;
        font-weight: 600;
        cursor: pointer;
    }
    .sipet-ui-template .table-excel .ifb-row-label {
        font-weight: 600;
        color: #0f172a;
    }
    .sipet-ui-template .table-excel .ifb-validation-cell {
        padding: 10px;
    }
    .sipet-ui-template .table-excel .ifb-validation-output {
        font-size: 0.85rem;
        line-height: 1.35;
    }
    .sipet-ui-template .dg-grid > label {
        border: 1px solid rgba(15, 23, 42, .10);
        border-radius: 14px;
        padding: 10px;
    }
    .sipet-ui-template .dg-grid > label:nth-child(odd) {
        background: #ffffff;
    }
    .sipet-ui-template .dg-grid > label:nth-child(even) {
        background: #ecfdf3;
    }
""")


def _require_secret(name: str) -> str:
    value = (os.environ.get(name) or "").strip()
    if not value:
        raise RuntimeError(
            f"{name} no está configurada. Define esta variable de entorno antes de iniciar la aplicación."
        )
    return value


AUTH_COOKIE_SECRET = _require_secret("AUTH_COOKIE_SECRET")
SENSITIVE_DATA_SECRET = (os.environ.get("SENSITIVE_DATA_SECRET") or AUTH_COOKIE_SECRET).strip()
PASSKEY_COOKIE_REGISTER = "passkey_register"
PASSKEY_COOKIE_AUTH = "passkey_auth"
PASSKEY_COOKIE_MFA_GATE = "passkey_mfa_gate"
PASSKEY_CHALLENGE_TTL_SECONDS = 300
NON_DATA_FIELD_TYPES = {"header", "paragraph", "html", "divider", "pagebreak"}
GEOIP_CACHE_TTL_SECONDS = 60 * 60 * 6
_GEOIP_CACHE: Dict[str, Dict[str, Any]] = {}
LOGIN_RATE_LIMIT_WINDOW_SECONDS = int((os.environ.get("LOGIN_RATE_LIMIT_WINDOW_SECONDS") or "300").strip() or "300")
LOGIN_RATE_LIMIT_MAX_ATTEMPTS = int((os.environ.get("LOGIN_RATE_LIMIT_MAX_ATTEMPTS") or "7").strip() or "7")
_LOGIN_ATTEMPTS: Dict[str, List[float]] = {}
IDENTITY_UPLOAD_MAX_BYTES = int((os.environ.get("IDENTITY_UPLOAD_MAX_BYTES") or str(5 * 1024 * 1024)).strip() or str(5 * 1024 * 1024))
TOTP_PERIOD_SECONDS = int((os.environ.get("TOTP_PERIOD_SECONDS") or "30").strip() or "30")
TOTP_ALLOWED_DRIFT_STEPS = int((os.environ.get("TOTP_ALLOWED_DRIFT_STEPS") or "1").strip() or "1")


ROLE_ALIASES = {
    "super_admin": "superadministrador",
    "superadministrador": "superadministrador",
    "admin": "administrador",
    "administrador": "administrador",
    "autoridad": "autoridades",
    "autoridades": "autoridades",
    "authority": "autoridades",
    "authorities": "autoridades",
    "strategic_manager": "departamento",
    "department_manager": "departamento",
    "departamento": "departamento",
    "team_leader": "usuario",
    "collaborator": "usuario",
    "viewer": "usuario",
    "usuario": "usuario",
}


def normalize_role_name(role_name: Optional[str]) -> str:
    normalized = (role_name or "").strip().lower()
    if not normalized:
        return "usuario"
    return ROLE_ALIASES.get(normalized, normalized)


def get_current_role(request: Request) -> str:
    role = getattr(request.state, "user_role", None)
    if role is None:
        role = request.cookies.get("user_role") or os.environ.get("DEFAULT_USER_ROLE") or ""
    return normalize_role_name(role)


def _normalize_tenant_id(value: Optional[str]) -> str:
    raw = (value or "").strip().lower()
    normalized = re.sub(r"[^a-z0-9._-]+", "-", raw).strip("-._")
    return normalized or "default"


def get_current_tenant(request: Request) -> str:
    tenant = getattr(request.state, "tenant_id", None)
    if tenant:
        return _normalize_tenant_id(tenant)
    cookie_tenant = request.cookies.get("tenant_id")
    if cookie_tenant:
        return _normalize_tenant_id(cookie_tenant)
    header_tenant = request.headers.get("x-tenant-id")
    if header_tenant and is_superadmin(request):
        return _normalize_tenant_id(header_tenant)
    return _normalize_tenant_id(os.environ.get("DEFAULT_TENANT_ID", "default"))


def is_superadmin(request: Request) -> bool:
    return get_current_role(request) == "superadministrador"


def is_admin(request: Request) -> bool:
    return get_current_role(request) == "administrador"


def is_admin_or_superadmin(request: Request) -> bool:
    return is_superadmin(request) or is_admin(request)


def require_superadmin(request: Request) -> None:
    if not is_superadmin(request):
        raise HTTPException(status_code=403, detail="Acceso restringido a superadministrador")


def require_admin_or_superadmin(request: Request) -> None:
    if not is_admin_or_superadmin(request):
        raise HTTPException(status_code=403, detail="Acceso restringido a administradores")


def can_assign_role(request: Request, role_name: str) -> bool:
    normalized = (role_name or "").strip().lower()
    if not normalized:
        return True
    if is_superadmin(request):
        return True
    if is_admin(request):
        return normalized != "superadministrador"
    return False


def get_visible_role_names(request: Request) -> List[str]:
    if is_superadmin(request):
        return [name for name, _ in DEFAULT_SYSTEM_ROLES]
    if is_admin(request):
        return [name for name, _ in DEFAULT_SYSTEM_ROLES if name != "superadministrador"]
    return []


def _sensitive_secret_bytes() -> bytes:
    return hashlib.sha256(SENSITIVE_DATA_SECRET.encode("utf-8")).digest()


def _sensitive_fernet() -> Fernet:
    key = base64.urlsafe_b64encode(_sensitive_secret_bytes())
    return Fernet(key)


def _sensitive_lookup_hash(value: str) -> str:
    normalized = (value or "").strip().lower()
    return hmac.new(_sensitive_secret_bytes(), normalized.encode("utf-8"), hashlib.sha256).hexdigest()


def _auth_client_key(request: Request) -> str:
    forwarded_for = (request.headers.get("x-forwarded-for") or "").split(",")[0].strip()
    if forwarded_for:
        return forwarded_for
    if request.client and request.client.host:
        return request.client.host
    return "unknown"


def _is_login_rate_limited(request: Request) -> bool:
    key = _auth_client_key(request)
    now = time.time()
    window_start = now - max(1, LOGIN_RATE_LIMIT_WINDOW_SECONDS)
    attempts = [ts for ts in _LOGIN_ATTEMPTS.get(key, []) if ts >= window_start]
    _LOGIN_ATTEMPTS[key] = attempts
    return len(attempts) >= max(1, LOGIN_RATE_LIMIT_MAX_ATTEMPTS)


def _register_failed_login_attempt(request: Request) -> None:
    key = _auth_client_key(request)
    now = time.time()
    window_start = now - max(1, LOGIN_RATE_LIMIT_WINDOW_SECONDS)
    attempts = [ts for ts in _LOGIN_ATTEMPTS.get(key, []) if ts >= window_start]
    attempts.append(now)
    _LOGIN_ATTEMPTS[key] = attempts


def _clear_failed_login_attempts(request: Request) -> None:
    _LOGIN_ATTEMPTS.pop(_auth_client_key(request), None)


def _is_same_origin_request(request: Request) -> bool:
    host = (request.headers.get("host") or "").strip().lower()
    forwarded_host = (request.headers.get("x-forwarded-host") or "").split(",")[0].strip().lower()
    effective_host = forwarded_host or host
    if not effective_host:
        return False
    forwarded_proto = (request.headers.get("x-forwarded-proto") or "").split(",")[0].strip().lower()
    effective_scheme = forwarded_proto or request.url.scheme
    current_origin = f"{effective_scheme}://{effective_host}"

    origin = (request.headers.get("origin") or "").strip().rstrip("/")
    if origin:
        origin_normalized = origin.lower()
        if origin_normalized == current_origin:
            return True
        try:
            parsed_origin = urlparse(origin_normalized)
            parsed_current = urlparse(current_origin)
            # Detrás de proxy puede diferir el esquema interno/externo.
            return parsed_origin.netloc == parsed_current.netloc and parsed_origin.scheme in {"http", "https"}
        except Exception:
            return False

    referer = (request.headers.get("referer") or "").strip()
    if referer:
        parsed = urlparse(referer)
        referer_origin = f"{parsed.scheme}://{parsed.netloc}".rstrip("/").lower()
        if referer_origin == current_origin:
            return True
        parsed_current = urlparse(current_origin)
        return parsed.netloc.lower() == parsed_current.netloc and parsed.scheme in {"http", "https"}

    return False


def _encrypt_sensitive(value: Optional[str]) -> Optional[str]:
    raw = (value or "").strip()
    if not raw:
        return value
    if raw.startswith("enc$"):
        return raw
    token = _sensitive_fernet().encrypt(raw.encode("utf-8")).decode("utf-8")
    return f"enc${token}"


def _decrypt_sensitive(value: Optional[str]) -> str:
    raw = (value or "").strip()
    if not raw:
        return ""
    if not raw.startswith("enc$"):
        return raw
    token = raw[4:]
    try:
        return _sensitive_fernet().decrypt(token.encode("utf-8")).decode("utf-8")
    except (InvalidToken, ValueError):
        return ""


def _ensure_login_identity_paths() -> None:
    os.makedirs(IDENTIDAD_LOGIN_IMAGE_DIR, exist_ok=True)


def _load_login_identity() -> Dict[str, str]:
    data = DEFAULT_LOGIN_IDENTITY.copy()
    if os.path.exists(IDENTIDAD_LOGIN_CONFIG_PATH):
        try:
            with open(IDENTIDAD_LOGIN_CONFIG_PATH, "r", encoding="utf-8") as fh:
                loaded = json.load(fh)
            if isinstance(loaded, dict):
                data.update({k: v for k, v in loaded.items() if isinstance(v, str)})
        except (OSError, json.JSONDecodeError):
            pass
    return data


def _save_login_identity(data: Dict[str, str]) -> None:
    with open(IDENTIDAD_LOGIN_CONFIG_PATH, "w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=2)


def _load_datos_preliminares_store() -> Dict[str, str]:
    data = dict(DEFAULT_DATOS_GENERALES)
    if not os.path.exists(DATOS_PRELIMINARES_STORE_PATH):
        return data
    try:
        with open(DATOS_PRELIMINARES_STORE_PATH, "r", encoding="utf-8") as fh:
            loaded = json.load(fh)
        if isinstance(loaded, dict):
            for key in data.keys():
                if key in loaded and loaded[key] is not None:
                    data[key] = str(loaded[key]).strip()
    except (OSError, json.JSONDecodeError):
        pass
    return data


def _save_datos_preliminares_store(data: Dict[str, str]) -> None:
    safe_payload: Dict[str, str] = {}
    for key, default_value in DEFAULT_DATOS_GENERALES.items():
        safe_payload[key] = str(data.get(key, default_value) or "").strip()
    with open(DATOS_PRELIMINARES_STORE_PATH, "w", encoding="utf-8") as fh:
        json.dump(safe_payload, fh, ensure_ascii=False, indent=2)


def _load_regiones_store() -> List[Dict[str, str]]:
    if not os.path.exists(REGIONES_STORE_PATH):
        return []
    try:
        with open(REGIONES_STORE_PATH, "r", encoding="utf-8") as fh:
            loaded = json.load(fh)
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(loaded, list):
        return []
    rows: List[Dict[str, str]] = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        codigo = str(item.get("codigo") or "").strip()
        descripcion = str(item.get("descripcion") or "").strip()
        if not nombre and not codigo and not descripcion:
            continue
        rows.append(
            {
                "nombre": nombre,
                "codigo": codigo,
                "descripcion": descripcion,
            }
        )
    return rows


def _save_regiones_store(rows: List[Dict[str, str]]) -> None:
    safe_rows: List[Dict[str, str]] = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        codigo = str(item.get("codigo") or "").strip()
        descripcion = str(item.get("descripcion") or "").strip()
        if not nombre and not codigo and not descripcion:
            continue
        safe_rows.append(
            {
                "nombre": nombre,
                "codigo": codigo,
                "descripcion": descripcion,
            }
        )
    with open(REGIONES_STORE_PATH, "w", encoding="utf-8") as fh:
        json.dump(safe_rows, fh, ensure_ascii=False, indent=2)


def _load_sucursales_store() -> List[Dict[str, str]]:
    if not os.path.exists(SUCURSALES_STORE_PATH):
        return []
    try:
        with open(SUCURSALES_STORE_PATH, "r", encoding="utf-8") as fh:
            loaded = json.load(fh)
    except (OSError, json.JSONDecodeError):
        return []
    if not isinstance(loaded, list):
        return []
    rows: List[Dict[str, str]] = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        region = str(item.get("region") or "").strip()
        codigo = str(item.get("codigo") or "").strip()
        descripcion = str(item.get("descripcion") or "").strip()
        if not nombre and not region and not codigo and not descripcion:
            continue
        rows.append(
            {
                "nombre": nombre,
                "region": region,
                "codigo": codigo,
                "descripcion": descripcion,
            }
        )
    return rows


def _save_sucursales_store(rows: List[Dict[str, str]]) -> None:
    safe_rows: List[Dict[str, str]] = []
    for item in rows:
        if not isinstance(item, dict):
            continue
        nombre = str(item.get("nombre") or "").strip()
        region = str(item.get("region") or "").strip()
        codigo = str(item.get("codigo") or "").strip()
        descripcion = str(item.get("descripcion") or "").strip()
        if not nombre and not region and not codigo and not descripcion:
            continue
        safe_rows.append(
            {
                "nombre": nombre,
                "region": region,
                "codigo": codigo,
                "descripcion": descripcion,
            }
        )
    with open(SUCURSALES_STORE_PATH, "w", encoding="utf-8") as fh:
        json.dump(safe_rows, fh, ensure_ascii=False, indent=2)


def _get_upload_ext(upload: UploadFile) -> str:
    filename = (upload.filename or "").lower()
    ext = os.path.splitext(filename)[1]
    if ext in {".png", ".jpg", ".jpeg", ".webp", ".svg"}:
        return ext
    content_type = (upload.content_type or "").lower()
    if "svg" in content_type:
        return ".svg"
    if "webp" in content_type:
        return ".webp"
    if "jpeg" in content_type or "jpg" in content_type:
        return ".jpg"
    return ".png"


def _remove_login_image_if_custom(filename: Optional[str]) -> None:
    if not filename or filename in {
        DEFAULT_LOGIN_IDENTITY["favicon_filename"],
        DEFAULT_LOGIN_IDENTITY["logo_filename"],
        DEFAULT_LOGIN_IDENTITY["desktop_bg_filename"],
        DEFAULT_LOGIN_IDENTITY["mobile_bg_filename"],
    }:
        return
    path = os.path.join(IDENTIDAD_LOGIN_IMAGE_DIR, filename)
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError:
        pass


async def _store_login_image(upload: UploadFile, prefix: str) -> Optional[str]:
    if not upload or not upload.filename:
        return None
    content_type = (upload.content_type or "").lower().strip()
    filename = (upload.filename or "").lower()
    ext = os.path.splitext(filename)[1]
    allowed_exts = {".png", ".jpg", ".jpeg", ".webp", ".svg"}
    # Algunos navegadores/proxys envían application/octet-stream para imágenes válidas.
    if content_type and not content_type.startswith("image/") and ext not in allowed_exts:
        raise HTTPException(status_code=400, detail="Solo se permiten imágenes para identidad institucional")
    data = await upload.read()
    if not data:
        return None
    if len(data) > max(1, IDENTITY_UPLOAD_MAX_BYTES):
        raise HTTPException(status_code=413, detail="La imagen supera el tamaño máximo permitido")
    _ensure_login_identity_paths()
    ext = _get_upload_ext(upload)
    new_filename = f"{prefix}_{secrets.token_hex(6)}{ext}"
    image_path = os.path.join(IDENTIDAD_LOGIN_IMAGE_DIR, new_filename)
    with open(image_path, "wb") as fh:
        fh.write(data)
    return new_filename


def _sanitize_document_name(value: str) -> str:
    normalized = re.sub(r"[^a-zA-Z0-9._-]+", "_", (value or "").strip())
    return normalized.strip("._") or "documento"


def _ensure_documents_dir() -> None:
    os.makedirs(DOCUMENTS_UPLOAD_DIR, exist_ok=True)


async def _store_evidence_file(upload: UploadFile) -> Dict[str, Any]:
    if not upload or not upload.filename:
        raise HTTPException(status_code=400, detail="Archivo requerido")

    data = await upload.read()
    if not data:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    if len(data) > 25 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="El archivo supera 25MB")

    _ensure_documents_dir()
    ext = os.path.splitext(upload.filename or "")[1].lower()
    safe_base = _sanitize_document_name(os.path.splitext(upload.filename or "documento")[0])
    final_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{safe_base}_{secrets.token_hex(4)}{ext}"
    final_path = os.path.join(DOCUMENTS_UPLOAD_DIR, final_name)
    with open(final_path, "wb") as f:
        f.write(data)

    return {
        "filename": upload.filename,
        "path": final_path,
        "mime": (upload.content_type or "").strip() or "application/octet-stream",
        "size": len(data),
    }


def _delete_evidence_file(path: Optional[str]) -> None:
    if not path:
        return
    safe_root = os.path.abspath(DOCUMENTS_UPLOAD_DIR)
    target = os.path.abspath(path)
    if not target.startswith(safe_root):
        return
    if os.path.exists(target):
        os.remove(target)


def _build_login_asset_url(filename: Optional[str], default_filename: str) -> str:
    selected = filename or default_filename
    selected_path = os.path.join(IDENTIDAD_LOGIN_IMAGE_DIR, selected)
    if not os.path.exists(selected_path):
        selected = default_filename
        selected_path = os.path.join(IDENTIDAD_LOGIN_IMAGE_DIR, selected)
    version = int(os.path.getmtime(selected_path)) if os.path.exists(selected_path) else 0
    return f"/templates/imagenes/{selected}?v={version}"


def _get_login_identity_context() -> Dict[str, str]:
    data = _load_login_identity()
    return {
        "login_favicon_url": _build_login_asset_url(
            data.get("favicon_filename"),
            DEFAULT_LOGIN_IDENTITY["favicon_filename"],
        ),
        "login_logo_url": _build_login_asset_url(
            data.get("logo_filename"),
            DEFAULT_LOGIN_IDENTITY["logo_filename"],
        ),
        "login_bg_desktop_url": _build_login_asset_url(
            data.get("desktop_bg_filename"),
            DEFAULT_LOGIN_IDENTITY["desktop_bg_filename"],
        ),
        "login_bg_mobile_url": _build_login_asset_url(
            data.get("mobile_bg_filename"),
            DEFAULT_LOGIN_IDENTITY["mobile_bg_filename"],
        ),
        "login_company_short_name": data.get("company_short_name") or DEFAULT_LOGIN_IDENTITY["company_short_name"],
        "login_message": data.get("login_message") or DEFAULT_LOGIN_IDENTITY["login_message"],
    }


def _load_plantillas_store() -> List[Dict[str, str]]:
    if not os.path.exists(PLANTILLAS_STORE_PATH):
        return _with_system_templates([])
    try:
        with open(PLANTILLAS_STORE_PATH, "r", encoding="utf-8") as fh:
            loaded = json.load(fh)
    except (OSError, json.JSONDecodeError):
        return _with_system_templates([])
    if not isinstance(loaded, list):
        return _with_system_templates([])
    templates = []
    for item in loaded:
        if not isinstance(item, dict):
            continue
        templates.append(
            {
                "id": str(item.get("id") or "").strip(),
                "nombre": str(item.get("nombre") or "").strip(),
                "html": str(item.get("html") or ""),
                "css": str(item.get("css") or ""),
                "created_at": str(item.get("created_at") or ""),
                "updated_at": str(item.get("updated_at") or ""),
            }
        )
    return _with_system_templates([tpl for tpl in templates if tpl["id"] and tpl["nombre"]])


def _build_default_report_header_template() -> Dict[str, str]:
    now_iso = datetime.utcnow().isoformat()
    return {
        "id": SYSTEM_REPORT_HEADER_TEMPLATE_ID,
        "nombre": "Encabezado",
        "html": (
            "<header class='reporte-encabezado'>"
            "<div class='reporte-encabezado__marca'>{{ empresa }}</div>"
            "<div class='reporte-encabezado__meta'>"
            "<h1>{{ titulo_reporte }}</h1>"
            "<p>{{ subtitulo_reporte }}</p>"
            "</div>"
            "<div class='reporte-encabezado__fecha'>Fecha: {{ fecha_reporte }}</div>"
            "</header>"
        ),
        "css": (
            ".reporte-encabezado { display:flex; align-items:center; justify-content:space-between; gap:16px; "
            "padding:16px 20px; border:1px solid #cbd5e1; border-radius:12px; background:#f8fafc; font-family:Arial,sans-serif; } "
            ".reporte-encabezado__marca { font-weight:800; color:#0f172a; font-size:1.1rem; letter-spacing:.04em; } "
            ".reporte-encabezado__meta h1 { margin:0; font-size:1.05rem; color:#0f172a; } "
            ".reporte-encabezado__meta p { margin:4px 0 0; color:#475569; font-size:.88rem; } "
            ".reporte-encabezado__fecha { color:#334155; font-size:.84rem; white-space:nowrap; }"
        ),
        "created_at": now_iso,
        "updated_at": now_iso,
    }


def _with_system_templates(templates: List[Dict[str, str]]) -> List[Dict[str, str]]:
    default_header = _build_default_report_header_template()
    has_header = any(
        str(tpl.get("id", "")).strip() == SYSTEM_REPORT_HEADER_TEMPLATE_ID
        or str(tpl.get("nombre", "")).strip().lower() == "encabezado"
        for tpl in templates
    )
    if has_header:
        return templates
    return [default_header, *templates]


def _get_report_header_template() -> Dict[str, str]:
    templates = _load_plantillas_store()
    for tpl in templates:
        if str(tpl.get("id", "")).strip() == SYSTEM_REPORT_HEADER_TEMPLATE_ID:
            return tpl
    for tpl in templates:
        if str(tpl.get("nombre", "")).strip().lower() == "encabezado":
            return tpl
    return _build_default_report_header_template()


def _apply_template_context(content: str, context: Dict[str, str]) -> str:
    rendered = content or ""
    for key, value in context.items():
        rendered = rendered.replace(f"{{{{ {key} }}}}", value)
        rendered = rendered.replace(f"{{{{{key}}}}}", value)
    return rendered


def _build_report_export_context() -> Dict[str, str]:
    identidad = _load_login_identity()
    return {
        "empresa": identidad.get("company_short_name") or "SIPET",
        "titulo_reporte": "Reporte consolidado",
        "subtitulo_reporte": "Avance, desempeno y seguimiento",
        "fecha_reporte": datetime.now().strftime("%Y-%m-%d %H:%M"),
    }


def _build_report_export_rows() -> List[Dict[str, str]]:
    return [
        {
            "reporte": "Reporte ejecutivo",
            "descripcion": "Resumen de estado estrategico",
            "formato": "PDF / Excel",
        },
        {
            "reporte": "Reporte operativo",
            "descripcion": "Actividades, avances y cumplimiento",
            "formato": "PDF / Excel",
        },
        {
            "reporte": "Reporte KPI",
            "descripcion": "Indicadores, metas y variaciones",
            "formato": "PDF / Excel",
        },
    ]


def _build_report_export_html_document() -> str:
    template = _get_report_header_template()
    context = _build_report_export_context()
    header_html = _apply_template_context(template.get("html", ""), context)
    header_css = template.get("css", "")
    rows = _build_report_export_rows()
    rows_html = "".join(
        (
            "<tr>"
            f"<td>{escape(row['reporte'])}</td>"
            f"<td>{escape(row['descripcion'])}</td>"
            f"<td>{escape(row['formato'])}</td>"
            "</tr>"
        )
        for row in rows
    )
    return (
        "<!doctype html><html lang='es'><head><meta charset='utf-8'>"
        "<title>Reporte consolidado</title>"
        "<style>"
        f"{header_css}"
        "body{font-family:Arial,sans-serif;background:#fff;color:#0f172a;padding:24px;}"
        ".reporte-bloque{margin-top:18px;}"
        "table{width:100%;border-collapse:collapse;}"
        "th,td{border:1px solid #cbd5e1;padding:10px;text-align:left;font-size:14px;}"
        "th{background:#f1f5f9;}"
        "</style></head><body>"
        f"{header_html}"
        "<section class='reporte-bloque'>"
        "<h2>Detalle de reportes</h2>"
        "<table><thead><tr><th>Reporte</th><th>Descripcion</th><th>Formato</th></tr></thead>"
        f"<tbody>{rows_html}</tbody></table>"
        "</section></body></html>"
    )


def _save_plantillas_store(templates: List[Dict[str, str]]) -> None:
    os.makedirs(os.path.dirname(PLANTILLAS_STORE_PATH), exist_ok=True)
    with open(PLANTILLAS_STORE_PATH, "w", encoding="utf-8") as fh:
        json.dump(templates, fh, ensure_ascii=False, indent=2)


def build_view_buttons_html(view_buttons: Optional[List[Dict]]) -> str:
    if not view_buttons:
        return ""
    pieces = []
    for button in view_buttons:
        label = button.get("label", "").strip()
        if not label:
            continue
        icon = button.get("icon")
        view = button.get("view")
        url = button.get("url")
        classes = "view-pill"
        if button.get("active"):
            classes += " active"
        attrs = []
        if view:
            attrs.append(f'data-view="{view}"')
        if url:
            attrs.append(f'data-url="{url}"')
        attr_str = f' {" ".join(attrs)}' if attrs else ""
        icon_html = ""
        if icon:
            icon_html = (
                f'<span class="view-pill-icon-mask" aria-hidden="true" '
                f'style="--view-pill-icon-url:url(\'{icon}\')"></span>'
            )
        pieces.append(f'<button class="{classes}" type="button"{attr_str}>{icon_html}<span>{label}</span></button>')
    return "".join(pieces)


def backend_screen(
    request: Request,
    title: str,
    subtitle: Optional[str] = None,
    description: Optional[str] = None,
    content: str = "",
    view_buttons: Optional[List[Dict]] = None,
    view_buttons_html: str = "",
    floating_buttons: Optional[List[Dict]] = None,
    hide_floating_actions: bool = False,
    show_page_header: bool = True,
    page_title: Optional[str] = None,
    page_description: Optional[str] = None,
):
    """
    Helper para renderizar una pantalla backend con panel flotante y botones de vistas.
    - view_buttons: lista de dicts {label, view?, url, icon}
    - floating_buttons: lista de dicts {label, onclick}
    """
    rendered_view_buttons = view_buttons_html or build_view_buttons_html(view_buttons)
    can_manage_personalization = is_superadmin(request)
    login_identity = _get_login_identity_context()
    context = {
        "request": request,
        "title": title,
        "subtitle": subtitle,
        "page_title": page_title or title,
        "page_description": page_description or description,
        "content": content,
        "view_buttons_html": rendered_view_buttons,
        "floating_buttons": floating_buttons,
        "hide_floating_actions": hide_floating_actions,
        "show_page_header": show_page_header,
        "colores": get_colores_context(),
        "can_manage_personalization": can_manage_personalization,
        "app_favicon_url": login_identity.get("login_favicon_url"),
    }
    return templates.TemplateResponse("base.html", context)

# Configuración de BD por entorno (Railway/Local)
def _resolve_database_url() -> str:
    raw_url = (os.environ.get("DATABASE_URL") or "").strip()
    if raw_url:
        if raw_url.startswith("postgres://"):
            return raw_url.replace("postgres://", "postgresql://", 1)
        return raw_url
    sqlite_db_path = (os.environ.get("SQLITE_DB_PATH") or "strategic_planning.db").strip()
    return f"sqlite:///./{sqlite_db_path}"


def _extract_sqlite_path(db_url: str) -> Optional[str]:
    if not db_url.startswith("sqlite:///"):
        return None
    path = db_url.replace("sqlite:///", "", 1).split("?", 1)[0]
    if path.startswith("./"):
        return path[2:]
    return path


DATABASE_URL = _resolve_database_url()
IS_SQLITE_DATABASE = DATABASE_URL.startswith("sqlite:///")
PRIMARY_DB_PATH = _extract_sqlite_path(DATABASE_URL)
APP_ENV = (os.environ.get("APP_ENV") or os.environ.get("ENVIRONMENT") or "development").strip().lower()
SESSION_MAX_AGE_SECONDS = int((os.environ.get("SESSION_MAX_AGE_SECONDS") or "28800").strip() or "28800")
COOKIE_SECURE = (os.environ.get("COOKIE_SECURE") or "").strip().lower() in {"1", "true", "yes", "on"} or APP_ENV in {
    "production",
    "prod",
}
ALLOW_LEGACY_PLAINTEXT_PASSWORDS = (os.environ.get("ALLOW_LEGACY_PLAINTEXT_PASSWORDS") or "").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
ENABLE_API_DOCS = (os.environ.get("ENABLE_API_DOCS") or "").strip().lower() in {"1", "true", "yes", "on"} or APP_ENV not in {
    "production",
    "prod",
}
HEALTH_INCLUDE_DETAILS = (os.environ.get("HEALTH_INCLUDE_DETAILS") or "").strip().lower() in {"1", "true", "yes", "on"}
CSRF_PROTECTION_ENABLED = (os.environ.get("CSRF_PROTECTION_ENABLED") or "true").strip().lower() in {
    "1",
    "true",
    "yes",
    "on",
}
ENGINE_CONNECT_ARGS = {"check_same_thread": False} if IS_SQLITE_DATABASE else {}
engine = create_engine(DATABASE_URL, connect_args=ENGINE_CONNECT_ARGS)
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class Colores(Base):
    __tablename__ = "colores"
    id = Column(Integer, primary_key=True, index=True)
    key = Column(String, unique=True, index=True)
    value = Column(String)

# --- NUEVO: Modelos para roles y usuarios ---
class Rol(Base):
    __tablename__ = "roles"
    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(String, unique=True, index=True)
    descripcion = Column(String)

class Usuario(Base):
    __tablename__ = "users"
    id = Column(Integer, primary_key=True, index=True)
    nombre = Column("full_name", String)
    usuario = Column("username", String, unique=True, index=True)
    usuario_hash = Column(String, index=True)
    correo = Column("email", String, unique=True, index=True)
    correo_hash = Column(String, index=True)
    celular = Column(String)
    contrasena = Column("password", String)
    departamento = Column(String)
    puesto = Column(String)
    jefe = Column(String)
    coach = Column(String)
    rol_id = Column(Integer)
    imagen = Column(String)
    role = Column(String)
    is_active = Column(Boolean, default=True)
    webauthn_credential_id = Column(String, unique=True, index=True)
    webauthn_public_key = Column(String)
    webauthn_sign_count = Column(Integer, default=0)
    totp_secret = Column(String)
    totp_enabled = Column(Boolean, default=False)


class StrategicAxisConfig(Base):
    __tablename__ = "strategic_axes_config"

    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(String, nullable=False)
    codigo = Column(String, default="")
    lider_departamento = Column(String, default="")
    descripcion = Column(String, default="")
    orden = Column(Integer, default=0, index=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    objetivos = relationship(
        "StrategicObjectiveConfig",
        back_populates="eje",
        cascade="all, delete-orphan",
        order_by="StrategicObjectiveConfig.orden",
    )


class StrategicObjectiveConfig(Base):
    __tablename__ = "strategic_objectives_config"

    id = Column(Integer, primary_key=True, index=True)
    eje_id = Column(Integer, ForeignKey("strategic_axes_config.id"), nullable=False, index=True)
    codigo = Column(String, default="")
    nombre = Column(String, nullable=False)
    lider = Column(String, default="")
    fecha_inicial = Column(Date)
    fecha_final = Column(Date)
    descripcion = Column(String, default="")
    orden = Column(Integer, default=0, index=True)
    is_active = Column(Boolean, default=True)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)

    eje = relationship("StrategicAxisConfig", back_populates="objetivos")


class POAActivity(Base):
    __tablename__ = "poa_activities"

    id = Column(Integer, primary_key=True, index=True)
    objective_id = Column(Integer, ForeignKey("strategic_objectives_config.id"), nullable=False, index=True)
    nombre = Column(String, nullable=False)
    codigo = Column(String, default="")
    responsable = Column(String, nullable=False)
    entregable = Column(String, default="")
    fecha_inicial = Column(Date)
    fecha_final = Column(Date)
    descripcion = Column(String, default="")
    entrega_estado = Column(String, default="ninguna")
    entrega_solicitada_por = Column(String, default="")
    entrega_solicitada_at = Column(DateTime)
    entrega_aprobada_por = Column(String, default="")
    entrega_aprobada_at = Column(DateTime)
    created_by = Column(String, default="")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class POASubactivity(Base):
    __tablename__ = "poa_subactivities"

    id = Column(Integer, primary_key=True, index=True)
    activity_id = Column(Integer, ForeignKey("poa_activities.id"), nullable=False, index=True)
    nombre = Column(String, nullable=False)
    codigo = Column(String, default="")
    responsable = Column(String, nullable=False)
    entregable = Column(String, default="")
    fecha_inicial = Column(Date)
    fecha_final = Column(Date)
    descripcion = Column(String, default="")
    assigned_by = Column(String, default="")
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class POADeliverableApproval(Base):
    __tablename__ = "poa_deliverable_approvals"

    id = Column(Integer, primary_key=True, index=True)
    activity_id = Column(Integer, ForeignKey("poa_activities.id"), nullable=False, index=True)
    objective_id = Column(Integer, ForeignKey("strategic_objectives_config.id"), nullable=False, index=True)
    process_owner = Column(String, nullable=False)
    requester = Column(String, nullable=False)
    status = Column(String, default="pendiente")
    comment = Column(String, default="")
    resolved_by = Column(String, default="")
    resolved_at = Column(DateTime)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class FormDefinition(Base):
    __tablename__ = "form_definitions"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    slug = Column(String, unique=True, index=True, nullable=False)
    tenant_id = Column(String, index=True, default="default")
    description = Column(String)
    config = Column(JSON, default=dict)
    allowed_roles = Column(JSON, default=list)
    created_at = Column(DateTime, default=datetime.utcnow)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)
    is_active = Column(Boolean, default=True)

    fields = relationship("FormField", back_populates="form", cascade="all, delete-orphan")
    submissions = relationship("FormSubmission", back_populates="form", cascade="all, delete-orphan")


class FormField(Base):
    __tablename__ = "form_fields"

    id = Column(Integer, primary_key=True, index=True)
    form_id = Column(Integer, ForeignKey("form_definitions.id"), nullable=False, index=True)
    field_type = Column(String, nullable=False)
    label = Column(String, nullable=False)
    name = Column(String, nullable=False)
    placeholder = Column(String)
    help_text = Column(String)
    default_value = Column(String)
    is_required = Column(Boolean, default=False)
    validation_rules = Column(JSON, default=dict)
    options = Column(JSON, default=list)
    order = Column(Integer, default=0)
    conditional_logic = Column(JSON, default=dict)

    form = relationship("FormDefinition", back_populates="fields")


class FormSubmission(Base):
    __tablename__ = "form_submissions"

    id = Column(Integer, primary_key=True, index=True)
    form_id = Column(Integer, ForeignKey("form_definitions.id"), nullable=False, index=True)
    data = Column(JSON, default=dict)
    submitted_at = Column(DateTime, default=datetime.utcnow)
    ip_address = Column(String)
    user_agent = Column(String)

    form = relationship("FormDefinition", back_populates="submissions")


class DocumentoEvidencia(Base):
    __tablename__ = "documentos_evidencia"

    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(String, index=True, default="default")
    titulo = Column(String, nullable=False)
    descripcion = Column(String)
    proceso = Column(String, default="envio")
    estado = Column(String, default="borrador")
    version = Column(Integer, default=1)
    archivo_nombre = Column(String, nullable=False)
    archivo_ruta = Column(String, nullable=False)
    archivo_tipo = Column(String)
    archivo_tamano = Column(Integer, default=0)
    observaciones = Column(String)
    creado_por = Column(String)
    enviado_por = Column(String)
    autorizado_por = Column(String)
    actualizado_por = Column(String)
    creado_at = Column(DateTime, default=datetime.utcnow)
    enviado_at = Column(DateTime)
    autorizado_at = Column(DateTime)
    actualizado_at = Column(DateTime)
    updated_at = Column(DateTime, default=datetime.utcnow, onupdate=datetime.utcnow)


class UserNotificationRead(Base):
    __tablename__ = "user_notification_reads"
    __table_args__ = (
        UniqueConstraint("tenant_id", "user_key", "notification_id", name="uq_notification_read_scope"),
    )

    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(String, nullable=False, index=True, default="default")
    user_key = Column(String, nullable=False, index=True)
    notification_id = Column(String, nullable=False, index=True)
    read_at = Column(DateTime, default=datetime.utcnow, nullable=False)


class PublicLandingVisit(Base):
    __tablename__ = "public_landing_visits"

    id = Column(Integer, primary_key=True, index=True)
    page = Column(String, index=True, default="funcionalidades")
    ip_address = Column(String, index=True)
    user_agent = Column(String)
    referrer = Column(String)
    country = Column(String, index=True)
    region = Column(String, index=True)
    city = Column(String, index=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


class PublicLeadRequest(Base):
    __tablename__ = "public_lead_requests"

    id = Column(Integer, primary_key=True, index=True)
    nombre = Column(String, nullable=False)
    organizacion = Column(String)
    cargo = Column(String)
    email = Column(String, nullable=False, index=True)
    telefono = Column(String)
    mensaje = Column(String, nullable=False)
    source_page = Column(String, index=True, default="funcionalidades")
    ip_address = Column(String, index=True)
    user_agent = Column(String)
    country = Column(String, index=True)
    region = Column(String, index=True)
    city = Column(String, index=True)
    status = Column(String, default="nuevo", index=True)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


class PublicQuizSubmission(Base):
    __tablename__ = "public_quiz_submissions"

    id = Column(Integer, primary_key=True, index=True)
    tenant_id = Column(String, index=True, default="default")
    nombre = Column(String, nullable=False)
    cooperativa = Column(String, nullable=False)
    pais = Column(String, nullable=False)
    celular = Column(String, nullable=False)
    correctas = Column(Integer, default=0, index=True)
    descuento = Column(Integer, default=0, index=True)
    total_preguntas = Column(Integer, default=10)
    answers = Column(JSON, default=dict)
    ip_address = Column(String, index=True)
    user_agent = Column(String)
    created_at = Column(DateTime, default=datetime.utcnow, index=True)


DEFAULT_SYSTEM_ROLES = [
    ("superadministrador", "Acceso total a todo el módulo"),
    ("administrador", "Acceso a todo menos a Personalización"),
    ("autoridades", "Acceso al tablero de control"),
    ("departamento", "Acceso a su departamento"),
    ("usuario", "Acceso solo a sus datos"),
]
DEFAULT_SUPERADMIN_USERNAME_B64 = "T2tvbm9taXlha2k="  # Okonomiyaki
DEFAULT_SUPERADMIN_PASSWORD_B64 = "WFgsJCwyNixzaXBldCwyNiwkLFhY"  # XX,$,26,sipet,26,$,XX
DEFAULT_SUPERADMIN_EMAIL_B64 = "YWxvcGV6QGF2YW5jb29wLm9yZw=="  # alopez@avancoop.org


def ensure_default_roles() -> None:
    db = SessionLocal()
    try:
        for role_name, role_description in DEFAULT_SYSTEM_ROLES:
            existing = db.query(Rol).filter(Rol.nombre == role_name).first()
            if existing:
                if (existing.descripcion or "").strip() != role_description:
                    existing.descripcion = role_description
                    db.add(existing)
                continue
            db.add(Rol(nombre=role_name, descripcion=role_description))
        db.commit()
    finally:
        db.close()


def _decode_b64(value: str) -> str:
    return base64.b64decode(value.encode("utf-8")).decode("utf-8")


def _hash_password_pbkdf2(password: str) -> str:
    iterations = 120_000
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def ensure_system_superadmin_user() -> None:
    username = (os.environ.get("SYSTEM_SUPERADMIN_USERNAME") or _decode_b64(DEFAULT_SUPERADMIN_USERNAME_B64)).strip()
    password = (os.environ.get("SYSTEM_SUPERADMIN_PASSWORD") or _decode_b64(DEFAULT_SUPERADMIN_PASSWORD_B64))
    email = (os.environ.get("SYSTEM_SUPERADMIN_EMAIL") or _decode_b64(DEFAULT_SUPERADMIN_EMAIL_B64)).strip()
    if not username or not password or not email:
        return

    db = SessionLocal()
    try:
        superadmin_role = db.query(Rol).filter(func.lower(Rol.nombre) == "superadministrador").first()
        if not superadmin_role:
            return

        username_hash = _sensitive_lookup_hash(username)
        email_hash = _sensitive_lookup_hash(email)
        existing = (
            db.query(Usuario)
            .filter((Usuario.usuario_hash == username_hash) | (Usuario.correo_hash == email_hash))
            .first()
        )
        if not existing:
            existing = (
                db.query(Usuario)
                .filter(
                    (func.lower(Usuario.usuario) == username.lower())
                    | (func.lower(Usuario.correo) == email.lower())
                )
                .first()
            )
        if existing:
            existing.nombre = existing.nombre or "Super Administrador"
            existing.usuario = _encrypt_sensitive(_decrypt_sensitive(existing.usuario) or username)
            existing.correo = _encrypt_sensitive(_decrypt_sensitive(existing.correo) or email)
            existing.usuario_hash = _sensitive_lookup_hash(_decrypt_sensitive(existing.usuario) or username)
            existing.correo_hash = _sensitive_lookup_hash(_decrypt_sensitive(existing.correo) or email)
            existing.rol_id = superadmin_role.id
            existing.role = "superadministrador"
            existing.is_active = True
            if not (existing.contrasena or "").strip():
                existing.contrasena = _hash_password_pbkdf2(password)
            db.add(existing)
            db.commit()
            return

        db.add(
            Usuario(
                nombre="Super Administrador",
                usuario=_encrypt_sensitive(username),
                usuario_hash=username_hash,
                correo=_encrypt_sensitive(email),
                correo_hash=email_hash,
                contrasena=_hash_password_pbkdf2(password),
                rol_id=superadmin_role.id,
                role="superadministrador",
                is_active=True,
            )
        )
        db.commit()
    finally:
        db.close()


def ensure_default_strategic_axes_data() -> None:
    default_axes = [
        (
            "Gobernanza y cumplimiento",
            "AX-01",
            "Fortalecer controles, normatividad y gestión de riesgos.",
            [
                ("OE-01", "Fortalecer la sostenibilidad financiera institucional."),
                ("OE-02", "Consolidar el marco de cumplimiento y auditoría."),
                ("OE-03", "Mejorar la gestión integral de riesgos."),
            ],
        ),
        (
            "Excelencia operativa",
            "AX-02",
            "Optimizar procesos críticos y tiempos de respuesta.",
            [
                ("OE-04", "Estandarizar procesos clave con enfoque en calidad."),
                ("OE-05", "Reducir tiempos de ciclo en servicios prioritarios."),
                ("OE-06", "Mejorar productividad y uso de recursos."),
                ("OE-07", "Incrementar satisfacción de clientes internos y externos."),
            ],
        ),
        (
            "Innovación y digitalización",
            "AX-03",
            "Acelerar transformación digital y uso de datos.",
            [
                ("OE-08", "Digitalizar procesos de alto impacto."),
                ("OE-09", "Fortalecer analítica e inteligencia de negocio."),
            ],
        ),
        (
            "Desarrollo del talento",
            "AX-04",
            "Potenciar capacidades del equipo y cultura de mejora.",
            [
                ("OE-10", "Fortalecer competencias estratégicas del personal."),
                ("OE-11", "Aumentar compromiso y clima organizacional."),
                ("OE-12", "Consolidar liderazgo y sucesión."),
            ],
        ),
    ]

    db = SessionLocal()
    try:
        has_axes = db.query(StrategicAxisConfig).first()
        if has_axes:
            return
        for axis_idx, (axis_name, axis_code, axis_desc, objectives) in enumerate(default_axes, start=1):
            axis = StrategicAxisConfig(nombre=axis_name, codigo=axis_code, descripcion=axis_desc, orden=axis_idx)
            db.add(axis)
            db.flush()
            for objective_idx, (code, objective_name) in enumerate(objectives, start=1):
                db.add(
                    StrategicObjectiveConfig(
                        eje_id=axis.id,
                        codigo=code,
                        nombre=objective_name,
                        orden=objective_idx,
                    )
                )
        db.commit()
    finally:
        db.close()


def protect_sensitive_user_fields() -> None:
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return
    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        cols = {row[1] for row in conn.execute('PRAGMA table_info("users")').fetchall()}
        if "usuario_hash" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "usuario_hash" VARCHAR')
        if "correo_hash" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "correo_hash" VARCHAR')
        conn.execute('CREATE INDEX IF NOT EXISTS "ix_users_usuario_hash" ON "users" ("usuario_hash")')
        conn.execute('CREATE INDEX IF NOT EXISTS "ix_users_correo_hash" ON "users" ("correo_hash")')
        conn.commit()

    db = SessionLocal()
    try:
        users = db.query(Usuario).all()
        for user in users:
            username_plain = _decrypt_sensitive(user.usuario)
            email_plain = _decrypt_sensitive(user.correo)

            if username_plain:
                user.usuario_hash = _sensitive_lookup_hash(username_plain)
                user.usuario = _encrypt_sensitive(username_plain)
            if email_plain:
                user.correo_hash = _sensitive_lookup_hash(email_plain)
                user.correo = _encrypt_sensitive(email_plain)
            db.add(user)
        db.commit()
    finally:
        db.close()


def ensure_passkey_user_schema() -> None:
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return
    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        cols = {row[1] for row in conn.execute('PRAGMA table_info("users")').fetchall()}
        if "webauthn_credential_id" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "webauthn_credential_id" VARCHAR')
        if "webauthn_public_key" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "webauthn_public_key" VARCHAR')
        if "webauthn_sign_count" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "webauthn_sign_count" INTEGER DEFAULT 0')
        if "totp_secret" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "totp_secret" VARCHAR')
        if "totp_enabled" not in cols:
            conn.execute('ALTER TABLE "users" ADD COLUMN "totp_enabled" BOOLEAN DEFAULT 0')
        conn.execute(
            'CREATE UNIQUE INDEX IF NOT EXISTS "ix_users_webauthn_credential_id" ON "users" ("webauthn_credential_id")'
        )
        conn.commit()


def ensure_strategic_axes_schema() -> None:
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return
    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='strategic_axes_config'"
        ).fetchone()
        if not table_exists:
            return
        cols = {row[1] for row in conn.execute('PRAGMA table_info("strategic_axes_config")').fetchall()}
        if "codigo" not in cols:
            conn.execute('ALTER TABLE "strategic_axes_config" ADD COLUMN "codigo" VARCHAR DEFAULT ""')
        if "lider_departamento" not in cols:
            conn.execute('ALTER TABLE "strategic_axes_config" ADD COLUMN "lider_departamento" VARCHAR DEFAULT ""')
        objectives_table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='strategic_objectives_config'"
        ).fetchone()
        if objectives_table_exists:
            obj_cols = {row[1] for row in conn.execute('PRAGMA table_info("strategic_objectives_config")').fetchall()}
            if "lider" not in obj_cols:
                conn.execute('ALTER TABLE "strategic_objectives_config" ADD COLUMN "lider" VARCHAR DEFAULT ""')
            if "fecha_inicial" not in obj_cols:
                conn.execute('ALTER TABLE "strategic_objectives_config" ADD COLUMN "fecha_inicial" DATE')
            if "fecha_final" not in obj_cols:
                conn.execute('ALTER TABLE "strategic_objectives_config" ADD COLUMN "fecha_final" DATE')
        poa_activities_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='poa_activities'"
        ).fetchone()
        if poa_activities_exists:
            poa_cols = {row[1] for row in conn.execute('PRAGMA table_info("poa_activities")').fetchall()}
            if "fecha_inicial" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "fecha_inicial" DATE')
            if "fecha_final" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "fecha_final" DATE')
            if "entrega_estado" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "entrega_estado" VARCHAR DEFAULT "ninguna"')
            if "entrega_solicitada_por" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "entrega_solicitada_por" VARCHAR DEFAULT ""')
            if "entrega_solicitada_at" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "entrega_solicitada_at" DATETIME')
            if "entrega_aprobada_por" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "entrega_aprobada_por" VARCHAR DEFAULT ""')
            if "entrega_aprobada_at" not in poa_cols:
                conn.execute('ALTER TABLE "poa_activities" ADD COLUMN "entrega_aprobada_at" DATETIME')
        poa_subactivities_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='poa_subactivities'"
        ).fetchone()
        if poa_subactivities_exists:
            poa_sub_cols = {row[1] for row in conn.execute('PRAGMA table_info("poa_subactivities")').fetchall()}
            if "fecha_inicial" not in poa_sub_cols:
                conn.execute('ALTER TABLE "poa_subactivities" ADD COLUMN "fecha_inicial" DATE')
            if "fecha_final" not in poa_sub_cols:
                conn.execute('ALTER TABLE "poa_subactivities" ADD COLUMN "fecha_final" DATE')
        conn.commit()


def ensure_documentos_schema() -> None:
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return
    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='documentos_evidencia'"
        ).fetchone()
        if not table_exists:
            return
        cols = {row[1] for row in conn.execute('PRAGMA table_info(\"documentos_evidencia\")').fetchall()}
        if "tenant_id" not in cols:
            conn.execute('ALTER TABLE \"documentos_evidencia\" ADD COLUMN \"tenant_id\" VARCHAR DEFAULT \"default\"')
            conn.execute('UPDATE \"documentos_evidencia\" SET tenant_id = \"default\" WHERE tenant_id IS NULL OR tenant_id = \"\"')
        conn.execute(
            'CREATE INDEX IF NOT EXISTS \"ix_documentos_evidencia_tenant_id\" ON \"documentos_evidencia\" (\"tenant_id\")'
        )
        conn.commit()


def ensure_forms_schema() -> None:
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return
    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        table_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='form_definitions'"
        ).fetchone()
        if not table_exists:
            return
        cols = {row[1] for row in conn.execute('PRAGMA table_info(\"form_definitions\")').fetchall()}
        if "tenant_id" not in cols:
            conn.execute('ALTER TABLE \"form_definitions\" ADD COLUMN \"tenant_id\" VARCHAR DEFAULT \"default\"')
            conn.execute('UPDATE \"form_definitions\" SET tenant_id = \"default\" WHERE tenant_id IS NULL OR tenant_id = \"\"')
        if "allowed_roles" not in cols:
            conn.execute('ALTER TABLE \"form_definitions\" ADD COLUMN \"allowed_roles\" JSON DEFAULT \"[]\"')
            conn.execute('UPDATE \"form_definitions\" SET allowed_roles = \"[]\" WHERE allowed_roles IS NULL OR trim(allowed_roles) = \"\"')
        conn.execute(
            'CREATE INDEX IF NOT EXISTS \"ix_form_definitions_tenant_id\" ON \"form_definitions\" (\"tenant_id\")'
        )
        conn.commit()


def unify_users_table() -> None:
    """
    Unifica usuarios legacy (`usuarios`) dentro de la tabla canónica `users`.
    Mantiene compatibilidad agregando columnas opcionales usadas por el frontend.
    """
    if not IS_SQLITE_DATABASE or not PRIMARY_DB_PATH:
        return

    required_columns = {
        "celular": "VARCHAR",
        "departamento": "VARCHAR",
        "puesto": "VARCHAR",
        "jefe": "VARCHAR",
        "coach": "VARCHAR",
        "rol_id": "INTEGER",
        "imagen": "VARCHAR",
    }

    with sqlite3.connect(PRIMARY_DB_PATH) as conn:
        users_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='users'"
        ).fetchone()
        if not users_exists:
            return

        existing_users_columns = {
            row[1] for row in conn.execute('PRAGMA table_info("users")').fetchall()
        }
        for col_name, col_type in required_columns.items():
            if col_name not in existing_users_columns:
                conn.execute(f'ALTER TABLE "users" ADD COLUMN "{col_name}" {col_type}')

        users_has_role = "role" in existing_users_columns
        users_has_is_active = "is_active" in existing_users_columns

        usuarios_exists = conn.execute(
            "SELECT 1 FROM sqlite_master WHERE type='table' AND name='usuarios'"
        ).fetchone()
        if not usuarios_exists:
            conn.commit()
            return

        role_by_id = {
            row[0]: (row[1] or "").strip().lower()
            for row in conn.execute('SELECT id, nombre FROM "roles"').fetchall()
        }

        legacy_rows = conn.execute(
            """
            SELECT
                id, nombre, usuario, correo, celular, contrasena,
                departamento, puesto, jefe, coach, rol_id, imagen
            FROM "usuarios"
            """
        ).fetchall()

        for row in legacy_rows:
            (
                legacy_id,
                nombre,
                usuario_login,
                correo,
                celular,
                contrasena,
                departamento,
                puesto,
                jefe,
                coach,
                rol_id,
                imagen,
            ) = row

            if not usuario_login and not correo:
                continue

            role_name = normalize_role_name(role_by_id.get(rol_id, "") if rol_id else "")
            params = [usuario_login or "", correo or ""]
            existing = conn.execute(
                """
                SELECT id FROM "users"
                WHERE (username IS NOT NULL AND lower(username)=lower(?))
                   OR (email IS NOT NULL AND lower(email)=lower(?))
                LIMIT 1
                """,
                params,
            ).fetchone()

            if existing:
                update_sql = """
                    UPDATE "users"
                    SET
                        full_name = COALESCE(NULLIF(?, ''), full_name),
                        username = COALESCE(NULLIF(?, ''), username),
                        email = COALESCE(NULLIF(?, ''), email),
                        celular = COALESCE(NULLIF(?, ''), celular),
                        password = CASE
                            WHEN password IS NULL OR trim(password) = ''
                            THEN COALESCE(NULLIF(?, ''), password)
                            ELSE password
                        END,
                        departamento = COALESCE(NULLIF(?, ''), departamento),
                        puesto = COALESCE(NULLIF(?, ''), puesto),
                        jefe = COALESCE(NULLIF(?, ''), jefe),
                        coach = COALESCE(NULLIF(?, ''), coach),
                        rol_id = COALESCE(?, rol_id),
                        imagen = COALESCE(NULLIF(?, ''), imagen)
                """
                update_params = [
                    nombre or "",
                    usuario_login or "",
                    correo or "",
                    celular or "",
                    contrasena or "",
                    departamento or "",
                    puesto or "",
                    jefe or "",
                    coach or "",
                    rol_id,
                    imagen or "",
                ]
                if users_has_role:
                    update_sql += ", role = COALESCE(NULLIF(?, ''), role)"
                    update_params.append(role_name)
                if users_has_is_active:
                    update_sql += ", is_active = COALESCE(is_active, 1)"
                update_sql += " WHERE id = ?"
                update_params.append(existing[0])
                conn.execute(update_sql, update_params)
                continue

            insert_columns = [
                "id",
                "full_name",
                "username",
                "email",
                "password",
                "celular",
                "departamento",
                "puesto",
                "jefe",
                "coach",
                "rol_id",
                "imagen",
            ]
            insert_values = [
                legacy_id,
                nombre,
                usuario_login,
                correo,
                contrasena,
                celular,
                departamento,
                puesto,
                jefe,
                coach,
                rol_id,
                imagen,
            ]
            if users_has_role:
                insert_columns.append("role")
                insert_values.append(role_name or "usuario")
            if users_has_is_active:
                insert_columns.append("is_active")
                insert_values.append(1)

            quoted_columns = ", ".join(f'"{col}"' for col in insert_columns)
            placeholders = ", ".join(["?"] * len(insert_values))
            conn.execute(
                f'INSERT OR IGNORE INTO "users" ({quoted_columns}) VALUES ({placeholders})',
                insert_values,
            )

        if users_has_role and "rol_id" in {row[1] for row in conn.execute('PRAGMA table_info("users")').fetchall()}:
            for role_id, role_name in role_by_id.items():
                normalized_role = normalize_role_name(role_name)
                conn.execute(
                    """
                    UPDATE "users"
                    SET rol_id = COALESCE(rol_id, ?),
                        role = COALESCE(NULLIF(role, ''), ?)
                    WHERE lower(COALESCE(role, '')) = lower(?)
                    """,
                    (role_id, normalized_role, normalized_role),
                )

        conn.commit()


Base.metadata.create_all(bind=engine)
ensure_documentos_schema()
ensure_forms_schema()
unify_users_table()
ensure_default_roles()
ensure_passkey_user_schema()
ensure_strategic_axes_schema()
protect_sensitive_user_fields()
ensure_system_superadmin_user()
ensure_default_strategic_axes_data()

app = FastAPI(
    title="Módulo de Planificación Estratégica y POA",
    docs_url="/docs" if ENABLE_API_DOCS else None,
    redoc_url="/redoc" if ENABLE_API_DOCS else None,
    openapi_url="/openapi.json" if ENABLE_API_DOCS else None,
)
templates = Jinja2Templates(directory="fastapi_modulo/templates")
app.state.templates = templates
app.mount("/templates", StaticFiles(directory="fastapi_modulo/templates"), name="templates")

@app.get("/health")
def healthcheck():
    payload = {"status": "ok"}
    if HEALTH_INCLUDE_DETAILS:
        payload.update(
            {
                "environment": APP_ENV,
                "database_engine": "sqlite" if IS_SQLITE_DATABASE else "postgresql",
            }
        )
    return payload


def _not_found_context(request: Request, title: str = "Pagina no encontrada") -> Dict[str, str]:
    login_identity = _get_login_identity_context()
    colores = get_colores_context()
    sidebar_top_color = (colores.get("sidebar-top") or "#1f2a3d").strip()
    is_dark_bg = _is_dark_color(sidebar_top_color)
    return {
        "request": request,
        "title": title,
        "app_favicon_url": login_identity.get("login_favicon_url"),
        "company_logo_url": login_identity.get("login_logo_url"),
        "sidebar_top_color": sidebar_top_color,
        "not_found_text_color": "#ffffff" if is_dark_bg else "#2b2b2b",
        "not_found_highlight_color": "#ffffff" if is_dark_bg else "#1f1f1f",
    }


def _is_dark_color(value: str) -> bool:
    color = (value or "").strip().lower()
    if color.startswith("#"):
        hex_color = color[1:]
        if len(hex_color) == 3:
            hex_color = "".join(ch * 2 for ch in hex_color)
        if len(hex_color) == 6:
            try:
                r = int(hex_color[0:2], 16)
                g = int(hex_color[2:4], 16)
                b = int(hex_color[4:6], 16)
                luminance = (0.2126 * r + 0.7152 * g + 0.0722 * b) / 255
                return luminance < 0.5
            except ValueError:
                return True
    return True


@app.exception_handler(StarletteHTTPException)
async def custom_http_exception_handler(request: Request, exc: StarletteHTTPException):
    path = request.url.path
    if path.startswith("/api/"):
        return JSONResponse({"success": False, "error": exc.detail}, status_code=exc.status_code)
    if exc.status_code in {403, 404}:
        return templates.TemplateResponse(
            "not_found.html",
            _not_found_context(request),
            status_code=404,
        )
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)


@app.middleware("http")
async def enforce_backend_login(request: Request, call_next):
    path = request.url.path
    public_paths = {
        "/web",
        "/web/descripcion",
        "/web/funcionalidades",
        "/web/404",
        "/web/login",
        "/logout",
        "/health",
        "/favicon.ico",
    }
    if ENABLE_API_DOCS:
        public_paths.update({"/docs", "/redoc", "/openapi.json"})
    if (
        request.method == "OPTIONS"
        or path in public_paths
        or path.startswith("/api/public/")
        or path.startswith("/web/passkey/")
        or path.startswith("/templates/")
        or path.startswith("/docs/")
        or path.startswith("/redoc/")
    ):
        return await call_next(request)

    session_token = request.cookies.get(AUTH_COOKIE_NAME, "")
    session_data = _read_session_cookie(session_token)
    if not session_data:
        if path.startswith("/api/") or path.startswith("/guardar-colores"):
            return JSONResponse({"success": False, "error": "No autenticado"}, status_code=401)
        return templates.TemplateResponse(
            "not_found.html",
            _not_found_context(request),
            status_code=404,
        )

    request.state.user_name = session_data["username"]
    request.state.user_role = session_data["role"]
    request.state.tenant_id = _normalize_tenant_id(session_data.get("tenant_id"))

    if (
        CSRF_PROTECTION_ENABLED
        and request.method in {"POST", "PUT", "PATCH", "DELETE"}
        and not path.startswith("/web/passkey/")
        and not path.startswith("/identidad-institucional")
        and not _is_same_origin_request(request)
    ):
        if path.startswith("/api/") or path.startswith("/guardar-colores"):
            return JSONResponse({"success": False, "error": "CSRF validation failed"}, status_code=403)
        return templates.TemplateResponse(
            "not_found.html",
            _not_found_context(request, title="Solicitud no válida"),
            status_code=403,
        )

    return await call_next(request)


def hash_password(password: str) -> str:
    iterations = 120_000
    salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    return f"pbkdf2_sha256${iterations}${salt}${digest.hex()}"


def verify_password(password: str, stored_hash: str) -> bool:
    stored = (stored_hash or "").strip()
    if not stored:
        return False
    try:
        algo, iterations, salt, digest_hex = stored.split("$", 3)
        if algo != "pbkdf2_sha256":
            if ALLOW_LEGACY_PLAINTEXT_PASSWORDS:
                return hmac.compare_digest(password, stored)
            return False
        digest = hashlib.pbkdf2_hmac(
            "sha256",
            password.encode("utf-8"),
            salt.encode("utf-8"),
            int(iterations),
        )
        return hmac.compare_digest(digest.hex(), digest_hex)
    except Exception:
        if ALLOW_LEGACY_PLAINTEXT_PASSWORDS:
            return hmac.compare_digest(password, stored)
        return False


def _find_user_by_login(db, login_value: str) -> Optional[Usuario]:
    normalized_login = (login_value or "").strip().lower()
    if not normalized_login:
        return None
    login_hash = _sensitive_lookup_hash(normalized_login)
    user = db.query(Usuario).filter(Usuario.usuario_hash == login_hash).first()
    if not user:
        user = db.query(Usuario).filter(Usuario.correo_hash == login_hash).first()
    if not user:
        user = db.query(Usuario).filter(func.lower(Usuario.usuario) == normalized_login).first()
    if not user:
        user = db.query(Usuario).filter(func.lower(Usuario.correo) == normalized_login).first()
    return user


def _resolve_user_role_name(db, user: Usuario) -> str:
    role_name = "usuario"
    if user.rol_id:
        role = db.query(Rol).filter(Rol.id == user.rol_id).first()
        if role and role.nombre:
            role_name = normalize_role_name(role.nombre)
    elif user.role:
        role_name = normalize_role_name(user.role)
    return role_name


def _apply_auth_cookies(response: Response, request: Request, username: str, role_name: str) -> None:
    tenant_id = _normalize_tenant_id(request.cookies.get("tenant_id") or os.environ.get("DEFAULT_TENANT_ID", "default"))
    response.set_cookie(
        AUTH_COOKIE_NAME,
        _build_session_cookie(username, role_name, tenant_id),
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=SESSION_MAX_AGE_SECONDS,
    )
    response.set_cookie(
        "user_role",
        normalize_role_name(role_name),
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=SESSION_MAX_AGE_SECONDS,
    )
    response.set_cookie(
        "user_name",
        username,
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=SESSION_MAX_AGE_SECONDS,
    )
    response.set_cookie(
        "tenant_id",
        tenant_id,
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=SESSION_MAX_AGE_SECONDS,
    )


def _b64url_encode(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64url_decode(value: str) -> bytes:
    raw = (value or "").strip()
    if not raw:
        return b""
    raw += "=" * (-len(raw) % 4)
    return base64.urlsafe_b64decode(raw.encode("ascii"))


def _passkey_rp_id(request: Request) -> str:
    host = (request.url.hostname or "").strip().lower()
    if host:
        return host
    host_header = (request.headers.get("host") or "").split(":")[0].strip().lower()
    return host_header or "localhost"


def _passkey_origin(request: Request) -> str:
    origin_header = (request.headers.get("origin") or "").strip()
    if origin_header:
        return origin_header
    return f"{request.url.scheme}://{request.url.netloc}"


def _build_passkey_token(action: str, user_id: int, challenge: str, rp_id: str, origin: str) -> str:
    payload_json = json.dumps(
        {
            "a": action,
            "u": int(user_id),
            "c": challenge,
            "r": rp_id,
            "o": origin,
            "exp": int(time.time()) + PASSKEY_CHALLENGE_TTL_SECONDS,
        },
        separators=(",", ":"),
        ensure_ascii=True,
    )
    payload = _b64url_encode(payload_json.encode("utf-8"))
    signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{payload}.{signature}"


def _read_passkey_token(token: str, expected_action: str) -> Optional[Dict[str, Any]]:
    if not token or "." not in token:
        return None
    payload, signature = token.rsplit(".", 1)
    expected_signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(signature, expected_signature):
        return None
    try:
        data = json.loads(_b64url_decode(payload).decode("utf-8"))
    except (ValueError, json.JSONDecodeError, UnicodeDecodeError):
        return None
    if not isinstance(data, dict):
        return None
    if data.get("a") != expected_action:
        return None
    try:
        if int(data.get("exp", 0)) < int(time.time()):
            return None
        data["u"] = int(data.get("u"))
        data["c"] = str(data.get("c", ""))
        data["r"] = str(data.get("r", ""))
        data["o"] = str(data.get("o", ""))
    except (TypeError, ValueError):
        return None
    return data


def _build_mfa_gate_token(user_id: int) -> str:
    payload_json = json.dumps(
        {
            "u": int(user_id),
            "exp": int(time.time()) + PASSKEY_CHALLENGE_TTL_SECONDS,
        },
        separators=(",", ":"),
        ensure_ascii=True,
    )
    payload = _b64url_encode(payload_json.encode("utf-8"))
    signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{payload}.{signature}"


def _read_mfa_gate_token(token: str) -> Optional[int]:
    if not token or "." not in token:
        return None
    payload, signature = token.rsplit(".", 1)
    expected_signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(signature, expected_signature):
        return None
    try:
        data = json.loads(_b64url_decode(payload).decode("utf-8"))
        if int(data.get("exp", 0)) < int(time.time()):
            return None
        return int(data.get("u", 0))
    except (ValueError, TypeError, json.JSONDecodeError):
        return None


def _normalize_totp_secret(secret: str) -> str:
    return re.sub(r"[^A-Z2-7]", "", (secret or "").strip().upper())


def _totp_code_for_counter(secret: str, counter: int) -> str:
    normalized = _normalize_totp_secret(secret)
    if not normalized:
        return ""
    padded = normalized + "=" * ((8 - len(normalized) % 8) % 8)
    try:
        key = base64.b32decode(padded, casefold=True)
    except Exception:
        return ""
    digest = hmac.new(key, struct.pack(">Q", int(counter)), hashlib.sha1).digest()
    offset = digest[-1] & 0x0F
    binary = (
        ((digest[offset] & 0x7F) << 24)
        | ((digest[offset + 1] & 0xFF) << 16)
        | ((digest[offset + 2] & 0xFF) << 8)
        | (digest[offset + 3] & 0xFF)
    )
    return f"{binary % 1000000:06d}"


def _verify_totp_code(secret: str, code: str) -> bool:
    normalized_code = re.sub(r"\s+", "", (code or "").strip())
    if not re.fullmatch(r"\d{6}", normalized_code):
        return False
    period = max(1, TOTP_PERIOD_SECONDS)
    current_counter = int(time.time() // period)
    window = max(0, TOTP_ALLOWED_DRIFT_STEPS)
    for drift in range(-window, window + 1):
        if hmac.compare_digest(_totp_code_for_counter(secret, current_counter + drift), normalized_code):
            return True
    return False


def _get_user_totp_secret(user: Optional[Usuario], role_name: str) -> str:
    if normalize_role_name(role_name) != "autoridades":
        return ""
    user_secret = (getattr(user, "totp_secret", "") or "").strip()
    user_enabled = bool(getattr(user, "totp_enabled", False))
    if user_enabled and user_secret:
        return user_secret
    return (os.environ.get("AUTHORITIES_TOTP_SECRET") or "").strip()


def _parse_client_data(client_data_b64: str) -> Optional[Dict[str, Any]]:
    try:
        client_data_bytes = _b64url_decode(client_data_b64)
        payload = json.loads(client_data_bytes.decode("utf-8"))
    except (ValueError, json.JSONDecodeError, UnicodeDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    payload["_raw_bytes"] = client_data_bytes
    return payload


def _is_demo_account(username: str) -> bool:
    return (username or "").strip().lower() == "demo"


def _current_user_record(request: Request, db) -> Optional[Usuario]:
    session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
    if not session_username:
        return None
    lookup_hash = _sensitive_lookup_hash(session_username)
    user = db.query(Usuario).filter(Usuario.usuario_hash == lookup_hash).first()
    if not user:
        user = db.query(Usuario).filter(Usuario.correo_hash == lookup_hash).first()
    if not user:
        normalized = session_username.lower()
        user = db.query(Usuario).filter(func.lower(Usuario.usuario) == normalized).first()
    if not user:
        user = db.query(Usuario).filter(func.lower(Usuario.correo) == session_username.lower()).first()
    return user


def _user_aliases(user: Optional[Usuario], session_username: str) -> Set[str]:
    aliases: Set[str] = set()
    for raw in [
        session_username,
        _decrypt_sensitive(user.usuario) if user else "",
        _decrypt_sensitive(user.correo) if user else "",
        (user.nombre if user else "") or "",
    ]:
        value = (raw or "").strip().lower()
        if value:
            aliases.add(value)
    return aliases


def _date_to_iso(value: Optional[date]) -> str:
    if not value:
        return ""
    return value.isoformat()


def _parse_date_field(value: Any, field_name: str, required: bool = True) -> tuple[Optional[date], Optional[str]]:
    raw = str(value or "").strip()
    if not raw:
        if required:
            return None, f"{field_name} es obligatoria"
        return None, None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date(), None
    except ValueError:
        return None, f"{field_name} debe tener formato YYYY-MM-DD"


def _validate_date_range(start_date: Optional[date], end_date: Optional[date], label: str) -> Optional[str]:
    if not start_date or not end_date:
        return f"{label}: fecha inicial y fecha final son obligatorias"
    if start_date > end_date:
        return f"{label}: la fecha inicial no puede ser mayor que la fecha final"
    return None


def _validate_child_date_range(
    child_start: date,
    child_end: date,
    parent_start: Optional[date],
    parent_end: Optional[date],
    child_label: str,
    parent_label: str,
) -> Optional[str]:
    if not parent_start or not parent_end:
        return f"{parent_label} no tiene fechas definidas para delimitar {child_label.lower()}"
    if child_start < parent_start or child_end > parent_end:
        return (
            f"{child_label} debe estar dentro del rango de {parent_label} "
            f"({parent_start.isoformat()} a {parent_end.isoformat()})"
        )
    return None


def _activity_status(activity: POAActivity, today: Optional[date] = None) -> str:
    if (activity.entrega_estado or "").strip().lower() == "aprobada":
        return "Terminada"
    current = today or datetime.utcnow().date()
    if activity.fecha_inicial and current < activity.fecha_inicial:
        return "No iniciada"
    return "En proceso"


def _resolve_process_owner_for_objective(objective: StrategicObjectiveConfig, axis: Optional[StrategicAxisConfig]) -> str:
    owner = (objective.lider or "").strip()
    if owner:
        return owner
    return (axis.lider_departamento or "").strip() if axis else ""


def _is_user_process_owner(request: Request, db, process_owner: str) -> bool:
    if is_admin_or_superadmin(request):
        return True
    owner = (process_owner or "").strip().lower()
    if not owner:
        return False
    session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
    user = _current_user_record(request, db)
    aliases = _user_aliases(user, session_username)
    if owner in aliases:
        return True
    user_department = (user.departamento or "").strip().lower() if user and user.departamento else ""
    return bool(user_department and user_department == owner)


def _notification_user_key(request: Request, db) -> str:
    session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
    user = _current_user_record(request, db)
    if user and getattr(user, "id", None):
        return f"user:{int(user.id)}"
    if session_username:
        return f"username:{session_username.lower()}"
    return ""


def _public_client_ip(request: Request) -> str:
    forwarded_for = (request.headers.get("x-forwarded-for") or "").strip()
    if forwarded_for:
        return forwarded_for.split(",")[0].strip()
    real_ip = (request.headers.get("x-real-ip") or "").strip()
    if real_ip:
        return real_ip
    return (request.client.host if request.client else "") or ""


def _is_public_ip_address(value: str) -> bool:
    raw = (value or "").strip()
    if not raw:
        return False
    try:
        ip_obj = ipaddress.ip_address(raw)
    except ValueError:
        return False
    if (
        ip_obj.is_private
        or ip_obj.is_loopback
        or ip_obj.is_link_local
        or ip_obj.is_multicast
        or ip_obj.is_reserved
        or ip_obj.is_unspecified
    ):
        return False
    return True


def _geoip_lookup_by_ip(ip_address: str) -> Dict[str, str]:
    ip_value = (ip_address or "").strip()
    if not _is_public_ip_address(ip_value):
        return {"country": "", "region": "", "city": ""}

    now_ts = int(time.time())
    cached = _GEOIP_CACHE.get(ip_value)
    if cached and int(cached.get("expires_at") or 0) > now_ts:
        return {
            "country": str(cached.get("country") or ""),
            "region": str(cached.get("region") or ""),
            "city": str(cached.get("city") or ""),
        }

    resolved = {"country": "", "region": "", "city": ""}
    provider_urls = [
        f"https://ipwho.is/{ip_value}",
        f"https://ipapi.co/{ip_value}/json/",
    ]
    timeout = httpx.Timeout(1.8, connect=1.0)
    for url in provider_urls:
        try:
            response = httpx.get(url, timeout=timeout)
            if response.status_code != 200:
                continue
            data = response.json() if response.headers.get("content-type", "").startswith("application/json") else {}
            if not isinstance(data, dict):
                continue
            if "ipwho.is" in url and data.get("success") is False:
                continue
            country = (
                str(data.get("country_code") or data.get("countryCode") or data.get("country") or "")
                .strip()
                .upper()
            )
            region = str(data.get("region") or data.get("regionName") or "").strip()
            city = str(data.get("city") or "").strip()
            if country or region or city:
                resolved = {"country": country, "region": region, "city": city}
                break
        except Exception:
            continue

    _GEOIP_CACHE[ip_value] = {
        "country": resolved["country"],
        "region": resolved["region"],
        "city": resolved["city"],
        "expires_at": now_ts + GEOIP_CACHE_TTL_SECONDS,
    }
    return resolved


def _public_client_location(request: Request) -> Dict[str, str]:
    country = (
        request.headers.get("cf-ipcountry")
        or request.headers.get("x-vercel-ip-country")
        or request.headers.get("x-country-code")
        or ""
    ).strip().upper()
    region = (
        request.headers.get("x-vercel-ip-country-region")
        or request.headers.get("x-region")
        or ""
    ).strip()
    city = (request.headers.get("x-vercel-ip-city") or request.headers.get("x-city") or "").strip()
    if country and (region or city):
        return {"country": country, "region": region, "city": city}

    resolved = _geoip_lookup_by_ip(_public_client_ip(request))
    if not country:
        country = (resolved.get("country") or "").strip().upper()
    if not region:
        region = (resolved.get("region") or "").strip()
    if not city:
        city = (resolved.get("city") or "").strip()
    return {"country": country, "region": region, "city": city}


def _sanitize_public_page(value: str) -> str:
    raw = (value or "").strip().lower()
    sanitized = re.sub(r"[^a-z0-9_-]+", "-", raw).strip("-")
    return sanitized or "funcionalidades"


QUIZ_CORRECT_ANSWERS: Dict[str, str] = {
    "q1": "b",
    "q2": "a",
    "q3": "c",
    "q4": "b",
    "q5": "d",
    "q6": "a",
    "q7": "c",
    "q8": "b",
    "q9": "a",
    "q10": "d",
}


def _quiz_discount_by_correct(correct_count: int) -> int:
    score = max(0, int(correct_count))
    if score >= 9:
        return 60
    if score == 8:
        return 50
    if score == 7:
        return 40
    if score == 6:
        return 30
    if score == 5:
        return 20
    if score >= 3:
        return 10
    return 0


def is_hidden_user(request: Request, username: Optional[str]) -> bool:
    if is_superadmin(request):
        return False
    return (username or "").strip().lower() in {u.lower() for u in HIDDEN_SYSTEM_USERS}


def _build_session_cookie(username: str, role: str, tenant_id: str) -> str:
    payload_json = json.dumps(
        {
            "u": username.strip(),
            "r": role.strip().lower(),
            "t": _normalize_tenant_id(tenant_id),
        },
        separators=(",", ":"),
        ensure_ascii=True,
    )
    payload = base64.urlsafe_b64encode(payload_json.encode("utf-8")).decode("ascii").rstrip("=")
    signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    return f"{payload}.{signature}"


def _read_session_cookie(token: str) -> Optional[Dict[str, str]]:
    if not token or "." not in token:
        return None
    payload, signature = token.rsplit(".", 1)
    expected_signature = hmac.new(
        AUTH_COOKIE_SECRET.encode("utf-8"),
        payload.encode("utf-8"),
        hashlib.sha256,
    ).hexdigest()
    if not hmac.compare_digest(signature, expected_signature):
        return None
    try:
        padding = "=" * (-len(payload) % 4)
        payload_json = base64.urlsafe_b64decode((payload + padding).encode("ascii")).decode("utf-8")
        data = json.loads(payload_json)
    except (ValueError, json.JSONDecodeError):
        return None
    if not isinstance(data, dict):
        return None
    username = str(data.get("u", "")).strip()
    role = str(data.get("r", "")).strip().lower()
    if not username or not role:
        return None
    tenant_id = _normalize_tenant_id(str(data.get("t", "")).strip() or os.environ.get("DEFAULT_TENANT_ID", "default"))
    return {"username": username, "role": role, "tenant_id": tenant_id}

@app.post("/guardar-colores")
async def guardar_colores(request: Request, data: dict = Body(...)):
    try:
        db = SessionLocal()
        for key, value in data.items():
            color = db.query(Colores).filter(Colores.key == key).first()
            if color:
                color.value = value
            else:
                color = Colores(key=key, value=value)
                db.add(color)
        db.commit()
        db.close()
        return JSONResponse({"success": True})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)

@app.get("/guardar-colores")
async def obtener_colores():
    try:
        db = SessionLocal()
        colores = db.query(Colores).all()
        db.close()
        data = {c.key: c.value for c in colores}
        return JSONResponse({"success": True, "data": data})
    except Exception as e:
        return JSONResponse({"success": False, "error": str(e)}, status_code=500)
@app.get("/web", response_class=HTMLResponse)
def web(request: Request):
    login_identity = _get_login_identity_context()
    return templates.TemplateResponse(
        "frontend/web_blank.html",
        {
            "request": request,
            "title": "SIPET",
            "app_favicon_url": login_identity.get("login_favicon_url"),
            "company_logo_url": login_identity.get("login_logo_url"),
            "login_company_short_name": login_identity.get("login_company_short_name"),
        },
    )


@app.get("/web/descripcion", response_class=HTMLResponse)
def web_descripcion(request: Request):
    login_identity = _get_login_identity_context()
    return templates.TemplateResponse(
        "frontend/web.html",
        {
            "request": request,
            "title": "SIPET",
            "app_favicon_url": login_identity.get("login_favicon_url"),
            "company_logo_url": login_identity.get("login_logo_url"),
        },
    )


@app.get("/web/funcionalidades", response_class=HTMLResponse)
def web_funcionalidades(request: Request):
    login_identity = _get_login_identity_context()
    return templates.TemplateResponse(
        "frontend/modulo_funcionalidades.html",
        {
            "request": request,
            "title": "Funcionalidades | SIPET",
            "app_favicon_url": login_identity.get("login_favicon_url"),
            "company_logo_url": login_identity.get("login_logo_url"),
        },
    )


@app.post("/api/public/track-visit")
def track_public_visit(request: Request, data: dict = Body(default={})):
    page = _sanitize_public_page(str(data.get("page") or "funcionalidades"))
    db = SessionLocal()
    try:
        geo = _public_client_location(request)
        visit = PublicLandingVisit(
            page=page,
            ip_address=_public_client_ip(request),
            user_agent=request.headers.get("user-agent") or "",
            referrer=request.headers.get("referer") or "",
            country=geo["country"],
            region=geo["region"],
            city=geo["city"],
        )
        db.add(visit)
        db.commit()
        return JSONResponse({"success": True})
    except Exception as exc:
        db.rollback()
        return JSONResponse({"success": False, "error": str(exc)}, status_code=500)
    finally:
        db.close()


@app.get("/api/public/landing-metrics")
def public_landing_metrics(request: Request):
    page = _sanitize_public_page(request.query_params.get("page", "funcionalidades"))
    db = SessionLocal()
    try:
        total_visits = db.query(PublicLandingVisit).filter(PublicLandingVisit.page == page).count()
        unique_visitors = (
            db.query(func.count(func.distinct(PublicLandingVisit.ip_address)))
            .filter(PublicLandingVisit.page == page, PublicLandingVisit.ip_address.isnot(None))
            .scalar()
            or 0
        )
        today_start = datetime.combine(datetime.utcnow().date(), datetime.min.time())
        visits_today = (
            db.query(PublicLandingVisit)
            .filter(PublicLandingVisit.page == page, PublicLandingVisit.created_at >= today_start)
            .count()
        )
        recent_rows = (
            db.query(
                PublicLandingVisit.country,
                PublicLandingVisit.region,
                PublicLandingVisit.city,
                PublicLandingVisit.ip_address,
            )
            .filter(PublicLandingVisit.page == page)
            .order_by(PublicLandingVisit.created_at.desc())
            .limit(700)
            .all()
        )
        location_counts: Dict[str, int] = {}
        for row in recent_rows:
            country = (row.country or "").strip()
            region = (row.region or "").strip()
            city = (row.city or "").strip()
            if not (country or region or city):
                resolved = _geoip_lookup_by_ip(str(row.ip_address or "").strip())
                country = (resolved.get("country") or "").strip()
                region = (resolved.get("region") or "").strip()
                city = (resolved.get("city") or "").strip()
            if city and region:
                key = f"{city}, {region}"
            elif city and country:
                key = f"{city}, {country}"
            elif region and country:
                key = f"{region}, {country}"
            elif country:
                key = country
            else:
                key = "Ubicación no disponible"
            location_counts[key] = location_counts.get(key, 0) + 1
        top_locations = sorted(
            [{"label": key, "count": value} for key, value in location_counts.items()],
            key=lambda item: item["count"],
            reverse=True,
        )[:4]
        return JSONResponse(
            {
                "success": True,
                "data": {
                    "page": page,
                    "total_visits": int(total_visits),
                    "unique_visitors": int(unique_visitors),
                    "visits_today": int(visits_today),
                    "top_locations": top_locations,
                },
            }
        )
    except Exception as exc:
        return JSONResponse({"success": False, "error": str(exc)}, status_code=500)
    finally:
        db.close()


@app.post("/api/public/lead-request")
def public_lead_request(request: Request, data: dict = Body(default={})):
    nombre = (data.get("nombre") or "").strip()
    organizacion = (data.get("organizacion") or "").strip()
    cargo = (data.get("cargo") or "").strip()
    email = (data.get("email") or "").strip().lower()
    telefono = (data.get("telefono") or "").strip()
    mensaje = (data.get("mensaje") or "").strip()
    source_page = _sanitize_public_page(str(data.get("source_page") or "funcionalidades"))

    if len(nombre) < 2:
        return JSONResponse({"success": False, "error": "Nombre requerido"}, status_code=400)
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", email):
        return JSONResponse({"success": False, "error": "Correo electrónico inválido"}, status_code=400)
    if len(mensaje) < 10:
        return JSONResponse({"success": False, "error": "Describe brevemente tu requerimiento"}, status_code=400)

    db = SessionLocal()
    try:
        geo = _public_client_location(request)
        record = PublicLeadRequest(
            nombre=nombre,
            organizacion=organizacion,
            cargo=cargo,
            email=email,
            telefono=telefono,
            mensaje=mensaje,
            source_page=source_page,
            ip_address=_public_client_ip(request),
            user_agent=request.headers.get("user-agent") or "",
            country=geo["country"],
            region=geo["region"],
            city=geo["city"],
        )
        db.add(record)
        db.commit()
        return JSONResponse(
            {
                "success": True,
                "message": "Gracias por tu interés. Nuestro equipo te contactará pronto.",
            }
        )
    except Exception as exc:
        db.rollback()
        return JSONResponse({"success": False, "error": str(exc)}, status_code=500)
    finally:
        db.close()


@app.post("/api/public/quiz-discount")
def public_quiz_discount(request: Request, data: dict = Body(default={})):
    nombre = (data.get("nombre") or "").strip()
    cooperativa = (data.get("cooperativa") or "").strip()
    pais = (data.get("pais") or "").strip()
    celular = (data.get("celular") or "").strip()
    raw_answers = data.get("answers") if isinstance(data.get("answers"), dict) else {}
    tenant_id = _normalize_tenant_id(str(data.get("tenant_id") or "default"))

    if len(nombre) < 2:
        return JSONResponse({"success": False, "error": "Nombre requerido"}, status_code=400)
    if len(cooperativa) < 2:
        return JSONResponse({"success": False, "error": "Cooperativa requerida"}, status_code=400)
    if len(pais) < 2:
        return JSONResponse({"success": False, "error": "País requerido"}, status_code=400)
    if len(celular) < 6:
        return JSONResponse({"success": False, "error": "Celular requerido"}, status_code=400)

    normalized_answers: Dict[str, str] = {}
    for key in QUIZ_CORRECT_ANSWERS.keys():
        value = str(raw_answers.get(key, "")).strip().lower()
        normalized_answers[key] = value
    answered = sum(1 for value in normalized_answers.values() if value)
    if answered < len(QUIZ_CORRECT_ANSWERS):
        return JSONResponse({"success": False, "error": "Responde las 10 preguntas"}, status_code=400)

    correct_count = sum(
        1 for key, expected in QUIZ_CORRECT_ANSWERS.items()
        if normalized_answers.get(key, "") == expected
    )
    discount = _quiz_discount_by_correct(correct_count)

    db = SessionLocal()
    try:
        current_count = db.query(PublicQuizSubmission).count()
        available_slots = max(0, 5 - int(current_count))
        promo_enabled = available_slots > 0
        if not promo_enabled:
            discount = 0

        record = PublicQuizSubmission(
            tenant_id=tenant_id,
            nombre=nombre,
            cooperativa=cooperativa,
            pais=pais,
            celular=celular,
            correctas=int(correct_count),
            descuento=int(discount),
            total_preguntas=len(QUIZ_CORRECT_ANSWERS),
            answers=normalized_answers,
            ip_address=_public_client_ip(request),
            user_agent=request.headers.get("user-agent") or "",
        )
        db.add(record)
        db.commit()
        db.refresh(record)
        return JSONResponse(
            {
                "success": True,
                "data": {
                    "id": record.id,
                    "correctas": int(correct_count),
                    "total": len(QUIZ_CORRECT_ANSWERS),
                    "descuento": int(discount),
                    "promo_aplicada": bool(promo_enabled),
                    "cupos_restantes": max(0, available_slots - 1),
                },
                "message": (
                    "Cuestionario enviado. Tu resultado fue calculado correctamente."
                    if promo_enabled
                    else "Cuestionario enviado. El cupo promocional de descuento ya fue cubierto."
                ),
            }
        )
    except Exception as exc:
        db.rollback()
        return JSONResponse({"success": False, "error": str(exc)}, status_code=500)
    finally:
        db.close()


@app.get("/web/login", response_class=HTMLResponse)
def web_login(request: Request):
    login_identity = _get_login_identity_context()
    return templates.TemplateResponse(
        "web_login.html",
        {
            "request": request,
            "title": "Login",
            "login_error": "",
            **login_identity,
        },
    )


@app.get("/web/404", response_class=HTMLResponse)
def web_not_found(request: Request):
    return templates.TemplateResponse(
        "not_found.html",
        _not_found_context(request),
    )


@app.post("/web/login")
def web_login_submit(
    request: Request,
    usuario: str = Form(""),
    contrasena: str = Form(""),
    codigo_autenticador: str = Form(""),
):
    login_identity = _get_login_identity_context()
    if _is_login_rate_limited(request):
        return templates.TemplateResponse(
            "web_login.html",
            {
                "request": request,
                "title": "Login",
                "login_error": "Demasiados intentos. Intenta de nuevo en unos minutos.",
                **login_identity,
            },
            status_code=429,
        )
    username = usuario.strip()
    password = contrasena or ""
    if not username or not password:
        _register_failed_login_attempt(request)
        return templates.TemplateResponse(
            "web_login.html",
            {
                "request": request,
                "title": "Login",
                "login_error": "Datos incorrectos, vuelva a intentarlo",
                **login_identity,
            },
            status_code=401,
        )

    db = SessionLocal()
    has_passkey = False
    totp_secret = ""
    try:
        user = _find_user_by_login(db, username)
        if not user or not verify_password(password, user.contrasena or ""):
            _register_failed_login_attempt(request)
            return templates.TemplateResponse(
                "web_login.html",
                {
                    "request": request,
                    "title": "Login",
                    "login_error": "Datos incorrectos, vuelva a intentarlo",
                    **login_identity,
                },
                status_code=401,
            )

        role_name = _resolve_user_role_name(db, user)
        session_username = _decrypt_sensitive(user.usuario) or username
        has_passkey = bool(user.webauthn_credential_id and user.webauthn_public_key)
        totp_secret = _get_user_totp_secret(user, role_name)
    finally:
        db.close()

    _clear_failed_login_attempts(request)
    if role_name == "autoridades":
        code_value = re.sub(r"\s+", "", codigo_autenticador or "")
        if code_value:
            if not totp_secret:
                _register_failed_login_attempt(request)
                return templates.TemplateResponse(
                    "web_login.html",
                    {
                        "request": request,
                        "title": "Login",
                        "login_error": "El código autenticador no está configurado para este usuario.",
                        **login_identity,
                    },
                    status_code=403,
                )
            if not _verify_totp_code(totp_secret, code_value):
                _register_failed_login_attempt(request)
                return templates.TemplateResponse(
                    "web_login.html",
                    {
                        "request": request,
                        "title": "Login",
                        "login_error": "Código de autenticador inválido.",
                        **login_identity,
                    },
                    status_code=401,
                )
            response = RedirectResponse(url="/inicio", status_code=303)
            _apply_auth_cookies(response, request, session_username, role_name)
            response.delete_cookie(PASSKEY_COOKIE_MFA_GATE)
            return response

        if not has_passkey and not totp_secret:
            return templates.TemplateResponse(
                "web_login.html",
                {
                    "request": request,
                    "title": "Login",
                    "login_error": "El rol Autoridades requiere segundo factor (biometría o autenticador) configurado.",
                    **login_identity,
                },
                status_code=403,
            )
        if has_passkey:
            response = RedirectResponse(url=f"/web/login?mfa=required&usuario={quote(username)}", status_code=303)
            response.set_cookie(
                PASSKEY_COOKIE_MFA_GATE,
                _build_mfa_gate_token(user.id),
                httponly=True,
                samesite="lax",
                secure=COOKIE_SECURE,
                max_age=PASSKEY_CHALLENGE_TTL_SECONDS,
            )
            return response
        return templates.TemplateResponse(
            "web_login.html",
            {
                "request": request,
                "title": "Login",
                "login_error": "Ingresa tu código de autenticador para completar el acceso.",
                **login_identity,
            },
            status_code=401,
        )

    response = RedirectResponse(url="/inicio", status_code=303)
    _apply_auth_cookies(response, request, session_username, role_name)
    response.delete_cookie(PASSKEY_COOKIE_MFA_GATE)
    return response


@app.post("/web/passkey/register/options")
def passkey_register_options(
    request: Request,
    payload: dict = Body(default={}),
):
    username = str(payload.get("usuario", "")).strip()
    password = str(payload.get("contrasena", ""))
    if not username or not password:
        return JSONResponse({"success": False, "error": "Usuario y contraseña son obligatorios"}, status_code=400)
    if _is_demo_account(username):
        return JSONResponse(
            {"success": False, "error": "La biometría no está habilitada para el usuario demo"},
            status_code=403,
        )

    db = SessionLocal()
    try:
        user = _find_user_by_login(db, username)
        if not user or not verify_password(password, user.contrasena or ""):
            return JSONResponse({"success": False, "error": "Credenciales inválidas"}, status_code=401)
        username_plain = _decrypt_sensitive(user.usuario) or username
        display_name = (user.nombre or "").strip() or username_plain
        challenge = _b64url_encode(secrets.token_bytes(32))
        rp_id = _passkey_rp_id(request)
        origin = _passkey_origin(request)
        token = _build_passkey_token("register", user.id, challenge, rp_id, origin)
        options: Dict[str, Any] = {
            "challenge": challenge,
            "rp": {"name": "SIPET", "id": rp_id},
            "user": {
                "id": _b64url_encode(f"user:{user.id}".encode("utf-8")),
                "name": username_plain,
                "displayName": display_name,
            },
            "pubKeyCredParams": [{"type": "public-key", "alg": -7}],
            "timeout": 60000,
            "attestation": "none",
            "authenticatorSelection": {
                "authenticatorAttachment": "platform",
                "residentKey": "preferred",
                "userVerification": "preferred",
            },
        }
        if user.webauthn_credential_id:
            options["excludeCredentials"] = [
                {
                    "id": user.webauthn_credential_id,
                    "type": "public-key",
                    "transports": ["internal"],
                }
            ]
    finally:
        db.close()

    response = JSONResponse({"success": True, "options": options})
    response.set_cookie(
        PASSKEY_COOKIE_REGISTER,
        token,
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=PASSKEY_CHALLENGE_TTL_SECONDS,
    )
    return response


@app.post("/web/passkey/register/verify")
def passkey_register_verify(
    request: Request,
    payload: dict = Body(default={}),
):
    token_data = _read_passkey_token(request.cookies.get(PASSKEY_COOKIE_REGISTER, ""), "register")
    if not token_data:
        return JSONResponse({"success": False, "error": "Solicitud biométrica expirada, inténtalo de nuevo"}, status_code=400)

    credential_id = str(payload.get("id", "")).strip()
    response_payload = payload.get("response") or {}
    if not credential_id or not isinstance(response_payload, dict):
        return JSONResponse({"success": False, "error": "Respuesta biométrica inválida"}, status_code=400)

    client_data = _parse_client_data(str(response_payload.get("clientDataJSON", "")))
    public_key_b64 = str(response_payload.get("publicKey", "")).strip()
    if not client_data or not public_key_b64:
        return JSONResponse({"success": False, "error": "No se pudo registrar la clave biométrica"}, status_code=400)
    if str(client_data.get("type", "")) != "webauthn.create":
        return JSONResponse({"success": False, "error": "Tipo de autenticación no válido"}, status_code=400)
    if str(client_data.get("challenge", "")) != token_data["c"]:
        return JSONResponse({"success": False, "error": "Desafío biométrico inválido"}, status_code=400)
    if str(client_data.get("origin", "")).rstrip("/") != token_data["o"].rstrip("/"):
        return JSONResponse({"success": False, "error": "Origen no permitido para biometría"}, status_code=400)

    try:
        public_key_der = _b64url_decode(public_key_b64)
        serialization.load_der_public_key(public_key_der)
    except Exception:
        return JSONResponse({"success": False, "error": "Llave pública biométrica inválida"}, status_code=400)

    db = SessionLocal()
    try:
        user = db.query(Usuario).filter(Usuario.id == token_data["u"]).first()
        if not user:
            return JSONResponse({"success": False, "error": "Usuario no encontrado"}, status_code=404)
        user.webauthn_credential_id = credential_id
        user.webauthn_public_key = _b64url_encode(public_key_der)
        user.webauthn_sign_count = 0
        db.add(user)
        db.commit()
    except Exception:
        db.rollback()
        return JSONResponse({"success": False, "error": "No se pudo guardar la biometría"}, status_code=500)
    finally:
        db.close()

    response = JSONResponse({"success": True, "message": "Biometría registrada correctamente"})
    response.delete_cookie(PASSKEY_COOKIE_REGISTER)
    return response


@app.post("/web/passkey/auth/options")
def passkey_auth_options(
    request: Request,
    payload: dict = Body(default={}),
):
    username = str(payload.get("usuario", "")).strip()
    if not username:
        return JSONResponse({"success": False, "error": "Ingresa tu usuario para autenticar con biometría"}, status_code=400)
    if _is_demo_account(username):
        return JSONResponse(
            {"success": False, "error": "La biometría no está habilitada para el usuario demo"},
            status_code=403,
        )

    db = SessionLocal()
    try:
        user = _find_user_by_login(db, username)
        if not user or not user.webauthn_credential_id or not user.webauthn_public_key:
            return JSONResponse({"success": False, "error": "Este usuario no tiene biometría registrada"}, status_code=404)
        role_name = _resolve_user_role_name(db, user)
        if role_name == "autoridades":
            gate_user_id = _read_mfa_gate_token(request.cookies.get(PASSKEY_COOKIE_MFA_GATE, ""))
            if not gate_user_id or gate_user_id != user.id:
                return JSONResponse(
                    {"success": False, "error": "Primero valida usuario y contraseña para continuar con doble autenticación"},
                    status_code=403,
                )
        challenge = _b64url_encode(secrets.token_bytes(32))
        rp_id = _passkey_rp_id(request)
        origin = _passkey_origin(request)
        token = _build_passkey_token("auth", user.id, challenge, rp_id, origin)
        options = {
            "challenge": challenge,
            "rpId": rp_id,
            "allowCredentials": [
                {
                    "id": user.webauthn_credential_id,
                    "type": "public-key",
                    "transports": ["internal"],
                }
            ],
            "timeout": 60000,
            "userVerification": "preferred",
        }
    finally:
        db.close()

    response = JSONResponse({"success": True, "options": options})
    response.set_cookie(
        PASSKEY_COOKIE_AUTH,
        token,
        httponly=True,
        samesite="lax",
        secure=COOKIE_SECURE,
        max_age=PASSKEY_CHALLENGE_TTL_SECONDS,
    )
    return response


@app.post("/web/passkey/auth/verify")
def passkey_auth_verify(
    request: Request,
    payload: dict = Body(default={}),
):
    token_data = _read_passkey_token(request.cookies.get(PASSKEY_COOKIE_AUTH, ""), "auth")
    if not token_data:
        return JSONResponse({"success": False, "error": "Solicitud biométrica expirada, inténtalo de nuevo"}, status_code=400)

    credential_id = str(payload.get("id", "")).strip()
    response_payload = payload.get("response") or {}
    if not credential_id or not isinstance(response_payload, dict):
        return JSONResponse({"success": False, "error": "Respuesta biométrica inválida"}, status_code=400)

    client_data_b64 = str(response_payload.get("clientDataJSON", "")).strip()
    auth_data_b64 = str(response_payload.get("authenticatorData", "")).strip()
    signature_b64 = str(response_payload.get("signature", "")).strip()
    if not client_data_b64 or not auth_data_b64 or not signature_b64:
        return JSONResponse({"success": False, "error": "Datos biométricos incompletos"}, status_code=400)

    client_data = _parse_client_data(client_data_b64)
    if not client_data:
        return JSONResponse({"success": False, "error": "No se pudo leer la respuesta del autenticador"}, status_code=400)
    if str(client_data.get("type", "")) != "webauthn.get":
        return JSONResponse({"success": False, "error": "Tipo de autenticación no válido"}, status_code=400)
    if str(client_data.get("challenge", "")) != token_data["c"]:
        return JSONResponse({"success": False, "error": "Desafío biométrico inválido"}, status_code=400)
    if str(client_data.get("origin", "")).rstrip("/") != token_data["o"].rstrip("/"):
        return JSONResponse({"success": False, "error": "Origen no permitido para biometría"}, status_code=400)

    try:
        authenticator_data = _b64url_decode(auth_data_b64)
        signature = _b64url_decode(signature_b64)
    except ValueError:
        return JSONResponse({"success": False, "error": "Formato biométrico inválido"}, status_code=400)

    if len(authenticator_data) < 37:
        return JSONResponse({"success": False, "error": "AuthenticatorData inválido"}, status_code=400)

    expected_rp_hash = hashlib.sha256(token_data["r"].encode("utf-8")).digest()
    rp_hash = authenticator_data[:32]
    flags = authenticator_data[32]
    sign_count = int.from_bytes(authenticator_data[33:37], "big")
    if not hmac.compare_digest(rp_hash, expected_rp_hash):
        return JSONResponse({"success": False, "error": "RP ID inválido para biometría"}, status_code=400)
    if not (flags & 0x01):
        return JSONResponse({"success": False, "error": "Se requiere presencia del usuario"}, status_code=400)

    client_data_hash = hashlib.sha256(client_data["_raw_bytes"]).digest()
    signed_payload = authenticator_data + client_data_hash

    db = SessionLocal()
    try:
        user = db.query(Usuario).filter(Usuario.id == token_data["u"]).first()
        if not user or not user.webauthn_credential_id or not user.webauthn_public_key:
            return JSONResponse({"success": False, "error": "Usuario sin biometría registrada"}, status_code=404)
        if user.webauthn_credential_id != credential_id:
            return JSONResponse({"success": False, "error": "Credencial biométrica no coincide"}, status_code=401)

        try:
            public_key = serialization.load_der_public_key(_b64url_decode(user.webauthn_public_key))
        except Exception:
            return JSONResponse({"success": False, "error": "Llave biométrica inválida"}, status_code=400)

        try:
            if isinstance(public_key, ec.EllipticCurvePublicKey):
                public_key.verify(signature, signed_payload, ec.ECDSA(hashes.SHA256()))
            elif isinstance(public_key, rsa.RSAPublicKey):
                public_key.verify(signature, signed_payload, padding.PKCS1v15(), hashes.SHA256())
            else:
                return JSONResponse({"success": False, "error": "Tipo de llave biométrica no soportado"}, status_code=400)
        except InvalidSignature:
            return JSONResponse({"success": False, "error": "Firma biométrica inválida"}, status_code=401)

        stored_sign_count = int(user.webauthn_sign_count or 0)
        if sign_count > 0 and stored_sign_count > 0 and sign_count <= stored_sign_count:
            return JSONResponse({"success": False, "error": "Contador biométrico inválido"}, status_code=401)
        if sign_count > stored_sign_count:
            user.webauthn_sign_count = sign_count
            db.add(user)
            db.commit()

        role_name = _resolve_user_role_name(db, user)
        session_username = _decrypt_sensitive(user.usuario) or _decrypt_sensitive(user.correo) or f"user-{user.id}"
    finally:
        db.close()

    response = JSONResponse({"success": True, "redirect": "/inicio"})
    _apply_auth_cookies(response, request, session_username, role_name)
    response.delete_cookie(PASSKEY_COOKIE_AUTH)
    response.delete_cookie(PASSKEY_COOKIE_MFA_GATE)
    return response


@app.get("/logout")
@app.get("/logout/")
def logout():
    response = RedirectResponse(url="/web/login", status_code=303)
    response.delete_cookie(AUTH_COOKIE_NAME)
    response.delete_cookie("user_role")
    response.delete_cookie("user_name")
    response.delete_cookie("tenant_id")
    response.delete_cookie(PASSKEY_COOKIE_AUTH)
    response.delete_cookie(PASSKEY_COOKIE_REGISTER)
    response.delete_cookie(PASSKEY_COOKIE_MFA_GATE)
    return response


def get_colores_context() -> Dict[str, str]:
    db = SessionLocal()
    colores = {c.key: c.value for c in db.query(Colores).all()}
    db.close()
    return colores


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


def slugify_value(value: str) -> str:
    base = (value or "").strip().lower()
    base = re.sub(r"[^a-z0-9]+", "-", base)
    base = re.sub(r"-+", "-", base).strip("-")
    return base or secrets.token_hex(4)


class FormFieldCreateSchema(BaseModel):
    field_type: str
    label: str
    name: str
    placeholder: Optional[str] = None
    help_text: Optional[str] = None
    default_value: Optional[str] = None
    is_required: bool = False
    validation_rules: Dict[str, Any] = Field(default_factory=dict)
    options: List[Dict[str, Any]] = Field(default_factory=list)
    order: int = 0
    conditional_logic: Dict[str, Any] = Field(default_factory=dict)


class FormDefinitionCreateSchema(BaseModel):
    name: str
    description: Optional[str] = None
    slug: Optional[str] = None
    tenant_id: Optional[str] = None
    is_active: bool = True
    config: Dict[str, Any] = Field(default_factory=dict)
    allowed_roles: List[str] = Field(default_factory=list)
    fields: List[FormFieldCreateSchema] = Field(default_factory=list)


class FormFieldResponseSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    form_id: int
    field_type: str
    label: str
    name: str
    placeholder: Optional[str]
    help_text: Optional[str]
    default_value: Optional[str]
    is_required: bool
    validation_rules: Dict[str, Any]
    options: List[Dict[str, Any]]
    order: int
    conditional_logic: Dict[str, Any]


class FormDefinitionResponseSchema(BaseModel):
    model_config = ConfigDict(from_attributes=True)

    id: int
    name: str
    slug: str
    tenant_id: str
    description: Optional[str]
    config: Dict[str, Any]
    allowed_roles: List[str]
    fields: List[FormFieldResponseSchema]
    is_active: bool


class FormRenderer:
    """Servicio para renderizar y validar formularios de manera dinámica."""

    @staticmethod
    def generate_pydantic_model(
        form_definition: FormDefinition,
        visible_field_names: Optional[Set[str]] = None,
    ):
        fields: Dict[str, Any] = {}
        for field in sorted(form_definition.fields, key=lambda item: item.order or 0):
            normalized_field_type = (field.field_type or "").strip().lower()
            if normalized_field_type in NON_DATA_FIELD_TYPES:
                continue
            field_type = FormRenderer._get_pydantic_type(field.field_type)
            is_visible = True if visible_field_names is None else field.name in visible_field_names
            if field.default_value not in (None, ""):
                default_value: Any = field.default_value
            elif field.is_required and is_visible:
                default_value = ...
            else:
                default_value = None
            fields[field.name] = (
                field_type,
                Field(default=default_value, description=field.help_text or ""),
            )

        model_name = f"FormModel_{form_definition.id}"
        return create_model(model_name, **fields)

    @staticmethod
    def _get_pydantic_type(field_type: str):
        type_map = {
            "text": str,
            "email": str,
            "password": str,
            "textarea": str,
            "number": float,
            "decimal": float,
            "integer": int,
            "checkbox": bool,
            "checkboxes": List[str],
            "select": str,
            "radio": str,
            "likert": int,
            "date": str,
            "time": str,
            "daterange": List[str],
            "signature": str,
            "url": str,
            "file": str,
            "header": str,
            "paragraph": str,
            "html": str,
            "divider": str,
            "pagebreak": str,
        }
        return type_map.get((field_type or "").strip().lower(), str)

    @staticmethod
    def _to_comparable(value: Any) -> Any:
        if isinstance(value, bool):
            return value
        if value is None:
            return None
        text = str(value).strip()
        low = text.lower()
        if low in {"true", "1", "yes", "si", "on"}:
            return True
        if low in {"false", "0", "no", "off"}:
            return False
        return text

    @staticmethod
    def _evaluate_single_condition(rule: Dict[str, Any], answers: Dict[str, Any]) -> bool:
        source_field = (
            rule.get("field")
            or rule.get("source")
            or rule.get("depends_on")
            or rule.get("name")
            or ""
        )
        source_field = str(source_field).strip()
        if not source_field:
            return True

        left = FormRenderer._to_comparable(answers.get(source_field))
        right = FormRenderer._to_comparable(rule.get("value"))
        operator = str(rule.get("operator") or rule.get("op") or "equals").strip().lower()

        if operator in {"equals", "eq", "=="}:
            return left == right
        if operator in {"not_equals", "neq", "!=", "<>"}:
            return left != right
        if operator in {"contains"}:
            if left is None:
                return False
            return str(right or "") in str(left)
        if operator in {"in"}:
            candidates = rule.get("values")
            if not isinstance(candidates, list):
                candidates = [rule.get("value")]
            normalized = [FormRenderer._to_comparable(item) for item in candidates]
            return left in normalized
        if operator in {"not_in"}:
            candidates = rule.get("values")
            if not isinstance(candidates, list):
                candidates = [rule.get("value")]
            normalized = [FormRenderer._to_comparable(item) for item in candidates]
            return left not in normalized
        if operator in {"truthy", "is_true"}:
            return bool(left)
        if operator in {"falsy", "is_false"}:
            return not bool(left)
        if operator in {"greater_than", "gt", ">"}:
            try:
                return float(left) > float(right)
            except (TypeError, ValueError):
                return False
        if operator in {"greater_or_equal", "gte", ">="}:
            try:
                return float(left) >= float(right)
            except (TypeError, ValueError):
                return False
        if operator in {"less_than", "lt", "<"}:
            try:
                return float(left) < float(right)
            except (TypeError, ValueError):
                return False
        if operator in {"less_or_equal", "lte", "<="}:
            try:
                return float(left) <= float(right)
            except (TypeError, ValueError):
                return False
        return True

    @staticmethod
    def evaluate_visibility(condition: Dict[str, Any], answers: Dict[str, Any]) -> bool:
        if not isinstance(condition, dict) or not condition:
            return True

        if "show_if" in condition and isinstance(condition.get("show_if"), dict):
            return FormRenderer.evaluate_visibility(condition["show_if"], answers)
        if "hide_if" in condition and isinstance(condition.get("hide_if"), dict):
            return not FormRenderer.evaluate_visibility(condition["hide_if"], answers)

        if isinstance(condition.get("all"), list):
            return all(
                FormRenderer.evaluate_visibility(item, answers)
                for item in condition["all"]
                if isinstance(item, dict)
            )
        if isinstance(condition.get("any"), list):
            return any(
                FormRenderer.evaluate_visibility(item, answers)
                for item in condition["any"]
                if isinstance(item, dict)
            )

        return FormRenderer._evaluate_single_condition(condition, answers)

    @staticmethod
    def visible_field_names(form_definition: FormDefinition, answers: Dict[str, Any]) -> Set[str]:
        visible: Set[str] = set()
        ordered_fields = sorted(form_definition.fields, key=lambda item: item.order or 0)
        # Evalua en orden para que las condiciones puedan depender de campos anteriores.
        for field in ordered_fields:
            condition = field.conditional_logic or {}
            if FormRenderer.evaluate_visibility(condition, answers):
                visible.add(field.name)
        return visible

    @staticmethod
    def _apply_custom_rules(
        form_definition: FormDefinition,
        data: Dict[str, Any],
        visible_field_names: Optional[Set[str]] = None,
    ) -> None:
        for field in form_definition.fields:
            if visible_field_names is not None and field.name not in visible_field_names:
                continue
            rules = field.validation_rules or {}
            value = data.get(field.name)
            if value is None:
                continue

            normalized_field_type = (field.field_type or "").strip().lower()
            if normalized_field_type in NON_DATA_FIELD_TYPES:
                continue
            if normalized_field_type == "daterange":
                if not isinstance(value, list):
                    raise ValueError(f"'{field.label}' debe ser un rango de fechas")
                cleaned = [str(item).strip() for item in value if str(item).strip()]
                if len(cleaned) != 2:
                    raise ValueError(f"'{field.label}' requiere fecha inicio y fecha fin")
                data[field.name] = cleaned
                value = cleaned
            elif normalized_field_type == "url":
                parsed = urlparse(str(value).strip())
                if parsed.scheme not in {"http", "https"} or not parsed.netloc:
                    raise ValueError(f"'{field.label}' debe ser una URL válida")
            elif normalized_field_type == "signature":
                signature_value = str(value).strip()
                if field.is_required and not signature_value:
                    raise ValueError(f"'{field.label}' es obligatoria")
                if signature_value and not (
                    signature_value.startswith("data:image/")
                    or signature_value.startswith("http://")
                    or signature_value.startswith("https://")
                ):
                    raise ValueError(f"'{field.label}' no tiene un formato de firma válido")

            if isinstance(value, str):
                if "min_length" in rules and len(value) < int(rules["min_length"]):
                    raise ValueError(f"'{field.label}' requiere longitud mínima de {rules['min_length']}")
                if "max_length" in rules and len(value) > int(rules["max_length"]):
                    raise ValueError(f"'{field.label}' supera la longitud máxima de {rules['max_length']}")
                pattern = rules.get("pattern")
                if pattern and not re.match(pattern, value):
                    raise ValueError(f"'{field.label}' tiene un formato inválido")
            elif isinstance(value, list):
                if "min_length" in rules and len(value) < int(rules["min_length"]):
                    raise ValueError(f"'{field.label}' requiere al menos {rules['min_length']} selección(es)")
                if "max_length" in rules and len(value) > int(rules["max_length"]):
                    raise ValueError(f"'{field.label}' supera el máximo de {rules['max_length']} selección(es)")

            if isinstance(value, (int, float)):
                if "min" in rules and value < float(rules["min"]):
                    raise ValueError(f"'{field.label}' es menor al mínimo permitido ({rules['min']})")
                if "max" in rules and value > float(rules["max"]):
                    raise ValueError(f"'{field.label}' excede el máximo permitido ({rules['max']})")

    @staticmethod
    def render_to_json(form_definition: FormDefinition) -> Dict[str, Any]:
        fields = []
        for field in sorted(form_definition.fields, key=lambda item: item.order or 0):
            fields.append(
                {
                    "type": field.field_type,
                    "name": field.name,
                    "label": field.label,
                    "placeholder": field.placeholder,
                    "helpText": field.help_text,
                    "required": field.is_required,
                    "defaultValue": field.default_value,
                    "validation": field.validation_rules or {},
                    "options": field.options if field.field_type in {"select", "radio", "checkboxes", "likert"} else [],
                    "conditional": field.conditional_logic or {},
                }
            )

        return {
            "id": form_definition.id,
            "name": form_definition.name,
            "slug": form_definition.slug,
            "tenant_id": _normalize_tenant_id(form_definition.tenant_id or "default"),
            "description": form_definition.description,
            "allowed_roles": form_definition.allowed_roles if isinstance(form_definition.allowed_roles, list) else [],
            "fields": fields,
            "config": form_definition.config or {},
        }


def _normalize_recipients(raw: Any) -> List[str]:
    if raw is None:
        return []
    if isinstance(raw, str):
        candidates = [part.strip() for part in re.split(r"[;,]", raw)]
    elif isinstance(raw, list):
        candidates = [str(item).strip() for item in raw]
    else:
        return []
    return [item for item in candidates if item and "@" in item]


def _to_bool(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "si", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return default


def _notification_email_settings(form_definition: FormDefinition) -> Dict[str, Any]:
    config = form_definition.config or {}
    notifications = config.get("notifications") if isinstance(config.get("notifications"), dict) else {}
    email_cfg = notifications.get("email") if isinstance(notifications.get("email"), dict) else {}
    enabled = _to_bool(
        email_cfg.get("enabled"),
        default=bool(email_cfg or config.get("notification_emails") or os.environ.get("FORM_NOTIFICATION_TO")),
    )
    to_list = _normalize_recipients(
        email_cfg.get("to")
        or config.get("notification_emails")
        or os.environ.get("FORM_NOTIFICATION_TO")
    )
    cc_list = _normalize_recipients(email_cfg.get("cc"))
    subject = (
        str(email_cfg.get("subject") or "").strip()
        or f"[SIPET] Nuevo envío: {form_definition.name}"
    )
    return {
        "enabled": enabled,
        "to": to_list,
        "cc": cc_list,
        "subject": subject,
    }


def send_form_submission_email_notification(
    form_definition: FormDefinition,
    submission: FormSubmission,
) -> Dict[str, Any]:
    settings = _notification_email_settings(form_definition)
    if not settings["enabled"]:
        return {"sent": False, "reason": "not_enabled"}
    recipients = [*settings["to"], *settings["cc"]]
    if not recipients:
        return {"sent": False, "reason": "no_recipients"}

    smtp_host = (os.environ.get("SMTP_HOST") or "").strip()
    smtp_port = int(os.environ.get("SMTP_PORT") or 587)
    smtp_user = (os.environ.get("SMTP_USER") or "").strip()
    smtp_password = os.environ.get("SMTP_PASSWORD") or ""
    smtp_from = (os.environ.get("SMTP_FROM") or smtp_user or "no-reply@sipet.local").strip()
    smtp_use_ssl = _to_bool(os.environ.get("SMTP_USE_SSL"), default=False)
    smtp_use_tls = _to_bool(os.environ.get("SMTP_USE_TLS"), default=not smtp_use_ssl)
    if not smtp_host:
        return {"sent": False, "reason": "smtp_not_configured"}

    message = EmailMessage()
    message["Subject"] = settings["subject"]
    message["From"] = smtp_from
    message["To"] = ", ".join(settings["to"])
    if settings["cc"]:
        message["Cc"] = ", ".join(settings["cc"])
    payload_text = json.dumps(submission.data or {}, ensure_ascii=False, indent=2)
    message.set_content(
        "\n".join(
            [
                "Se recibió un nuevo envío de formulario.",
                f"Formulario: {form_definition.name} ({form_definition.slug})",
                f"Submission ID: {submission.id}",
                f"Fecha: {submission.submitted_at.isoformat() if submission.submitted_at else ''}",
                f"IP: {submission.ip_address or ''}",
                f"User-Agent: {submission.user_agent or ''}",
                "",
                "Datos enviados:",
                payload_text,
            ]
        )
    )

    try:
        if smtp_use_ssl:
            with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=10) as server:
                if smtp_user:
                    server.login(smtp_user, smtp_password)
                server.send_message(message)
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=10) as server:
                if smtp_use_tls:
                    server.starttls()
                if smtp_user:
                    server.login(smtp_user, smtp_password)
                server.send_message(message)
    except Exception as exc:
        return {"sent": False, "reason": "smtp_error", "detail": str(exc)}
        return {"sent": True}


def _normalize_form_submission_payload(
    form_definition: FormDefinition,
    payload: Dict[str, Any],
) -> Dict[str, Any]:
    normalized = dict(payload)
    fields_by_name = {field.name: field for field in form_definition.fields}
    for field_name, field in fields_by_name.items():
        if field_name not in normalized:
            continue
        raw_value = normalized[field_name]
        field_type = (field.field_type or "").strip().lower()
        if field_type in NON_DATA_FIELD_TYPES:
            normalized.pop(field_name, None)
            continue

        if field_type == "file":
            if isinstance(raw_value, UploadFile):
                normalized[field_name] = raw_value.filename or ""
            elif isinstance(raw_value, list):
                filenames = []
                for item in raw_value:
                    if isinstance(item, UploadFile):
                        filenames.append(item.filename or "")
                    elif item is not None:
                        filenames.append(str(item))
                normalized[field_name] = filenames[0] if len(filenames) == 1 else filenames
            elif raw_value is None:
                normalized[field_name] = ""
            else:
                normalized[field_name] = str(raw_value)
            continue

        if field_type in {"checkboxes", "daterange"}:
            if isinstance(raw_value, list):
                normalized[field_name] = [str(item).strip() for item in raw_value if str(item).strip()]
            elif raw_value is None:
                normalized[field_name] = []
            else:
                cleaned = str(raw_value).strip()
                normalized[field_name] = [cleaned] if cleaned else []
            continue

        if isinstance(raw_value, list):
            normalized[field_name] = str(raw_value[0]) if raw_value else ""
        elif isinstance(raw_value, UploadFile):
            normalized[field_name] = raw_value.filename or ""
        elif raw_value is None:
            normalized[field_name] = ""
        else:
            normalized[field_name] = str(raw_value) if not isinstance(raw_value, (int, float, bool, dict)) else raw_value
    return normalized


def _normalize_webhook_configs(form_definition: FormDefinition) -> List[Dict[str, Any]]:
    config = form_definition.config or {}
    notifications = config.get("notifications") if isinstance(config.get("notifications"), dict) else {}
    raw_webhooks = notifications.get("webhooks", config.get("webhooks"))
    items: List[Any]
    if isinstance(raw_webhooks, list):
        items = raw_webhooks
    elif isinstance(raw_webhooks, dict):
        items = [raw_webhooks]
    elif isinstance(raw_webhooks, str):
        items = [raw_webhooks]
    else:
        items = []

    if not items:
        env_hook = (os.environ.get("FORM_WEBHOOK_URL") or "").strip()
        if env_hook:
            items = [env_hook]

    normalized: List[Dict[str, Any]] = []
    for item in items:
        if isinstance(item, str):
            url = item.strip()
            if not url:
                continue
            normalized.append(
                {
                    "url": url,
                    "method": "POST",
                    "headers": {},
                    "timeout": 10,
                    "payload_mode": "full",
                }
            )
            continue
        if not isinstance(item, dict):
            continue
        url = str(item.get("url") or "").strip()
        if not url:
            continue
        method = str(item.get("method") or "POST").strip().upper()
        if method not in {"POST", "PUT", "PATCH", "GET"}:
            method = "POST"
        headers = item.get("headers") if isinstance(item.get("headers"), dict) else {}
        timeout_raw = item.get("timeout")
        try:
            timeout = float(timeout_raw) if timeout_raw is not None else 10.0
        except (TypeError, ValueError):
            timeout = 10.0
        payload_mode = str(item.get("payload_mode") or "full").strip().lower()
        if payload_mode not in {"full", "data_only"}:
            payload_mode = "full"
        normalized.append(
            {
                "url": url,
                "method": method,
                "headers": headers,
                "timeout": timeout,
                "payload_mode": payload_mode,
            }
        )
    return normalized


def send_form_submission_webhooks(
    form_definition: FormDefinition,
    submission: FormSubmission,
) -> Dict[str, Any]:
    hooks = _normalize_webhook_configs(form_definition)
    if not hooks:
        return {"attempted": 0, "succeeded": 0, "failed": 0, "results": []}

    base_payload = {
        "event": "form_submission.created",
        "form": {
            "id": form_definition.id,
            "name": form_definition.name,
            "slug": form_definition.slug,
        },
        "submission": {
            "id": submission.id,
            "submitted_at": submission.submitted_at.isoformat() if submission.submitted_at else "",
            "ip_address": submission.ip_address,
            "user_agent": submission.user_agent,
        },
        "data": submission.data or {},
    }

    results: List[Dict[str, Any]] = []
    succeeded = 0
    for hook in hooks:
        payload = (
            base_payload.get("data", {})
            if hook["payload_mode"] == "data_only"
            else base_payload
        )
        try:
            if hook["method"] == "GET":
                response = httpx.request(
                    method=hook["method"],
                    url=hook["url"],
                    params=payload if isinstance(payload, dict) else {},
                    headers=hook["headers"],
                    timeout=hook["timeout"],
                )
            else:
                response = httpx.request(
                    method=hook["method"],
                    url=hook["url"],
                    json=payload,
                    headers=hook["headers"],
                    timeout=hook["timeout"],
                )
            ok = 200 <= response.status_code < 300
            if ok:
                succeeded += 1
            results.append(
                {
                    "url": hook["url"],
                    "method": hook["method"],
                    "status_code": response.status_code,
                    "ok": ok,
                }
            )
        except Exception as exc:
            results.append(
                {
                    "url": hook["url"],
                    "method": hook["method"],
                    "ok": False,
                    "error": str(exc),
                }
            )

    attempted = len(hooks)
    failed = attempted - succeeded
    return {"attempted": attempted, "succeeded": succeeded, "failed": failed, "results": results}


def _normalize_submission_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _build_submission_export_columns(
    form_definition: FormDefinition,
    submissions: List[FormSubmission],
) -> List[str]:
    field_names = [
        field.name
        for field in sorted(form_definition.fields, key=lambda item: item.order or 0)
        if field.name and (field.field_type or "").strip().lower() not in NON_DATA_FIELD_TYPES
    ]
    known = set(field_names)
    extra = []
    for submission in submissions:
        data = submission.data if isinstance(submission.data, dict) else {}
        for key in data.keys():
            if key and key not in known:
                known.add(key)
                extra.append(key)
    return [*field_names, *extra]


def render_backend_page(
    request: Request,
    title: str,
    description: str = "",
    content: str = "",
    subtitle: Optional[str] = None,
    hide_floating_actions: bool = True,
    view_buttons: Optional[List[Dict]] = None,
    view_buttons_html: str = "",
    floating_actions_html: str = "",
    floating_actions_screen: str = "personalization",
    show_page_header: bool = True,
) -> HTMLResponse:
    rendered_view_buttons = view_buttons_html or build_view_buttons_html(view_buttons)
    can_manage_personalization = is_superadmin(request)
    login_identity = _get_login_identity_context()
    context = {
        "request": request,
        "title": title,
        "content": content,
        "subtitle": subtitle,
        "page_title": title,
        "page_description": description,
        "hide_floating_actions": hide_floating_actions,
        "floating_actions_html": floating_actions_html,
        "floating_actions_screen": floating_actions_screen,
        "view_buttons_html": rendered_view_buttons,
        "show_page_header": show_page_header,
        "can_manage_personalization": can_manage_personalization,
        "app_favicon_url": login_identity.get("login_favicon_url"),
    }
    context.update({"colores": get_colores_context()})
    return templates.TemplateResponse("base.html", context)



# Almacenamiento simple en archivo para el contenido editable de Avance
AVANCE_CONTENT_FILE = "fastapi_modulo/avance_content.txt"
def get_avance_content():
    if os.path.exists(AVANCE_CONTENT_FILE):
        with open(AVANCE_CONTENT_FILE, "r", encoding="utf-8") as f:
            stored = f.read()
            if stored.strip():
                return stored
    return "<p>Sin contenido personalizado aún.</p><p>Inicio.<br>bienvenido al tablero</p>"

def set_avance_content(new_content: str):
    with open(AVANCE_CONTENT_FILE, "w", encoding="utf-8") as f:
        f.write(new_content)

@app.get("/avan", response_class=HTMLResponse)
def avan(request: Request, edit: Optional[bool] = False):
    # Solo mostrar contenido, sin edición
    meta = {"title": "Avance", "subtitle": "Progreso y métricas clave", "description": "Resumen del estado del sistema"}
    content = get_avance_content()
    return backend_screen(
        request,
        title=meta["title"],
        subtitle=meta["subtitle"],
        description=meta["description"],
        content=content,
        floating_buttons=None,
        hide_floating_actions=True,
    )




PERSONALIZACION_HTML = dedent("""
    <section class="personalization-panel" aria-labelledby="personalizacion-title">
        <div>
            <h2 id="personalizacion-title">Personalizar pantalla</h2>
            <p>Define los colores principales que se aplicarán en todo el sitio para mantener la identidad institucional.</p>
        </div>
        <div class="color-group">
            <h3>Navbar</h3>
            <div class="color-grid">
                <article class="color-option">
                    <label for="navbar-bg">Fondo</label>
                    <input type="color" id="navbar-bg" name="navbar-bg" value="#ffffff">
                    <div class="color-preview" id="navbar-bg-preview" style="background:#ffffff;"></div>
                </article>
                <article class="color-option">
                    <label for="navbar-text">Texto</label>
                    <input type="color" id="navbar-text" name="navbar-text" value="#0f172a">
                    <div class="color-preview" id="navbar-text-preview" style="background:#0f172a;"></div>
                </article>
            </div>
        </div>
        <div class="color-group">
            <h3>Sidebar</h3>
            <div class="color-grid">
                <article class="color-option">
                    <label for="sidebar-top">Color superior</label>
                    <input type="color" id="sidebar-top" name="sidebar-top" value="#1f2a3d">
                    <div class="color-preview" id="sidebar-top-preview" style="background:#1f2a3d;"></div>
                </article>
                <article class="color-option">
                    <label for="sidebar-bottom">Color inferior</label>
                    <input type="color" id="sidebar-bottom" name="sidebar-bottom" value="#0f172a">
                    <div class="color-preview" id="sidebar-bottom-preview" style="background:#0f172a;"></div>
                </article>
                <article class="color-option">
                    <label for="sidebar-text">Texto</label>
                    <input type="color" id="sidebar-text" name="sidebar-text" value="#ffffff">
                    <div class="color-preview" id="sidebar-text-preview" style="background:#ffffff;"></div>
                </article>
                <article class="color-option">
                    <label for="sidebar-icon">Iconos</label>
                    <input type="color" id="sidebar-icon" name="sidebar-icon" value="#f5f7fb">
                    <div class="color-preview" id="sidebar-icon-preview" style="background:#f5f7fb;"></div>
                </article>
                <article class="color-option">
                    <label for="sidebar-hover">Hover</label>
                    <input type="color" id="sidebar-hover" name="sidebar-hover" value="#2a3a52">
                    <div class="color-preview" id="sidebar-hover-preview" style="background:#2a3a52;"></div>
                </article>
                <article class="color-option">
                    <label for="field-color">Color del campo</label>
                    <input type="color" id="field-color" name="field-color" value="#e7d7da">
                    <div class="color-preview" id="field-color-preview" style="background:#e7d7da;"></div>
                </article>
            </div>
        </div>
        <div class="color-field-note">
            <p><strong>Clase disponible:</strong> usa <code>campo-personalizado</code> para aplicar el color de campo configurado por el usuario.</p>
        </div>
    </section>
""")

FODA_HTML = dedent("""
    <section class="foda-page pe-foda" id="foda-builder">
        <style>
            .pe-foda {
                --bg: #f6f8fc;
                --surface: rgba(255,255,255,.88);
                --text: #0f172a;
                --muted: #64748b;
                --border: rgba(148,163,184,.38);
                --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                --primary: #0f3d2e;
                --ok: #16a34a;
                --warn: #f59e0b;
                --crit: #ef4444;
                width: 100%;
                font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                color: var(--text);
                background:
                  radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                  radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                  var(--bg);
                border-radius: 18px;
                padding: 18px 0 34px;
            }
            .pe-foda .foda-input-card,
            .pe-foda .foda-quadrant {
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 22px;
                box-shadow: var(--shadow-soft);
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
            }
            .pe-foda .foda-input-card {
                padding: 16px;
            }
            .pe-foda .foda-input-card h2 {
                margin: 0;
                font-size: 20px;
                letter-spacing: -0.02em;
            }
            .pe-foda .foda-input-card p {
                margin: 6px 0 12px;
                color: var(--muted);
                font-size: 13px;
            }
            .pe-foda .foda-input-grid {
                grid-template-columns: 1.4fr .8fr .8fr;
                gap: 12px;
            }
            .pe-foda .foda-input-grid textarea,
            .pe-foda .foda-input-grid select {
                border: 1px solid var(--border);
                border-radius: 14px;
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-foda .foda-input-actions button {
                border-radius: 14px;
                padding: 10px 14px;
                font-weight: 800;
                border: 1px solid rgba(15,61,46,.35);
                background: linear-gradient(135deg, #0f3d2e, #1f6f52);
                color: #fff;
                cursor: pointer;
            }
            .pe-foda .foda-matrix {
                grid-template-columns: repeat(2, minmax(0, 1fr));
                gap: 14px;
                margin-top: 14px;
            }
            .pe-foda .foda-quadrant {
                padding: 14px;
            }
            .pe-foda .foda-quadrant header {
                margin-bottom: 10px;
            }
            .pe-foda .foda-quadrant header span {
                border-radius: 999px;
                padding: 6px 10px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.70);
            }
            .pe-foda .foda-quadrant ul {
                min-height: 120px;
                gap: 10px;
            }
            .pe-foda .foda-item {
                border-radius: 16px;
                border: 1px solid rgba(148,163,184,.30);
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-foda .foda-item button {
                border-radius: 999px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                padding: 6px 10px;
                font-size: 12px;
            }
            .pe-foda .foda-strengths { border-color: rgba(22,163,74,.28); }
            .pe-foda .foda-weaknesses { border-color: rgba(245,158,11,.30); }
            .pe-foda .foda-opportunities { border-color: rgba(37,99,235,.30); }
            .pe-foda .foda-threats { border-color: rgba(239,68,68,.30); }
            @media (max-width: 1100px) {
                .pe-foda .foda-input-grid { grid-template-columns: 1fr; }
            }
            @media (max-width: 640px) {
                .pe-foda .foda-matrix { grid-template-columns: 1fr; }
            }
        </style>

        <article class="foda-input-card">
            <h2>Matriz FODA</h2>
            <p>Registra factores y clasifícalos como fortalezas, debilidades, oportunidades o amenazas.</p>
            <div class="foda-input-grid">
                <div class="form-field">
                    <label for="foda-description">Descripción</label>
                    <textarea id="foda-description" placeholder="Describe el factor"></textarea>
                </div>
                <div class="form-field">
                    <label for="foda-category">Categoría</label>
                    <select id="foda-category">
                        <option value="strengths">Fortalezas</option>
                        <option value="weaknesses">Debilidades</option>
                        <option value="opportunities">Oportunidades</option>
                        <option value="threats">Amenazas</option>
                    </select>
                </div>
                <div class="form-field">
                    <label for="foda-impact">Impacto</label>
                    <select id="foda-impact">
                        <option value="high">Alto</option>
                        <option value="medium">Medio</option>
                        <option value="low">Bajo</option>
                    </select>
                </div>
            </div>
            <div class="foda-input-actions">
                <button type="button" id="foda-add">Agregar factor</button>
            </div>
        </article>
        <section class="foda-matrix">
            <article class="foda-quadrant foda-strengths">
                <header>
                    <h3>Fortalezas</h3>
                    <span id="foda-count-strengths">0</span>
                </header>
                <ul id="foda-list-strengths"></ul>
            </article>
            <article class="foda-quadrant foda-weaknesses">
                <header>
                    <h3>Debilidades</h3>
                    <span id="foda-count-weaknesses">0</span>
                </header>
                <ul id="foda-list-weaknesses"></ul>
            </article>
            <article class="foda-quadrant foda-opportunities">
                <header>
                    <h3>Oportunidades</h3>
                    <span id="foda-count-opportunities">0</span>
                </header>
                <ul id="foda-list-opportunities"></ul>
            </article>
            <article class="foda-quadrant foda-threats">
                <header>
                    <h3>Amenazas</h3>
                    <span id="foda-count-threats">0</span>
                </header>
                <ul id="foda-list-threats"></ul>
            </article>
        </section>
    </section>
""")

PESTEL_HTML = dedent("""
    <section class="pestel-page pe-pestel" id="pestel-builder">
        <style>
            .pe-pestel {
                --bg: #f6f8fc;
                --surface: rgba(255,255,255,.88);
                --text: #0f172a;
                --muted: #64748b;
                --border: rgba(148,163,184,.38);
                --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                width: 100%;
                font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                color: var(--text);
                background:
                  radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                  radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                  var(--bg);
                border-radius: 18px;
                padding: 18px 0 34px;
            }
            .pe-pestel .pestel-input-card,
            .pe-pestel .pestel-column {
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 22px;
                box-shadow: var(--shadow-soft);
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
            }
            .pe-pestel .pestel-input-card { padding: 16px; }
            .pe-pestel .pestel-input-card h2 { margin: 0; font-size: 20px; letter-spacing: -0.02em; }
            .pe-pestel .pestel-input-card p { margin: 6px 0 12px; color: var(--muted); font-size: 13px; }
            .pe-pestel .pestel-input-grid { grid-template-columns: 1.4fr .8fr .8fr; gap: 12px; }
            .pe-pestel .pestel-input-grid textarea,
            .pe-pestel .pestel-input-grid select {
                border: 1px solid var(--border);
                border-radius: 14px;
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-pestel .pestel-input-actions button {
                border-radius: 14px;
                padding: 10px 14px;
                font-weight: 800;
                border: 1px solid rgba(15,61,46,.35);
                background: linear-gradient(135deg, #0f3d2e, #1f6f52);
                color: #fff;
                cursor: pointer;
            }
            .pe-pestel .pestel-board {
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 14px;
                margin-top: 14px;
            }
            .pe-pestel .pestel-column { padding: 14px; }
            .pe-pestel .pestel-column header { margin-bottom: 10px; }
            .pe-pestel .pestel-column header span {
                border-radius: 999px;
                padding: 6px 10px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.70);
            }
            .pe-pestel .pestel-column ul { min-height: 120px; gap: 10px; }
            .pe-pestel .pestel-item {
                border-radius: 16px;
                border: 1px solid rgba(148,163,184,.30);
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-pestel .pestel-item button {
                border-radius: 999px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                padding: 6px 10px;
                font-size: 12px;
            }
            @media (max-width: 1100px) { .pe-pestel .pestel-input-grid { grid-template-columns: 1fr; } }
            @media (max-width: 900px) { .pe-pestel .pestel-board { grid-template-columns: 1fr; } }
        </style>
        <article class="pestel-input-card">
            <h2>Análisis PESTEL</h2>
            <p>Registra factores externos por dimensión: Política, Económica, Social, Tecnológica, Ecológica y Legal.</p>
            <div class="pestel-input-grid">
                <div class="form-field">
                    <label for="pestel-description">Factor</label>
                    <textarea id="pestel-description" placeholder="Describe el factor PESTEL"></textarea>
                </div>
                <div class="form-field">
                    <label for="pestel-factor">Dimensión</label>
                    <select id="pestel-factor">
                        <option value="political">Política</option>
                        <option value="economic">Económica</option>
                        <option value="social">Social</option>
                        <option value="technological">Tecnológica</option>
                        <option value="ecological">Ecológica</option>
                        <option value="legal">Legal</option>
                    </select>
                </div>
                <div class="form-field">
                    <label for="pestel-impact">Impacto</label>
                    <select id="pestel-impact">
                        <option value="high">Alto</option>
                        <option value="medium">Medio</option>
                        <option value="low">Bajo</option>
                    </select>
                </div>
            </div>
            <div class="pestel-input-actions">
                <button type="button" id="pestel-add">Agregar factor</button>
            </div>
        </article>
        <section class="pestel-board">
            <article class="pestel-column">
                <header><h3>Política</h3><span id="pestel-count-political">0</span></header>
                <ul id="pestel-list-political"></ul>
            </article>
            <article class="pestel-column">
                <header><h3>Económica</h3><span id="pestel-count-economic">0</span></header>
                <ul id="pestel-list-economic"></ul>
            </article>
            <article class="pestel-column">
                <header><h3>Social</h3><span id="pestel-count-social">0</span></header>
                <ul id="pestel-list-social"></ul>
            </article>
            <article class="pestel-column">
                <header><h3>Tecnológica</h3><span id="pestel-count-technological">0</span></header>
                <ul id="pestel-list-technological"></ul>
            </article>
            <article class="pestel-column">
                <header><h3>Ecológica</h3><span id="pestel-count-ecological">0</span></header>
                <ul id="pestel-list-ecological"></ul>
            </article>
            <article class="pestel-column">
                <header><h3>Legal</h3><span id="pestel-count-legal">0</span></header>
                <ul id="pestel-list-legal"></ul>
            </article>
        </section>
    </section>
""")

PORTER_HTML = dedent("""
    <section class="porter-page pe-porter" id="porter-builder">
        <style>
            .pe-porter {
                --bg: #f6f8fc;
                --surface: rgba(255,255,255,.88);
                --text: #0f172a;
                --muted: #64748b;
                --border: rgba(148,163,184,.38);
                --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                width: 100%;
                font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                color: var(--text);
                background:
                  radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                  radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                  var(--bg);
                border-radius: 18px;
                padding: 18px 0 34px;
            }
            .pe-porter .porter-input-card,
            .pe-porter .porter-column {
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 22px;
                box-shadow: var(--shadow-soft);
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
            }
            .pe-porter .porter-input-card { padding: 16px; }
            .pe-porter .porter-input-card h2 { margin: 0; font-size: 20px; letter-spacing: -0.02em; }
            .pe-porter .porter-input-card p { margin: 6px 0 12px; color: var(--muted); font-size: 13px; }
            .pe-porter .porter-input-grid { grid-template-columns: 1.4fr .8fr .8fr; gap: 12px; }
            .pe-porter .porter-input-grid textarea,
            .pe-porter .porter-input-grid select {
                border: 1px solid var(--border);
                border-radius: 14px;
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-porter .porter-input-actions button {
                border-radius: 14px;
                padding: 10px 14px;
                font-weight: 800;
                border: 1px solid rgba(15,61,46,.35);
                background: linear-gradient(135deg, #0f3d2e, #1f6f52);
                color: #fff;
                cursor: pointer;
            }
            .pe-porter .porter-board {
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 14px;
                margin-top: 14px;
            }
            .pe-porter .porter-column { padding: 14px; }
            .pe-porter .porter-column header { margin-bottom: 10px; }
            .pe-porter .porter-column header span {
                border-radius: 999px;
                padding: 6px 10px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.70);
            }
            .pe-porter .porter-column ul { min-height: 120px; gap: 10px; }
            .pe-porter .porter-item {
                border-radius: 16px;
                border: 1px solid rgba(148,163,184,.30);
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-porter .porter-item button {
                border-radius: 999px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                padding: 6px 10px;
                font-size: 12px;
            }
            @media (max-width: 1100px) { .pe-porter .porter-input-grid { grid-template-columns: 1fr; } }
            @media (max-width: 900px) { .pe-porter .porter-board { grid-template-columns: 1fr; } }
        </style>
        <article class="porter-input-card">
            <h2>5 Fuerzas de Porter</h2>
            <p>Analiza la intensidad competitiva del entorno y su impacto en la estrategia.</p>
            <div class="porter-input-grid">
                <div class="form-field">
                    <label for="porter-description">Factor</label>
                    <textarea id="porter-description" placeholder="Describe el factor de la fuerza"></textarea>
                </div>
                <div class="form-field">
                    <label for="porter-force">Fuerza</label>
                    <select id="porter-force">
                        <option value="competitors">Rivalidad entre competidores</option>
                        <option value="entrants">Amenaza de nuevos entrantes</option>
                        <option value="suppliers">Poder de negociación de proveedores</option>
                        <option value="buyers">Poder de negociación de compradores</option>
                        <option value="substitutes">Amenaza de productos sustitutos</option>
                    </select>
                </div>
                <div class="form-field">
                    <label for="porter-impact">Impacto</label>
                    <select id="porter-impact">
                        <option value="high">Alto</option>
                        <option value="medium">Medio</option>
                        <option value="low">Bajo</option>
                    </select>
                </div>
            </div>
            <div class="porter-input-actions">
                <button type="button" id="porter-add">Agregar factor</button>
            </div>
        </article>
        <section class="porter-board">
            <article class="porter-column">
                <header><h3>Competidores</h3><span id="porter-count-competitors">0</span></header>
                <ul id="porter-list-competitors"></ul>
            </article>
            <article class="porter-column">
                <header><h3>Nuevos entrantes</h3><span id="porter-count-entrants">0</span></header>
                <ul id="porter-list-entrants"></ul>
            </article>
            <article class="porter-column">
                <header><h3>Proveedores</h3><span id="porter-count-suppliers">0</span></header>
                <ul id="porter-list-suppliers"></ul>
            </article>
            <article class="porter-column">
                <header><h3>Compradores</h3><span id="porter-count-buyers">0</span></header>
                <ul id="porter-list-buyers"></ul>
            </article>
            <article class="porter-column">
                <header><h3>Sustitutos</h3><span id="porter-count-substitutes">0</span></header>
                <ul id="porter-list-substitutes"></ul>
            </article>
        </section>
    </section>
""")

PERCEPCION_CLIENTE_HTML = dedent("""
    <section class="percepcion-page pe-percepcion" id="percepcion-builder">
        <style>
            .pe-percepcion {
                --bg: #f6f8fc;
                --surface: rgba(255,255,255,.88);
                --text: #0f172a;
                --muted: #64748b;
                --border: rgba(148,163,184,.38);
                --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                width: 100%;
                font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                color: var(--text);
                background:
                  radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                  radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                  var(--bg);
                border-radius: 18px;
                padding: 18px 0 34px;
            }
            .pe-percepcion .percepcion-input-card,
            .pe-percepcion .percepcion-column {
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 22px;
                box-shadow: var(--shadow-soft);
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
            }
            .pe-percepcion .percepcion-input-card { padding: 16px; }
            .pe-percepcion .percepcion-input-card h2 { margin: 0; font-size: 20px; letter-spacing: -0.02em; }
            .pe-percepcion .percepcion-input-card p { margin: 6px 0 12px; color: var(--muted); font-size: 13px; }
            .pe-percepcion .percepcion-input-grid { grid-template-columns: 1.4fr .8fr .8fr; gap: 12px; }
            .pe-percepcion .percepcion-input-grid textarea,
            .pe-percepcion .percepcion-input-grid select {
                border: 1px solid var(--border);
                border-radius: 14px;
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-percepcion .percepcion-input-actions button {
                border-radius: 14px;
                padding: 10px 14px;
                font-weight: 800;
                border: 1px solid rgba(15,61,46,.35);
                background: linear-gradient(135deg, #0f3d2e, #1f6f52);
                color: #fff;
                cursor: pointer;
            }
            .pe-percepcion .percepcion-board {
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 14px;
                margin-top: 14px;
            }
            .pe-percepcion .percepcion-column { padding: 14px; }
            .pe-percepcion .percepcion-column header { margin-bottom: 10px; }
            .pe-percepcion .percepcion-column header span {
                border-radius: 999px;
                padding: 6px 10px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.70);
            }
            .pe-percepcion .percepcion-column ul { min-height: 120px; gap: 10px; }
            .pe-percepcion .percepcion-item {
                border-radius: 16px;
                border: 1px solid rgba(148,163,184,.30);
                background: rgba(255,255,255,.82);
                box-shadow: 0 10px 20px rgba(15,23,42,.04);
                padding: 12px;
            }
            .pe-percepcion .percepcion-item button {
                border-radius: 999px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                padding: 6px 10px;
                font-size: 12px;
            }
            @media (max-width: 1100px) { .pe-percepcion .percepcion-input-grid { grid-template-columns: 1fr; } }
            @media (max-width: 900px) { .pe-percepcion .percepcion-board { grid-template-columns: 1fr; } }
        </style>
        <article class="percepcion-input-card">
            <h2>Percepción del cliente</h2>
            <p>Registra percepciones y clasifícalas para orientar mejoras del servicio.</p>
            <div class="percepcion-input-grid">
                <div class="form-field">
                    <label for="percepcion-description">Comentario</label>
                    <textarea id="percepcion-description" placeholder="Describe la percepción detectada"></textarea>
                </div>
                <div class="form-field">
                    <label for="percepcion-category">Categoría</label>
                    <select id="percepcion-category">
                        <option value="positive">Percepción positiva</option>
                        <option value="neutral">Percepción neutral</option>
                        <option value="negative">Percepción negativa</option>
                    </select>
                </div>
                <div class="form-field">
                    <label for="percepcion-priority">Prioridad</label>
                    <select id="percepcion-priority">
                        <option value="high">Alta</option>
                        <option value="medium">Media</option>
                        <option value="low">Baja</option>
                    </select>
                </div>
            </div>
            <div class="percepcion-input-actions">
                <button type="button" id="percepcion-add">Agregar percepción</button>
            </div>
        </article>
        <section class="percepcion-board">
            <article class="percepcion-column">
                <header><h3>Positivas</h3><span id="percepcion-count-positive">0</span></header>
                <ul id="percepcion-list-positive"></ul>
            </article>
            <article class="percepcion-column">
                <header><h3>Neutrales</h3><span id="percepcion-count-neutral">0</span></header>
                <ul id="percepcion-list-neutral"></ul>
            </article>
            <article class="percepcion-column">
                <header><h3>Negativas</h3><span id="percepcion-count-negative">0</span></header>
                <ul id="percepcion-list-negative"></ul>
            </article>
        </section>
    </section>
""")

PLAN_ESTRATEGICO_HTML = dedent("""
    <section class="pe-page">
      <style>
        .pe-page{
          --bg: #f6f8fc;
          --surface: rgba(255,255,255,.88);
          --card: #ffffff;
          --text: #0f172a;
          --muted: #64748b;
          --border: rgba(148,163,184,.38);
          --shadow: 0 18px 40px rgba(15,23,42,.08);
          --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
          --radius: 18px;
          --primary: #0f3d2e;
          --primary-2: #1f6f52;
          --accent: #2563eb;
          --ok: #16a34a;
          --warn: #f59e0b;
          --crit: #ef4444;
          width: 100%;
          font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
          color: var(--text);
          background:
            radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
            radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
            var(--bg);
          border-radius: 18px;
        }
        .pe-wrap{
          width: 100%;
          margin: 0 auto;
          padding: 18px 0 34px;
        }
        .pe-btn{
          border-radius: 14px;
          padding: 10px 14px;
          font-weight: 800;
          border: 1px solid var(--border);
          background: rgba(255,255,255,.75);
          cursor:pointer;
          box-shadow: var(--shadow-soft);
          transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
        }
        .pe-btn:hover{ transform: translateY(-1px); box-shadow: var(--shadow); background: rgba(255,255,255,.95); }
        .pe-btn--primary{ background: linear-gradient(135deg, var(--primary), var(--primary-2)); color:#fff; border-color: rgba(15,61,46,.35); }
        .pe-btn--ghost{ background: rgba(255,255,255,.78); }
        .pe-btn--soft{ background: rgba(15,61,46,.10); border-color: rgba(15,61,46,.18); color:#0b2a20; }
        .pe-kpis{
          display:grid;
          grid-template-columns: repeat(4, minmax(0, 1fr));
          gap: 14px;
          margin: 16px 0 18px;
        }
        .pe-kpi{
          background: var(--surface);
          border: 1px solid var(--border);
          border-radius: var(--radius);
          box-shadow: var(--shadow-soft);
          padding: 16px;
          backdrop-filter: blur(10px);
          -webkit-backdrop-filter: blur(10px);
          position: relative;
          overflow:hidden;
        }
        .pe-kpi:before{
          content:"";
          position:absolute;
          inset:-1px;
          background: linear-gradient(135deg, rgba(15,61,46,.12), transparent 35%, rgba(37,99,235,.10));
          opacity:.9;
          pointer-events:none;
        }
        .pe-kpi > *{ position: relative; }
        .pe-kpi__label{ color: var(--muted); font-size: 14px; font-weight: 600; }
        .pe-kpi__value{ margin-top: 10px; font-size: 34px; font-weight: 900; letter-spacing: -0.03em; }
        .pe-kpi__meta{ margin-top: 10px; display:flex; gap: 8px; flex-wrap: wrap; }
        .pe-chip{
          font-size: 12px;
          padding: 6px 10px;
          border-radius: 999px;
          background: rgba(15,23,42,.05);
          border: 1px solid rgba(15,23,42,.08);
          color: rgba(15,23,42,.72);
        }
        .pe-chip--info{ background: rgba(37,99,235,.10); border-color: rgba(37,99,235,.18); color: #1d4ed8; }
        .pe-chip--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
        .pe-kpi--progress{ display:flex; flex-direction:column; }
        .pe-kpi__progress{ margin-top: 8px; display:flex; align-items:center; justify-content:flex-start; }
        .pe-ring{
          --p: 74;
          width: 86px;
          height: 86px;
          border-radius: 999px;
          background: conic-gradient(rgba(15,61,46,1) calc(var(--p) * 1%), rgba(148,163,184,.25) 0);
          display:grid;
          place-items:center;
          border: 1px solid rgba(148,163,184,.25);
          box-shadow: 0 12px 24px rgba(15,23,42,.06);
        }
        .pe-ring__inner{
          width: 66px;
          height: 66px;
          border-radius: 999px;
          background: rgba(255,255,255,.90);
          display:flex;
          flex-direction:column;
          align-items:center;
          justify-content:center;
          border: 1px solid rgba(148,163,184,.25);
        }
        .pe-ring__val{ font-weight: 900; font-size: 16px; letter-spacing: -0.02em; }
        .pe-ring__sub{ font-size: 11px; color: var(--muted); }
        .pe-grid{
          display:grid;
          grid-template-columns: 1.35fr .85fr;
          gap: 14px;
          align-items:start;
        }
        .pe-card{
          background: var(--surface);
          border: 1px solid var(--border);
          border-radius: 22px;
          box-shadow: var(--shadow-soft);
          padding: 16px;
          backdrop-filter: blur(10px);
          -webkit-backdrop-filter: blur(10px);
          overflow:hidden;
        }
        .pe-card__head{
          display:flex;
          align-items:flex-start;
          justify-content:space-between;
          gap: 12px;
          margin-bottom: 14px;
        }
        .pe-card__head h2{ margin:0; font-size: 20px; letter-spacing: -0.02em; }
        .pe-card__head p{ margin: 6px 0 0; color: var(--muted); font-size: 13px; }
        .pe-card__tools{ display:flex; align-items:center; gap: 10px; }
        .pe-pill{
          font-size: 12px;
          padding: 6px 10px;
          border-radius: 999px;
          background: rgba(255,255,255,.70);
          border: 1px solid var(--border);
          color: rgba(15,23,42,.72);
          white-space: nowrap;
        }
        .pe-pill--soft{ background: rgba(15,61,46,.10); border-color: rgba(15,61,46,.18); color: #0b2a20; }
        .pe-list{ display:flex; flex-direction:column; gap: 12px; }
        .pe-item{
          background: rgba(255,255,255,.80);
          border: 1px solid rgba(148,163,184,.30);
          border-radius: 18px;
          padding: 14px 14px 12px;
          box-shadow: 0 10px 20px rgba(15,23,42,.04);
          transition: transform .15s ease, box-shadow .15s ease, border-color .15s ease;
        }
        .pe-item:hover{ transform: translateY(-1px); box-shadow: 0 16px 30px rgba(15,23,42,.07); border-color: rgba(37,99,235,.22); }
        .pe-item--active{ outline: 1px solid rgba(15,61,46,.22); background: rgba(15,61,46,.06); }
        .pe-item__top{ display:flex; align-items:flex-start; justify-content:space-between; gap: 12px; }
        .pe-item__title{ display:flex; gap: 10px; align-items:flex-start; line-height: 1.35; }
        .pe-code{
          font-weight: 900;
          color: #0b2a20;
          background: rgba(15,61,46,.10);
          border: 1px solid rgba(15,61,46,.18);
          padding: 6px 10px;
          border-radius: 999px;
          font-size: 12px;
          white-space: nowrap;
        }
        .pe-item__meta{ margin-top: 10px; display:flex; gap: 10px; flex-wrap:wrap; color: var(--muted); }
        .pe-mini{
          font-size: 12px;
          padding: 6px 10px;
          border-radius: 999px;
          background: rgba(15,23,42,.04);
          border: 1px solid rgba(15,23,42,.08);
        }
        .pe-status{
          font-size: 12px;
          font-weight: 800;
          padding: 6px 10px;
          border-radius: 999px;
          border: 1px solid var(--border);
          background: rgba(255,255,255,.70);
          color: rgba(15,23,42,.72);
          white-space: nowrap;
        }
        .pe-status--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }
        .pe-status--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
        .pe-status--neutral{ background: rgba(15,23,42,.05); border-color: rgba(15,23,42,.10); color: rgba(15,23,42,.70); }
        .pe-bar{
          margin-top: 12px;
          height: 10px;
          border-radius: 999px;
          background: rgba(148,163,184,.25);
          border: 1px solid rgba(148,163,184,.25);
          overflow:hidden;
        }
        .pe-bar__fill{ height: 100%; width: 50%; border-radius: 999px; background: linear-gradient(90deg, rgba(37,99,235,1), rgba(96,165,250,1)); }
        .pe-bar__fill--ok{ background: linear-gradient(90deg, rgba(15,61,46,1), rgba(31,111,82,1)); }
        .pe-bar__fill--warn{ background: linear-gradient(90deg, rgba(245,158,11,1), rgba(253,230,138,1)); }
        .pe-axis{ display:flex; flex-direction:column; gap: 12px; }
        .pe-axis__btn{
          width:100%;
          text-align:left;
          border: 1px solid rgba(148,163,184,.30);
          background: rgba(255,255,255,.80);
          border-radius: 18px;
          padding: 14px;
          display:flex;
          align-items:center;
          gap: 12px;
          cursor:pointer;
          box-shadow: 0 10px 20px rgba(15,23,42,.04);
          transition: transform .15s ease, box-shadow .15s ease, border-color .15s ease;
        }
        .pe-axis__btn:hover{ transform: translateY(-1px); box-shadow: 0 16px 30px rgba(15,23,42,.07); border-color: rgba(37,99,235,.22); }
        .pe-axis__btn--active{ background: rgba(15,61,46,.06); outline: 1px solid rgba(15,61,46,.22); }
        .pe-axis__dot{ width: 10px; height: 10px; border-radius: 999px; background: var(--primary); box-shadow: 0 0 0 6px rgba(15,61,46,.10); }
        .pe-axis__dot--alt{ background: #2563eb; box-shadow: 0 0 0 6px rgba(37,99,235,.12); }
        .pe-axis__dot--alt2{ background: #7c3aed; box-shadow: 0 0 0 6px rgba(124,58,237,.12); }
        .pe-axis__dot--alt3{ background: #f59e0b; box-shadow: 0 0 0 6px rgba(245,158,11,.12); }
        .pe-axis__count{
          margin-left:auto;
          font-weight: 900;
          color: rgba(15,23,42,.72);
          background: rgba(15,23,42,.04);
          border: 1px solid rgba(15,23,42,.08);
          padding: 6px 10px;
          border-radius: 999px;
        }
        .pe-sidebox{
          margin-top: 14px;
          background: rgba(255,255,255,.80);
          border: 1px solid rgba(148,163,184,.30);
          border-radius: 18px;
          padding: 14px;
        }
        .pe-sidebox__head{ display:flex; align-items:center; justify-content:space-between; gap: 10px; }
        .pe-sidebox__grid{ margin-top: 12px; display:grid; grid-template-columns: 1fr 1fr; gap: 10px; }
        .pe-metric{
          background: rgba(15,23,42,.03);
          border: 1px solid rgba(15,23,42,.08);
          border-radius: 16px;
          padding: 12px;
        }
        .pe-metric__k{ font-size: 12px; color: var(--muted); font-weight: 700; }
        .pe-metric__v{ margin-top: 6px; font-size: 18px; font-weight: 900; letter-spacing: -0.02em; }
        .pe-metric__v--warn{ color: #92400e; }
        .pe-roadmap{ margin-top: 14px; }
        .pe-timeline{ display:grid; grid-template-columns: repeat(4, minmax(0,1fr)); gap: 12px; }
        .pe-tlcol{
          background: rgba(255,255,255,.70);
          border: 1px solid rgba(148,163,184,.30);
          border-radius: 18px;
          padding: 12px;
        }
        .pe-tlcol__head{
          font-weight: 900;
          color: rgba(15,23,42,.78);
          padding: 8px 10px;
          border-radius: 999px;
          background: rgba(15,23,42,.04);
          border: 1px solid rgba(15,23,42,.08);
          display:inline-block;
          margin-bottom: 10px;
        }
        .pe-tlitem{
          background: rgba(255,255,255,.82);
          border: 1px solid rgba(148,163,184,.30);
          border-radius: 16px;
          padding: 12px;
          box-shadow: 0 10px 20px rgba(15,23,42,.04);
          margin-bottom: 10px;
        }
        .pe-tlitem:last-child{ margin-bottom: 0; }
        .pe-tlitem__top{ display:flex; align-items:flex-start; justify-content:space-between; gap: 10px; }
        .pe-tlitem p{ margin: 8px 0 0; color: var(--muted); font-size: 12.5px; line-height: 1.4; }
        .pe-tlmeta{ margin-top: 10px; display:flex; gap: 8px; flex-wrap:wrap; }
        .pe-tlitem--ok{ outline: 1px solid rgba(22,163,74,.18); background: rgba(22,163,74,.06); }
        .pe-tlitem--warn{ outline: 1px solid rgba(245,158,11,.18); background: rgba(245,158,11,.07); }
        .pe-tlitem--soft{ background: rgba(15,61,46,.06); outline: 1px solid rgba(15,61,46,.18); }
        @media (max-width: 1100px){
          .pe-kpis{ grid-template-columns: repeat(2, minmax(0,1fr)); }
          .pe-grid{ grid-template-columns: 1fr; }
          .pe-timeline{ grid-template-columns: repeat(2, minmax(0,1fr)); }
        }
        @media (max-width: 640px){
          .pe-kpis{ grid-template-columns: 1fr; }
          .pe-timeline{ grid-template-columns: 1fr; }
        }
      </style>

      <div class="pe-wrap">
        __CREATE_PLAN_BUTTON__
        <section class="pe-kpis">
          <article class="pe-kpi">
            <div class="pe-kpi__label">Objetivos estratégicos</div>
            <div class="pe-kpi__value">12</div>
            <div class="pe-kpi__meta"><span class="pe-chip pe-chip--info">4 por eje</span></div>
          </article>
          <article class="pe-kpi">
            <div class="pe-kpi__label">Iniciativas activas</div>
            <div class="pe-kpi__value">24</div>
            <div class="pe-kpi__meta"><span class="pe-chip">18 en ejecución</span></div>
          </article>
          <article class="pe-kpi">
            <div class="pe-kpi__label">Indicadores vinculados</div>
            <div class="pe-kpi__value">36</div>
            <div class="pe-kpi__meta"><span class="pe-chip pe-chip--warn">6 sin línea base</span></div>
          </article>
          <article class="pe-kpi pe-kpi--progress">
            <div class="pe-kpi__label">Avance del plan</div>
            <div class="pe-kpi__progress">
              <div class="pe-ring" style="--p:74;">
                <div class="pe-ring__inner">
                  <div class="pe-ring__val">74%</div>
                  <div class="pe-ring__sub">global</div>
                </div>
              </div>
            </div>
          </article>
        </section>

        <section class="pe-grid">
          <article class="pe-card">
            <div class="pe-card__head">
              <div>
                <h2>Objetivos priorizados</h2>
                <p>Portafolio vigente con estado y trazabilidad.</p>
              </div>
              <div class="pe-card__tools">
                <span class="pe-pill pe-pill--soft">12 objetivos</span>
              </div>
            </div>

            <div class="pe-list">
              <article class="pe-item pe-item--active">
                <div class="pe-item__top">
                  <div class="pe-item__title"><span class="pe-code">OE-01</span><strong>Fortalecer la sostenibilidad financiera institucional.</strong></div>
                  <span class="pe-status pe-status--ok">En meta</span>
                </div>
                <div class="pe-item__meta"><span class="pe-mini">3 iniciativas</span><span class="pe-mini">8 KPIs</span><span class="pe-mini">Líder: Dirección</span></div>
                <div class="pe-bar"><div class="pe-bar__fill pe-bar__fill--ok" style="width:81%"></div></div>
              </article>

              <article class="pe-item">
                <div class="pe-item__top">
                  <div class="pe-item__title"><span class="pe-code">OE-02</span><strong>Mejorar la satisfacción de clientes y usuarios finales.</strong></div>
                  <span class="pe-status pe-status--warn">En riesgo</span>
                </div>
                <div class="pe-item__meta"><span class="pe-mini">2 iniciativas</span><span class="pe-mini">9 KPIs</span><span class="pe-mini">Líder: Servicios</span></div>
                <div class="pe-bar"><div class="pe-bar__fill pe-bar__fill--warn" style="width:63%"></div></div>
              </article>

              <article class="pe-item">
                <div class="pe-item__top">
                  <div class="pe-item__title"><span class="pe-code">OE-03</span><strong>Optimizar procesos críticos y tiempos de respuesta.</strong></div>
                  <span class="pe-status pe-status--neutral">Seguimiento</span>
                </div>
                <div class="pe-item__meta"><span class="pe-mini">4 iniciativas</span><span class="pe-mini">11 KPIs</span><span class="pe-mini">Líder: Operaciones</span></div>
                <div class="pe-bar"><div class="pe-bar__fill" style="width:71%"></div></div>
              </article>
            </div>
          </article>

          <aside class="pe-card">
            <div class="pe-card__head">
              <div>
                <h2>Ejes estratégicos</h2>
                <p>Distribución del portafolio por eje.</p>
              </div>
            </div>

            <div class="pe-axis">
              <button class="pe-axis__btn pe-axis__btn--active" type="button" onclick="window.location.href='/ejes-estrategicos'"><span class="pe-axis__dot"></span><span>Gobernanza y cumplimiento</span><span class="pe-axis__count">3</span></button>
              <button class="pe-axis__btn" type="button" onclick="window.location.href='/ejes-estrategicos'"><span class="pe-axis__dot pe-axis__dot--alt"></span><span>Excelencia operativa</span><span class="pe-axis__count">4</span></button>
              <button class="pe-axis__btn" type="button" onclick="window.location.href='/ejes-estrategicos'"><span class="pe-axis__dot pe-axis__dot--alt2"></span><span>Innovación y digitalización</span><span class="pe-axis__count">2</span></button>
              <button class="pe-axis__btn" type="button" onclick="window.location.href='/ejes-estrategicos'"><span class="pe-axis__dot pe-axis__dot--alt3"></span><span>Desarrollo del talento</span><span class="pe-axis__count">3</span></button>
            </div>

            <div class="pe-sidebox">
              <div class="pe-sidebox__head"><strong>Resumen rápido</strong><span class="pe-pill">Corte mensual</span></div>
              <div class="pe-sidebox__grid">
                <div class="pe-metric"><div class="pe-metric__k">Hitos del trimestre</div><div class="pe-metric__v">28</div></div>
                <div class="pe-metric"><div class="pe-metric__k">Desviaciones</div><div class="pe-metric__v pe-metric__v--warn">6</div></div>
                <div class="pe-metric"><div class="pe-metric__k">Cumplimiento</div><div class="pe-metric__v">74%</div></div>
                <div class="pe-metric"><div class="pe-metric__k">Acciones correctivas</div><div class="pe-metric__v">9</div></div>
              </div>
            </div>
          </aside>
        </section>

        <article class="pe-card pe-roadmap">
          <div class="pe-card__head">
            <div>
              <h2>Hoja de ruta anual</h2>
              <p>Fases clave del ciclo estratégico.</p>
            </div>
            <div class="pe-card__tools">
              <button class="pe-btn pe-btn--soft" type="button">Exportar</button>
              <button class="pe-btn pe-btn--ghost" type="button">Vista Gantt</button>
            </div>
          </div>

          <div class="pe-timeline">
            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q1</div>
              <article class="pe-tlitem pe-tlitem--ok">
                <div class="pe-tlitem__top"><strong>Diagnóstico estratégico</strong><span class="pe-status pe-status--ok">Completo</span></div>
                <p>Actualización de línea base y priorización inicial.</p>
                <div class="pe-tlmeta"><span class="pe-mini">8 entregables</span><span class="pe-mini">Líder PMO</span></div>
              </article>
            </div>

            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q2</div>
              <article class="pe-tlitem pe-tlitem--warn">
                <div class="pe-tlitem__top"><strong>Definición de metas</strong><span class="pe-status pe-status--warn">Ajuste</span></div>
                <p>Alineación final de KPIs y responsables por objetivo.</p>
                <div class="pe-tlmeta"><span class="pe-mini">6 entregables</span><span class="pe-mini">Comité</span></div>
              </article>
            </div>

            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q3</div>
              <article class="pe-tlitem pe-tlitem--soft">
                <div class="pe-tlitem__top"><strong>Implementación</strong><span class="pe-status pe-status--neutral">En curso</span></div>
                <p>Ejecución de iniciativas con control de hitos mensuales.</p>
                <div class="pe-tlmeta"><span class="pe-mini">10 entregables</span><span class="pe-mini">Áreas</span></div>
              </article>
            </div>

            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q4</div>
              <article class="pe-tlitem">
                <div class="pe-tlitem__top"><strong>Cierre y evaluación</strong><span class="pe-status pe-status--neutral">Programado</span></div>
                <p>Lecciones aprendidas y propuesta de ajustes para el siguiente ciclo.</p>
                <div class="pe-tlmeta"><span class="pe-mini">4 entregables</span><span class="pe-mini">Dirección</span></div>
              </article>
            </div>
          </div>
        </article>
      </div>
    </section>
""")

INICIO_BSC_HTML = dedent("""
    <section class="poa-dashboard">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');

            :root{
              --bg: #f6f8fc;
              --surface: rgba(255,255,255,.86);
              --card: #ffffff;
              --text: #0f172a;
              --muted: #64748b;
              --border: rgba(148,163,184,.35);
              --shadow: 0 18px 40px rgba(15,23,42,.08);
              --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
              --radius: 18px;

              --primary: #2563eb;
              --primary-2: #60a5fa;
              --ok: #16a34a;
              --warn: #f59e0b;
              --crit: #ef4444;

              --chip: rgba(37,99,235,.10);
              --chip-text: #1d4ed8;
            }

            .poa-dashboard * { box-sizing: border-box; }
            .poa-dashboard{
              font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
              color: var(--text);
              background:
                radial-gradient(1200px 600px at 20% 0%, rgba(37,99,235,.10), transparent 60%),
                radial-gradient(1000px 500px at 90% 10%, rgba(96,165,250,.12), transparent 55%),
                var(--bg);
              border-radius: 18px;
              padding: 14px;
            }

            .wrap{
              width: 100%;
              padding: 18px 18px 28px;
            }

            .topbar{
              display:flex;
              gap: 12px;
              align-items:center;
              flex-wrap: wrap;
              justify-content:space-between;
              margin-bottom: 14px;
            }
            .title{
              display:flex;
              flex-direction:column;
              gap: 4px;
            }
            .title h1{
              font-size: 18px;
              margin:0;
              letter-spacing: -0.02em;
            }
            .title p{
              margin:0;
              color: var(--muted);
              font-size: 13px;
            }

            .actions{
              display:flex;
              gap:10px;
              align-items:center;
              flex-wrap: wrap;
            }
            .btn{
              border: 1px solid var(--border);
              background: rgba(255,255,255,.7);
              padding: 10px 12px;
              border-radius: 12px;
              color: var(--text);
              display:flex;
              gap:10px;
              align-items:center;
              flex-wrap: wrap;
              box-shadow: var(--shadow-soft);
              cursor:pointer;
              user-select:none;
              transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
            }
            .btn:hover{ transform: translateY(-1px); box-shadow: var(--shadow); background: rgba(255,255,255,.95); }
            .btn .dot{
              width: 10px; height:10px; border-radius:50%;
              background: var(--primary);
              box-shadow: 0 0 0 6px rgba(37,99,235,.12);
            }

            .kpis{
              display:grid;
              grid-template-columns: repeat(4, minmax(0, 1fr));
              gap: 12px;
              margin: 12px 0 16px;
            }
            .kpi{
              background: var(--surface);
              border: 1px solid var(--border);
              border-radius: var(--radius);
              box-shadow: var(--shadow-soft);
              padding: 14px 14px 12px;
              backdrop-filter: blur(10px);
              -webkit-backdrop-filter: blur(10px);
              position: relative;
              overflow:hidden;
            }
            .kpi:before{
              content:"";
              position:absolute;
              inset:-1px;
              background: linear-gradient(135deg, rgba(37,99,235,.10), transparent 35%, rgba(96,165,250,.12));
              pointer-events:none;
              opacity:.8;
            }
            .kpi > *{ position: relative; }
            .kpi .label{
              font-size: 13px;
              color: var(--muted);
              margin-bottom: 10px;
              display:flex;
              align-items:center;
              justify-content:space-between;
              flex-wrap: wrap;
              gap:10px;
            }
            .kpi .value{
              font-size: 32px;
              font-weight: 700;
              letter-spacing: -0.03em;
              line-height: 1.05;
            }
            .kpi .sub{
              margin-top: 8px;
              font-size: 12px;
              color: var(--muted);
              display:flex;
              justify-content:space-between;
              flex-wrap: wrap;
              gap:10px;
            }

            .pill{
              font-size: 12px;
              padding: 6px 10px;
              border-radius: 999px;
              background: var(--chip);
              color: var(--chip-text);
              border: 1px solid rgba(37,99,235,.18);
              white-space: nowrap;
            }

            .state{
              display:inline-flex;
              gap:8px;
              align-items:center;
              font-size: 12px;
              padding: 6px 10px;
              border-radius: 999px;
              border: 1px solid var(--border);
              background: rgba(255,255,255,.65);
              color: var(--muted);
            }
            .state i{ width:8px; height:8px; border-radius:50%; display:inline-block; }
            .ok i{ background: var(--ok); box-shadow: 0 0 0 6px rgba(22,163,74,.10); }
            .warn i{ background: var(--warn); box-shadow: 0 0 0 6px rgba(245,158,11,.12); }
            .crit i{ background: var(--crit); box-shadow: 0 0 0 6px rgba(239,68,68,.12); }

            .grid{
              display:grid;
              grid-template-columns: 1.2fr 1fr;
              gap: 12px;
              align-items:start;
            }

            .panel{
              background: var(--surface);
              border: 1px solid var(--border);
              border-radius: var(--radius);
              box-shadow: var(--shadow-soft);
              padding: 14px;
              backdrop-filter: blur(10px);
              -webkit-backdrop-filter: blur(10px);
              overflow:hidden;
            }

            .panel-header{
              display:flex;
              align-items:center;
              justify-content:space-between;
              flex-wrap: wrap;
              gap: 10px;
              margin-bottom: 12px;
            }
            .panel-title{
              display:flex;
              flex-direction:column;
              gap: 4px;
            }
            .panel-title h2{
              margin:0;
              font-size: 16px;
              letter-spacing: -0.02em;
            }
            .panel-title small{
              color: var(--muted);
              font-size: 12px;
            }

            .meta-pill{
              display:inline-flex;
              align-items:center;
              gap: 8px;
              padding: 6px 10px;
              border-radius: 999px;
              background: rgba(2,132,199,.10);
              border: 1px solid rgba(2,132,199,.18);
              color: #0369a1;
              font-size: 12px;
              white-space:nowrap;
            }

            .rows{
              display:flex;
              flex-direction:column;
              gap: 10px;
            }
            .row{
              background: rgba(255,255,255,.75);
              border: 1px solid var(--border);
              border-radius: 14px;
              padding: 10px 12px;
              display:flex;
              flex-direction:column;
              gap: 10px;
              box-shadow: 0 10px 20px rgba(15,23,42,.04);
            }
            .row-top{
              display:flex;
              align-items:center;
              justify-content:space-between;
              flex-wrap: wrap;
              gap: 12px;
            }
            .row-label{
              font-weight: 600;
              font-size: 13px;
              color: #0b1220;
            }
            .row-value{
              font-weight: 700;
              font-size: 13px;
              color: #0b1220;
              white-space:nowrap;
            }
            .hint{
              font-size: 12px;
              color: var(--muted);
            }

            .progress{
              width: 100%;
              height: 10px;
              border-radius: 999px;
              background: rgba(148,163,184,.25);
              overflow:hidden;
              border: 1px solid rgba(148,163,184,.25);
            }
            .bar{
              height: 100%;
              width: 50%;
              border-radius: 999px;
              background: linear-gradient(90deg, var(--primary), var(--primary-2));
            }
            .bar.ok{ background: linear-gradient(90deg, #16a34a, #86efac); }
            .bar.warn{ background: linear-gradient(90deg, #f59e0b, #fde68a); }
            .bar.crit{ background: linear-gradient(90deg, #ef4444, #fecaca); }

            .charts{
              display:grid;
              grid-template-columns: 1fr 1fr;
              gap: 12px;
              margin-top: 12px;
            }
            .chart-card{
              background: rgba(255,255,255,.75);
              border: 1px solid var(--border);
              border-radius: 16px;
              padding: 12px;
              box-shadow: 0 10px 20px rgba(15,23,42,.04);
              min-height: 240px;
              display:flex;
              flex-direction:column;
              gap: 10px;
            }
            .chart-card h3{
              margin:0;
              font-size: 13px;
              color: #0b1220;
              letter-spacing: -0.01em;
              display:flex;
              align-items:center;
              justify-content:space-between;
              gap: 10px;
            }
            .chart-card h3 span{
              font-weight: 500;
              color: var(--muted);
              font-size: 12px;
            }
            .poa-dashboard canvas{ width:100% !important; height: 180px !important; }

            @media (max-width: 1100px){
              .wrap{ width: 90%; }
              .kpis{ grid-template-columns: repeat(2, minmax(0, 1fr)); }
              .grid{ grid-template-columns: 1fr; }
            }
            @media (max-width: 900px){
              .wrap{
                width: 100%;
                padding: 14px 10px 22px;
              }
              .poa-dashboard{
                border-radius: 14px;
                padding: 10px;
              }
              .title h1{
                font-size: 16px;
              }
              .title p{
                font-size: 12px;
              }
              .panel{
                padding: 12px;
              }
            }
            @media (max-width: 560px){
              .kpis{ grid-template-columns: 1fr; }
              .charts{ grid-template-columns: 1fr; }
              .kpi .value{ font-size: 28px; }
              .btn{
                width: 100%;
                justify-content: flex-start;
                padding: 10px;
              }
              .actions{
                width: 100%;
              }
              .chart-card{
                min-height: 210px;
              }
              .poa-dashboard canvas{
                height: 165px !important;
              }
            }
        </style>

        <div class="wrap">
            <div class="topbar">
              <div class="title">
                <h1>Tablero POA / BSC</h1>
                <p>Seguimiento ejecutivo con metas, desviaciones y tendencias (estilo premium).</p>
              </div>
              <div class="actions">
                <div class="btn" title="Filtrar">
                  <span class="dot"></span>
                  <strong style="font-size:13px;">Filtros</strong>
                  <span style="color:var(--muted);font-size:12px;">(Mes / Área)</span>
                </div>
                <div class="btn" title="Exportar">
                  <span class="dot" style="background:#16a34a; box-shadow:0 0 0 6px rgba(22,163,74,.12);"></span>
                  <strong style="font-size:13px;">Exportar</strong>
                  <span style="color:var(--muted);font-size:12px;">PDF/Excel</span>
                </div>
              </div>
            </div>

            <section class="kpis">
              <article class="kpi">
                <div class="label">
                  <span>Cumplimiento global POA</span>
                  <span class="state ok"><i></i>En rango</span>
                </div>
                <div class="value">78%</div>
                <div class="sub">
                  <span class="pill">Meta: 85%</span>
                  <span class="hint">+2.1 pts vs mes anterior</span>
                </div>
              </article>

              <article class="kpi">
                <div class="label">
                  <span>Objetivos estratégicos en meta</span>
                  <span class="state ok"><i></i>Estable</span>
                </div>
                <div class="value">9 <span style="font-weight:600;color:var(--muted);font-size:18px;">/ 12</span></div>
                <div class="sub">
                  <span class="pill">Pendientes: 3</span>
                  <span class="hint">2 en riesgo (alerta)</span>
                </div>
              </article>

              <article class="kpi">
                <div class="label">
                  <span>Iniciativas activas</span>
                  <span class="state warn"><i></i>Alta carga</span>
                </div>
                <div class="value">24</div>
                <div class="sub">
                  <span class="pill">Capacidad: 20</span>
                  <span class="hint">Reasignar responsables</span>
                </div>
              </article>

              <article class="kpi">
                <div class="label">
                  <span>Desviaciones críticas</span>
                  <span class="state crit"><i></i>Prioridad</span>
                </div>
                <div class="value">3</div>
                <div class="sub">
                  <span class="pill">SLA: 7 días</span>
                  <span class="hint">2 sin plan de acción</span>
                </div>
              </article>
            </section>

            <section class="grid">
              <div class="panel">
                <div class="panel-header">
                  <div class="panel-title">
                    <h2>Perspectiva Financiera</h2>
                    <small>Resultados financieros, ejecución y disciplina presupuestaria</small>
                  </div>
                  <div class="meta-pill">Meta 85%</div>
                </div>

                <div class="rows">
                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Ejecución presupuestaria</div>
                      <div class="row-value">81%</div>
                    </div>
                    <div class="progress"><div class="bar ok" style="width:81%"></div></div>
                    <div class="hint">En rango; falta optimizar calendario de gasto.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Reducción de costos operativos</div>
                      <div class="row-value">6.2%</div>
                    </div>
                    <div class="progress"><div class="bar warn" style="width:62%"></div></div>
                    <div class="hint">En avance; validar quick wins y renegociaciones.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Proyectos dentro de presupuesto</div>
                      <div class="row-value">14 / 18</div>
                    </div>
                    <div class="progress"><div class="bar" style="width:78%"></div></div>
                    <div class="hint">4 con sobrecosto: revisar alcance y compras.</div>
                  </div>
                </div>

                <div class="charts">
                  <div class="chart-card">
                    <h3>Trend: Cumplimiento POA <span>ultimos 6 meses</span></h3>
                    <canvas id="chartCumplimiento"></canvas>
                  </div>
                  <div class="chart-card">
                    <h3>Distribucion de desviaciones <span>por severidad</span></h3>
                    <canvas id="chartDesviaciones"></canvas>
                  </div>
                </div>
              </div>

              <div class="panel">
                <div class="panel-header">
                  <div class="panel-title">
                    <h2>Clientes / Usuarios</h2>
                    <small>Calidad de servicio, satisfaccion y tiempos de atencion</small>
                  </div>
                  <div class="meta-pill">Meta 90%</div>
                </div>

                <div class="rows">
                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Satisfaccion general</div>
                      <div class="row-value">88%</div>
                    </div>
                    <div class="progress"><div class="bar ok" style="width:88%"></div></div>
                    <div class="hint">Muy cerca de meta; reforzar calidad y comunicacion.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Tiempo de respuesta a solicitudes</div>
                      <div class="row-value">42h</div>
                    </div>
                    <div class="progress"><div class="bar warn" style="width:70%"></div></div>
                    <div class="hint">Meta sugerida: <= 36h. Ajustar colas y SLA por tipo.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Servicios con nivel alto</div>
                      <div class="row-value">11 / 13</div>
                    </div>
                    <div class="progress"><div class="bar" style="width:85%"></div></div>
                    <div class="hint">2 servicios en medio: plan de mejora y auditoria.</div>
                  </div>
                </div>

                <div class="charts">
                  <div class="chart-card" style="grid-column: 1 / -1;">
                    <h3>Atencion: Volumen vs SLA <span>por semana</span></h3>
                    <canvas id="chartSLA"></canvas>
                  </div>
                </div>
              </div>

              <div class="panel">
                <div class="panel-header">
                  <div class="panel-title">
                    <h2>Procesos Internos</h2>
                    <small>Estandarizacion, hitos, control operativo</small>
                  </div>
                  <div class="meta-pill">Meta 80%</div>
                </div>

                <div class="rows">
                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Procesos criticos estandarizados</div>
                      <div class="row-value">73%</div>
                    </div>
                    <div class="progress"><div class="bar warn" style="width:73%"></div></div>
                    <div class="hint">Priorizar procesos de alto impacto y riesgos.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Hitos del POA cumplidos en fecha</div>
                      <div class="row-value">31 / 40</div>
                    </div>
                    <div class="progress"><div class="bar" style="width:78%"></div></div>
                    <div class="hint">Reforzar seguimiento semanal y responsables.</div>
                  </div>
                </div>
              </div>

              <div class="panel">
                <div class="panel-header">
                  <div class="panel-title">
                    <h2>Aprendizaje y Crecimiento</h2>
                    <small>Capacitacion, competencias y cultura</small>
                  </div>
                  <div class="meta-pill">Meta 75%</div>
                </div>

                <div class="rows">
                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Capacitaciones ejecutadas</div>
                      <div class="row-value">18 / 22</div>
                    </div>
                    <div class="progress"><div class="bar ok" style="width:82%"></div></div>
                    <div class="hint">Buen avance; asegurar evidencia y evaluacion.</div>
                  </div>

                  <div class="row">
                    <div class="row-top">
                      <div class="row-label">Competencias criticas certificadas</div>
                      <div class="row-value">69%</div>
                    </div>
                    <div class="progress"><div class="bar warn" style="width:69%"></div></div>
                    <div class="hint">Definir ruta de certificacion por rol y prioridad.</div>
                  </div>
                </div>
              </div>
            </section>
        </div>

        <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
        <script>
            (function () {
                if (!window.Chart) return;
                const baseOpts = {
                  responsive: true,
                  maintainAspectRatio: false,
                  plugins: {
                    legend: { display: false },
                    tooltip: { mode: 'index', intersect: false }
                  },
                  interaction: { mode: 'index', intersect: false },
                  scales: {
                    x: { grid: { display: false }, ticks: { maxRotation: 0 } },
                    y: { grid: { color: 'rgba(148,163,184,.25)' }, ticks: { precision: 0 } }
                  }
                };

                if (document.getElementById('chartCumplimiento')) {
                    new Chart(document.getElementById('chartCumplimiento'), {
                      type: 'line',
                      data: {
                        labels: ['Ago','Sep','Oct','Nov','Dic','Ene'],
                        datasets: [{
                          label: 'Cumplimiento',
                          data: [70, 72, 74, 76, 77, 78],
                          borderWidth: 2,
                          tension: 0.35,
                          pointRadius: 3,
                          fill: true,
                          backgroundColor: 'rgba(37,99,235,.12)'
                        }]
                      },
                      options: {
                        ...baseOpts,
                        scales: {
                          ...baseOpts.scales,
                          y: { ...baseOpts.scales.y, min: 0, max: 100 }
                        }
                      }
                    });
                }

                if (document.getElementById('chartDesviaciones')) {
                    new Chart(document.getElementById('chartDesviaciones'), {
                      type: 'doughnut',
                      data: {
                        labels: ['Criticas','Advertencias','Menores'],
                        datasets: [{
                          data: [3, 6, 11],
                          borderWidth: 1
                        }]
                      },
                      options: {
                        responsive: true,
                        maintainAspectRatio: false,
                        plugins: { legend: { display: true, position: 'bottom' } },
                        cutout: '68%'
                      }
                    });
                }

                if (document.getElementById('chartSLA')) {
                    new Chart(document.getElementById('chartSLA'), {
                      type: 'bar',
                      data: {
                        labels: ['S1','S2','S3','S4','S5','S6'],
                        datasets: [
                          { label: 'Solicitudes', data: [120, 138, 110, 150, 162, 140], borderWidth: 1 },
                          { label: 'Fuera de SLA', data: [14, 18, 12, 22, 26, 19], borderWidth: 1 }
                        ]
                      },
                      options: {
                        ...baseOpts,
                        plugins: { legend: { display: true, position: 'bottom' } }
                      }
                    });
                }
            })();
        </script>
    </section>
""")

PROYECTANDO_HTML = dedent("""
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <style>
      .fp-shell{
        --bg: #f6f8fc;
        --surface: rgba(255,255,255,.88);
        --text: #0f172a;
        --muted: #64748b;
        --border: rgba(148,163,184,.35);
        --shadow: 0 18px 40px rgba(15,23,42,.08);
        --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
        --radius: 18px;
        --primary: #2563eb;
        --ok: #16a34a;
        --warn: #f59e0b;
        --crit: #ef4444;
      }
      .fp-shell *{ box-sizing:border-box; }
      .fp-shell{
        width: 100%;
        padding: 18px 0 28px;
        font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
        color: var(--text);
        background:
          radial-gradient(1200px 600px at 20% 0%, rgba(37,99,235,.10), transparent 60%),
          radial-gradient(1000px 500px at 90% 10%, rgba(96,165,250,.12), transparent 55%),
          var(--bg);
        border-radius: 18px;
      }
      .fp-topbar{
        display:flex;
        align-items:flex-start;
        justify-content:space-between;
        gap: 14px;
        margin-bottom: 14px;
      }
      .fp-title{
        display:flex;
        gap: 12px;
        align-items:flex-start;
      }
      .fp-mark{
        width: 36px;
        height: 36px;
        border-radius: 12px;
        display:flex;
        align-items:center;
        justify-content:center;
        background: rgba(37,99,235,.12);
        color: #1d4ed8;
        border: 1px solid rgba(37,99,235,.18);
        box-shadow: var(--shadow-soft);
      }
      .fp-title h1{
        margin:0;
        font-size: 20px;
        letter-spacing: -.02em;
      }
      .fp-sub{ font-weight:600; color: var(--muted); font-size: 14px; }
      .fp-title p{
        margin: 6px 0 0;
        font-size: 13px;
        color: var(--muted);
      }
      .fp-actions{ display:flex; align-items:center; gap: 10px; }
      .fp-chip{
        padding: 8px 12px;
        border-radius: 999px;
        border: 1px solid var(--border);
        background: rgba(255,255,255,.75);
        box-shadow: var(--shadow-soft);
        font-weight: 800;
      }
      .fp-chip--count{ width: 40px; display:flex; justify-content:center; }
      .fp-icon{
        border: 1px solid var(--border);
        background: rgba(255,255,255,.75);
        box-shadow: var(--shadow-soft);
        border-radius: 14px;
        padding: 10px 12px;
        cursor:pointer;
        transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
      }
      .fp-icon:hover{ transform: translateY(-1px); box-shadow: var(--shadow); background: rgba(255,255,255,.95); }
      .fp-btn{
        border-radius: 14px;
        padding: 10px 14px;
        font-weight: 800;
        border: 1px solid var(--border);
        background: rgba(255,255,255,.75);
        cursor:pointer;
        transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
      }
      .fp-btn:hover{ transform: translateY(-1px); box-shadow: var(--shadow-soft); background: rgba(255,255,255,.95); }
      .fp-btn--primary{
        background: var(--primary);
        border-color: rgba(37,99,235,.65);
        color: #fff;
      }
      .fp-btn--soft{
        background: rgba(37,99,235,.10);
        color: #1d4ed8;
        border-color: rgba(37,99,235,.18);
      }
      .fp-btn--ghost{
        background: rgba(255,255,255,.78);
        color: var(--text);
      }
      .fp-kpis{
        display:grid;
        grid-template-columns: repeat(4, minmax(0, 1fr));
        gap: 12px;
        margin: 12px 0 16px;
      }
      .fp-kpi{
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        box-shadow: var(--shadow-soft);
        padding: 14px;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        position: relative;
        overflow:hidden;
      }
      .fp-kpi:before{
        content:"";
        position:absolute;
        inset:-1px;
        background: linear-gradient(135deg, rgba(37,99,235,.10), transparent 35%, rgba(96,165,250,.12));
        opacity:.8;
        pointer-events:none;
      }
      .fp-kpi > *{ position:relative; }
      .fp-kpi__top{
        display:flex;
        align-items:center;
        justify-content:space-between;
        gap: 10px;
        color: var(--muted);
        font-size: 13px;
      }
      .fp-kpi__value{
        margin-top: 10px;
        font-size: 32px;
        font-weight: 900;
        letter-spacing: -0.03em;
        line-height: 1.05;
      }
      .fp-kpi__sub{ margin-top: 10px; color: var(--muted); font-size: 12px; }
      .fp-tag{
        font-size: 12px;
        padding: 6px 10px;
        border-radius: 999px;
        border: 1px solid var(--border);
        background: rgba(255,255,255,.70);
        color: rgba(15,23,42,.70);
        white-space:nowrap;
      }
      .fp-tag--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }
      .fp-tag--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
      .fp-tag--crit{ background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.22); color: #991b1b; }
      .fp-tag--neutral{ background: rgba(15,23,42,.06); border-color: rgba(15,23,42,.10); color: rgba(15,23,42,.70); }
      .fp-tag--info{ background: rgba(37,99,235,.10); border-color: rgba(37,99,235,.18); color: #1d4ed8; }
      .fp-grid{
        display:grid;
        grid-template-columns: 1.35fr .85fr;
        gap: 12px;
        align-items:start;
      }
      .fp-card{
        background: var(--surface);
        border: 1px solid var(--border);
        border-radius: var(--radius);
        box-shadow: var(--shadow-soft);
        padding: 14px;
        backdrop-filter: blur(10px);
        -webkit-backdrop-filter: blur(10px);
        overflow:hidden;
      }
      .fp-card__head{
        display:flex;
        align-items:flex-start;
        justify-content:space-between;
        gap: 12px;
        margin-bottom: 14px;
      }
      .fp-card__head h2{ margin:0; font-size: 16px; letter-spacing: -0.02em; }
      .fp-card__head p{ margin: 6px 0 0; color: var(--muted); font-size: 12px; }
      .fp-badge{
        font-size: 12px;
        padding: 6px 10px;
        border-radius: 999px;
        background: rgba(2,132,199,.10);
        border: 1px solid rgba(2,132,199,.18);
        color: #0369a1;
        white-space:nowrap;
      }
      .fp-badge--pill{
        background: rgba(15,23,42,.06);
        border-color: rgba(15,23,42,.10);
        color: rgba(15,23,42,.70);
      }
      .fp-form{ display:grid; grid-template-columns: 1fr 1fr; gap: 12px; }
      .fp-field{ display:flex; flex-direction:column; gap: 8px; }
      .fp-field--full{ grid-column: 1 / -1; }
      .fp-field label{ font-size: 12px; font-weight: 800; color: rgba(15,23,42,.78); }
      .fp-field select{
        width: 100%;
        border: 1px solid var(--border);
        border-radius: 14px;
        padding: 12px;
        font-size: 14px;
        background: rgba(255,255,255,.82);
        outline:none;
        box-shadow: 0 10px 20px rgba(15,23,42,.04);
      }
      .fp-money{
        display:flex;
        align-items:center;
        border: 1px solid var(--border);
        border-radius: 14px;
        background: rgba(255,255,255,.82);
        box-shadow: 0 10px 20px rgba(15,23,42,.04);
        overflow:hidden;
      }
      .fp-money__sym{ padding: 0 12px; color: var(--muted); font-weight: 900; }
      .fp-money input{
        border:0;
        outline:0;
        padding: 12px;
        width:100%;
        font-size: 14px;
        background: transparent;
      }
      .fp-range{ display:flex; gap: 10px; align-items:center; }
      .fp-range input[type="range"]{ width: 100%; }
      .fp-range__val{
        min-width: 92px;
        text-align:right;
        font-weight: 900;
        color: rgba(15,23,42,.85);
        background: rgba(15,23,42,.04);
        border: 1px solid rgba(15,23,42,.08);
        padding: 8px 10px;
        border-radius: 12px;
      }
      .fp-cta{ display:flex; flex-wrap:wrap; gap: 10px; margin-top: 14px; }
      .fp-charts2{ display:grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 14px; }
      .fp-mini{
        background: rgba(255,255,255,.78);
        border: 1px solid var(--border);
        border-radius: 16px;
        padding: 12px;
        box-shadow: 0 10px 20px rgba(15,23,42,.04);
        min-height: 260px;
        display:flex;
        flex-direction:column;
        gap: 10px;
      }
      .fp-mini__head{ display:flex; align-items:baseline; justify-content:space-between; gap: 10px; }
      .fp-mini__head strong{ font-size: 13px; }
      .fp-mini__hint{ color: var(--muted); font-size: 12px; }
      .fp-mini canvas{ width:100% !important; height: 190px !important; }
      .fp-table{
        border: 1px solid var(--border);
        border-radius: 16px;
        overflow:hidden;
        background: rgba(255,255,255,.65);
      }
      .fp-table__head{
        display:grid;
        grid-template-columns: 1fr 160px 140px;
        gap: 10px;
        align-items:center;
        padding: 12px;
        font-size: 12px;
        color: var(--muted);
        font-weight: 900;
        background: rgba(15,23,42,.03);
        border-bottom: 1px solid var(--border);
      }
      .fp-row{
        width:100%;
        border:0;
        background: transparent;
        display:grid;
        grid-template-columns: 1fr 160px 140px;
        gap: 10px;
        align-items:center;
        padding: 12px;
        border-bottom: 1px solid rgba(148,163,184,.25);
        cursor:pointer;
        text-align:left;
      }
      .fp-row:last-child{ border-bottom:0; }
      .fp-row:hover{ background: rgba(37,99,235,.06); }
      .fp-row--active{ background: rgba(37,99,235,.10); outline: 1px solid rgba(37,99,235,.16); }
      .fp-right{ text-align:right; }
      .fp-name{ display:flex; flex-direction:column; gap: 2px; }
      .fp-name strong{ display:flex; align-items:center; gap: 10px; font-size: 14px; }
      .fp-name small{ color: var(--muted); font-size: 12px; }
      .fp-dot{ width: 10px; height: 10px; border-radius: 999px; display:inline-block; }
      .fp-dot--ok{ background: var(--ok); box-shadow: 0 0 0 6px rgba(22,163,74,.10); }
      .fp-dot--info{ background: var(--primary); box-shadow: 0 0 0 6px rgba(37,99,235,.12); }
      .fp-dot--warn{ background: var(--warn); box-shadow: 0 0 0 6px rgba(245,158,11,.12); }
      .fp-pill{
        display:inline-flex;
        align-items:center;
        justify-content:center;
        padding: 6px 10px;
        border-radius: 999px;
        font-size: 12px;
        font-weight: 900;
        border: 1px solid var(--border);
      }
      .fp-pill--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color:#166534; }
      .fp-pill--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color:#92400e; }
      .fp-pill--crit{ background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.22); color:#991b1b; }
      .fp-sidecharts{ display:grid; grid-template-columns: 1fr; gap: 12px; margin-top: 14px; }
      @media (max-width: 1200px){
        .fp-kpis{ grid-template-columns: repeat(2, minmax(0, 1fr)); }
        .fp-grid{ grid-template-columns: 1fr; }
        .fp-charts2{ grid-template-columns: 1fr; }
      }
      @media (max-width: 640px){
        .fp-kpis{ grid-template-columns: 1fr; }
        .fp-form{ grid-template-columns: 1fr; }
        .fp-table__head, .fp-row{ grid-template-columns: 1fr 120px 100px; }
      }
    </style>
    <div class="fp-shell">
      <header class="fp-topbar">
        <div class="fp-title">
          <div class="fp-mark">▶</div>
          <div>
            <h1>Proyección Financiera <span class="fp-sub">(12 meses)</span></h1>
            <p>Ajusta parámetros y observa el recalculo en tiempo real.</p>
          </div>
        </div>

        <div class="fp-actions">
          <button class="fp-btn fp-btn--ghost" type="button" id="btnGuardar">Guardar</button>
          <button class="fp-btn fp-btn--soft" type="button" id="btnExportar">Exportar</button>
          <div class="fp-chip fp-chip--count" title="Escenarios activos">3</div>
          <button class="fp-icon" type="button" title="Ajustes">🎚️</button>
        </div>
      </header>

      <section class="fp-kpis">
        <article class="fp-kpi">
          <div class="fp-kpi__top">
            <span>Ingresos proyectados</span>
            <span class="fp-tag fp-tag--ok" id="kpiIngTag">↑ 2.5% / mes</span>
          </div>
          <div class="fp-kpi__value" id="kpiIngresos">$0</div>
          <div class="fp-kpi__sub">Total acumulado en 12 meses</div>
        </article>

        <article class="fp-kpi">
          <div class="fp-kpi__top">
            <span>Costos proyectados</span>
            <span class="fp-tag fp-tag--neutral" id="kpiCostTag">Inflación +4.0%</span>
          </div>
          <div class="fp-kpi__value" id="kpiCostos">$0</div>
          <div class="fp-kpi__sub">Total acumulado en 12 meses</div>
        </article>

        <article class="fp-kpi">
          <div class="fp-kpi__top">
            <span>Flujo neto estimado</span>
            <span class="fp-tag fp-tag--info" id="kpiFlujoTag">Base</span>
          </div>
          <div class="fp-kpi__value" id="kpiFlujo">$0</div>
          <div class="fp-kpi__sub">Ingresos − costos (12m)</div>
        </article>

        <article class="fp-kpi">
          <div class="fp-kpi__top">
            <span>Margen estimado</span>
            <span class="fp-tag fp-tag--warn" id="kpiMargTag">Objetivo ≥ 30%</span>
          </div>
          <div class="fp-kpi__value" id="kpiMargen">0%</div>
          <div class="fp-kpi__sub">Flujo / ingresos</div>
        </article>
      </section>

      <main class="fp-grid">
        <section class="fp-card">
          <div class="fp-card__head">
            <div>
              <h2>Parámetros de proyección</h2>
              <p>Modifica supuestos y recalcula automáticamente.</p>
            </div>
            <span class="fp-badge" id="badgeHorizonte">Horizonte 12m</span>
          </div>

          <div class="fp-form">
            <div class="fp-field">
              <label>Horizonte</label>
              <select id="horizonte">
                <option value="12" selected>12 meses</option>
                <option value="24">24 meses</option>
                <option value="36">36 meses</option>
              </select>
            </div>

            <div class="fp-field">
              <label>Crecimiento mensual</label>
              <div class="fp-range">
                <input id="crec" type="range" min="0" max="10" step="0.1" value="2.5"/>
                <div class="fp-range__val"><span id="crecVal">2.5</span>%</div>
              </div>
            </div>

            <div class="fp-field">
              <label>Inflación anual</label>
              <div class="fp-range">
                <input id="infl" type="range" min="0" max="20" step="0.1" value="4.0"/>
                <div class="fp-range__val"><span id="inflVal">4.0</span>%</div>
              </div>
            </div>

            <div class="fp-field">
              <label>Tasa de descuento anual</label>
              <div class="fp-range">
                <input id="desc" type="range" min="0" max="30" step="0.1" value="10.0"/>
                <div class="fp-range__val"><span id="descVal">10.0</span>%</div>
              </div>
            </div>

            <div class="fp-field fp-field--full">
              <label>Ingresos base / mes</label>
              <div class="fp-money">
                <span class="fp-money__sym">$</span>
                <input id="ingBase" type="number" min="0" step="100" value="85000"/>
              </div>
            </div>

            <div class="fp-field fp-field--full">
              <label>Costos base / mes</label>
              <div class="fp-money">
                <span class="fp-money__sym">$</span>
                <input id="costBase" type="number" min="0" step="100" value="62000"/>
              </div>
            </div>
          </div>

          <div class="fp-cta">
            <button class="fp-btn fp-btn--primary" type="button" id="btnRecalcular">Calcular proyección</button>
            <button class="fp-btn fp-btn--ghost" type="button" id="btnGuardarEsc">Guardar escenario</button>
          </div>

          <div class="fp-charts2">
            <div class="fp-mini">
              <div class="fp-mini__head">
                <strong>Proyección mensual</strong>
                <span class="fp-mini__hint">ingresos, costos, flujo</span>
              </div>
              <canvas id="chartMensual"></canvas>
            </div>

            <div class="fp-mini">
              <div class="fp-mini__head">
                <strong>Flujo acumulado</strong>
                <span class="fp-mini__hint">caja acumulada</span>
              </div>
              <canvas id="chartAcum"></canvas>
            </div>
          </div>
        </section>

        <aside class="fp-card fp-card--side">
          <div class="fp-card__head">
            <div>
              <h2>Escenarios</h2>
              <p>Comparación base vs conservador vs agresivo.</p>
            </div>
            <span class="fp-badge fp-badge--pill">3 escenarios</span>
          </div>

          <div class="fp-table">
            <div class="fp-table__head">
              <span>Escenario</span>
              <span class="fp-right">Flujo neto</span>
              <span class="fp-right">Riesgo</span>
            </div>

            <button class="fp-row" type="button" data-scn="cons">
              <div class="fp-name">
                <span class="fp-dot fp-dot--ok"></span>
                <strong>Conservador</strong>
                <small>crec −30%</small>
              </div>
              <div class="fp-right" id="scnConsVal">$0</div>
              <div class="fp-right"><span class="fp-pill fp-pill--ok">Bajo</span></div>
            </button>

            <button class="fp-row fp-row--active" type="button" data-scn="base">
              <div class="fp-name">
                <span class="fp-dot fp-dot--info"></span>
                <strong>Base</strong>
                <small>actual</small>
              </div>
              <div class="fp-right" id="scnBaseVal">$0</div>
              <div class="fp-right"><span class="fp-pill fp-pill--warn">Medio</span></div>
            </button>

            <button class="fp-row" type="button" data-scn="agr">
              <div class="fp-name">
                <span class="fp-dot fp-dot--warn"></span>
                <strong>Agresivo</strong>
                <small>crec +30%</small>
              </div>
              <div class="fp-right" id="scnAgrVal">$0</div>
              <div class="fp-right"><span class="fp-pill fp-pill--crit">Alto</span></div>
            </button>
          </div>

          <div class="fp-sidecharts">
            <div class="fp-mini">
              <div class="fp-mini__head">
                <strong>Flujo neto por escenario</strong>
                <span class="fp-mini__hint">comparativo</span>
              </div>
              <canvas id="chartEsc"></canvas>
            </div>

            <div class="fp-mini">
              <div class="fp-mini__head">
                <strong>Mix Ingresos/Costos</strong>
                <span class="fp-mini__hint">12 meses</span>
              </div>
              <canvas id="chartMix"></canvas>
            </div>
          </div>
        </aside>
      </main>
    </div>

    <script>
      (function () {
        if (typeof Chart === "undefined") return;
        const $ = (id) => document.getElementById(id);

        const fmtMoney = (n) => {
          const v = Math.round(n);
          return v.toLocaleString("es-MX", { style: "currency", currency: "USD", maximumFractionDigits: 0 });
        };

        const clamp = (v, min, max) => Math.max(min, Math.min(max, v));

        function proyectar({ months, growthM, inflA, ing0, cost0 }) {
          const inflM = Math.pow(1 + inflA / 100, 1 / 12) - 1;

          const labels = [];
          const ingresos = [];
          const costos = [];
          const flujo = [];
          const acum = [];

          let acc = 0;
          for (let i = 1; i <= months; i += 1) {
            labels.push(`M${i}`);
            const ing = ing0 * Math.pow(1 + growthM / 100, i - 1);
            const cos = cost0 * Math.pow(1 + inflM, i - 1);
            const f = ing - cos;

            ingresos.push(ing);
            costos.push(cos);
            flujo.push(f);

            acc += f;
            acum.push(acc);
          }

          const totalIng = ingresos.reduce((a, b) => a + b, 0);
          const totalCos = costos.reduce((a, b) => a + b, 0);
          const totalF = totalIng - totalCos;
          const margen = totalIng > 0 ? (totalF / totalIng) * 100 : 0;

          return { labels, ingresos, costos, flujo, acum, totalIng, totalCos, totalF, margen };
        }

        function scenarioParams(base, factorGrowth) {
          return { ...base, growthM: base.growthM * factorGrowth };
        }

        const gridColor = "rgba(148,163,184,.25)";
        const tickColor = "rgba(100,116,139,.95)";

        const baseOpts = {
          responsive: true,
          maintainAspectRatio: false,
          plugins: {
            legend: { position: "bottom", labels: { color: tickColor, boxWidth: 10, boxHeight: 10 } },
            tooltip: { mode: "index", intersect: false },
          },
          interaction: { mode: "index", intersect: false },
          scales: {
            x: { grid: { display: false }, ticks: { color: tickColor } },
            y: { grid: { color: gridColor }, ticks: { color: tickColor } },
          },
        };

        const chartMensual = new Chart($("chartMensual"), {
          type: "bar",
          data: {
            labels: [],
            datasets: [
              { label: "Ingresos", data: [], borderWidth: 1 },
              { label: "Costos", data: [], borderWidth: 1 },
              { label: "Flujo neto", data: [], borderWidth: 1 },
            ],
          },
          options: {
            ...baseOpts,
            scales: {
              ...baseOpts.scales,
              y: { ...baseOpts.scales.y, ticks: { ...baseOpts.scales.y.ticks, callback: (v) => "$" + Number(v).toLocaleString("es-MX") } },
            },
          },
        });

        const chartAcum = new Chart($("chartAcum"), {
          type: "line",
          data: {
            labels: [],
            datasets: [{
              label: "Acumulado",
              data: [],
              borderWidth: 2,
              tension: 0.35,
              pointRadius: 2,
              fill: true,
              backgroundColor: "rgba(37,99,235,.12)",
            }],
          },
          options: {
            ...baseOpts,
            plugins: { ...baseOpts.plugins, legend: { display: false } },
            scales: {
              ...baseOpts.scales,
              y: { ...baseOpts.scales.y, ticks: { ...baseOpts.scales.y.ticks, callback: (v) => "$" + Number(v).toLocaleString("es-MX") } },
            },
          },
        });

        const chartEsc = new Chart($("chartEsc"), {
          type: "bar",
          data: {
            labels: ["Conservador", "Base", "Agresivo"],
            datasets: [{ label: "Flujo neto", data: [0, 0, 0], borderWidth: 1 }],
          },
          options: {
            ...baseOpts,
            indexAxis: "y",
            plugins: { ...baseOpts.plugins, legend: { display: false } },
            scales: {
              ...baseOpts.scales,
              x: { ...baseOpts.scales.x, ticks: { ...baseOpts.scales.x.ticks, callback: (v) => "$" + Number(v).toLocaleString("es-MX") } },
            },
          },
        });

        const chartMix = new Chart($("chartMix"), {
          type: "doughnut",
          data: {
            labels: ["Ingresos", "Costos"],
            datasets: [{ data: [0, 0], borderWidth: 1 }],
          },
          options: {
            responsive: true,
            maintainAspectRatio: false,
            plugins: { legend: { position: "bottom", labels: { color: tickColor, boxWidth: 10, boxHeight: 10 } } },
            cutout: "68%",
          },
        });

        const state = {
          months: 12,
          growthM: 2.5,
          inflA: 4.0,
          discA: 10.0,
          ing0: 85000,
          cost0: 62000,
          activeScenario: "base",
        };

        function syncLabels() {
          $("crecVal").textContent = Number(state.growthM).toFixed(1);
          $("inflVal").textContent = Number(state.inflA).toFixed(1);
          $("descVal").textContent = Number(state.discA).toFixed(1);

          $("kpiIngTag").textContent = `↑ ${Number(state.growthM).toFixed(1)}% / mes`;
          $("kpiCostTag").textContent = `Inflación +${Number(state.inflA).toFixed(1)}%`;
          $("badgeHorizonte").textContent = `Horizonte ${state.months}m`;
        }

        function setMarginTag(m) {
          const el = $("kpiMargTag");
          el.classList.remove("fp-tag--ok", "fp-tag--warn", "fp-tag--crit");
          if (m >= 30) { el.textContent = "En meta"; el.classList.add("fp-tag--ok"); }
          else if (m >= 20) { el.textContent = "Cerca de meta"; el.classList.add("fp-tag--warn"); }
          else { el.textContent = "Bajo"; el.classList.add("fp-tag--crit"); }
        }

        function recalc() {
          syncLabels();

          const baseParams = {
            months: state.months,
            growthM: state.growthM,
            inflA: state.inflA,
            ing0: state.ing0,
            cost0: state.cost0,
          };

          const resBase = proyectar(baseParams);
          const resCons = proyectar(scenarioParams(baseParams, 0.7));
          const resAgr = proyectar(scenarioParams(baseParams, 1.3));

          $("kpiIngresos").textContent = fmtMoney(resBase.totalIng);
          $("kpiCostos").textContent = fmtMoney(resBase.totalCos);
          $("kpiFlujo").textContent = fmtMoney(resBase.totalF);
          $("kpiMargen").textContent = `${clamp(resBase.margen, -999, 999).toFixed(1)}%`;
          setMarginTag(resBase.margen);

          $("scnConsVal").textContent = fmtMoney(resCons.totalF);
          $("scnBaseVal").textContent = fmtMoney(resBase.totalF);
          $("scnAgrVal").textContent = fmtMoney(resAgr.totalF);

          const showMonths = Math.min(state.months, 12);
          chartMensual.data.labels = resBase.labels.slice(0, showMonths);
          chartMensual.data.datasets[0].data = resBase.ingresos.slice(0, showMonths);
          chartMensual.data.datasets[1].data = resBase.costos.slice(0, showMonths);
          chartMensual.data.datasets[2].data = resBase.flujo.slice(0, showMonths);
          chartMensual.update();

          chartAcum.data.labels = resBase.labels.slice(0, showMonths);
          chartAcum.data.datasets[0].data = resBase.acum.slice(0, showMonths);
          chartAcum.update();

          chartEsc.data.datasets[0].data = [resCons.totalF, resBase.totalF, resAgr.totalF];
          chartEsc.update();

          chartMix.data.datasets[0].data = [resBase.totalIng, resBase.totalCos];
          chartMix.update();
        }

        function bind() {
          $("horizonte").addEventListener("change", (e) => {
            state.months = Number(e.target.value);
            recalc();
          });

          $("crec").addEventListener("input", (e) => {
            state.growthM = Number(e.target.value);
            recalc();
          });

          $("infl").addEventListener("input", (e) => {
            state.inflA = Number(e.target.value);
            recalc();
          });

          $("desc").addEventListener("input", (e) => {
            state.discA = Number(e.target.value);
            recalc();
          });

          $("ingBase").addEventListener("input", (e) => {
            state.ing0 = Number(e.target.value || 0);
            recalc();
          });

          $("costBase").addEventListener("input", (e) => {
            state.cost0 = Number(e.target.value || 0);
            recalc();
          });

          $("btnRecalcular").addEventListener("click", recalc);

          document.querySelectorAll(".fp-row").forEach((btn) => {
            btn.addEventListener("click", () => {
              document.querySelectorAll(".fp-row").forEach((b) => b.classList.remove("fp-row--active"));
              btn.classList.add("fp-row--active");
              state.activeScenario = btn.dataset.scn;
              $("kpiFlujoTag").textContent =
                state.activeScenario === "base" ? "Base" :
                state.activeScenario === "cons" ? "Conservador" : "Agresivo";
            });
          });

          $("btnGuardar").addEventListener("click", () => alert("Demo: Guardar (conectar a backend)"));
          $("btnExportar").addEventListener("click", () => alert("Demo: Exportar (PDF/Excel)"));
          $("btnGuardarEsc").addEventListener("click", () => alert("Demo: Guardar escenario"));
        }

        bind();
        recalc();
      })();
    </script>
""")

POA_HTML = dedent("""
    <section class="pe-page">
      <style>
        .pe-page{
          --bg: #f6f8fc;
          --surface: rgba(255,255,255,.88);
          --card: #ffffff;
          --text: #0f172a;
          --muted: #64748b;
          --border: rgba(148,163,184,.38);
          --shadow: 0 18px 40px rgba(15,23,42,.08);
          --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
          --radius: 18px;
          --primary: #0f3d2e;
          --primary-2: #1f6f52;
          --accent: #2563eb;
          --ok: #16a34a;
          --warn: #f59e0b;
          --crit: #ef4444;
          width: 100%;
          font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
          color: var(--text);
          background:
            radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
            radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
            var(--bg);
          border-radius: 18px;
        }
        .pe-wrap{ width: 100%; margin: 0 auto; padding: 18px 0 34px; }
        .pe-kpis{ display:grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 14px; margin: 16px 0 18px; }
        .pe-kpi{
          background: var(--surface);
          border: 1px solid var(--border);
          border-radius: var(--radius);
          box-shadow: var(--shadow-soft);
          padding: 16px;
          backdrop-filter: blur(10px);
          -webkit-backdrop-filter: blur(10px);
          position: relative;
          overflow:hidden;
        }
        .pe-kpi:before{
          content:"";
          position:absolute;
          inset:-1px;
          background: linear-gradient(135deg, rgba(15,61,46,.12), transparent 35%, rgba(37,99,235,.10));
          opacity:.9;
          pointer-events:none;
        }
        .pe-kpi > *{ position: relative; }
        .pe-kpi__label{ color: var(--muted); font-size: 14px; font-weight: 600; }
        .pe-kpi__value{ margin-top: 10px; font-size: 34px; font-weight: 900; letter-spacing: -0.03em; }
        .pe-kpi__meta{ margin-top: 10px; display:flex; gap: 8px; flex-wrap: wrap; }
        .pe-chip{ font-size: 12px; padding: 6px 10px; border-radius: 999px; background: rgba(15,23,42,.05); border: 1px solid rgba(15,23,42,.08); color: rgba(15,23,42,.72); }
        .pe-chip--info{ background: rgba(37,99,235,.10); border-color: rgba(37,99,235,.18); color: #1d4ed8; }
        .pe-chip--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
        .pe-kpi--progress{ display:flex; flex-direction:column; }
        .pe-kpi__progress{ margin-top: 8px; display:flex; align-items:center; justify-content:flex-start; }
        .pe-ring{
          --p: 76;
          width: 86px;
          height: 86px;
          border-radius: 999px;
          background: conic-gradient(rgba(15,61,46,1) calc(var(--p) * 1%), rgba(148,163,184,.25) 0);
          display:grid;
          place-items:center;
          border: 1px solid rgba(148,163,184,.25);
          box-shadow: 0 12px 24px rgba(15,23,42,.06);
        }
        .pe-ring__inner{ width: 66px; height: 66px; border-radius: 999px; background: rgba(255,255,255,.90); display:flex; flex-direction:column; align-items:center; justify-content:center; border: 1px solid rgba(148,163,184,.25); }
        .pe-ring__val{ font-weight: 900; font-size: 16px; letter-spacing: -0.02em; }
        .pe-ring__sub{ font-size: 11px; color: var(--muted); }
        .pe-grid{ display:grid; grid-template-columns: 1.35fr .85fr; gap: 14px; align-items:start; }
        .pe-card{
          background: var(--surface);
          border: 1px solid var(--border);
          border-radius: 22px;
          box-shadow: var(--shadow-soft);
          padding: 16px;
          backdrop-filter: blur(10px);
          -webkit-backdrop-filter: blur(10px);
          overflow:hidden;
        }
        .pe-card__head{ display:flex; align-items:flex-start; justify-content:space-between; gap: 12px; margin-bottom: 14px; }
        .pe-card__head h2{ margin:0; font-size: 20px; letter-spacing: -0.02em; }
        .pe-card__head p{ margin: 6px 0 0; color: var(--muted); font-size: 13px; }
        .pe-card__tools{ display:flex; align-items:center; gap: 10px; }
        .pe-pill{ font-size: 12px; padding: 6px 10px; border-radius: 999px; background: rgba(255,255,255,.70); border: 1px solid var(--border); color: rgba(15,23,42,.72); white-space: nowrap; }
        .pe-pill--soft{ background: rgba(15,61,46,.10); border-color: rgba(15,61,46,.18); color: #0b2a20; }
        .pe-list{ display:flex; flex-direction:column; gap: 12px; }
        .pe-item{ background: rgba(255,255,255,.80); border: 1px solid rgba(148,163,184,.30); border-radius: 18px; padding: 14px 14px 12px; box-shadow: 0 10px 20px rgba(15,23,42,.04); }
        .pe-item--active{ outline: 1px solid rgba(15,61,46,.22); background: rgba(15,61,46,.06); }
        .pe-item__top{ display:flex; align-items:flex-start; justify-content:space-between; gap: 12px; }
        .pe-item__title{ display:flex; gap: 10px; align-items:flex-start; line-height: 1.35; }
        .pe-code{ font-weight: 900; color: #0b2a20; background: rgba(15,61,46,.10); border: 1px solid rgba(15,61,46,.18); padding: 6px 10px; border-radius: 999px; font-size: 12px; white-space: nowrap; }
        .pe-item__meta{ margin-top: 10px; display:flex; gap: 10px; flex-wrap:wrap; color: var(--muted); }
        .pe-mini{ font-size: 12px; padding: 6px 10px; border-radius: 999px; background: rgba(15,23,42,.04); border: 1px solid rgba(15,23,42,.08); }
        .pe-status{ font-size: 12px; font-weight: 800; padding: 6px 10px; border-radius: 999px; border: 1px solid var(--border); background: rgba(255,255,255,.70); color: rgba(15,23,42,.72); white-space: nowrap; }
        .pe-status--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }
        .pe-status--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
        .pe-status--neutral{ background: rgba(15,23,42,.05); border-color: rgba(15,23,42,.10); color: rgba(15,23,42,.70); }
        .pe-bar{ margin-top: 12px; height: 10px; border-radius: 999px; background: rgba(148,163,184,.25); border: 1px solid rgba(148,163,184,.25); overflow:hidden; }
        .pe-bar__fill{ height: 100%; width: 50%; border-radius: 999px; background: linear-gradient(90deg, rgba(37,99,235,1), rgba(96,165,250,1)); }
        .pe-bar__fill--ok{ background: linear-gradient(90deg, rgba(15,61,46,1), rgba(31,111,82,1)); }
        .pe-bar__fill--warn{ background: linear-gradient(90deg, rgba(245,158,11,1), rgba(253,230,138,1)); }
        .pe-axis{ display:flex; flex-direction:column; gap: 12px; }
        .pe-axis__btn{ width:100%; text-align:left; border: 1px solid rgba(148,163,184,.30); background: rgba(255,255,255,.80); border-radius: 18px; padding: 14px; display:flex; align-items:center; gap: 12px; }
        .pe-axis__btn--active{ background: rgba(15,61,46,.06); outline: 1px solid rgba(15,61,46,.22); }
        .pe-axis__dot{ width: 10px; height: 10px; border-radius: 999px; background: var(--primary); box-shadow: 0 0 0 6px rgba(15,61,46,.10); }
        .pe-axis__dot--alt{ background: #2563eb; box-shadow: 0 0 0 6px rgba(37,99,235,.12); }
        .pe-axis__dot--alt2{ background: #7c3aed; box-shadow: 0 0 0 6px rgba(124,58,237,.12); }
        .pe-axis__dot--alt3{ background: #f59e0b; box-shadow: 0 0 0 6px rgba(245,158,11,.12); }
        .pe-axis__count{ margin-left:auto; font-weight: 900; color: rgba(15,23,42,.72); background: rgba(15,23,42,.04); border: 1px solid rgba(15,23,42,.08); padding: 6px 10px; border-radius: 999px; }
        .pe-roadmap{ margin-top: 14px; }
        .pe-timeline{ display:grid; grid-template-columns: repeat(4, minmax(0,1fr)); gap: 12px; }
        .pe-tlcol{ background: rgba(255,255,255,.70); border: 1px solid rgba(148,163,184,.30); border-radius: 18px; padding: 12px; }
        .pe-tlcol__head{ font-weight: 900; color: rgba(15,23,42,.78); padding: 8px 10px; border-radius: 999px; background: rgba(15,23,42,.04); border: 1px solid rgba(15,23,42,.08); display:inline-block; margin-bottom: 10px; }
        .pe-tlitem{ background: rgba(255,255,255,.82); border: 1px solid rgba(148,163,184,.30); border-radius: 16px; padding: 12px; box-shadow: 0 10px 20px rgba(15,23,42,.04); margin-bottom: 10px; }
        .pe-tlitem:last-child{ margin-bottom: 0; }
        .pe-tlitem__top{ display:flex; align-items:flex-start; justify-content:space-between; gap: 10px; }
        .pe-tlitem p{ margin: 8px 0 0; color: var(--muted); font-size: 12.5px; line-height: 1.4; }
        .pe-tlmeta{ margin-top: 10px; display:flex; gap: 8px; flex-wrap:wrap; }
        .pe-tlitem--ok{ outline: 1px solid rgba(22,163,74,.18); background: rgba(22,163,74,.06); }
        .pe-tlitem--warn{ outline: 1px solid rgba(245,158,11,.18); background: rgba(245,158,11,.07); }
        .pe-tlitem--soft{ background: rgba(15,61,46,.06); outline: 1px solid rgba(15,61,46,.18); }
        @media (max-width: 1100px){
          .pe-kpis{ grid-template-columns: repeat(2, minmax(0,1fr)); }
          .pe-grid{ grid-template-columns: 1fr; }
          .pe-timeline{ grid-template-columns: repeat(2, minmax(0,1fr)); }
        }
        @media (max-width: 640px){
          .pe-kpis{ grid-template-columns: 1fr; }
          .pe-timeline{ grid-template-columns: 1fr; }
        }
      </style>

      <div class="pe-wrap">
        __CREATE_POA_BUTTON__
        <section class="pe-kpis">
          <article class="pe-kpi">
            <div class="pe-kpi__label">Actividades programadas</div>
            <div class="pe-kpi__value">64</div>
            <div class="pe-kpi__meta"><span class="pe-chip pe-chip--info">POA anual</span></div>
          </article>
          <article class="pe-kpi">
            <div class="pe-kpi__label">Actividades en ejecución</div>
            <div class="pe-kpi__value">39</div>
            <div class="pe-kpi__meta"><span class="pe-chip">61% activas</span></div>
          </article>
          <article class="pe-kpi">
            <div class="pe-kpi__label">Desviaciones críticas</div>
            <div class="pe-kpi__value">5</div>
            <div class="pe-kpi__meta"><span class="pe-chip pe-chip--warn">2 sin plan</span></div>
          </article>
          <article class="pe-kpi pe-kpi--progress">
            <div class="pe-kpi__label">Cumplimiento trimestral</div>
            <div class="pe-kpi__progress">
              <div class="pe-ring" style="--p:76;">
                <div class="pe-ring__inner">
                  <div class="pe-ring__val">76%</div>
                  <div class="pe-ring__sub">avance</div>
                </div>
              </div>
            </div>
          </article>
        </section>

        <section class="pe-grid">
          <article class="pe-card">
            <div class="pe-card__head">
              <div>
                <h2>Metas operativas priorizadas</h2>
                <p>Seguimiento por meta con estado y avance.</p>
              </div>
              <div class="pe-card__tools"><span class="pe-pill pe-pill--soft">4 metas clave</span></div>
            </div>
            <div class="pe-list">
              <article class="pe-item pe-item--active">
                <div class="pe-item__top"><div class="pe-item__title"><span class="pe-code">MO-01</span><strong>Reducir tiempos de atención de solicitudes en 20%.</strong></div><span class="pe-status pe-status--ok">En meta</span></div>
                <div class="pe-item__meta"><span class="pe-mini">Responsable: Operaciones</span><span class="pe-mini">Q2</span></div>
                <div class="pe-bar"><div class="pe-bar__fill pe-bar__fill--ok" style="width:82%"></div></div>
              </article>
              <article class="pe-item">
                <div class="pe-item__top"><div class="pe-item__title"><span class="pe-code">MO-02</span><strong>Estandarizar procesos clave en áreas operativas.</strong></div><span class="pe-status pe-status--warn">En riesgo</span></div>
                <div class="pe-item__meta"><span class="pe-mini">Responsable: Calidad</span><span class="pe-mini">Q3</span></div>
                <div class="pe-bar"><div class="pe-bar__fill pe-bar__fill--warn" style="width:64%"></div></div>
              </article>
              <article class="pe-item">
                <div class="pe-item__top"><div class="pe-item__title"><span class="pe-code">MO-03</span><strong>Incrementar productividad por célula de trabajo.</strong></div><span class="pe-status pe-status--neutral">Seguimiento</span></div>
                <div class="pe-item__meta"><span class="pe-mini">Responsable: Dirección</span><span class="pe-mini">Q4</span></div>
                <div class="pe-bar"><div class="pe-bar__fill" style="width:70%"></div></div>
              </article>
            </div>
          </article>

          <aside class="pe-card">
            <div class="pe-card__head">
              <div>
                <h2>Responsables por frente</h2>
                <p>Asignación de liderazgo operativo.</p>
              </div>
            </div>
            <div class="pe-axis">
              <button class="pe-axis__btn pe-axis__btn--active" type="button"><span class="pe-axis__dot"></span><span>Operaciones: Dirección Operativa</span><span class="pe-axis__count">16</span></button>
              <button class="pe-axis__btn" type="button"><span class="pe-axis__dot pe-axis__dot--alt"></span><span>Calidad: Mejora Continua</span><span class="pe-axis__count">14</span></button>
              <button class="pe-axis__btn" type="button"><span class="pe-axis__dot pe-axis__dot--alt2"></span><span>Tecnología: Transformación Digital</span><span class="pe-axis__count">18</span></button>
              <button class="pe-axis__btn" type="button"><span class="pe-axis__dot pe-axis__dot--alt3"></span><span>Control: PMO / Planeación</span><span class="pe-axis__count">16</span></button>
            </div>
          </aside>
        </section>

        <article class="pe-card pe-roadmap">
          <div class="pe-card__head">
            <div>
              <h2>Seguimiento trimestral y cronograma</h2>
              <p>Hitos, avance y estado por trimestre.</p>
            </div>
          </div>
          <div class="pe-timeline">
            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q1</div>
              <article class="pe-tlitem pe-tlitem--ok">
                <div class="pe-tlitem__top"><strong>18 hitos</strong><span class="pe-status pe-status--ok">16 / 18</span></div>
                <p>Arranque operativo y ajustes de línea base.</p>
                <div class="pe-tlmeta"><span class="pe-mini">Estado: En meta</span></div>
              </article>
            </div>
            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q2</div>
              <article class="pe-tlitem pe-tlitem--warn">
                <div class="pe-tlitem__top"><strong>16 hitos</strong><span class="pe-status pe-status--warn">12 / 16</span></div>
                <p>Escalamiento de iniciativas y control de costos.</p>
                <div class="pe-tlmeta"><span class="pe-mini">Estado: En riesgo</span></div>
              </article>
            </div>
            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q3</div>
              <article class="pe-tlitem pe-tlitem--soft">
                <div class="pe-tlitem__top"><strong>15 hitos</strong><span class="pe-status pe-status--neutral">0 / 15</span></div>
                <p>Consolidación de mejoras y auditoría de avance.</p>
                <div class="pe-tlmeta"><span class="pe-mini">Estado: Pendiente</span></div>
              </article>
            </div>
            <div class="pe-tlcol">
              <div class="pe-tlcol__head">Q4</div>
              <article class="pe-tlitem">
                <div class="pe-tlitem__top"><strong>15 hitos</strong><span class="pe-status pe-status--neutral">0 / 15</span></div>
                <p>Cierre, evaluación de resultados y lecciones aprendidas.</p>
                <div class="pe-tlmeta"><span class="pe-mini">Estado: Programado</span></div>
              </article>
            </div>
          </div>
        </article>
      </div>
    </section>
""")

POA_CREAR_HTML = dedent("""
    <section class="poac-wrap">
      <style>
        .poac-wrap *{ box-sizing:border-box; }
        .poac-wrap{
          font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
          color:#0f172a;
        }
        .poac-head{
          display:flex;
          justify-content:space-between;
          align-items:center;
          gap:10px;
          margin-bottom: 12px;
        }
        .poac-btn{
          border:1px solid rgba(148,163,184,.38);
          border-radius: 12px;
          padding: 9px 12px;
          background:#fff;
          cursor:pointer;
          font-weight:700;
        }
        .poac-btn.primary{
          background:#0f3d2e;
          color:#fff;
          border-color:#0f3d2e;
        }
        .poac-grid{
          display:grid;
          grid-template-columns: minmax(300px, 360px) minmax(0, 1fr);
          gap: 12px;
        }
        .poac-card{
          background: rgba(255,255,255,.92);
          border:1px solid rgba(148,163,184,.32);
          border-radius: 16px;
          padding: 12px;
        }
        .poac-list{ display:flex; flex-direction:column; gap:8px; max-height: 68vh; overflow:auto; }
        .poac-item{
          border:1px solid rgba(148,163,184,.32);
          border-radius: 12px;
          background:#fff;
          padding:10px;
          cursor:pointer;
          text-align:left;
        }
        .poac-item.active{ background: rgba(15,61,46,.08); border-color: rgba(15,61,46,.30); }
        .poac-sub{ color:#64748b; font-size: 12px; margin-top: 4px; }
        .poac-field{ display:flex; flex-direction:column; gap:6px; margin-top:10px; }
        .poac-field label{ font-size:12px; color:#475569; font-weight:700; }
        .poac-input, .poac-textarea{
          width:100%;
          border:1px solid rgba(148,163,184,.42);
          border-radius: 10px;
          padding:10px 12px;
          background:#fff;
          font-size:14px;
        }
        .poac-textarea{ min-height: 80px; resize:vertical; }
        .poac-row{ display:grid; grid-template-columns: 1fr 1fr; gap:8px; }
        .poac-activity{ border:1px solid rgba(148,163,184,.32); border-radius: 12px; background:#fff; padding: 10px; margin-top: 10px; }
        .poac-activity-head{ display:flex; justify-content:space-between; gap:8px; align-items:flex-start; }
        .poac-inline-actions{ display:flex; gap:6px; flex-wrap:wrap; margin-top:8px; }
        .poac-subactivity{ margin-top:8px; border-top:1px dashed rgba(148,163,184,.32); padding-top:8px; }
        .poac-msg{ margin-top:10px; min-height:1.2em; font-size:13px; color:#0f3d2e; }
        .poac-status{
          display:inline-flex;
          align-items:center;
          gap:4px;
          font-size:11px;
          font-weight:800;
          border-radius:999px;
          padding:4px 8px;
          border:1px solid rgba(148,163,184,.35);
          background: rgba(255,255,255,.82);
          margin-top:6px;
        }
        .poac-status.ni{ background: rgba(148,163,184,.18); color:#475569; }
        .poac-status.ep{ background: rgba(245,158,11,.16); color:#92400e; }
        .poac-status.te{ background: rgba(22,163,74,.15); color:#166534; }
        .poac-approval-card{
          margin-top: 12px;
          border:1px solid rgba(148,163,184,.32);
          border-radius: 12px;
          padding: 10px;
          background:#fff;
        }
        @media (max-width: 980px){
          .poac-grid{ grid-template-columns: 1fr; }
          .poac-row{ grid-template-columns: 1fr; }
        }
      </style>

      <div class="poac-head">
        <div>
          <h2 style="margin:0;">Tablero de creación de POA</h2>
          <p style="margin:4px 0 0;color:#64748b;">Trabaja sobre ejes y objetivos estratégicos asignados a tu usuario.</p>
        </div>
        <button class="poac-btn" type="button" onclick="window.location.href='/poa'">Volver a POA</button>
      </div>

      <div class="poac-grid">
        <aside class="poac-card">
          <h3 style="margin:0;font-size:16px;">Ejes y objetivos asignados</h3>
          <div class="poac-list" id="poac-objective-list"></div>
        </aside>
        <section class="poac-card">
          <h3 style="margin:0;font-size:16px;">Actividad POA</h3>
          <div class="poac-row">
            <div class="poac-field">
              <label for="poac-act-name">Nombre</label>
              <input id="poac-act-name" class="poac-input" type="text" placeholder="Nombre de la actividad">
            </div>
            <div class="poac-field">
              <label for="poac-act-code">Código</label>
              <input id="poac-act-code" class="poac-input" type="text" placeholder="Código">
            </div>
          </div>
          <div class="poac-row">
            <div class="poac-field">
              <label for="poac-act-owner">Responsable</label>
              <input id="poac-act-owner" class="poac-input" type="text" placeholder="Responsable">
            </div>
            <div class="poac-field">
              <label for="poac-act-deliverable">Entregable</label>
              <input id="poac-act-deliverable" class="poac-input" type="text" placeholder="Entregable">
            </div>
          </div>
          <div class="poac-row">
            <div class="poac-field">
              <label for="poac-act-start">Fecha inicial</label>
              <input id="poac-act-start" class="poac-input" type="date">
            </div>
            <div class="poac-field">
              <label for="poac-act-end">Fecha final</label>
              <input id="poac-act-end" class="poac-input" type="date">
            </div>
          </div>
          <div class="poac-field">
            <label for="poac-act-desc">Descripción</label>
            <textarea id="poac-act-desc" class="poac-textarea" placeholder="Descripción"></textarea>
          </div>
          <div class="poac-inline-actions">
            <button class="poac-btn primary" id="poac-add-activity" type="button">Agregar actividad</button>
          </div>

          <div id="poac-activities"></div>
          <div class="poac-approval-card">
            <h4 style="margin:0 0 6px 0; font-size:14px;">Aprobaciones pendientes de entregables</h4>
            <div id="poac-approvals"></div>
          </div>
          <div class="poac-msg" id="poac-msg" aria-live="polite"></div>
        </section>
      </div>

      <script>
        (() => {
          const objectiveListEl = document.getElementById("poac-objective-list");
          const activitiesEl = document.getElementById("poac-activities");
          const approvalsEl = document.getElementById("poac-approvals");
          const msgEl = document.getElementById("poac-msg");
          const addActivityBtn = document.getElementById("poac-add-activity");
          const actNameEl = document.getElementById("poac-act-name");
          const actCodeEl = document.getElementById("poac-act-code");
          const actOwnerEl = document.getElementById("poac-act-owner");
          const actDeliverableEl = document.getElementById("poac-act-deliverable");
          const actStartEl = document.getElementById("poac-act-start");
          const actEndEl = document.getElementById("poac-act-end");
          const actDescEl = document.getElementById("poac-act-desc");

          let objectives = [];
          let activities = [];
          let pendingApprovals = [];
          let selectedObjectiveId = null;

          const showMsg = (text, isError = false) => {
            if (!msgEl) return;
            msgEl.style.color = isError ? "#b91c1c" : "#0f3d2e";
            msgEl.textContent = text || "";
          };

          const requestJson = async (url, options = {}) => {
            const response = await fetch(url, {
              headers: { "Content-Type": "application/json" },
              credentials: "same-origin",
              ...options,
            });
            const payload = await response.json().catch(() => ({}));
            if (!response.ok || payload.success === false) {
              throw new Error(payload.error || "No se pudo completar la operación.");
            }
            return payload;
          };

          const selectedObjective = () => objectives.find((item) => item.id === selectedObjectiveId) || null;
          const statusClass = (status) => {
            if (status === "Terminada") return "te";
            if (status === "No iniciada") return "ni";
            return "ep";
          };
          const visualRangeError = (start, end, label) => {
            if (!start || !end) return `${label}: fecha inicial y fecha final son obligatorias.`;
            if (start > end) return `${label}: la fecha inicial no puede ser mayor que la final.`;
            return "";
          };
          const visualChildRangeError = (childStart, childEnd, parentStart, parentEnd, childLabel, parentLabel) => {
            if (!parentStart || !parentEnd) return `${parentLabel} no tiene fechas definidas para delimitar ${childLabel.toLowerCase()}.`;
            if (childStart < parentStart || childEnd > parentEnd) {
              return `${childLabel} debe quedar dentro de ${parentLabel} (${parentStart} a ${parentEnd}).`;
            }
            return "";
          };

          const renderObjectives = () => {
            if (!objectiveListEl) return;
            objectiveListEl.innerHTML = objectives.map((obj) => `
              <button class="poac-item ${obj.id === selectedObjectiveId ? "active" : ""}" type="button" data-obj-id="${obj.id}">
                <strong>${obj.codigo || "OBJ"} - ${obj.nombre}</strong>
                <div class="poac-sub">${obj.axis_name || "Sin eje"} | ${obj.fecha_inicial || "-"} a ${obj.fecha_final || "-"}</div>
              </button>
            `).join("");
            objectiveListEl.querySelectorAll("[data-obj-id]").forEach((button) => {
              button.addEventListener("click", () => {
                selectedObjectiveId = Number(button.getAttribute("data-obj-id"));
                renderObjectives();
                renderActivities();
              });
            });
          };

          const renderActivities = () => {
            const objective = selectedObjective();
            if (!activitiesEl || !objective) {
              if (activitiesEl) activitiesEl.innerHTML = "";
              return;
            }
            const objectiveActivities = activities.filter((item) => item.objective_id === objective.id);
            activitiesEl.innerHTML = objectiveActivities.map((activity) => `
              <article class="poac-activity" data-activity-id="${activity.id}">
                <div class="poac-activity-head">
                  <div>
                    <strong>${activity.codigo || "ACT"} - ${activity.nombre}</strong>
                    <div class="poac-sub">Responsable: ${activity.responsable || "-"} | Entregable: ${activity.entregable || "-"}</div>
                    <div class="poac-sub">Fechas: ${activity.fecha_inicial || "-"} a ${activity.fecha_final || "-"}</div>
                    <span class="poac-status ${statusClass(activity.status)}">${activity.status}</span>
                    ${(activity.entrega_estado === "pendiente") ? '<div class="poac-sub">Aprobación de entregable pendiente.</div>' : ''}
                    ${(activity.entrega_estado === "rechazada") ? '<div class="poac-sub">Entregable rechazado, requiere ajustes.</div>' : ''}
                  </div>
                </div>
                <div class="poac-sub">${activity.descripcion || ""}</div>
                <div class="poac-inline-actions">
                  <button class="poac-btn" type="button" data-act="edit">Editar</button>
                  <button class="poac-btn" type="button" data-act="delete">Eliminar</button>
                  ${(activity.status !== "Terminada" && activity.entrega_estado !== "pendiente") ? '<button class="poac-btn primary" type="button" data-act="finish">Terminar</button>' : ''}
                </div>
                <div class="poac-subactivity">
                  <strong style="font-size:13px;">Subactividades</strong>
                  ${(activity.subactivities || []).map((sub) => `
                    <div class="poac-sub" data-sub-id="${sub.id}">
                      ${sub.codigo || "SUB"} - ${sub.nombre} | Responsable: ${sub.responsable || "-"} | Entregable: ${sub.entregable || "-"} | ${sub.fecha_inicial || "-"} a ${sub.fecha_final || "-"}
                    </div>
                  `).join("") || '<div class="poac-sub">Sin subactividades.</div>'}
                  <div class="poac-row" style="margin-top:8px;">
                    <input class="poac-input" data-sub-field="nombre" type="text" placeholder="Nombre subactividad">
                    <input class="poac-input" data-sub-field="codigo" type="text" placeholder="Código">
                  </div>
                  <div class="poac-row" style="margin-top:8px;">
                    <input class="poac-input" data-sub-field="responsable" type="text" placeholder="Responsable">
                    <input class="poac-input" data-sub-field="entregable" type="text" placeholder="Entregable">
                  </div>
                  <div class="poac-row" style="margin-top:8px;">
                    <input class="poac-input" data-sub-field="fecha_inicial" type="date" placeholder="Fecha inicial">
                    <input class="poac-input" data-sub-field="fecha_final" type="date" placeholder="Fecha final">
                  </div>
                  <textarea class="poac-textarea" data-sub-field="descripcion" placeholder="Descripción" style="margin-top:8px;"></textarea>
                  <div class="poac-inline-actions">
                    <button class="poac-btn primary" type="button" data-act="add-sub">Agregar subactividad</button>
                  </div>
                </div>
              </article>
            `).join("");

            activitiesEl.querySelectorAll("[data-act='delete']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-activity-id]");
                const id = Number(card.getAttribute("data-activity-id"));
                if (!window.confirm("¿Eliminar esta actividad y sus subactividades?")) return;
                try {
                  await requestJson(`/api/poa/activities/${id}`, { method: "DELETE" });
                  await loadBoard();
                  showMsg("Actividad eliminada.");
                } catch (err) {
                  showMsg(err.message || "No se pudo eliminar.", true);
                }
              });
            });

            activitiesEl.querySelectorAll("[data-act='edit']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-activity-id]");
                const id = Number(card.getAttribute("data-activity-id"));
                const source = objectiveActivities.find((a) => a.id === id);
                if (!source) return;
                actNameEl.value = source.nombre || "";
                actCodeEl.value = source.codigo || "";
                actOwnerEl.value = source.responsable || "";
                actDeliverableEl.value = source.entregable || "";
                actStartEl.value = source.fecha_inicial || "";
                actEndEl.value = source.fecha_final || "";
                actDescEl.value = source.descripcion || "";
                addActivityBtn.textContent = "Guardar actividad";
                addActivityBtn.dataset.editId = String(id);
              });
            });

            activitiesEl.querySelectorAll("[data-act='finish']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-activity-id]");
                const id = Number(card.getAttribute("data-activity-id"));
                if (!window.confirm("¿Enviar entregable a aprobación para terminar la actividad?")) return;
                try {
                  await requestJson(`/api/poa/activities/${id}/request-completion`, { method: "POST" });
                  await loadBoard();
                  showMsg("Solicitud de aprobación enviada al dueño del proceso.");
                } catch (err) {
                  showMsg(err.message || "No se pudo solicitar aprobación.", true);
                }
              });
            });

            activitiesEl.querySelectorAll("[data-act='add-sub']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-activity-id]");
                const activityId = Number(card.getAttribute("data-activity-id"));
                const body = {
                  nombre: (card.querySelector("[data-sub-field='nombre']")?.value || "").trim(),
                  codigo: (card.querySelector("[data-sub-field='codigo']")?.value || "").trim(),
                  responsable: (card.querySelector("[data-sub-field='responsable']")?.value || "").trim(),
                  entregable: (card.querySelector("[data-sub-field='entregable']")?.value || "").trim(),
                  fecha_inicial: (card.querySelector("[data-sub-field='fecha_inicial']")?.value || "").trim(),
                  fecha_final: (card.querySelector("[data-sub-field='fecha_final']")?.value || "").trim(),
                  descripcion: (card.querySelector("[data-sub-field='descripcion']")?.value || "").trim(),
                };
                if (!body.nombre || !body.responsable) {
                  showMsg("Subactividad: nombre y responsable son obligatorios.", true);
                  return;
                }
                const subDateError = visualRangeError(body.fecha_inicial, body.fecha_final, "Subactividad");
                if (subDateError) {
                  showMsg(subDateError, true);
                  return;
                }
                const source = objectiveActivities.find((item) => item.id === activityId);
                const subParentError = visualChildRangeError(
                  body.fecha_inicial,
                  body.fecha_final,
                  source?.fecha_inicial || "",
                  source?.fecha_final || "",
                  "Subactividad",
                  "Actividad",
                );
                if (subParentError) {
                  showMsg(subParentError, true);
                  return;
                }
                try {
                  await requestJson(`/api/poa/activities/${activityId}/subactivities`, { method: "POST", body: JSON.stringify(body) });
                  await loadBoard();
                  showMsg("Subactividad agregada.");
                } catch (err) {
                  showMsg(err.message || "No se pudo agregar la subactividad.", true);
                }
              });
            });
          };

          const renderApprovals = () => {
            if (!approvalsEl) return;
            if (!pendingApprovals.length) {
              approvalsEl.innerHTML = '<div class="poac-sub">Sin aprobaciones pendientes.</div>';
              return;
            }
            approvalsEl.innerHTML = pendingApprovals.map((approval) => `
              <div class="poac-subactivity" data-approval-id="${approval.id}" style="border-top:0; padding-top:0; margin-top:0;">
                <div class="poac-sub"><strong>${approval.activity_codigo || "ACT"} - ${approval.activity_nombre || ""}</strong></div>
                <div class="poac-sub">Objetivo: ${approval.objective_codigo || "OBJ"} - ${approval.objective_nombre || ""}</div>
                <div class="poac-sub">Solicitó: ${approval.requester || "-"}</div>
                <textarea class="poac-textarea" data-f="comentario" placeholder="Comentario de validación (opcional)" style="margin-top:6px;"></textarea>
                <div class="poac-inline-actions">
                  <button class="poac-btn primary" type="button" data-act="approve">Autorizar</button>
                  <button class="poac-btn" type="button" data-act="reject">Rechazar</button>
                </div>
              </div>
            `).join("");

            approvalsEl.querySelectorAll("[data-act='approve']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-approval-id]");
                const id = Number(card.getAttribute("data-approval-id"));
                const comentario = (card.querySelector("[data-f='comentario']")?.value || "").trim();
                try {
                  await requestJson(`/api/poa/approvals/${id}/decision`, {
                    method: "POST",
                    body: JSON.stringify({ accion: "autorizar", comentario }),
                  });
                  await loadBoard();
                  showMsg("Entregable autorizado. Actividad terminada.");
                } catch (err) {
                  showMsg(err.message || "No se pudo autorizar.", true);
                }
              });
            });

            approvalsEl.querySelectorAll("[data-act='reject']").forEach((button) => {
              button.addEventListener("click", async () => {
                const card = button.closest("[data-approval-id]");
                const id = Number(card.getAttribute("data-approval-id"));
                const comentario = (card.querySelector("[data-f='comentario']")?.value || "").trim();
                try {
                  await requestJson(`/api/poa/approvals/${id}/decision`, {
                    method: "POST",
                    body: JSON.stringify({ accion: "rechazar", comentario }),
                  });
                  await loadBoard();
                  showMsg("Entregable rechazado.");
                } catch (err) {
                  showMsg(err.message || "No se pudo rechazar.", true);
                }
              });
            });
          };

          const loadBoard = async () => {
            const payload = await requestJson("/api/poa/board-data");
            objectives = Array.isArray(payload.objectives) ? payload.objectives : [];
            activities = Array.isArray(payload.activities) ? payload.activities : [];
            pendingApprovals = Array.isArray(payload.pending_approvals) ? payload.pending_approvals : [];
            if (!selectedObjectiveId || !objectives.some((item) => item.id === selectedObjectiveId)) {
              selectedObjectiveId = objectives.length ? objectives[0].id : null;
            }
            renderObjectives();
            renderActivities();
            renderApprovals();
          };

          addActivityBtn && addActivityBtn.addEventListener("click", async () => {
            const objective = selectedObjective();
            if (!objective) {
              showMsg("No tienes objetivos asignados para crear actividades.", true);
              return;
            }
            const body = {
              objective_id: objective.id,
              nombre: (actNameEl.value || "").trim(),
              codigo: (actCodeEl.value || "").trim(),
              responsable: (actOwnerEl.value || "").trim(),
              entregable: (actDeliverableEl.value || "").trim(),
              fecha_inicial: (actStartEl.value || "").trim(),
              fecha_final: (actEndEl.value || "").trim(),
              descripcion: (actDescEl.value || "").trim(),
            };
            if (!body.nombre || !body.responsable) {
              showMsg("Actividad: nombre y responsable son obligatorios.", true);
              return;
            }
            if (!body.entregable) {
              showMsg("Actividad: el entregable es obligatorio.", true);
              return;
            }
            const activityDateError = visualRangeError(body.fecha_inicial, body.fecha_final, "Actividad");
            if (activityDateError) {
              showMsg(activityDateError, true);
              return;
            }
            const activityParentError = visualChildRangeError(
              body.fecha_inicial,
              body.fecha_final,
              objective.fecha_inicial || "",
              objective.fecha_final || "",
              "Actividad",
              "Objetivo",
            );
            if (activityParentError) {
              showMsg(activityParentError, true);
              return;
            }
            try {
              const editId = Number(addActivityBtn.dataset.editId || 0);
              if (editId) {
                await requestJson(`/api/poa/activities/${editId}`, { method: "PUT", body: JSON.stringify(body) });
                showMsg("Actividad actualizada.");
              } else {
                await requestJson("/api/poa/activities", { method: "POST", body: JSON.stringify(body) });
                showMsg("Actividad agregada.");
              }
              actNameEl.value = "";
              actCodeEl.value = "";
              actOwnerEl.value = "";
              actDeliverableEl.value = "";
              actStartEl.value = "";
              actEndEl.value = "";
              actDescEl.value = "";
              addActivityBtn.textContent = "Agregar actividad";
              delete addActivityBtn.dataset.editId;
              await loadBoard();
            } catch (err) {
              showMsg(err.message || "No se pudo guardar la actividad.", true);
            }
          });

          loadBoard().catch((err) => showMsg(err.message || "No se pudo cargar tablero POA.", true));
        })();
      </script>
    </section>
""")

KPI_HTML = dedent("""
    <section class="kpi-page">
        <style>
            .kpi-page{
                --bg: #f6f8fc;
                --surface: rgba(255,255,255,.88);
                --card: #ffffff;
                --text: #0f172a;
                --muted: #64748b;
                --border: rgba(148,163,184,.38);
                --shadow: 0 18px 40px rgba(15,23,42,.08);
                --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                --radius: 18px;
                --primary: #0f3d2e;
                --primary-2: #1f6f52;
                --ok: #16a34a;
                --warn: #f59e0b;
                --crit: #ef4444;
                width: 100%;
                font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                color: var(--text);
                background:
                  radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                  radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                  var(--bg);
                border-radius: 18px;
            }
            .kpi-wrap{
                width: 100%;
                margin: 0 auto;
                padding: 18px 0 34px;
            }
            .kpi-btn{
                border-radius: 14px;
                padding: 10px 14px;
                font-weight: 800;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                cursor: pointer;
                box-shadow: var(--shadow-soft);
                transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
            }
            .kpi-btn:hover{ transform: translateY(-1px); box-shadow: var(--shadow); background: rgba(255,255,255,.95); }
            .kpi-btn--primary{
                background: linear-gradient(135deg, var(--primary), var(--primary-2));
                color: #fff;
                border-color: rgba(15,61,46,.35);
            }
            .kpi-btn--primary2{
                background: rgba(37,99,235,.12);
                color: #1d4ed8;
                border-color: rgba(37,99,235,.24);
            }
            .kpi-btn--ghost{ background: rgba(255,255,255,.78); }
            .kpi-btn--soft{
                background: rgba(15,61,46,.10);
                border-color: rgba(15,61,46,.18);
                color: #0b2a20;
            }
            .kpi-controls{
                display: grid;
                grid-template-columns: 180px 200px 220px 1fr auto;
                gap: 12px;
                margin-bottom: 14px;
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 18px;
                box-shadow: var(--shadow-soft);
                padding: 14px;
            }
            .kpi-control,
            .kpi-search{
                display: flex;
                flex-direction: column;
                gap: 7px;
            }
            .kpi-control label,
            .kpi-search label{
                font-size: 12px;
                font-weight: 700;
                color: var(--muted);
            }
            .kpi-control select{
                border: 1px solid var(--border);
                border-radius: 12px;
                padding: 10px 12px;
                background: #fff;
            }
            .kpi-search__box{
                border: 1px solid var(--border);
                border-radius: 12px;
                background: #fff;
                padding: 0 12px;
                display: flex;
                align-items: center;
                gap: 8px;
            }
            .kpi-search__icon{ color: #64748b; font-size: 15px; }
            .kpi-search__box input{
                border: 0;
                outline: 0;
                width: 100%;
                padding: 10px 0;
                background: transparent;
            }
            .kpi-actions{
                display: flex;
                gap: 8px;
                align-items: end;
                justify-content: flex-end;
            }
            .kpi-kpis{
                display: grid;
                grid-template-columns: repeat(5, minmax(0, 1fr));
                gap: 12px;
                margin-bottom: 14px;
            }
            .kpi-stat{
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: var(--radius);
                box-shadow: var(--shadow-soft);
                padding: 14px;
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
                position: relative;
                overflow: hidden;
            }
            .kpi-stat::before{
                content: "";
                position: absolute;
                inset: -1px;
                background: linear-gradient(135deg, rgba(15,61,46,.12), transparent 35%, rgba(37,99,235,.10));
                opacity: .85;
                pointer-events: none;
            }
            .kpi-stat > *{ position: relative; }
            .kpi-stat__label{ color: var(--muted); font-size: 13px; font-weight: 700; }
            .kpi-stat__value{ margin-top: 8px; font-size: 30px; font-weight: 900; letter-spacing: -0.03em; }
            .kpi-stat__meta{ margin-top: 8px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }
            .kpi-chip{
                font-size: 12px;
                padding: 6px 10px;
                border-radius: 999px;
                background: rgba(15,23,42,.05);
                border: 1px solid rgba(15,23,42,.08);
                color: rgba(15,23,42,.72);
            }
            .kpi-chip--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }
            .kpi-chip--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
            .kpi-chip--crit{ background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.22); color: #991b1b; }
            .kpi-bar{
                height: 10px;
                flex: 1 1 auto;
                min-width: 120px;
                border-radius: 999px;
                background: rgba(148,163,184,.25);
                border: 1px solid rgba(148,163,184,.25);
                overflow: hidden;
            }
            .kpi-bar__fill{ height: 100%; width: 50%; border-radius: 999px; background: linear-gradient(90deg, rgba(37,99,235,1), rgba(96,165,250,1)); }
            .kpi-bar__fill--ok{ background: linear-gradient(90deg, rgba(15,61,46,1), rgba(31,111,82,1)); }
            .kpi-grid{ display: grid; grid-template-columns: 1.35fr .85fr; gap: 14px; align-items: start; }
            .kpi-card{
                background: var(--surface);
                border: 1px solid var(--border);
                border-radius: 22px;
                box-shadow: var(--shadow-soft);
                padding: 16px;
                backdrop-filter: blur(10px);
                -webkit-backdrop-filter: blur(10px);
                overflow: hidden;
            }
            .kpi-card__head{
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 12px;
                margin-bottom: 14px;
            }
            .kpi-card__head h2{ margin: 0; font-size: 20px; letter-spacing: -0.02em; }
            .kpi-card__head p{ margin: 6px 0 0; color: var(--muted); font-size: 13px; }
            .kpi-card__tools{ display: flex; align-items: center; gap: 10px; }
            .kpi-pill{
                font-size: 12px;
                padding: 6px 10px;
                border-radius: 999px;
                background: rgba(255,255,255,.70);
                border: 1px solid var(--border);
                color: rgba(15,23,42,.72);
                white-space: nowrap;
            }
            .kpi-pill--soft{
                background: rgba(15,61,46,.10);
                border-color: rgba(15,61,46,.18);
                color: #0b2a20;
            }
            .kpi-iconbtn{
                width: 38px;
                height: 38px;
                border-radius: 14px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.75);
                box-shadow: var(--shadow-soft);
                cursor: pointer;
                transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
            }
            .kpi-iconbtn:hover{
                transform: translateY(-1px);
                box-shadow: var(--shadow);
                background: rgba(255,255,255,.95);
            }
            .kpi-table{
                border: 1px solid var(--border);
                border-radius: 16px;
                overflow: hidden;
                background: rgba(255,255,255,.65);
            }
            .kpi-table__head{
                display: grid;
                grid-template-columns: 1.7fr 1fr .7fr .7fr .45fr .75fr;
                gap: 10px;
                align-items: center;
                padding: 12px;
                font-size: 12px;
                color: var(--muted);
                font-weight: 900;
                background: rgba(15,23,42,.03);
                border-bottom: 1px solid var(--border);
            }
            .kpi-row{
                display: grid;
                grid-template-columns: 1.7fr 1fr .7fr .7fr .45fr .75fr;
                gap: 10px;
                align-items: center;
                padding: 12px;
                border-bottom: 1px solid rgba(148,163,184,.25);
                font-size: 14px;
            }
            .kpi-row:last-child{ border-bottom: 0; }
            .kpi-row:hover{ background: rgba(37,99,235,.05); }
            .kpi-row--active{ background: rgba(15,61,46,.06); outline: 1px solid rgba(15,61,46,.18); }
            .kpi-row__main{ display: flex; flex-direction: column; gap: 2px; }
            .kpi-row__main strong{ font-size: 13px; }
            .kpi-sub{ font-size: 12px; color: var(--muted); }
            .kpi-right{ text-align: right; }
            .kpi-status{
                font-size: 12px;
                font-weight: 800;
                padding: 6px 10px;
                border-radius: 999px;
                border: 1px solid var(--border);
                background: rgba(255,255,255,.70);
                white-space: nowrap;
            }
            .kpi-status--warn{ background: rgba(245,158,11,.12); border-color: rgba(245,158,11,.22); color: #92400e; }
            .kpi-status--crit{ background: rgba(239,68,68,.10); border-color: rgba(239,68,68,.22); color: #991b1b; }
            .kpi-status--ok{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }
            .kpi-trend{
                display: inline-flex;
                align-items: center;
                justify-content: center;
                width: 30px;
                height: 30px;
                border-radius: 12px;
                border: 1px solid rgba(148,163,184,.25);
                background: rgba(15,23,42,.03);
                font-weight: 900;
                color: #64748b;
            }
            .kpi-trend--up{
                background: rgba(22,163,74,.10);
                border-color: rgba(22,163,74,.20);
                color: #166534;
            }
            .kpi-trend--down{
                background: rgba(239,68,68,.10);
                border-color: rgba(239,68,68,.20);
                color: #991b1b;
            }
            .kpi-foot{
                margin-top: 12px;
                display: flex;
                align-items: center;
                justify-content: space-between;
                gap: 10px;
                flex-wrap: wrap;
            }
            .kpi-foot__note{ display: flex; align-items: center; gap: 10px; color: #475569; font-size: 12px; }
            .kpi-dot{ width: 9px; height: 9px; border-radius: 999px; display: inline-block; margin-right: 5px; }
            .kpi-dot--ok{ background: var(--ok); }
            .kpi-dot--warn{ background: var(--warn); }
            .kpi-dot--crit{ background: var(--crit); }
            .kpi-detail{ display: flex; flex-direction: column; gap: 12px; }
            .kpi-detail__top{
                display: flex;
                align-items: flex-start;
                justify-content: space-between;
                gap: 10px;
            }
            .kpi-detail__name{ font-size: 17px; font-weight: 800; line-height: 1.25; }
            .kpi-detail__meta{ margin-top: 8px; display: flex; flex-wrap: wrap; gap: 8px; }
            .kpi-score{
                min-width: 94px;
                text-align: center;
                border: 1px solid var(--border);
                border-radius: 14px;
                background: rgba(255,255,255,.75);
                padding: 8px 10px;
            }
            .kpi-score__k{ font-size: 11px; color: var(--muted); }
            .kpi-score__v{ margin-top: 4px; font-size: 22px; font-weight: 900; }
            .kpi-detail__cards{
                display: grid;
                grid-template-columns: repeat(3, minmax(0, 1fr));
                gap: 10px;
            }
            .kpi-miniCard{
                background: rgba(15,23,42,.03);
                border: 1px solid rgba(15,23,42,.08);
                border-radius: 14px;
                padding: 10px;
            }
            .kpi-miniCard__k{ font-size: 12px; color: var(--muted); font-weight: 700; }
            .kpi-miniCard__v{ margin-top: 6px; font-size: 20px; font-weight: 900; }
            .kpi-miniCard__v--warn{ color: #92400e; }
            .kpi-chart{
                background: rgba(255,255,255,.8);
                border: 1px solid rgba(148,163,184,.30);
                border-radius: 16px;
                padding: 12px;
            }
            .kpi-chart__head{
                display: flex;
                justify-content: space-between;
                align-items: center;
                gap: 8px;
                margin-bottom: 10px;
            }
            .kpi-spark__bars{
                height: 120px;
                display: grid;
                grid-template-columns: repeat(8, 1fr);
                gap: 8px;
                align-items: end;
            }
            .kpi-spark__bars span{
                display: block;
                border-radius: 8px 8px 4px 4px;
                background: linear-gradient(180deg, rgba(37,99,235,.95), rgba(15,61,46,.95));
            }
            .kpi-spark__axis{
                margin-top: 8px;
                display: grid;
                grid-template-columns: repeat(5, 1fr);
                font-size: 11px;
                color: var(--muted);
            }
            .kpi-notes{
                display: flex;
                flex-direction: column;
                gap: 8px;
            }
            .kpi-notes label{ font-size: 12px; font-weight: 700; color: var(--muted); }
            .kpi-notes textarea{
                min-height: 90px;
                border: 1px solid var(--border);
                border-radius: 14px;
                padding: 10px 12px;
                background: #fff;
                resize: vertical;
            }
            .kpi-detail__actions{ display: flex; gap: 8px; justify-content: flex-end; }
            @media (max-width: 1200px){
                .kpi-controls{ grid-template-columns: 1fr 1fr 1fr; }
                .kpi-search{ grid-column: 1 / -1; }
                .kpi-actions{ grid-column: 1 / -1; justify-content: flex-start; }
                .kpi-kpis{ grid-template-columns: repeat(3, minmax(0, 1fr)); }
            }
            @media (max-width: 1100px){
                .kpi-grid{ grid-template-columns: 1fr; }
                .kpi-table__head,
                .kpi-row{ grid-template-columns: 1.6fr .9fr .7fr .7fr .5fr .8fr; }
            }
            @media (max-width: 640px){
                .kpi-controls{ grid-template-columns: 1fr; }
                .kpi-kpis{ grid-template-columns: 1fr; }
                .kpi-detail__cards{ grid-template-columns: 1fr; }
                .kpi-table__head{ display: none; }
                .kpi-row{
                    grid-template-columns: 1fr;
                    gap: 8px;
                    font-size: 12px;
                }
                .kpi-right{ text-align: left; }
                .kpi-foot{ flex-direction: column; align-items: flex-start; }
            }
        </style>
        <div class="kpi-wrap">
            <section class="kpi-controls">
                <div class="kpi-control">
                    <label>Periodo</label>
                    <select>
                        <option>Mensual</option>
                        <option>Trimestral</option>
                        <option>Anual</option>
                    </select>
                </div>
                <div class="kpi-control">
                    <label>Area</label>
                    <select>
                        <option>Todas</option>
                        <option>Operaciones</option>
                        <option>Finanzas</option>
                        <option>Comercial</option>
                        <option>TI</option>
                        <option>Calidad</option>
                    </select>
                </div>
                <div class="kpi-control">
                    <label>Perspectiva</label>
                    <select>
                        <option>Todas</option>
                        <option>Financiera</option>
                        <option>Clientes</option>
                        <option>Procesos</option>
                        <option>Aprendizaje</option>
                    </select>
                </div>
                <div class="kpi-search">
                    <label>Buscar</label>
                    <div class="kpi-search__box">
                        <span class="kpi-search__icon">⌕</span>
                        <input type="search" placeholder="Buscar KPI, codigo o responsable...">
                    </div>
                </div>
                <div class="kpi-actions">
                    <button class="kpi-btn kpi-btn--soft" type="button">Limpiar</button>
                    <button class="kpi-btn kpi-btn--primary2" type="button">Aplicar</button>
                </div>
            </section>
            <section class="kpi-kpis">
                <article class="kpi-stat">
                    <div class="kpi-stat__label">KPIs monitoreados</div>
                    <div class="kpi-stat__value">32</div>
                    <div class="kpi-stat__meta"><span class="kpi-chip">Activos</span></div>
                </article>
                <article class="kpi-stat kpi-stat--progress">
                    <div class="kpi-stat__label">Cumplimiento global</div>
                    <div class="kpi-stat__value">78%</div>
                    <div class="kpi-stat__meta">
                        <div class="kpi-bar"><div class="kpi-bar__fill kpi-bar__fill--ok" style="width:78%"></div></div>
                        <span class="kpi-chip kpi-chip--ok">Meta 85%</span>
                    </div>
                </article>
                <article class="kpi-stat">
                    <div class="kpi-stat__label">En meta</div>
                    <div class="kpi-stat__value">21</div>
                    <div class="kpi-stat__meta"><span class="kpi-chip kpi-chip--ok">+2 vs periodo</span></div>
                </article>
                <article class="kpi-stat">
                    <div class="kpi-stat__label">En riesgo</div>
                    <div class="kpi-stat__value">7</div>
                    <div class="kpi-stat__meta"><span class="kpi-chip kpi-chip--warn">Requiere accion</span></div>
                </article>
                <article class="kpi-stat">
                    <div class="kpi-stat__label">Criticos</div>
                    <div class="kpi-stat__value">4</div>
                    <div class="kpi-stat__meta"><span class="kpi-chip kpi-chip--crit">Prioridad alta</span></div>
                </article>
            </section>
            <section class="kpi-grid">
                <section class="kpi-card">
                    <header class="kpi-card__head">
                        <div>
                            <h2>Indicadores</h2>
                            <p>Vista ejecutiva con estado, meta, valor actual y tendencia.</p>
                        </div>
                        <div class="kpi-card__tools">
                            <span class="kpi-pill">32 KPIs</span>
                            <button class="kpi-iconbtn" type="button" title="Ordenar">⇅</button>
                            <button class="kpi-iconbtn" type="button" title="Columnas">☰</button>
                        </div>
                    </header>
                    <div class="kpi-table">
                        <div class="kpi-table__head">
                            <span>KPI</span>
                            <span>Responsable</span>
                            <span>Meta</span>
                            <span>Actual</span>
                            <span>Tendencia</span>
                            <span class="kpi-right">Estado</span>
                        </div>
                        <div class="kpi-row kpi-row--active">
                            <div class="kpi-row__main"><strong>KPI-OPS-01</strong><span class="kpi-sub">Tiempo de respuesta a solicitudes</span></div>
                            <div>Operaciones</div>
                            <div>&le; 24h</div>
                            <div>42h</div>
                            <div><span class="kpi-trend kpi-trend--down">▼</span></div>
                            <div class="kpi-right"><span class="kpi-status kpi-status--warn">En riesgo</span></div>
                        </div>
                        <div class="kpi-row">
                            <div class="kpi-row__main"><strong>KPI-FIN-02</strong><span class="kpi-sub">Ejecucion presupuestaria</span></div>
                            <div>Finanzas</div>
                            <div>&ge; 85%</div>
                            <div>81%</div>
                            <div><span class="kpi-trend">-</span></div>
                            <div class="kpi-right"><span class="kpi-status kpi-status--warn">En riesgo</span></div>
                        </div>
                        <div class="kpi-row">
                            <div class="kpi-row__main"><strong>KPI-CLI-01</strong><span class="kpi-sub">Satisfaccion general</span></div>
                            <div>Atencion</div>
                            <div>&ge; 90%</div>
                            <div>88%</div>
                            <div><span class="kpi-trend kpi-trend--up">▲</span></div>
                            <div class="kpi-right"><span class="kpi-status kpi-status--warn">En riesgo</span></div>
                        </div>
                        <div class="kpi-row">
                            <div class="kpi-row__main"><strong>KPI-PRO-03</strong><span class="kpi-sub">Procesos criticos estandarizados</span></div>
                            <div>Calidad</div>
                            <div>&ge; 80%</div>
                            <div>73%</div>
                            <div><span class="kpi-trend kpi-trend--down">▼</span></div>
                            <div class="kpi-right"><span class="kpi-status kpi-status--crit">Critico</span></div>
                        </div>
                        <div class="kpi-row">
                            <div class="kpi-row__main"><strong>KPI-TAL-02</strong><span class="kpi-sub">Competencias criticas certificadas</span></div>
                            <div>Talento</div>
                            <div>&ge; 75%</div>
                            <div>69%</div>
                            <div><span class="kpi-trend kpi-trend--down">▼</span></div>
                            <div class="kpi-right"><span class="kpi-status kpi-status--warn">En riesgo</span></div>
                        </div>
                    </div>
                    <div class="kpi-foot">
                        <div class="kpi-foot__note">
                            <span><span class="kpi-dot kpi-dot--ok"></span> En meta</span>
                            <span><span class="kpi-dot kpi-dot--warn"></span> En riesgo</span>
                            <span><span class="kpi-dot kpi-dot--crit"></span> Critico</span>
                        </div>
                        <button class="kpi-btn kpi-btn--ghost" type="button">Ver acciones correctivas</button>
                    </div>
                </section>
                <aside class="kpi-card">
                    <header class="kpi-card__head">
                        <div>
                            <h2>Detalle del KPI</h2>
                            <p>Selecciona un indicador para ver su ficha y evolucion.</p>
                        </div>
                        <div class="kpi-card__tools"><span class="kpi-pill kpi-pill--soft">KPI-OPS-01</span></div>
                    </header>
                    <section class="kpi-detail">
                        <div class="kpi-detail__top">
                            <div>
                                <div class="kpi-detail__name">Tiempo de respuesta a solicitudes</div>
                                <div class="kpi-detail__meta">
                                    <span class="kpi-chip">Operaciones</span>
                                    <span class="kpi-chip kpi-chip--warn">En riesgo</span>
                                    <span class="kpi-chip">Frecuencia: semanal</span>
                                </div>
                            </div>
                            <div class="kpi-score"><div class="kpi-score__k">Cumplimiento</div><div class="kpi-score__v">62%</div></div>
                        </div>
                        <div class="kpi-detail__cards">
                            <div class="kpi-miniCard"><div class="kpi-miniCard__k">Meta</div><div class="kpi-miniCard__v">&le; 24h</div></div>
                            <div class="kpi-miniCard"><div class="kpi-miniCard__k">Actual</div><div class="kpi-miniCard__v kpi-miniCard__v--warn">42h</div></div>
                            <div class="kpi-miniCard"><div class="kpi-miniCard__k">Brecha</div><div class="kpi-miniCard__v kpi-miniCard__v--warn">+18h</div></div>
                        </div>
                        <div class="kpi-chart">
                            <div class="kpi-chart__head"><strong>Evolucion (ultimas 8 mediciones)</strong><span class="kpi-pill">Mock</span></div>
                            <div class="kpi-spark" aria-label="Grafico de tendencia">
                                <div class="kpi-spark__bars">
                                    <span style="height:38%"></span><span style="height:46%"></span><span style="height:52%"></span><span style="height:58%"></span>
                                    <span style="height:63%"></span><span style="height:70%"></span><span style="height:78%"></span><span style="height:84%"></span>
                                </div>
                                <div class="kpi-spark__axis"><span>-8</span><span>-6</span><span>-4</span><span>-2</span><span>Hoy</span></div>
                            </div>
                        </div>
                        <div class="kpi-notes">
                            <label>Notas / acciones</label>
                            <textarea placeholder="Registra hallazgos, causas y acciones correctivas..."></textarea>
                            <div class="kpi-detail__actions">
                                <button class="kpi-btn kpi-btn--soft" type="button">Guardar nota</button>
                                <button class="kpi-btn kpi-btn--primary" type="button">Crear accion</button>
                            </div>
                        </div>
                    </section>
                </aside>
            </section>
        </div>
    </section>
""")

REPORTES_HTML = dedent("""
    <section class="pe-page">
        <style>
            .pe-page{--bg:#f6f8fc;--surface:rgba(255,255,255,.88);--text:#0f172a;--muted:#64748b;--border:rgba(148,163,184,.38);--shadow-soft:0 10px 22px rgba(15,23,42,.06);--radius:18px;width:100%;font-family:Inter,system-ui,-apple-system,Segoe UI,Roboto,Arial,sans-serif;color:var(--text);background:radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),var(--bg);border-radius:18px}
            .pe-wrap{width:100%;padding:18px 0 34px}
            .pe-grid{display:grid;grid-template-columns:1.1fr .9fr;gap:14px;align-items:start}
            .pe-card{background:var(--surface);border:1px solid var(--border);border-radius:22px;box-shadow:var(--shadow-soft);padding:16px;overflow:hidden}
            .pe-card__head h2{margin:0;font-size:20px;letter-spacing:-.02em}
            .pe-card__head p{margin:6px 0 12px;color:var(--muted);font-size:13px}
            .pe-list{display:flex;flex-direction:column;gap:12px}
            .pe-item{background:rgba(255,255,255,.80);border:1px solid rgba(148,163,184,.30);border-radius:18px;padding:14px 14px 12px;box-shadow:0 10px 20px rgba(15,23,42,.04)}
            .pe-item strong{display:block}
            .pe-item small{color:var(--muted)}
            .pe-chip{font-size:12px;padding:6px 10px;border-radius:999px;background:rgba(37,99,235,.10);border:1px solid rgba(37,99,235,.18);color:#1d4ed8}
            .pe-table{width:100%;border-collapse:collapse;font-size:14px}
            .pe-table th,.pe-table td{text-align:left;padding:10px;border-bottom:1px solid rgba(148,163,184,.25)}
            @media (max-width:1000px){.pe-grid{grid-template-columns:1fr}}
        </style>
        <div class="pe-wrap">
            <section class="pe-grid">
                <article class="pe-card">
                    <div class="pe-card__head"><h2>Reportes disponibles</h2><p>Consolida avance, desempeño y seguimiento institucional.</p></div>
                    <div class="pe-list">
                        <article class="pe-item"><strong>Reporte ejecutivo</strong><small>Resumen de estado estratégico por eje y objetivo.</small></article>
                        <article class="pe-item"><strong>Reporte operativo</strong><small>Actividades, avances, cumplimientos y desviaciones.</small></article>
                        <article class="pe-item"><strong>Reporte KPI</strong><small>Indicadores, metas, semáforos y variaciones.</small></article>
                    </div>
                </article>
                <aside class="pe-card">
                    <div class="pe-card__head"><h2>Exportaciones</h2><p>Salidas disponibles para distribución y análisis.</p></div>
                    <table class="pe-table">
                        <thead><tr><th>Formato</th><th>Uso</th></tr></thead>
                        <tbody>
                            <tr><td><span class="pe-chip">PDF</span></td><td>Presentación institucional</td></tr>
                            <tr><td><span class="pe-chip">Excel</span></td><td>Análisis detallado</td></tr>
                            <tr><td><span class="pe-chip">HTML</span></td><td>Vista compartible</td></tr>
                        </tbody>
                    </table>
                </aside>
            </section>
        </div>
    </section>
""")

DOCUMENTOS_HTML = dedent("""
    <section class="pe-page">
        <style>
            .doc-wrap{display:grid;gap:14px}
            .doc-grid{display:grid;grid-template-columns:1fr 1fr;gap:14px}
            .doc-card{background:rgba(255,255,255,.88);border:1px solid rgba(148,163,184,.35);border-radius:16px;padding:14px}
            .doc-card h3{margin:0 0 10px}
            .doc-form{display:grid;gap:10px}
            .doc-form input,.doc-form select,.doc-form textarea{width:100%;padding:10px;border:1px solid #cbd5e1;border-radius:10px}
            .doc-actions{display:flex;gap:8px;flex-wrap:wrap}
            .doc-btn{border:1px solid #cbd5e1;background:#fff;border-radius:10px;padding:8px 10px;cursor:pointer;font-weight:600}
            .doc-btn.primary{background:#0f3d2e;color:#fff;border-color:#0f3d2e}
            .doc-btn.warn{background:#fff7ed;color:#9a3412;border-color:#fdba74}
            .doc-btn.danger{background:#fef2f2;color:#991b1b;border-color:#fecaca}
            .doc-table{width:100%;border-collapse:collapse}
            .doc-table th,.doc-table td{padding:10px;border-bottom:1px solid #e2e8f0;text-align:left;vertical-align:top}
            .doc-pill{display:inline-block;padding:4px 8px;border-radius:999px;font-size:12px;font-weight:700}
            .doc-pill.borrador{background:#f1f5f9;color:#334155}
            .doc-pill.enviado{background:#eff6ff;color:#1d4ed8}
            .doc-pill.autorizado{background:#ecfdf5;color:#166534}
            .doc-pill.actualizado{background:#fff7ed;color:#9a3412}
            .doc-pill.rechazado{background:#fef2f2;color:#991b1b}
            .doc-message{display:none;padding:10px;border-radius:10px;font-size:13px}
            .doc-message.ok{display:block;background:#ecfdf5;color:#166534;border:1px solid #86efac}
            .doc-message.error{display:block;background:#fef2f2;color:#991b1b;border:1px solid #fecaca}
            @media (max-width: 980px){.doc-grid{grid-template-columns:1fr}}
        </style>
        <div class="doc-wrap">
            <section class="doc-grid">
                <article class="doc-card">
                    <h3>Nuevo documento de evidencia</h3>
                    <form id="doc-create-form" class="doc-form">
                        <input type="text" name="titulo" placeholder="Título" required>
                        <select name="proceso" required>
                            <option value="envio">Envio</option>
                            <option value="autorizacion">Autorizacion</option>
                            <option value="actualizacion">Actualizacion</option>
                        </select>
                        <textarea name="descripcion" placeholder="Descripción"></textarea>
                        <textarea name="observaciones" placeholder="Observaciones"></textarea>
                        <input type="file" name="archivo" required>
                        <div class="doc-actions">
                            <button class="doc-btn primary" type="submit">Subir y guardar</button>
                        </div>
                    </form>
                </article>
                <article class="doc-card">
                    <h3>Editar documento</h3>
                    <form id="doc-edit-form" class="doc-form">
                        <input type="hidden" name="id">
                        <input type="text" name="titulo" placeholder="Título" required>
                        <select name="proceso" required>
                            <option value="envio">Envio</option>
                            <option value="autorizacion">Autorizacion</option>
                            <option value="actualizacion">Actualizacion</option>
                        </select>
                        <textarea name="descripcion" placeholder="Descripción"></textarea>
                        <textarea name="observaciones" placeholder="Observaciones"></textarea>
                        <div class="doc-actions">
                            <button class="doc-btn" type="button" id="doc-edit-clear">Limpiar</button>
                            <button class="doc-btn primary" type="submit">Guardar cambios</button>
                        </div>
                    </form>
                </article>
            </section>
            <section class="doc-card">
                <h3>Documentos de evidencia</h3>
                <div id="doc-message" class="doc-message"></div>
                <table class="doc-table">
                    <thead>
                        <tr>
                            <th>Titulo</th>
                            <th>Proceso</th>
                            <th>Estado</th>
                            <th>Version</th>
                            <th>Archivo</th>
                            <th>Acciones</th>
                        </tr>
                    </thead>
                    <tbody id="doc-table-body"></tbody>
                </table>
            </section>
        </div>
        <script>
            const docBody = document.getElementById('doc-table-body');
            const docMsg = document.getElementById('doc-message');
            const createForm = document.getElementById('doc-create-form');
            const editForm = document.getElementById('doc-edit-form');
            const editClear = document.getElementById('doc-edit-clear');
            let docsCache = [];
            let permissions = { can_create: true, can_authorize: false, can_reject: false };
            const setDocMessage = (kind, text) => {
                docMsg.className = `doc-message ${kind}`;
                docMsg.textContent = text || '';
            };
            const fillEdit = (item) => {
                editForm.id.value = item.id;
                editForm.titulo.value = item.titulo || '';
                editForm.proceso.value = item.proceso || 'envio';
                editForm.descripcion.value = item.descripcion || '';
                editForm.observaciones.value = item.observaciones || '';
            };
            const clearEdit = () => {
                editForm.reset();
                editForm.id.value = '';
            };
            editClear.addEventListener('click', clearEdit);

            const api = async (url, opts = {}) => {
                const res = await fetch(url, opts);
                const json = await res.json().catch(() => ({}));
                if (!res.ok || json.success === false) throw new Error(json.detail || json.error || 'Operacion fallida');
                return json;
            };

            const renderDocs = (items) => {
                docsCache = Array.isArray(items) ? items : [];
                docBody.innerHTML = (items || []).map((item) => `
                    <tr>
                        <td><strong>${item.titulo || ''}</strong><br><small>${item.descripcion || ''}</small></td>
                        <td>${item.proceso || ''}</td>
                        <td><span class="doc-pill ${item.estado || 'borrador'}">${item.estado || 'borrador'}</span></td>
                        <td>v${item.version || 1}</td>
                        <td><a href="/api/documentos/${item.id}/download" target="_blank">${item.archivo_nombre || 'archivo'}</a></td>
                        <td>
                            <div class="doc-actions">
                                ${item.actions?.can_edit ? `<button class="doc-btn" data-act="edit" data-id="${item.id}">Editar</button>` : ''}
                                ${item.actions?.can_send ? `<button class="doc-btn" data-act="send" data-id="${item.id}">Enviar</button>` : ''}
                                ${item.actions?.can_authorize ? `<button class="doc-btn warn" data-act="authorize" data-id="${item.id}">Autorizar</button>` : ''}
                                ${item.actions?.can_reject ? `<button class="doc-btn warn" data-act="reject" data-id="${item.id}">Rechazar</button>` : ''}
                                ${item.actions?.can_update ? `<button class="doc-btn" data-act="update" data-id="${item.id}">Actualizar</button>` : ''}
                                ${item.actions?.can_delete ? `<button class="doc-btn danger" data-act="delete" data-id="${item.id}">Eliminar</button>` : ''}
                            </div>
                        </td>
                    </tr>
                `).join('');
            };

            const loadDocs = async () => {
                try {
                    const json = await api('/api/documentos');
                    permissions = json.permissions || permissions;
                    createForm.querySelectorAll('input,select,textarea,button').forEach((el) => {
                        el.disabled = !permissions.can_create;
                    });
                    renderDocs(json.data || []);
                } catch (error) {
                    setDocMessage('error', error.message || 'No se pudieron cargar documentos');
                }
            };

            createForm.addEventListener('submit', async (event) => {
                event.preventDefault();
                try {
                    const fd = new FormData(createForm);
                    await api('/api/documentos', { method: 'POST', body: fd });
                    createForm.reset();
                    setDocMessage('ok', 'Documento creado.');
                    await loadDocs();
                } catch (error) {
                    setDocMessage('error', error.message);
                }
            });

            editForm.addEventListener('submit', async (event) => {
                event.preventDefault();
                const id = editForm.id.value;
                if (!id) return setDocMessage('error', 'Selecciona un documento para editar.');
                try {
                    const payload = {
                        titulo: editForm.titulo.value,
                        proceso: editForm.proceso.value,
                        descripcion: editForm.descripcion.value,
                        observaciones: editForm.observaciones.value,
                    };
                    await api(`/api/documentos/${id}`, {
                        method: 'PUT',
                        headers: { 'Content-Type': 'application/json' },
                        body: JSON.stringify(payload),
                    });
                    setDocMessage('ok', 'Documento actualizado.');
                    await loadDocs();
                } catch (error) {
                    setDocMessage('error', error.message);
                }
            });

            docBody.addEventListener('click', async (event) => {
                const btn = event.target.closest('button[data-id][data-act]');
                if (!btn) return;
                const id = btn.dataset.id;
                const action = btn.dataset.act;
                try {
                    if (action === 'edit') {
                        const found = docsCache.find((item) => String(item.id) === String(id));
                        if (found) fillEdit(found);
                        return;
                    }
                    if (action === 'send') {
                        await api(`/api/documentos/${id}/enviar`, { method: 'POST' });
                        setDocMessage('ok', 'Documento enviado para autorizacion.');
                    }
                    if (action === 'authorize') {
                        await api(`/api/documentos/${id}/autorizar`, {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ accion: 'autorizar' }),
                        });
                        setDocMessage('ok', 'Documento autorizado.');
                    }
                    if (action === 'reject') {
                        const motivo = prompt('Motivo de rechazo (opcional):') || '';
                        await api(`/api/documentos/${id}/autorizar`, {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ accion: 'rechazar', observaciones: motivo }),
                        });
                        setDocMessage('ok', 'Documento rechazado.');
                    }
                    if (action === 'update') {
                        const fileInput = document.createElement('input');
                        fileInput.type = 'file';
                        fileInput.onchange = async () => {
                            if (!fileInput.files || !fileInput.files.length) return;
                            const fd = new FormData();
                            fd.append('archivo', fileInput.files[0]);
                            await api(`/api/documentos/${id}/actualizar`, { method: 'POST', body: fd });
                            setDocMessage('ok', 'Documento actualizado a nueva version.');
                            await loadDocs();
                        };
                        fileInput.click();
                        return;
                    }
                    if (action === 'delete') {
                        if (!confirm('¿Eliminar este documento?')) return;
                        await api(`/api/documentos/${id}`, { method: 'DELETE' });
                        setDocMessage('ok', 'Documento eliminado.');
                    }
                    await loadDocs();
                } catch (error) {
                    setDocMessage('error', error.message);
                }
            });

            loadDocs();
        </script>
    </section>
""")


def render_personalizacion_page(request: Request) -> HTMLResponse:
    return render_backend_page(
        request,
        title="Personalizar",
        description="Agrega tu imagen corporativa",
        content=PERSONALIZACION_HTML,
        hide_floating_actions=False,
        floating_actions_screen="personalization",
    )


@app.get("/personalizar-pantalla", response_class=HTMLResponse)
def personalizar_pantalla(request: Request):
    require_superadmin(request)
    return backend_screen(
        request,
        title="Personalizar",
        subtitle="Colores institucionales",
        description="Agrega tu imagen corporativa",
        content=PERSONALIZACION_HTML,
        view_buttons=[
            {"label": "Formulario", "view": "form", "icon": "/templates/icon/formulario.svg"},
            {"label": "Lista", "view": "list", "icon": "/templates/icon/list.svg"},
            {"label": "Colores", "view": "colores", "icon": "/templates/icon/personalizacion.svg", "active": True},
        ],
        hide_floating_actions=False,
    )


@app.get("/backend-template", response_class=HTMLResponse)
def backend_template(request: Request):
    placeholder = "<section style='min-height:300px;display:flex;align-items:center;justify-content:center;'><strong>Template listo para montar una página de backend.</strong></section>"
    return render_backend_page(
        request,
            title="Backend Template",
            description="Cascarón para nuevas pantallas",
            content=placeholder,
            hide_floating_actions=False,
            view_buttons=[
                {"label": "Formulario", "icon": "/templates/icon/formulario.svg", "view": "form"},
                {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list"},
                {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
                {"label": "Gráfica", "icon": "/templates/icon/grid.svg", "view": "grafica"},
                {"label": "Gantt", "icon": "/templates/icon/grid.svg", "view": "gantt"},
                {"label": "Dashboard", "icon": "/templates/icon/tablero.svg", "view": "dashboard"},
                {"label": "Calendario", "icon": "/templates/icon/calendario.svg", "view": "calendario"},
            ],
        )


@app.get("/reportes", response_class=HTMLResponse)
def reportes(request: Request):
    return render_backend_page(
        request,
        title="Reportes",
        description="Resumen de indicadores, exportaciones y archivos generados",
        content=REPORTES_HTML,
        hide_floating_actions=False,
        floating_actions_html="""
            <div class="floating-actions-group" data-floating-screen="reportes">
                <button class="action-button" type="button" id="action-export-reportes" aria-label="Exportar">
                    <img src="/templates/icon/guardar.svg" alt="Exportar">
                    <span class="action-label">Exportar</span>
                </button>
                <button class="action-button" type="button" id="action-export-pdf" aria-label="Generar PDF">
                    <img src="/templates/icon/guardar.svg" alt="PDF">
                    <span class="action-label">PDF</span>
                </button>
                <button class="action-button" type="button" id="action-export-excel" aria-label="Generar Excel">
                    <img src="/templates/icon/list.svg" alt="Excel">
                    <span class="action-label">Excel</span>
                </button>
            </div>
        """,
        floating_actions_screen="reportes",
    )


def _can_authorize_documents(request: Request) -> bool:
    return is_admin_or_superadmin(request)


def _can_edit_document(request: Request, doc: DocumentoEvidencia) -> bool:
    if is_admin_or_superadmin(request):
        return True
    current_user = (getattr(request.state, "user_name", "") or request.cookies.get("user_name", "") or "").strip().lower()
    return bool(current_user and current_user == (doc.creado_por or "").strip().lower())


def _document_actions_for_request(request: Request, doc: DocumentoEvidencia) -> Dict[str, bool]:
    can_edit = _can_edit_document(request, doc)
    can_authorize = _can_authorize_documents(request)
    return {
        "can_edit": can_edit,
        "can_send": can_edit,
        "can_update": can_edit,
        "can_delete": can_edit,
        "can_authorize": can_authorize,
        "can_reject": can_authorize,
    }


def _get_document_tenant(request: Request) -> str:
    return _normalize_tenant_id(get_current_tenant(request))


def _get_document_for_request(db, request: Request, doc_id: int) -> Optional[DocumentoEvidencia]:
    tenant_id = _get_document_tenant(request)
    query = db.query(DocumentoEvidencia).filter(DocumentoEvidencia.id == doc_id)
    if is_superadmin(request):
        header_tenant = request.headers.get("x-tenant-id")
        if header_tenant and _normalize_tenant_id(header_tenant) != "all":
            query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == _normalize_tenant_id(header_tenant).lower())
        elif not header_tenant:
            query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
    else:
        query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
    return query.first()


def _serialize_documento(doc: DocumentoEvidencia, request: Optional[Request] = None) -> Dict[str, Any]:
    payload = {
        "id": doc.id,
        "tenant_id": _normalize_tenant_id(doc.tenant_id or "default"),
        "titulo": doc.titulo,
        "descripcion": doc.descripcion or "",
        "proceso": doc.proceso or "envio",
        "estado": doc.estado or "borrador",
        "version": int(doc.version or 1),
        "archivo_nombre": doc.archivo_nombre,
        "archivo_tipo": doc.archivo_tipo or "",
        "archivo_tamano": int(doc.archivo_tamano or 0),
        "observaciones": doc.observaciones or "",
        "creado_por": doc.creado_por or "",
        "enviado_por": doc.enviado_por or "",
        "autorizado_por": doc.autorizado_por or "",
        "actualizado_por": doc.actualizado_por or "",
        "creado_at": doc.creado_at.isoformat() if doc.creado_at else "",
        "enviado_at": doc.enviado_at.isoformat() if doc.enviado_at else "",
        "autorizado_at": doc.autorizado_at.isoformat() if doc.autorizado_at else "",
        "actualizado_at": doc.actualizado_at.isoformat() if doc.actualizado_at else "",
    }
    if request is not None:
        payload["actions"] = _document_actions_for_request(request, doc)
    return payload


def _serialize_document_permissions(request: Request) -> Dict[str, bool]:
    can_authorize = _can_authorize_documents(request)
    return {
        "can_create": True,
        "can_authorize": can_authorize,
        "can_reject": can_authorize,
    }

@app.get("/reportes/documentos", response_class=HTMLResponse)
def reportes_documentos(request: Request):
    return render_backend_page(
        request,
        title="Documentos",
        description="Carga y gestión de evidencias con envío, autorización y actualización.",
        content=DOCUMENTOS_HTML,
        hide_floating_actions=True,
        floating_actions_screen="reportes",
    )


@app.get("/api/documentos")
def listar_documentos(request: Request):
    db = SessionLocal()
    try:
        tenant_id = _get_document_tenant(request)
        query = db.query(DocumentoEvidencia)
        if is_superadmin(request):
            header_tenant = request.headers.get("x-tenant-id")
            if header_tenant and _normalize_tenant_id(header_tenant) != "all":
                query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == _normalize_tenant_id(header_tenant).lower())
            elif not header_tenant:
                query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
        else:
            query = query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
        docs = query.order_by(DocumentoEvidencia.updated_at.desc(), DocumentoEvidencia.id.desc()).all()
        return {
            "success": True,
            "tenant_id": tenant_id,
            "permissions": _serialize_document_permissions(request),
            "data": [_serialize_documento(doc, request=request) for doc in docs],
        }
    finally:
        db.close()


@app.post("/api/documentos")
async def crear_documento(
    request: Request,
    titulo: str = Form(""),
    descripcion: str = Form(""),
    proceso: str = Form("envio"),
    observaciones: str = Form(""),
    archivo: UploadFile = File(...),
):
    clean_title = (titulo or "").strip()
    if not clean_title:
        return JSONResponse({"success": False, "error": "El título es obligatorio"}, status_code=422)
    if (proceso or "").strip().lower() not in {"envio", "autorizacion", "actualizacion"}:
        return JSONResponse({"success": False, "error": "Proceso inválido"}, status_code=422)

    stored = await _store_evidence_file(archivo)
    db = SessionLocal()
    try:
        doc = DocumentoEvidencia(
            tenant_id=_get_document_tenant(request),
            titulo=clean_title,
            descripcion=(descripcion or "").strip(),
            proceso=(proceso or "envio").strip().lower(),
            estado="borrador",
            version=1,
            archivo_nombre=stored["filename"],
            archivo_ruta=stored["path"],
            archivo_tipo=stored["mime"],
            archivo_tamano=stored["size"],
            observaciones=(observaciones or "").strip(),
            creado_por=getattr(request.state, "user_name", "") or request.cookies.get("user_name", ""),
            creado_at=datetime.utcnow(),
        )
        db.add(doc)
        db.commit()
        db.refresh(doc)
        return {"success": True, "data": _serialize_documento(doc, request=request)}
    except Exception:
        _delete_evidence_file(stored.get("path"))
        raise
    finally:
        db.close()


@app.put("/api/documentos/{doc_id}")
def editar_documento(request: Request, doc_id: int, data: dict = Body(...)):
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            return JSONResponse({"success": False, "error": "Documento no encontrado"}, status_code=404)
        if not _can_edit_document(request, doc):
            return JSONResponse({"success": False, "error": "Sin permisos para editar este documento"}, status_code=403)
        titulo = (data.get("titulo") or "").strip()
        if not titulo:
            return JSONResponse({"success": False, "error": "El título es obligatorio"}, status_code=422)
        proceso = (data.get("proceso") or doc.proceso or "envio").strip().lower()
        if proceso not in {"envio", "autorizacion", "actualizacion"}:
            return JSONResponse({"success": False, "error": "Proceso inválido"}, status_code=422)
        doc.titulo = titulo
        doc.descripcion = (data.get("descripcion") or "").strip()
        doc.proceso = proceso
        doc.observaciones = (data.get("observaciones") or "").strip()
        doc.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(doc)
        return {"success": True, "data": _serialize_documento(doc, request=request)}
    finally:
        db.close()


@app.post("/api/documentos/{doc_id}/enviar")
def enviar_documento(request: Request, doc_id: int):
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            return JSONResponse({"success": False, "error": "Documento no encontrado"}, status_code=404)
        if not _can_edit_document(request, doc):
            return JSONResponse({"success": False, "error": "Sin permisos para enviar este documento"}, status_code=403)
        if doc.estado not in {"borrador", "actualizado", "rechazado"}:
            return JSONResponse({"success": False, "error": "El documento no puede enviarse en su estado actual"}, status_code=409)
        doc.estado = "enviado"
        doc.enviado_por = getattr(request.state, "user_name", "") or request.cookies.get("user_name", "")
        doc.enviado_at = datetime.utcnow()
        doc.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(doc)
        return {"success": True, "data": _serialize_documento(doc, request=request)}
    finally:
        db.close()


@app.post("/api/documentos/{doc_id}/autorizar")
def autorizar_documento(request: Request, doc_id: int, data: dict = Body(default={})):
    require_admin_or_superadmin(request)
    accion = (data.get("accion") or "autorizar").strip().lower()
    observaciones = (data.get("observaciones") or "").strip()
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            return JSONResponse({"success": False, "error": "Documento no encontrado"}, status_code=404)
        if accion == "rechazar":
            doc.estado = "rechazado"
            doc.observaciones = observaciones or doc.observaciones
        else:
            if doc.estado not in {"enviado", "actualizado"}:
                return JSONResponse({"success": False, "error": "Solo se autorizan documentos enviados o actualizados"}, status_code=409)
            doc.estado = "autorizado"
            doc.autorizado_por = getattr(request.state, "user_name", "") or request.cookies.get("user_name", "")
            doc.autorizado_at = datetime.utcnow()
        doc.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(doc)
        return {"success": True, "data": _serialize_documento(doc, request=request)}
    finally:
        db.close()


@app.post("/api/documentos/{doc_id}/actualizar")
async def actualizar_documento_archivo(
    request: Request,
    doc_id: int,
    archivo: UploadFile = File(...),
    descripcion: str = Form(""),
    observaciones: str = Form(""),
):
    stored = await _store_evidence_file(archivo)
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            _delete_evidence_file(stored.get("path"))
            return JSONResponse({"success": False, "error": "Documento no encontrado"}, status_code=404)
        if not _can_edit_document(request, doc):
            _delete_evidence_file(stored.get("path"))
            return JSONResponse({"success": False, "error": "Sin permisos para actualizar este documento"}, status_code=403)
        previous_path = doc.archivo_ruta
        doc.archivo_nombre = stored["filename"]
        doc.archivo_ruta = stored["path"]
        doc.archivo_tipo = stored["mime"]
        doc.archivo_tamano = stored["size"]
        doc.version = int(doc.version or 1) + 1
        doc.estado = "actualizado"
        if (descripcion or "").strip():
            doc.descripcion = descripcion.strip()
        if (observaciones or "").strip():
            doc.observaciones = observaciones.strip()
        doc.actualizado_por = getattr(request.state, "user_name", "") or request.cookies.get("user_name", "")
        doc.actualizado_at = datetime.utcnow()
        doc.updated_at = datetime.utcnow()
        db.commit()
        db.refresh(doc)
        _delete_evidence_file(previous_path)
        return {"success": True, "data": _serialize_documento(doc, request=request)}
    finally:
        db.close()


@app.get("/api/documentos/{doc_id}/download")
def descargar_documento(request: Request, doc_id: int):
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            raise HTTPException(status_code=404, detail="Documento no encontrado")
        if not doc.archivo_ruta or not os.path.exists(doc.archivo_ruta):
            raise HTTPException(status_code=404, detail="Archivo no disponible")
        with open(doc.archivo_ruta, "rb") as f:
            content = f.read()
        return Response(
            content=content,
            media_type=doc.archivo_tipo or "application/octet-stream",
            headers={"Content-Disposition": f'inline; filename="{_sanitize_document_name(doc.archivo_nombre)}"'},
        )
    finally:
        db.close()


@app.delete("/api/documentos/{doc_id}")
def eliminar_documento(request: Request, doc_id: int):
    db = SessionLocal()
    try:
        doc = _get_document_for_request(db, request, doc_id)
        if not doc:
            return JSONResponse({"success": False, "error": "Documento no encontrado"}, status_code=404)
        if not _can_edit_document(request, doc):
            return JSONResponse({"success": False, "error": "Sin permisos para eliminar este documento"}, status_code=403)
        path = doc.archivo_ruta
        db.delete(doc)
        db.commit()
        _delete_evidence_file(path)
        return {"success": True}
    finally:
        db.close()


@app.get("/api/reportes/export/{formato}")
def exportar_reporte(formato: str):
    formato_normalizado = (formato or "").strip().lower()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    context = _build_report_export_context()
    rows = _build_report_export_rows()

    if formato_normalizado == "html":
        document = _build_report_export_html_document()
        filename = f"reporte_consolidado_{timestamp}.html"
        return Response(
            content=document,
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if formato_normalizado == "pdf":
        filename = f"reporte_consolidado_{timestamp}.pdf"
        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=letter)
        width, height = letter
        y = height - 56
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(48, y, context["empresa"])
        y -= 20
        pdf.setFont("Helvetica-Bold", 12)
        pdf.drawString(48, y, context["titulo_reporte"])
        y -= 16
        pdf.setFont("Helvetica", 10)
        pdf.drawString(48, y, context["subtitulo_reporte"])
        y -= 16
        pdf.drawString(48, y, f"Fecha: {context['fecha_reporte']}")
        y -= 28
        pdf.setFont("Helvetica-Bold", 11)
        pdf.drawString(48, y, "Detalle de reportes")
        y -= 18
        pdf.setFont("Helvetica", 10)
        for row in rows:
            line = f"- {row['reporte']}: {row['descripcion']} ({row['formato']})"
            pdf.drawString(48, y, line[:110])
            y -= 14
            if y < 60:
                pdf.showPage()
                y = height - 56
                pdf.setFont("Helvetica", 10)
        pdf.save()
        content = buffer.getvalue()
        buffer.close()
        return Response(
            content=content,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if formato_normalizado == "excel":
        filename = f"reporte_consolidado_{timestamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"
        ws["A1"] = context["empresa"]
        ws["A2"] = context["titulo_reporte"]
        ws["A3"] = context["subtitulo_reporte"]
        ws["A4"] = f"Fecha: {context['fecha_reporte']}"
        ws["A6"] = "Reporte"
        ws["B6"] = "Descripcion"
        ws["C6"] = "Formato"
        row_index = 7
        for row in rows:
            ws[f"A{row_index}"] = row["reporte"]
            ws[f"B{row_index}"] = row["descripcion"]
            ws[f"C{row_index}"] = row["formato"]
            row_index += 1
        output = BytesIO()
        wb.save(output)
        content = output.getvalue()
        output.close()
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    return JSONResponse({"success": False, "error": "Formato no soportado"}, status_code=400)


@app.get("/kpis", response_class=HTMLResponse)
def kpis_page(request: Request):
    return render_backend_page(
        request,
        title="KPIs",
        description="Gestión y seguimiento de indicadores clave de desempeño.",
        content=KPI_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )



# --- NUEVO: Endpoint dinámico para usuarios desde BD ---
def _render_usuarios_page(
    request: Request,
    title: str = "Usuarios",
    description: str = "Gestiona usuarios, roles y permisos desde la misma pantalla",
) -> HTMLResponse:
    db = SessionLocal()
    usuarios = db.query(Usuario).all()
    db.close()
    usuarios_content = f"""
        <section id="usuario-panel" class="usuario-panel">
            <div id="usuario-view"></div>
        </section>
    """
    return render_backend_page(
        request,
        title=title,
        description=description,
        content=usuarios_content,
        hide_floating_actions=False,
        view_buttons=[
            {"label": "Form", "icon": "/templates/icon/formulario.svg", "view": "form", "active": True},
            {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list"},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
            {"label": "Organigrama", "view": "organigrama"},
        ],
        floating_actions_screen="usuarios",
    )


def _serialize_strategic_objective(obj: StrategicObjectiveConfig) -> Dict[str, Any]:
    return {
        "id": obj.id,
        "eje_id": obj.eje_id,
        "codigo": obj.codigo or "",
        "nombre": obj.nombre or "",
        "lider": obj.lider or "",
        "fecha_inicial": _date_to_iso(obj.fecha_inicial),
        "fecha_final": _date_to_iso(obj.fecha_final),
        "descripcion": obj.descripcion or "",
        "orden": obj.orden or 0,
    }


def _serialize_strategic_axis(axis: StrategicAxisConfig) -> Dict[str, Any]:
    objetivos = sorted(axis.objetivos or [], key=lambda item: (item.orden or 0, item.id or 0))
    return {
        "id": axis.id,
        "nombre": axis.nombre or "",
        "codigo": axis.codigo or "",
        "lider_departamento": axis.lider_departamento or "",
        "descripcion": axis.descripcion or "",
        "orden": axis.orden or 0,
        "objetivos_count": len(objetivos),
        "objetivos": [_serialize_strategic_objective(obj) for obj in objetivos],
    }


def _serialize_poa_subactivity(item: POASubactivity) -> Dict[str, Any]:
    return {
        "id": item.id,
        "activity_id": item.activity_id,
        "nombre": item.nombre or "",
        "codigo": item.codigo or "",
        "responsable": item.responsable or "",
        "entregable": item.entregable or "",
        "fecha_inicial": _date_to_iso(item.fecha_inicial),
        "fecha_final": _date_to_iso(item.fecha_final),
        "descripcion": item.descripcion or "",
    }


def _serialize_poa_activity(item: POAActivity, subactivities: List[POASubactivity]) -> Dict[str, Any]:
    return {
        "id": item.id,
        "objective_id": item.objective_id,
        "nombre": item.nombre or "",
        "codigo": item.codigo or "",
        "responsable": item.responsable or "",
        "entregable": item.entregable or "",
        "fecha_inicial": _date_to_iso(item.fecha_inicial),
        "fecha_final": _date_to_iso(item.fecha_final),
        "status": _activity_status(item),
        "entrega_estado": item.entrega_estado or "ninguna",
        "entrega_solicitada_por": item.entrega_solicitada_por or "",
        "entrega_solicitada_at": item.entrega_solicitada_at.isoformat() if item.entrega_solicitada_at else "",
        "entrega_aprobada_por": item.entrega_aprobada_por or "",
        "entrega_aprobada_at": item.entrega_aprobada_at.isoformat() if item.entrega_aprobada_at else "",
        "descripcion": item.descripcion or "",
        "subactivities": [_serialize_poa_subactivity(sub) for sub in subactivities],
    }


EJES_ESTRATEGICOS_HTML = dedent("""
    <section class="axm-wrap">
      <style>
        .axm-wrap *{ box-sizing:border-box; }
        .axm-wrap{
          font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
          color:#0f172a;
          padding: 10px;
        }
        .axm-grid{
          display:grid;
          grid-template-columns: minmax(280px, 360px) minmax(0, 1fr);
          gap: 14px;
        }
        .axm-card{
          background: rgba(255,255,255,.92);
          border: 1px solid rgba(148,163,184,.30);
          border-radius: 18px;
          padding: 14px;
          box-shadow: 0 8px 20px rgba(15,23,42,.06);
        }
        .axm-title{ margin:0; font-size: 20px; letter-spacing: -0.02em; }
        .axm-sub{ margin: 6px 0 0; color:#64748b; font-size:13px; }
        .axm-list{ margin-top: 12px; display:flex; flex-direction:column; gap: 10px; max-height: 65vh; overflow:auto; }
        .axm-axis-btn{
          width: 100%;
          border: 1px solid rgba(148,163,184,.32);
          background: rgba(255,255,255,.96);
          border-radius: 14px;
          padding: 12px;
          text-align: left;
          display:flex;
          align-items:center;
          justify-content:space-between;
          gap: 10px;
          cursor:pointer;
        }
        .axm-axis-btn.active{
          background: rgba(15,61,46,.08);
          border-color: rgba(15,61,46,.32);
        }
        .axm-axis-meta{ color:#64748b; font-size: 12px; }
        .axm-count{
          font-size: 12px;
          font-weight: 800;
          border:1px solid rgba(15,23,42,.14);
          border-radius: 999px;
          padding: 4px 8px;
          background: rgba(15,23,42,.04);
        }
        .axm-row{ display:grid; grid-template-columns: 1fr 1fr; gap: 10px; }
        .axm-field{ display:flex; flex-direction:column; gap: 6px; margin-top: 10px; }
        .axm-field label{ font-size: 12px; color:#475569; font-weight:700; letter-spacing: .02em; }
        .axm-input, .axm-textarea{
          width:100%;
          border:1px solid rgba(148,163,184,.42);
          border-radius: 12px;
          padding: 10px 12px;
          font-size: 14px;
          background: #fff;
        }
        .axm-textarea{ min-height: 82px; resize: vertical; }
        .axm-actions{ display:flex; gap:8px; flex-wrap:wrap; margin-top: 12px; }
        .axm-btn{
          border:1px solid rgba(148,163,184,.42);
          border-radius: 12px;
          padding: 9px 12px;
          background:#fff;
          cursor:pointer;
          font-weight:700;
          font-size: 13px;
        }
        .axm-btn.primary{ background:#0f3d2e; border-color:#0f3d2e; color:#fff; }
        .axm-btn.warn{ background:#ef4444; border-color:#ef4444; color:#fff; }
        .axm-obj-layout{
          margin-top: 12px;
          display:grid;
          grid-template-columns: minmax(280px, 0.95fr) minmax(460px, 1.35fr);
          gap: 14px;
          align-items: start;
        }
        .axm-obj-list{ display:flex; flex-direction:column; gap: 8px; max-height: 320px; overflow:auto; }
        .axm-obj-btn{
          width: 100%;
          text-align: left;
          border:1px solid rgba(148,163,184,.32);
          border-radius: 12px;
          padding: 10px;
          background: rgba(255,255,255,.95);
          cursor: pointer;
        }
        .axm-obj-btn.active{
          background: rgba(15,61,46,.08);
          border-color: rgba(15,61,46,.30);
        }
        .axm-obj-form{
          border:1px solid rgba(148,163,184,.32);
          border-radius: 12px;
          padding: 14px;
          background: rgba(255,255,255,.95);
        }
        .axm-obj-form .axm-row{
          grid-template-columns: minmax(0, 1fr) minmax(0, 1fr);
        }
        .axm-obj-form .axm-input,
        .axm-obj-form .axm-textarea{
          min-width: 0;
          width: 100%;
        }
        .axm-obj-form .axm-textarea{
          min-height: 116px;
        }
        .axm-obj-grid{ display:grid; grid-template-columns: 150px 1fr; gap: 8px; }
        .axm-msg{ margin-top: 10px; font-size: 13px; color:#0f3d2e; min-height: 1.2em; }
        @media (max-width: 1200px){
          .axm-obj-layout{ grid-template-columns: 1fr; }
          .axm-obj-form .axm-row{ grid-template-columns: 1fr; }
        }
        @media (max-width: 980px){
          .axm-grid{ grid-template-columns: 1fr; }
          .axm-list{ max-height: 36vh; }
          .axm-row{ grid-template-columns: 1fr; }
          .axm-obj-layout{ grid-template-columns: 1fr; }
          .axm-obj-grid{ grid-template-columns: 1fr; }
        }
      </style>

      <div class="axm-grid">
        <aside class="axm-card">
          <h2 class="axm-title">Ejes estratégicos</h2>
          <p class="axm-sub">Selecciona un eje para editarlo o crea uno nuevo.</p>
          <div class="axm-actions">
            <button class="axm-btn primary" id="axm-add-axis" type="button">Agregar eje</button>
          </div>
          <div class="axm-list" id="axm-axis-list"></div>
        </aside>

        <section class="axm-card">
          <h2 class="axm-title">Gestión de ejes y objetivos</h2>
          <p class="axm-sub">Edita, guarda o elimina ejes estratégicos y sus objetivos.</p>
          <div class="axm-row">
            <div class="axm-field">
              <label for="axm-axis-name">Nombre del eje</label>
              <input id="axm-axis-name" class="axm-input" type="text" placeholder="Ej. Gobernanza y cumplimiento">
            </div>
            <div class="axm-field">
              <label for="axm-axis-code">Código</label>
              <input id="axm-axis-code" class="axm-input" type="text" placeholder="Ej. AX-01">
            </div>
          </div>
          <div class="axm-field">
            <label for="axm-axis-leader">Líder del eje estratégico</label>
            <select id="axm-axis-leader" class="axm-input">
              <option value="">Selecciona departamento</option>
            </select>
          </div>
          <div class="axm-field">
            <label for="axm-axis-desc">Descripción</label>
            <textarea id="axm-axis-desc" class="axm-textarea" placeholder="Describe el propósito del eje"></textarea>
          </div>
          <div class="axm-actions">
            <button class="axm-btn primary" id="axm-save-axis" type="button">Guardar eje</button>
            <button class="axm-btn warn" id="axm-delete-axis" type="button">Eliminar eje</button>
          </div>

          <hr style="border:0;border-top:1px solid rgba(148,163,184,.28);margin:14px 0;">

          <h3 style="margin:0;font-size:16px;">Objetivos del eje</h3>
          <div class="axm-obj-layout">
            <div>
              <div class="axm-actions" style="margin-top:0;">
                <button class="axm-btn primary" id="axm-add-obj" type="button">Agregar objetivo</button>
              </div>
              <div class="axm-obj-list" id="axm-obj-list"></div>
            </div>
            <div class="axm-obj-form">
              <div class="axm-field" style="margin-top:0;">
                <label for="axm-obj-name">Nombre</label>
                <input id="axm-obj-name" class="axm-input" type="text" placeholder="Nombre del objetivo">
              </div>
              <div class="axm-field">
                <label for="axm-obj-code">Código</label>
                <input id="axm-obj-code" class="axm-input" type="text" placeholder="Ej. OE-13">
              </div>
              <div class="axm-field">
                <label for="axm-obj-leader">Lider</label>
                <select id="axm-obj-leader" class="axm-input">
                  <option value="">Selecciona colaborador</option>
                </select>
              </div>
              <div class="axm-row">
                <div class="axm-field">
                  <label for="axm-obj-start">Fecha inicial</label>
                  <input id="axm-obj-start" class="axm-input" type="date">
                </div>
                <div class="axm-field">
                  <label for="axm-obj-end">Fecha final</label>
                  <input id="axm-obj-end" class="axm-input" type="date">
                </div>
              </div>
              <div class="axm-field">
                <label for="axm-obj-desc">Descripción</label>
                <textarea id="axm-obj-desc" class="axm-textarea" placeholder="Descripción del objetivo"></textarea>
              </div>
              <div class="axm-actions">
                <button class="axm-btn primary" id="axm-save-obj" type="button">Guardar objetivo</button>
                <button class="axm-btn warn" id="axm-delete-obj" type="button">Eliminar objetivo</button>
              </div>
            </div>
          </div>
          <div class="axm-msg" id="axm-msg" aria-live="polite"></div>
        </section>
      </div>

      <script>
        (() => {
          const axisListEl = document.getElementById("axm-axis-list");
          const objListEl = document.getElementById("axm-obj-list");
          const axisNameEl = document.getElementById("axm-axis-name");
          const axisCodeEl = document.getElementById("axm-axis-code");
          const axisLeaderEl = document.getElementById("axm-axis-leader");
          const axisDescEl = document.getElementById("axm-axis-desc");
          const objNameEl = document.getElementById("axm-obj-name");
          const objCodeEl = document.getElementById("axm-obj-code");
          const objLeaderEl = document.getElementById("axm-obj-leader");
          const objStartEl = document.getElementById("axm-obj-start");
          const objEndEl = document.getElementById("axm-obj-end");
          const objDescEl = document.getElementById("axm-obj-desc");
          const msgEl = document.getElementById("axm-msg");
          const addAxisBtn = document.getElementById("axm-add-axis");
          const saveAxisBtn = document.getElementById("axm-save-axis");
          const deleteAxisBtn = document.getElementById("axm-delete-axis");
          const addObjBtn = document.getElementById("axm-add-obj");
          const saveObjBtn = document.getElementById("axm-save-obj");
          const deleteObjBtn = document.getElementById("axm-delete-obj");

          let axes = [];
          let departments = [];
          let collaborators = [];
          let selectedAxisId = null;
          let selectedObjectiveId = null;

          const showMsg = (text, isError = false) => {
            if (!msgEl) return;
            msgEl.style.color = isError ? "#b91c1c" : "#0f3d2e";
            msgEl.textContent = text || "";
          };

          const requestJson = async (url, options = {}) => {
            const response = await fetch(url, {
              headers: { "Content-Type": "application/json" },
              credentials: "same-origin",
              ...options,
            });
            const payload = await response.json().catch(() => ({}));
            if (!response.ok || payload.success === false) {
              throw new Error(payload.error || "No se pudo completar la operación.");
            }
            return payload;
          };

          const selectedAxis = () => axes.find((axis) => axis.id === selectedAxisId) || null;
          const selectedObjective = () => {
            const axis = selectedAxis();
            if (!axis) return null;
            return (axis.objetivos || []).find((obj) => obj.id === selectedObjectiveId) || null;
          };
          const visualRangeError = (start, end, label) => {
            if (!start && !end) return "";
            if (!start || !end) return `${label}: completa fecha inicial y fecha final.`;
            if (start > end) return `${label}: la fecha inicial no puede ser mayor que la final.`;
            return "";
          };

          const renderDepartmentOptions = (selectedValue = "") => {
            if (!axisLeaderEl) return;
            const options = ['<option value="">Selecciona departamento</option>']
              .concat(
                departments.map((name) => {
                  const selected = name === selectedValue ? "selected" : "";
                  return `<option value="${name}" ${selected}>${name}</option>`;
                })
              );
            axisLeaderEl.innerHTML = options.join("");
          };

          const renderCollaboratorOptions = (selectedValue = "") => {
            if (!objLeaderEl) return;
            const options = ['<option value="">Selecciona colaborador</option>']
              .concat(
                collaborators.map((name) => {
                  const selected = name === selectedValue ? "selected" : "";
                  return `<option value="${name}" ${selected}>${name}</option>`;
                })
              );
            objLeaderEl.innerHTML = options.join("");
          };

          const renderAxisList = () => {
            if (!axisListEl) return;
            axisListEl.innerHTML = axes.map((axis) => `
              <button class="axm-axis-btn ${axis.id === selectedAxisId ? "active" : ""}" type="button" data-axis-id="${axis.id}">
                <span>
                  <strong>${axis.nombre}</strong>
                  <div class="axm-axis-meta">${axis.codigo || "Sin código"} • ${axis.lider_departamento || "Sin líder"}</div>
                </span>
                <span class="axm-count">${axis.objetivos_count || 0}</span>
              </button>
            `).join("");
            axisListEl.querySelectorAll("[data-axis-id]").forEach((button) => {
              button.addEventListener("click", async () => {
                selectedAxisId = Number(button.getAttribute("data-axis-id"));
                selectedObjectiveId = null;
                await loadCollaborators();
                renderAll();
              });
            });
          };

          const renderAxisEditor = () => {
            const axis = selectedAxis();
            if (!axis) {
              axisNameEl.value = "";
              axisCodeEl.value = "";
              renderDepartmentOptions("");
              axisDescEl.value = "";
              return;
            }
            axisNameEl.value = axis.nombre || "";
            axisCodeEl.value = axis.codigo || "";
            renderDepartmentOptions(axis.lider_departamento || "");
            axisDescEl.value = axis.descripcion || "";
          };

          const renderObjectives = () => {
            const axis = selectedAxis();
            if (!axis || !objListEl) {
              if (objListEl) objListEl.innerHTML = "";
              selectedObjectiveId = null;
              if (objNameEl) objNameEl.value = "";
              if (objCodeEl) objCodeEl.value = "";
              if (objDescEl) objDescEl.value = "";
              if (objStartEl) objStartEl.value = "";
              if (objEndEl) objEndEl.value = "";
              renderCollaboratorOptions("");
              return;
            }
            if (!selectedObjectiveId || !(axis.objetivos || []).some((obj) => obj.id === selectedObjectiveId)) {
              selectedObjectiveId = (axis.objetivos || [])[0]?.id || null;
            }
            objListEl.innerHTML = (axis.objetivos || []).map((obj) => `
              <button class="axm-obj-btn ${obj.id === selectedObjectiveId ? "active" : ""}" type="button" data-obj-id="${obj.id}">
                <strong>${obj.codigo || "OBJ"} - ${obj.nombre || "Sin nombre"}</strong>
              </button>
            `).join("");

            objListEl.querySelectorAll("[data-obj-id]").forEach((button) => {
              button.addEventListener("click", () => {
                selectedObjectiveId = Number(button.getAttribute("data-obj-id"));
                renderAll();
              });
            });

            const objective = selectedObjective();
            if (!objective) return;
            if (objNameEl) objNameEl.value = objective.nombre || "";
            if (objCodeEl) objCodeEl.value = objective.codigo || "";
            if (objDescEl) objDescEl.value = objective.descripcion || "";
            if (objStartEl) objStartEl.value = objective.fecha_inicial || "";
            if (objEndEl) objEndEl.value = objective.fecha_final || "";
            renderCollaboratorOptions(objective.lider || "");
          };

          const renderAll = () => {
            renderAxisList();
            renderAxisEditor();
            renderObjectives();
          };

          const loadAxes = async () => {
            const payload = await requestJson("/api/strategic-axes");
            axes = Array.isArray(payload.data) ? payload.data : [];
            if (!selectedAxisId || !axes.some((axis) => axis.id === selectedAxisId)) {
              selectedAxisId = axes.length ? axes[0].id : null;
            }
            renderAll();
          };

          const loadDepartments = async () => {
            const payload = await requestJson("/api/strategic-axes/departments");
            departments = Array.isArray(payload.data) ? payload.data : [];
            renderDepartmentOptions(selectedAxis()?.lider_departamento || "");
          };

          const loadCollaborators = async () => {
            const axis = selectedAxis();
            if (!axis || !axis.id) {
              collaborators = [];
              renderCollaboratorOptions("");
              return;
            }
            const payload = await requestJson(`/api/strategic-axes/${axis.id}/collaborators`);
            collaborators = Array.isArray(payload.data) ? payload.data : [];
            renderCollaboratorOptions(selectedObjective()?.lider || "");
          };

          addAxisBtn && addAxisBtn.addEventListener("click", async () => {
            try {
              const payload = await requestJson("/api/strategic-axes", {
                method: "POST",
                body: JSON.stringify({
                  nombre: "Nuevo eje estratégico",
                  codigo: "",
                  lider_departamento: "",
                  descripcion: "",
                  orden: axes.length + 1,
                }),
              });
              selectedAxisId = payload.data?.id || null;
              await loadAxes();
              await loadCollaborators();
              showMsg("Eje agregado.");
            } catch (err) {
              showMsg(err.message || "No se pudo crear el eje.", true);
            }
          });

          saveAxisBtn && saveAxisBtn.addEventListener("click", async () => {
            const axis = selectedAxis();
            if (!axis) {
              showMsg("Selecciona un eje para guardar.", true);
              return;
            }
            const body = {
              nombre: axisNameEl.value.trim(),
              codigo: axisCodeEl.value.trim(),
              lider_departamento: axisLeaderEl && axisLeaderEl.value ? axisLeaderEl.value.trim() : "",
              descripcion: axisDescEl.value.trim(),
              orden: Number(axis.orden || 1),
            };
            if (!body.nombre) {
              showMsg("El nombre del eje es obligatorio.", true);
              return;
            }
            try {
              await requestJson(`/api/strategic-axes/${axis.id}`, { method: "PUT", body: JSON.stringify(body) });
              await loadAxes();
              await loadCollaborators();
              showMsg("Eje guardado correctamente.");
            } catch (err) {
              showMsg(err.message || "No se pudo guardar el eje.", true);
            }
          });

          deleteAxisBtn && deleteAxisBtn.addEventListener("click", async () => {
            const axis = selectedAxis();
            if (!axis) return;
            if (!window.confirm("¿Eliminar este eje y todos sus objetivos?")) return;
            try {
              await requestJson(`/api/strategic-axes/${axis.id}`, { method: "DELETE" });
              selectedAxisId = null;
              await loadAxes();
              await loadCollaborators();
              showMsg("Eje eliminado.");
            } catch (err) {
              showMsg(err.message || "No se pudo eliminar el eje.", true);
            }
          });

          addObjBtn && addObjBtn.addEventListener("click", async () => {
            const axis = selectedAxis();
            if (!axis) {
              showMsg("Primero selecciona un eje.", true);
              return;
            }
            const body = {
              codigo: "",
              nombre: "Nuevo objetivo",
              lider: "",
              descripcion: "",
            };
            try {
              const payload = await requestJson(`/api/strategic-axes/${axis.id}/objectives`, { method: "POST", body: JSON.stringify(body) });
              await loadAxes();
              selectedObjectiveId = payload.data?.id || selectedObjectiveId;
              renderAll();
              showMsg("Objetivo agregado.");
            } catch (err) {
              showMsg(err.message || "No se pudo agregar el objetivo.", true);
            }
          });

          saveObjBtn && saveObjBtn.addEventListener("click", async () => {
            const objective = selectedObjective();
            if (!objective) {
              showMsg("Selecciona un objetivo.", true);
              return;
            }
            const body = {
              nombre: objNameEl && objNameEl.value ? objNameEl.value.trim() : "",
              codigo: objCodeEl && objCodeEl.value ? objCodeEl.value.trim() : "",
              lider: objLeaderEl && objLeaderEl.value ? objLeaderEl.value.trim() : "",
              fecha_inicial: objStartEl && objStartEl.value ? objStartEl.value : "",
              fecha_final: objEndEl && objEndEl.value ? objEndEl.value : "",
              descripcion: objDescEl && objDescEl.value ? objDescEl.value.trim() : "",
            };
            if (!body.nombre) {
              showMsg("El nombre del objetivo es obligatorio.", true);
              return;
            }
            const objectiveDateError = visualRangeError(body.fecha_inicial, body.fecha_final, "Objetivo");
            if (objectiveDateError) {
              showMsg(objectiveDateError, true);
              return;
            }
            try {
              await requestJson(`/api/strategic-objectives/${objective.id}`, { method: "PUT", body: JSON.stringify(body) });
              await loadAxes();
              renderAll();
              showMsg("Objetivo guardado correctamente.");
            } catch (err) {
              showMsg(err.message || "No se pudo guardar el objetivo.", true);
            }
          });

          deleteObjBtn && deleteObjBtn.addEventListener("click", async () => {
            const objective = selectedObjective();
            if (!objective) return;
            if (!window.confirm("¿Eliminar este objetivo?")) return;
            try {
              await requestJson(`/api/strategic-objectives/${objective.id}`, { method: "DELETE" });
              selectedObjectiveId = null;
              await loadAxes();
              renderAll();
              showMsg("Objetivo eliminado.");
            } catch (err) {
              showMsg(err.message || "No se pudo eliminar el objetivo.", true);
            }
          });

          Promise.all([loadDepartments(), loadAxes()]).then(loadCollaborators).catch((err) => {
            showMsg(err.message || "No se pudieron cargar los ejes.", true);
          });
        })();
      </script>
    </section>
""")


@app.get("/usuarios", response_class=HTMLResponse)
@app.get("/usuarios-sistema", response_class=HTMLResponse)
def usuarios_page(request: Request):
    return _render_usuarios_page(request)


@app.get("/ejes-estrategicos", response_class=HTMLResponse)
def ejes_estrategicos_page(request: Request):
    return render_backend_page(
        request,
        title="Ejes estratégicos",
        description="Edición y administración de ejes y objetivos estratégicos.",
        content=EJES_ESTRATEGICOS_HTML,
        hide_floating_actions=True,
        show_page_header=True,
        view_buttons=[
            {"label": "Form", "icon": "/templates/icon/formulario.svg", "view": "form", "active": True},
        ],
    )


@app.get("/api/strategic-axes")
def list_strategic_axes(request: Request):
    db = SessionLocal()
    try:
        axes = (
            db.query(StrategicAxisConfig)
            .filter(StrategicAxisConfig.is_active == True)
            .order_by(StrategicAxisConfig.orden.asc(), StrategicAxisConfig.id.asc())
            .all()
        )
        return JSONResponse({"success": True, "data": [_serialize_strategic_axis(axis) for axis in axes]})
    finally:
        db.close()


@app.get("/api/strategic-axes/departments")
def list_strategic_axis_departments():
    db = SessionLocal()
    try:
        departments = []
        rows = (
            db.query(Usuario.departamento)
            .filter(Usuario.departamento.isnot(None))
            .all()
        )
        for row in rows:
            value = (row[0] or "").strip()
            if value:
                departments.append(value)
        unique_departments = sorted(set(departments), key=lambda item: item.lower())
        return JSONResponse({"success": True, "data": unique_departments})
    finally:
        db.close()


@app.get("/api/strategic-axes/{axis_id}/collaborators")
def list_strategic_axis_collaborators(axis_id: int):
    db = SessionLocal()
    try:
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == axis_id).first()
        if not axis:
            return JSONResponse({"success": False, "error": "Eje no encontrado"}, status_code=404)
        department = (axis.lider_departamento or "").strip()
        if not department:
            return JSONResponse({"success": True, "data": []})
        rows = (
            db.query(Usuario.nombre)
            .filter(Usuario.departamento == department)
            .all()
        )
        collaborators = []
        for row in rows:
            value = (row[0] or "").strip()
            if value:
                collaborators.append(value)
        unique_collaborators = sorted(set(collaborators), key=lambda item: item.lower())
        return JSONResponse({"success": True, "data": unique_collaborators})
    finally:
        db.close()


@app.post("/api/strategic-axes")
def create_strategic_axis(request: Request, data: dict = Body(...)):
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return JSONResponse({"success": False, "error": "El nombre del eje es obligatorio"}, status_code=400)

    db = SessionLocal()
    try:
        max_order = db.query(func.max(StrategicAxisConfig.orden)).scalar() or 0
        axis = StrategicAxisConfig(
            nombre=nombre,
            codigo=(data.get("codigo") or "").strip(),
            lider_departamento=(data.get("lider_departamento") or "").strip(),
            descripcion=(data.get("descripcion") or "").strip(),
            orden=int(data.get("orden") or (max_order + 1)),
            is_active=True,
        )
        db.add(axis)
        db.commit()
        db.refresh(axis)
        return JSONResponse({"success": True, "data": _serialize_strategic_axis(axis)})
    finally:
        db.close()


@app.put("/api/strategic-axes/{axis_id}")
def update_strategic_axis(axis_id: int, data: dict = Body(...)):
    db = SessionLocal()
    try:
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == axis_id).first()
        if not axis:
            return JSONResponse({"success": False, "error": "Eje no encontrado"}, status_code=404)
        nombre = (data.get("nombre") or "").strip()
        if not nombre:
            return JSONResponse({"success": False, "error": "El nombre del eje es obligatorio"}, status_code=400)
        axis.nombre = nombre
        axis.codigo = (data.get("codigo") or "").strip()
        axis.lider_departamento = (data.get("lider_departamento") or "").strip()
        axis.descripcion = (data.get("descripcion") or "").strip()
        axis.orden = int(data.get("orden") or axis.orden or 1)
        db.add(axis)
        db.commit()
        db.refresh(axis)
        return JSONResponse({"success": True, "data": _serialize_strategic_axis(axis)})
    finally:
        db.close()


@app.delete("/api/strategic-axes/{axis_id}")
def delete_strategic_axis(axis_id: int):
    db = SessionLocal()
    try:
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == axis_id).first()
        if not axis:
            return JSONResponse({"success": False, "error": "Eje no encontrado"}, status_code=404)
        db.delete(axis)
        db.commit()
        return JSONResponse({"success": True})
    finally:
        db.close()


@app.post("/api/strategic-axes/{axis_id}/objectives")
def create_strategic_objective(axis_id: int, data: dict = Body(...)):
    nombre = (data.get("nombre") or "").strip()
    if not nombre:
        return JSONResponse({"success": False, "error": "El nombre del objetivo es obligatorio"}, status_code=400)
    db = SessionLocal()
    try:
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == axis_id).first()
        if not axis:
            return JSONResponse({"success": False, "error": "Eje no encontrado"}, status_code=404)
        start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=False)
        if start_error:
            return JSONResponse({"success": False, "error": start_error}, status_code=400)
        end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=False)
        if end_error:
            return JSONResponse({"success": False, "error": end_error}, status_code=400)
        if (start_date and not end_date) or (end_date and not start_date):
            return JSONResponse(
                {"success": False, "error": "Objetivo: fecha inicial y fecha final deben definirse juntas"},
                status_code=400,
            )
        if start_date and end_date:
            range_error = _validate_date_range(start_date, end_date, "Objetivo")
            if range_error:
                return JSONResponse({"success": False, "error": range_error}, status_code=400)
        max_order = (
            db.query(func.max(StrategicObjectiveConfig.orden))
            .filter(StrategicObjectiveConfig.eje_id == axis_id)
            .scalar()
            or 0
        )
        objective = StrategicObjectiveConfig(
            eje_id=axis_id,
            codigo=(data.get("codigo") or "").strip(),
            nombre=nombre,
            lider=(data.get("lider") or "").strip(),
            fecha_inicial=start_date,
            fecha_final=end_date,
            descripcion=(data.get("descripcion") or "").strip(),
            orden=int(data.get("orden") or (max_order + 1)),
            is_active=True,
        )
        db.add(objective)
        db.commit()
        db.refresh(objective)
        return JSONResponse({"success": True, "data": _serialize_strategic_objective(objective)})
    finally:
        db.close()


@app.put("/api/strategic-objectives/{objective_id}")
def update_strategic_objective(objective_id: int, data: dict = Body(...)):
    db = SessionLocal()
    try:
        objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == objective_id).first()
        if not objective:
            return JSONResponse({"success": False, "error": "Objetivo no encontrado"}, status_code=404)
        nombre = (data.get("nombre") or "").strip()
        if not nombre:
            return JSONResponse({"success": False, "error": "El nombre del objetivo es obligatorio"}, status_code=400)
        start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=False)
        if start_error:
            return JSONResponse({"success": False, "error": start_error}, status_code=400)
        end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=False)
        if end_error:
            return JSONResponse({"success": False, "error": end_error}, status_code=400)
        if (start_date and not end_date) or (end_date and not start_date):
            return JSONResponse(
                {"success": False, "error": "Objetivo: fecha inicial y fecha final deben definirse juntas"},
                status_code=400,
            )
        if start_date and end_date:
            range_error = _validate_date_range(start_date, end_date, "Objetivo")
            if range_error:
                return JSONResponse({"success": False, "error": range_error}, status_code=400)
        objective.codigo = (data.get("codigo") or "").strip()
        objective.nombre = nombre
        objective.lider = (data.get("lider") or "").strip()
        objective.fecha_inicial = start_date
        objective.fecha_final = end_date
        objective.descripcion = (data.get("descripcion") or "").strip()
        if data.get("orden") is not None:
            objective.orden = int(data.get("orden"))
        db.add(objective)
        db.commit()
        db.refresh(objective)
        return JSONResponse({"success": True, "data": _serialize_strategic_objective(objective)})
    finally:
        db.close()


@app.delete("/api/strategic-objectives/{objective_id}")
def delete_strategic_objective(objective_id: int):
    db = SessionLocal()
    try:
        objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == objective_id).first()
        if not objective:
            return JSONResponse({"success": False, "error": "Objetivo no encontrado"}, status_code=404)
        db.delete(objective)
        db.commit()
        return JSONResponse({"success": True})
    finally:
        db.close()


def _allowed_objectives_for_user(request: Request, db) -> List[StrategicObjectiveConfig]:
    if is_admin_or_superadmin(request):
        return (
            db.query(StrategicObjectiveConfig)
            .filter(StrategicObjectiveConfig.is_active == True)
            .order_by(StrategicObjectiveConfig.orden.asc(), StrategicObjectiveConfig.id.asc())
            .all()
        )

    session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
    user = _current_user_record(request, db)
    aliases = _user_aliases(user, session_username)
    user_department = (user.departamento or "").strip().lower() if user and user.departamento else ""

    objectives = (
        db.query(StrategicObjectiveConfig)
        .join(StrategicAxisConfig, StrategicAxisConfig.id == StrategicObjectiveConfig.eje_id)
        .filter(StrategicObjectiveConfig.is_active == True)
        .order_by(StrategicObjectiveConfig.orden.asc(), StrategicObjectiveConfig.id.asc())
        .all()
    )
    allowed: List[StrategicObjectiveConfig] = []
    for obj in objectives:
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == obj.eje_id).first()
        objective_leader = (obj.lider or "").strip().lower()
        axis_department = (axis.lider_departamento or "").strip().lower() if axis else ""
        if objective_leader and objective_leader in aliases:
            allowed.append(obj)
            continue
        if user_department and axis_department and axis_department == user_department:
            allowed.append(obj)
    return allowed


@app.get("/api/poa/board-data")
def poa_board_data(request: Request):
    db = SessionLocal()
    try:
        objectives = _allowed_objectives_for_user(request, db)
        objective_ids = [obj.id for obj in objectives]
        objective_axis_map = {obj.id: obj.eje_id for obj in objectives}
        axis_ids = sorted(set(objective_axis_map.values()))
        axes = (
            db.query(StrategicAxisConfig)
            .filter(StrategicAxisConfig.id.in_(axis_ids))
            .all()
            if axis_ids else []
        )
        axis_name_map = {axis.id: axis.nombre for axis in axes}

        activities = (
            db.query(POAActivity)
            .filter(POAActivity.objective_id.in_(objective_ids))
            .order_by(POAActivity.id.asc())
            .all()
            if objective_ids else []
        )
        activity_ids = [item.id for item in activities]
        subactivities = (
            db.query(POASubactivity)
            .filter(POASubactivity.activity_id.in_(activity_ids))
            .order_by(POASubactivity.id.asc())
            .all()
            if activity_ids else []
        )
        sub_by_activity: Dict[int, List[POASubactivity]] = {}
        for sub in subactivities:
            sub_by_activity.setdefault(sub.activity_id, []).append(sub)

        pending_approvals = (
            db.query(POADeliverableApproval)
            .filter(POADeliverableApproval.status == "pendiente")
            .order_by(POADeliverableApproval.created_at.desc())
            .all()
        )
        approvals_for_user = []
        for approval in pending_approvals:
            if not _is_user_process_owner(request, db, approval.process_owner):
                continue
            activity = next((item for item in activities if item.id == approval.activity_id), None)
            if not activity:
                activity = db.query(POAActivity).filter(POAActivity.id == approval.activity_id).first()
            objective = next((item for item in objectives if item.id == approval.objective_id), None)
            if not objective:
                objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == approval.objective_id).first()
            approvals_for_user.append(
                {
                    "id": approval.id,
                    "activity_id": approval.activity_id,
                    "objective_id": approval.objective_id,
                    "process_owner": approval.process_owner or "",
                    "requester": approval.requester or "",
                    "created_at": approval.created_at.isoformat() if approval.created_at else "",
                    "activity_nombre": (activity.nombre if activity else ""),
                    "activity_codigo": (activity.codigo if activity else ""),
                    "objective_nombre": (objective.nombre if objective else ""),
                    "objective_codigo": (objective.codigo if objective else ""),
                }
            )

        return JSONResponse(
            {
                "success": True,
                "objectives": [
                    {
                        **_serialize_strategic_objective(obj),
                        "axis_name": axis_name_map.get(obj.eje_id, ""),
                    }
                    for obj in objectives
                ],
                "activities": [
                    _serialize_poa_activity(activity, sub_by_activity.get(activity.id, []))
                    for activity in activities
                ],
                "pending_approvals": approvals_for_user,
            }
        )
    finally:
        db.close()


@app.get("/api/notificaciones/resumen")
def notifications_summary(request: Request):
    db = SessionLocal()
    try:
        now = datetime.utcnow()
        today = now.date()
        tenant_id = _normalize_tenant_id(get_current_tenant(request))
        user_key = _notification_user_key(request, db)
        items: List[Dict[str, Any]] = []

        pending_approvals = (
            db.query(POADeliverableApproval)
            .filter(POADeliverableApproval.status == "pendiente")
            .order_by(POADeliverableApproval.created_at.desc())
            .all()
        )
        for approval in pending_approvals:
            if not _is_user_process_owner(request, db, approval.process_owner):
                continue
            activity = db.query(POAActivity).filter(POAActivity.id == approval.activity_id).first()
            objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == approval.objective_id).first()
            items.append(
                {
                    "id": f"poa-approval-{approval.id}",
                    "kind": "poa_aprobacion",
                    "title": "Aprobación de entregable pendiente",
                    "message": (
                        f"Actividad {(activity.nombre if activity else 'sin nombre')} "
                        f"({activity.codigo if activity else ''}) - Objetivo {(objective.nombre if objective else '')}"
                    ).strip(),
                    "created_at": (approval.created_at or now).isoformat(),
                    "href": "/poa/crear",
                }
            )

        if _can_authorize_documents(request):
            tenant_id = _get_document_tenant(request)
            docs_query = db.query(DocumentoEvidencia).filter(DocumentoEvidencia.estado.in_(["enviado", "actualizado"]))
            if is_superadmin(request):
                header_tenant = request.headers.get("x-tenant-id")
                if header_tenant and _normalize_tenant_id(header_tenant) != "all":
                    docs_query = docs_query.filter(
                        func.lower(DocumentoEvidencia.tenant_id) == _normalize_tenant_id(header_tenant).lower()
                    )
                elif not header_tenant:
                    docs_query = docs_query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
            else:
                docs_query = docs_query.filter(func.lower(DocumentoEvidencia.tenant_id) == tenant_id.lower())
            docs_pending = docs_query.order_by(DocumentoEvidencia.updated_at.desc()).limit(20).all()
            for doc in docs_pending:
                items.append(
                    {
                        "id": f"doc-approval-{doc.id}",
                        "kind": "documento_autorizacion",
                        "title": "Documento pendiente de autorización",
                        "message": f"{(doc.titulo or '').strip()} · Estado: {(doc.estado or '').strip()}",
                        "created_at": (doc.updated_at or doc.enviado_at or doc.creado_at or now).isoformat(),
                        "href": "/reportes/documentos",
                    }
                )

        if is_superadmin(request):
            quiz_rows = (
                db.query(PublicQuizSubmission)
                .order_by(PublicQuizSubmission.created_at.desc(), PublicQuizSubmission.id.desc())
                .limit(20)
                .all()
            )
            for quiz in quiz_rows:
                items.append(
                    {
                        "id": f"quiz-submission-{quiz.id}",
                        "kind": "quiz_descuento",
                        "title": "Nuevo cuestionario de descuento",
                        "message": (
                            f"{(quiz.nombre or '').strip()} · {(quiz.cooperativa or '').strip()} · "
                            f"{int(quiz.correctas or 0)}/10 correctas · {int(quiz.descuento or 0)}% de descuento"
                        ),
                        "created_at": (quiz.created_at or now).isoformat(),
                        "href": "/usuarios",
                    }
                )

        session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        user = _current_user_record(request, db)
        aliases = sorted(_user_aliases(user, session_username))
        if aliases:
            lookahead = today + timedelta(days=2)
            own_activities = (
                db.query(POAActivity)
                .filter(func.lower(POAActivity.responsable).in_(aliases))
                .order_by(POAActivity.fecha_final.asc(), POAActivity.id.asc())
                .all()
            )
            for activity in own_activities:
                if not activity.fecha_final:
                    continue
                if (activity.entrega_estado or "").strip().lower() == "aprobada":
                    continue
                if activity.fecha_final > lookahead:
                    continue
                delta_days = (activity.fecha_final - today).days
                if delta_days < 0:
                    title = "Actividad vencida"
                    message = f"{activity.nombre} venció el {activity.fecha_final.isoformat()}"
                elif delta_days == 0:
                    title = "Actividad vence hoy"
                    message = f"{activity.nombre} vence hoy"
                else:
                    title = "Actividad por vencer"
                    message = f"{activity.nombre} vence el {activity.fecha_final.isoformat()}"
                items.append(
                    {
                        "id": f"activity-deadline-{activity.id}",
                        "kind": "actividad_fecha",
                        "title": title,
                        "message": message,
                        "created_at": datetime.combine(activity.fecha_final, datetime.min.time()).isoformat(),
                        "href": "/poa/crear",
                    }
                )

        items.sort(key=lambda item: item.get("created_at") or "", reverse=True)
        limited_items = items[:25]
        notification_ids = [str(item.get("id") or "").strip() for item in limited_items if str(item.get("id") or "").strip()]
        read_ids: Set[str] = set()
        if user_key and notification_ids:
            read_rows = (
                db.query(UserNotificationRead.notification_id)
                .filter(
                    UserNotificationRead.tenant_id == tenant_id,
                    UserNotificationRead.user_key == user_key,
                    UserNotificationRead.notification_id.in_(notification_ids),
                )
                .all()
            )
            read_ids = {str(row[0]) for row in read_rows}
        for item in limited_items:
            item["read"] = str(item.get("id") or "") in read_ids

        counts = {
            "poa_aprobacion": 0,
            "documento_autorizacion": 0,
            "actividad_fecha": 0,
            "quiz_descuento": 0,
        }
        for item in limited_items:
            kind = str(item.get("kind") or "")
            if kind in counts:
                counts[kind] += 0 if item.get("read") else 1
        unread = sum(0 if item.get("read") else 1 for item in limited_items)

        return JSONResponse(
            {
                "success": True,
                "total": len(limited_items),
                "unread": unread,
                "counts": counts,
                "items": limited_items,
            }
        )
    finally:
        db.close()


@app.post("/api/notificaciones/marcar-leida")
def mark_notification_read(request: Request, data: dict = Body(default={})):
    notification_id = (data.get("id") or "").strip()
    if not notification_id:
        return JSONResponse({"success": False, "error": "ID de notificación requerido"}, status_code=400)

    db = SessionLocal()
    try:
        tenant_id = _normalize_tenant_id(get_current_tenant(request))
        user_key = _notification_user_key(request, db)
        if not user_key:
            return JSONResponse({"success": False, "error": "Usuario no autenticado"}, status_code=401)

        row = (
            db.query(UserNotificationRead)
            .filter(
                UserNotificationRead.tenant_id == tenant_id,
                UserNotificationRead.user_key == user_key,
                UserNotificationRead.notification_id == notification_id,
            )
            .first()
        )
        if row:
            row.read_at = datetime.utcnow()
            db.add(row)
        else:
            db.add(
                UserNotificationRead(
                    tenant_id=tenant_id,
                    user_key=user_key,
                    notification_id=notification_id,
                    read_at=datetime.utcnow(),
                )
            )
        db.commit()
        return JSONResponse({"success": True})
    finally:
        db.close()


@app.post("/api/notificaciones/marcar-todas-leidas")
def mark_all_notifications_read(request: Request, data: dict = Body(default={})):
    raw_ids = data.get("ids")
    ids = [str(value).strip() for value in (raw_ids if isinstance(raw_ids, list) else [])]
    ids = [value for value in ids if value][:200]

    db = SessionLocal()
    try:
        tenant_id = _normalize_tenant_id(get_current_tenant(request))
        user_key = _notification_user_key(request, db)
        if not user_key:
            return JSONResponse({"success": False, "error": "Usuario no autenticado"}, status_code=401)
        if not ids:
            return JSONResponse({"success": True, "updated": 0})

        existing = (
            db.query(UserNotificationRead)
            .filter(
                UserNotificationRead.tenant_id == tenant_id,
                UserNotificationRead.user_key == user_key,
                UserNotificationRead.notification_id.in_(ids),
            )
            .all()
        )
        existing_by_id = {row.notification_id: row for row in existing}
        now = datetime.utcnow()
        updates = 0
        for notif_id in ids:
            row = existing_by_id.get(notif_id)
            if row:
                row.read_at = now
                db.add(row)
            else:
                db.add(
                    UserNotificationRead(
                        tenant_id=tenant_id,
                        user_key=user_key,
                        notification_id=notif_id,
                        read_at=now,
                    )
                )
            updates += 1
        db.commit()
        return JSONResponse({"success": True, "updated": updates})
    finally:
        db.close()


@app.post("/api/poa/activities")
def create_poa_activity(request: Request, data: dict = Body(...)):
    objective_id = int(data.get("objective_id") or 0)
    nombre = (data.get("nombre") or "").strip()
    responsable = (data.get("responsable") or "").strip()
    entregable = (data.get("entregable") or "").strip()
    if not objective_id or not nombre or not responsable or not entregable:
        return JSONResponse(
            {"success": False, "error": "Objetivo, nombre, responsable y entregable son obligatorios"},
            status_code=400,
        )
    start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=True)
    if start_error:
        return JSONResponse({"success": False, "error": start_error}, status_code=400)
    end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=True)
    if end_error:
        return JSONResponse({"success": False, "error": end_error}, status_code=400)
    range_error = _validate_date_range(start_date, end_date, "Actividad")
    if range_error:
        return JSONResponse({"success": False, "error": range_error}, status_code=400)

    db = SessionLocal()
    try:
        allowed_ids = {obj.id for obj in _allowed_objectives_for_user(request, db)}
        if objective_id not in allowed_ids and not is_admin_or_superadmin(request):
            return JSONResponse({"success": False, "error": "No autorizado para este objetivo"}, status_code=403)
        objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == objective_id).first()
        if not objective:
            return JSONResponse({"success": False, "error": "Objetivo no encontrado"}, status_code=404)
        parent_error = _validate_child_date_range(
            start_date,
            end_date,
            objective.fecha_inicial,
            objective.fecha_final,
            "Actividad",
            "Objetivo",
        )
        if parent_error:
            return JSONResponse({"success": False, "error": parent_error}, status_code=400)
        created_by = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        activity = POAActivity(
            objective_id=objective_id,
            nombre=nombre,
            codigo=(data.get("codigo") or "").strip(),
            responsable=responsable,
            entregable=entregable,
            fecha_inicial=start_date,
            fecha_final=end_date,
            descripcion=(data.get("descripcion") or "").strip(),
            created_by=created_by,
        )
        db.add(activity)
        db.commit()
        db.refresh(activity)
        return JSONResponse({"success": True, "data": _serialize_poa_activity(activity, [])})
    finally:
        db.close()


@app.put("/api/poa/activities/{activity_id}")
def update_poa_activity(request: Request, activity_id: int, data: dict = Body(...)):
    db = SessionLocal()
    try:
        activity = db.query(POAActivity).filter(POAActivity.id == activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        allowed_ids = {obj.id for obj in _allowed_objectives_for_user(request, db)}
        if activity.objective_id not in allowed_ids and not is_admin_or_superadmin(request):
            return JSONResponse({"success": False, "error": "No autorizado para editar esta actividad"}, status_code=403)
        nombre = (data.get("nombre") or "").strip()
        responsable = (data.get("responsable") or "").strip()
        entregable = (data.get("entregable") or "").strip()
        if not nombre or not responsable or not entregable:
            return JSONResponse(
                {"success": False, "error": "Nombre, responsable y entregable son obligatorios"},
                status_code=400,
            )
        start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=True)
        if start_error:
            return JSONResponse({"success": False, "error": start_error}, status_code=400)
        end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=True)
        if end_error:
            return JSONResponse({"success": False, "error": end_error}, status_code=400)
        range_error = _validate_date_range(start_date, end_date, "Actividad")
        if range_error:
            return JSONResponse({"success": False, "error": range_error}, status_code=400)
        objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == activity.objective_id).first()
        if not objective:
            return JSONResponse({"success": False, "error": "Objetivo no encontrado"}, status_code=404)
        parent_error = _validate_child_date_range(
            start_date,
            end_date,
            objective.fecha_inicial,
            objective.fecha_final,
            "Actividad",
            "Objetivo",
        )
        if parent_error:
            return JSONResponse({"success": False, "error": parent_error}, status_code=400)
        activity.nombre = nombre
        activity.codigo = (data.get("codigo") or "").strip()
        activity.responsable = responsable
        activity.entregable = entregable
        activity.fecha_inicial = start_date
        activity.fecha_final = end_date
        activity.descripcion = (data.get("descripcion") or "").strip()
        db.add(activity)
        db.commit()
        db.refresh(activity)
        subs = db.query(POASubactivity).filter(POASubactivity.activity_id == activity.id).all()
        return JSONResponse({"success": True, "data": _serialize_poa_activity(activity, subs)})
    finally:
        db.close()


@app.delete("/api/poa/activities/{activity_id}")
def delete_poa_activity(request: Request, activity_id: int):
    db = SessionLocal()
    try:
        activity = db.query(POAActivity).filter(POAActivity.id == activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        allowed_ids = {obj.id for obj in _allowed_objectives_for_user(request, db)}
        if activity.objective_id not in allowed_ids and not is_admin_or_superadmin(request):
            return JSONResponse({"success": False, "error": "No autorizado para eliminar esta actividad"}, status_code=403)
        db.query(POASubactivity).filter(POASubactivity.activity_id == activity.id).delete()
        db.delete(activity)
        db.commit()
        return JSONResponse({"success": True})
    finally:
        db.close()


@app.post("/api/poa/activities/{activity_id}/request-completion")
def request_poa_activity_completion(request: Request, activity_id: int):
    db = SessionLocal()
    try:
        activity = db.query(POAActivity).filter(POAActivity.id == activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        allowed_ids = {obj.id for obj in _allowed_objectives_for_user(request, db)}
        if activity.objective_id not in allowed_ids and not is_admin_or_superadmin(request):
            return JSONResponse({"success": False, "error": "No autorizado para esta actividad"}, status_code=403)
        session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        user = _current_user_record(request, db)
        aliases = _user_aliases(user, session_username)
        is_activity_owner = (activity.responsable or "").strip().lower() in aliases
        if not (is_activity_owner or is_admin_or_superadmin(request)):
            return JSONResponse({"success": False, "error": "Solo el responsable puede solicitar terminación"}, status_code=403)
        if _activity_status(activity) == "No iniciada":
            return JSONResponse(
                {"success": False, "error": "La actividad no ha iniciado; no se puede solicitar terminación"},
                status_code=409,
            )
        if (activity.entrega_estado or "").strip().lower() == "aprobada":
            return JSONResponse({"success": False, "error": "La actividad ya fue aprobada y terminada"}, status_code=409)

        pending = (
            db.query(POADeliverableApproval)
            .filter(
                POADeliverableApproval.activity_id == activity.id,
                POADeliverableApproval.status == "pendiente",
            )
            .first()
        )
        if pending:
            return JSONResponse({"success": False, "error": "Ya existe una aprobación pendiente para esta actividad"}, status_code=409)

        objective = db.query(StrategicObjectiveConfig).filter(StrategicObjectiveConfig.id == activity.objective_id).first()
        if not objective:
            return JSONResponse({"success": False, "error": "Objetivo no encontrado"}, status_code=404)
        axis = db.query(StrategicAxisConfig).filter(StrategicAxisConfig.id == objective.eje_id).first()
        process_owner = _resolve_process_owner_for_objective(objective, axis)
        if not process_owner:
            return JSONResponse(
                {"success": False, "error": "No se pudo identificar dueño del proceso (líder objetivo/departamento eje)"},
                status_code=400,
            )

        approval = POADeliverableApproval(
            activity_id=activity.id,
            objective_id=objective.id,
            process_owner=process_owner,
            requester=session_username or (activity.responsable or ""),
            status="pendiente",
        )
        activity.entrega_estado = "pendiente"
        activity.entrega_solicitada_por = session_username or (activity.responsable or "")
        activity.entrega_solicitada_at = datetime.utcnow()
        activity.entrega_aprobada_por = ""
        activity.entrega_aprobada_at = None
        db.add(approval)
        db.add(activity)
        db.commit()
        return JSONResponse({"success": True, "message": "Solicitud de aprobación enviada al dueño del proceso"})
    finally:
        db.close()


@app.post("/api/poa/approvals/{approval_id}/decision")
def decide_poa_deliverable_approval(request: Request, approval_id: int, data: dict = Body(default={})):
    action = (data.get("accion") or "").strip().lower()
    if action not in {"autorizar", "rechazar"}:
        return JSONResponse({"success": False, "error": "Acción inválida. Usa autorizar o rechazar"}, status_code=400)
    comment = (data.get("comentario") or "").strip()

    db = SessionLocal()
    try:
        approval = db.query(POADeliverableApproval).filter(POADeliverableApproval.id == approval_id).first()
        if not approval:
            return JSONResponse({"success": False, "error": "Solicitud de aprobación no encontrada"}, status_code=404)
        if approval.status != "pendiente":
            return JSONResponse({"success": False, "error": "Esta solicitud ya fue resuelta"}, status_code=409)
        if not _is_user_process_owner(request, db, approval.process_owner):
            return JSONResponse({"success": False, "error": "No autorizado para resolver esta aprobación"}, status_code=403)

        activity = db.query(POAActivity).filter(POAActivity.id == approval.activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        resolver_user = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()

        approval.status = "autorizada" if action == "autorizar" else "rechazada"
        approval.comment = comment
        approval.resolved_by = resolver_user
        approval.resolved_at = datetime.utcnow()
        db.add(approval)

        if action == "autorizar":
            activity.entrega_estado = "aprobada"
            activity.entrega_aprobada_por = resolver_user
            activity.entrega_aprobada_at = datetime.utcnow()
        else:
            activity.entrega_estado = "rechazada"
            activity.entrega_aprobada_por = ""
            activity.entrega_aprobada_at = None
        db.add(activity)
        db.commit()
        return JSONResponse({"success": True, "message": "Aprobación procesada correctamente"})
    finally:
        db.close()


@app.post("/api/poa/activities/{activity_id}/subactivities")
def create_poa_subactivity(request: Request, activity_id: int, data: dict = Body(...)):
    nombre = (data.get("nombre") or "").strip()
    responsable = (data.get("responsable") or "").strip()
    if not nombre or not responsable:
        return JSONResponse({"success": False, "error": "Nombre y responsable son obligatorios"}, status_code=400)
    start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=True)
    if start_error:
        return JSONResponse({"success": False, "error": start_error}, status_code=400)
    end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=True)
    if end_error:
        return JSONResponse({"success": False, "error": end_error}, status_code=400)
    range_error = _validate_date_range(start_date, end_date, "Subactividad")
    if range_error:
        return JSONResponse({"success": False, "error": range_error}, status_code=400)

    db = SessionLocal()
    try:
        activity = db.query(POAActivity).filter(POAActivity.id == activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        user = _current_user_record(request, db)
        aliases = _user_aliases(user, session_username)
        is_activity_owner = (activity.responsable or "").strip().lower() in aliases
        if not (is_activity_owner or is_admin_or_superadmin(request)):
            return JSONResponse(
                {"success": False, "error": "Solo el responsable de la actividad puede asignar subactividades"},
                status_code=403,
            )
        parent_error = _validate_child_date_range(
            start_date,
            end_date,
            activity.fecha_inicial,
            activity.fecha_final,
            "Subactividad",
            "Actividad",
        )
        if parent_error:
            return JSONResponse({"success": False, "error": parent_error}, status_code=400)
        assigned_by = session_username
        sub = POASubactivity(
            activity_id=activity.id,
            nombre=nombre,
            codigo=(data.get("codigo") or "").strip(),
            responsable=responsable,
            entregable=(data.get("entregable") or "").strip(),
            fecha_inicial=start_date,
            fecha_final=end_date,
            descripcion=(data.get("descripcion") or "").strip(),
            assigned_by=assigned_by,
        )
        db.add(sub)
        db.commit()
        db.refresh(sub)
        return JSONResponse({"success": True, "data": _serialize_poa_subactivity(sub)})
    finally:
        db.close()


@app.put("/api/poa/subactivities/{subactivity_id}")
def update_poa_subactivity(request: Request, subactivity_id: int, data: dict = Body(...)):
    db = SessionLocal()
    try:
        sub = db.query(POASubactivity).filter(POASubactivity.id == subactivity_id).first()
        if not sub:
            return JSONResponse({"success": False, "error": "Subactividad no encontrada"}, status_code=404)
        activity = db.query(POAActivity).filter(POAActivity.id == sub.activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        user = _current_user_record(request, db)
        aliases = _user_aliases(user, session_username)
        is_activity_owner = (activity.responsable or "").strip().lower() in aliases
        if not (is_activity_owner or is_admin_or_superadmin(request)):
            return JSONResponse({"success": False, "error": "No autorizado para editar subactividad"}, status_code=403)
        nombre = (data.get("nombre") or "").strip()
        responsable = (data.get("responsable") or "").strip()
        if not nombre or not responsable:
            return JSONResponse({"success": False, "error": "Nombre y responsable son obligatorios"}, status_code=400)
        start_date, start_error = _parse_date_field(data.get("fecha_inicial"), "Fecha inicial", required=True)
        if start_error:
            return JSONResponse({"success": False, "error": start_error}, status_code=400)
        end_date, end_error = _parse_date_field(data.get("fecha_final"), "Fecha final", required=True)
        if end_error:
            return JSONResponse({"success": False, "error": end_error}, status_code=400)
        range_error = _validate_date_range(start_date, end_date, "Subactividad")
        if range_error:
            return JSONResponse({"success": False, "error": range_error}, status_code=400)
        parent_error = _validate_child_date_range(
            start_date,
            end_date,
            activity.fecha_inicial,
            activity.fecha_final,
            "Subactividad",
            "Actividad",
        )
        if parent_error:
            return JSONResponse({"success": False, "error": parent_error}, status_code=400)
        sub.nombre = nombre
        sub.codigo = (data.get("codigo") or "").strip()
        sub.responsable = responsable
        sub.entregable = (data.get("entregable") or "").strip()
        sub.fecha_inicial = start_date
        sub.fecha_final = end_date
        sub.descripcion = (data.get("descripcion") or "").strip()
        db.add(sub)
        db.commit()
        db.refresh(sub)
        return JSONResponse({"success": True, "data": _serialize_poa_subactivity(sub)})
    finally:
        db.close()


@app.delete("/api/poa/subactivities/{subactivity_id}")
def delete_poa_subactivity(request: Request, subactivity_id: int):
    db = SessionLocal()
    try:
        sub = db.query(POASubactivity).filter(POASubactivity.id == subactivity_id).first()
        if not sub:
            return JSONResponse({"success": False, "error": "Subactividad no encontrada"}, status_code=404)
        activity = db.query(POAActivity).filter(POAActivity.id == sub.activity_id).first()
        if not activity:
            return JSONResponse({"success": False, "error": "Actividad no encontrada"}, status_code=404)
        session_username = (getattr(request.state, "user_name", None) or request.cookies.get("user_name") or "").strip()
        user = _current_user_record(request, db)
        aliases = _user_aliases(user, session_username)
        is_activity_owner = (activity.responsable or "").strip().lower() in aliases
        if not (is_activity_owner or is_admin_or_superadmin(request)):
            return JSONResponse({"success": False, "error": "No autorizado para eliminar subactividad"}, status_code=403)
        db.delete(sub)
        db.commit()
        return JSONResponse({"success": True})
    finally:
        db.close()


@app.get("/plantillas", response_class=HTMLResponse)
def plantillas_page(request: Request):
    plantillas_content = """
        <section id="plantillas-page" class="plantillas-page">
            <div class="plantillas-layout">
                <aside class="plantillas-list-card">
                    <h3>Plantillas</h3>
                    <p class="plantillas-hint">Selecciona una plantilla existente o crea una nueva desde la barra flotante. "Encabezado" es la base para todos los reportes.</p>
                    <ul id="plantillas-list" class="plantillas-list"></ul>
                </aside>
                <section class="plantillas-editor-card">
                    <div class="form-field">
                        <label for="template-name">Nombre de plantilla</label>
                        <input type="text" id="template-name" placeholder="Ej: Tarjeta institucional">
                    </div>
                    <div class="plantillas-editor-grid">
                        <div class="form-field">
                            <label for="template-html">HTML</label>
                            <textarea id="template-html" class="plantilla-code" placeholder="<section class='card'>...</section>"></textarea>
                        </div>
                        <div class="form-field">
                            <label for="template-css">CSS</label>
                            <textarea id="template-css" class="plantilla-code" placeholder=".card { padding: 16px; border-radius: 12px; }"></textarea>
                        </div>
                    </div>
                    <div class="plantillas-preview-head">
                        <h4>Vista previa</h4>
                        <div style="display:flex; gap:8px;">
                            <button type="button" id="template-new-btn">Nuevo</button>
                            <button type="button" id="template-edit-btn">Editar</button>
                            <button type="button" id="template-builder-btn">Construir formulario</button>
                            <button type="button" id="template-preview-btn">Previsualizar</button>
                            <button type="button" id="template-save-btn">Guardar</button>
                            <button type="button" id="template-delete-btn">Eliminar</button>
                        </div>
                    </div>
                    <iframe id="template-preview" class="plantilla-preview-frame" title="Vista previa de plantilla"></iframe>
                    <section id="form-builder-panel" class="form-builder-panel hidden" aria-label="Constructor de formularios">
                        <div class="form-builder-head">
                            <h4>Constructor de formularios</h4>
                            <p>Crea formularios dinámicos para usuarios finales.</p>
                        </div>
                        <div class="form-builder-grid">
                            <div class="form-field">
                                <label for="builder-form-select">Formularios existentes</label>
                                <select id="builder-form-select" class="campo-personalizado">
                                    <option value="">Nuevo formulario</option>
                                </select>
                            </div>
                            <div class="form-field">
                                <label for="builder-form-name">Nombre</label>
                                <input type="text" id="builder-form-name" class="campo-personalizado" placeholder="Ej: Solicitud de crédito">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-slug">Slug (opcional)</label>
                                <input type="text" id="builder-form-slug" class="campo-personalizado" placeholder="solicitud-credito">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-active">Estado</label>
                                <select id="builder-form-active" class="campo-personalizado">
                                    <option value="true">Activo</option>
                                    <option value="false">Inactivo</option>
                                </select>
                            </div>
                            <div class="form-field">
                                <label for="builder-form-tenant">Tenant</label>
                                <input type="text" id="builder-form-tenant" class="campo-personalizado" placeholder="default">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-roles">Roles permitidos</label>
                                <select id="builder-form-roles" class="campo-personalizado" multiple>
                                </select>
                            </div>
                            <div class="form-field form-builder-description">
                                <label for="builder-form-description">Descripción</label>
                                <textarea id="builder-form-description" class="campo-personalizado" placeholder="Descripción del formulario"></textarea>
                            </div>
                            <div class="form-field form-builder-description">
                                <label for="builder-form-config">Configuración (JSON)</label>
                                <textarea id="builder-form-config" class="campo-personalizado" placeholder='{"wizard":{"steps":[{"title":"Paso 1","fields":["nombre","email"]},{"title":"Paso 2","fields":["tipo","detalle"]}]},"notifications":{"email":{"enabled":true,"to":["equipo@empresa.com"],"cc":[],"subject":"Nuevo envio"},"webhooks":[{"url":"https://api.empresa.com/hook/forms","method":"POST"}]}}'></textarea>
                            </div>
                        </div>
                        <div class="form-builder-field-editor">
                            <h5>Agregar campo</h5>
                            <div class="form-builder-grid field-grid">
                                <div class="form-field">
                                    <label for="builder-field-type">Tipo</label>
                                    <select id="builder-field-type" class="campo-personalizado">
                                        <option value="text">Texto corto</option>
                                        <option value="textarea">Texto largo</option>
                                        <option value="email">Email</option>
                                        <option value="password">Contraseña</option>
                                        <option value="number">Número (decimal)</option>
                                        <option value="integer">Número (entero)</option>
                                        <option value="select">Desplegable (Dropdown)</option>
                                        <option value="checkboxes">Opción múltiple (Checkboxes)</option>
                                        <option value="radio">Opción única (Radio)</option>
                                        <option value="likert">Escala Likert</option>
                                        <option value="checkbox">Checkbox</option>
                                        <option value="date">Selector de fecha</option>
                                        <option value="time">Selector de hora</option>
                                        <option value="daterange">Rango de fechas</option>
                                        <option value="file">Carga de archivos</option>
                                        <option value="signature">Firma digital</option>
                                        <option value="url">Enlace (URL)</option>
                                        <option value="header">Encabezado</option>
                                        <option value="paragraph">Texto estático (Paragraph)</option>
                                        <option value="html">Texto estático (HTML)</option>
                                        <option value="divider">Separador</option>
                                        <option value="pagebreak">Salto de página</option>
                                    </select>
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-label">Etiqueta</label>
                                    <input type="text" id="builder-field-label" class="campo-personalizado" placeholder="Nombre completo">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-name">Nombre técnico</label>
                                    <input type="text" id="builder-field-name" class="campo-personalizado" placeholder="nombre_completo">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-placeholder">Placeholder</label>
                                    <input type="text" id="builder-field-placeholder" class="campo-personalizado" placeholder="Escribe aquí">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-help">Ayuda</label>
                                    <input type="text" id="builder-field-help" class="campo-personalizado" placeholder="Texto de ayuda">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-options">Opciones (select/radio)</label>
                                    <input type="text" id="builder-field-options" class="campo-personalizado" placeholder="Opción A, Opción B, Opción C">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-conditional">Condicional (JSON)</label>
                                    <input type="text" id="builder-field-conditional" class="campo-personalizado" placeholder='{"field":"tipo","operator":"equals","value":"staff"}'>
                                </div>
                                <div class="form-field form-field-inline">
                                    <label for="builder-field-required">Obligatorio</label>
                                    <input type="checkbox" id="builder-field-required">
                                </div>
                            </div>
                            <div class="form-builder-actions">
                                <button type="button" id="builder-add-field-btn">Agregar campo</button>
                                <button type="button" id="builder-clear-form-btn">Limpiar</button>
                                <button type="button" id="builder-save-form-btn">Guardar formulario</button>
                                <button type="button" id="builder-delete-form-btn">Eliminar formulario</button>
                            </div>
                        </div>
                        <div class="form-builder-list-wrap">
                            <h5>Campos del formulario</h5>
                            <div id="builder-fields-list" class="form-builder-fields-list"></div>
                        </div>
                    </section>
                </section>
            </div>
        </section>
    """
    return render_backend_page(
        request,
        title="Plantillas",
        description="Crea y guarda plantillas con HTML y CSS.",
        content=plantillas_content,
        hide_floating_actions=False,
        floating_actions_screen="plantillas",
    )


@app.get("/plantillas/constructor", response_class=HTMLResponse)
def plantillas_constructor_page(request: Request):
    constructor_content = """
        <section id="plantillas-page" class="plantillas-page">
            <div class="plantillas-layout">
                <aside class="plantillas-list-card">
                    <h3>Constructor de formularios</h3>
                    <p class="plantillas-hint">Pantalla dedicada para crear, editar y publicar formularios dinámicos.</p>
                    <div class="form-builder-actions">
                        <button type="button" id="builder-back-to-templates">Volver a plantillas</button>
                    </div>
                </aside>
                <section class="plantillas-editor-card">
                    <section id="form-builder-panel" class="form-builder-panel" aria-label="Constructor de formularios">
                        <div class="form-builder-head">
                            <h4>Constructor de formularios</h4>
                            <p>Crea formularios dinámicos para usuarios finales.</p>
                        </div>
                        <div class="form-builder-grid">
                            <div class="form-field">
                                <label for="builder-form-select">Formulario</label>
                                <select id="builder-form-select">
                                    <option value="">Nuevo formulario</option>
                                </select>
                            </div>
                            <div class="form-field">
                                <label for="builder-form-name">Nombre</label>
                                <input type="text" id="builder-form-name" placeholder="Ej: Solicitud de apoyo">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-slug">Slug</label>
                                <input type="text" id="builder-form-slug" placeholder="solicitud-apoyo">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-active">Estado</label>
                                <select id="builder-form-active">
                                    <option value="true">Activo</option>
                                    <option value="false">Inactivo</option>
                                </select>
                            </div>
                            <div class="form-field">
                                <label for="builder-form-tenant">Tenant</label>
                                <input type="text" id="builder-form-tenant" placeholder="default">
                            </div>
                            <div class="form-field">
                                <label for="builder-form-roles">Roles permitidos</label>
                                <select id="builder-form-roles" multiple>
                                </select>
                            </div>
                            <div class="form-field form-builder-description">
                                <label for="builder-form-description">Descripción</label>
                                <textarea id="builder-form-description" placeholder="Describe el objetivo del formulario"></textarea>
                            </div>
                            <div class="form-field form-builder-description">
                                <label for="builder-form-config">Configuración JSON (opcional)</label>
                                <textarea id="builder-form-config" placeholder='{"submitLabel":"Enviar"}'></textarea>
                            </div>
                        </div>
                        <div class="form-builder-field-editor">
                            <h5>Campos del formulario</h5>
                            <div class="form-builder-grid field-grid">
                                <div class="form-field">
                                    <label for="builder-field-type">Tipo</label>
                                    <select id="builder-field-type">
                                        <option value="text">Texto</option>
                                        <option value="email">Email</option>
                                        <option value="number">Número</option>
                                        <option value="date">Fecha</option>
                                        <option value="select">Selección</option>
                                        <option value="radio">Opciones (radio)</option>
                                        <option value="checkboxes">Opciones (checkbox)</option>
                                        <option value="textarea">Texto largo</option>
                                        <option value="file">Archivo</option>
                                        <option value="password">Contraseña</option>
                                        <option value="likert">Likert</option>
                                    </select>
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-label">Etiqueta</label>
                                    <input type="text" id="builder-field-label" placeholder="Nombre completo">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-name">Nombre técnico</label>
                                    <input type="text" id="builder-field-name" placeholder="nombre_completo">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-placeholder">Placeholder</label>
                                    <input type="text" id="builder-field-placeholder" placeholder="Escribe aquí...">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-help">Ayuda</label>
                                    <input type="text" id="builder-field-help" placeholder="Texto de apoyo">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-options">Opciones (coma separadas)</label>
                                    <input type="text" id="builder-field-options" placeholder="A, B, C">
                                </div>
                                <div class="form-field">
                                    <label for="builder-field-conditional">Condición JSON</label>
                                    <input type="text" id="builder-field-conditional" placeholder='{"depends_on":"campo","equals":"valor"}'>
                                </div>
                                <div class="form-field form-field-inline">
                                    <input type="checkbox" id="builder-field-required">
                                    <label for="builder-field-required">Obligatorio</label>
                                </div>
                            </div>
                            <div class="form-builder-actions">
                                <button type="button" id="builder-add-field-btn">Agregar campo</button>
                                <button type="button" id="builder-clear-form-btn">Limpiar</button>
                                <button type="button" id="builder-save-form-btn">Guardar formulario</button>
                                <button type="button" id="builder-delete-form-btn">Eliminar formulario</button>
                            </div>
                        </div>
                        <div class="form-builder-list-wrap">
                            <h5>Campos agregados</h5>
                            <div id="builder-fields-list" class="form-builder-fields-list"></div>
                        </div>
                    </section>
                </section>
            </div>
        </section>
    """
    return render_backend_page(
        request,
        title="Constructor de formularios",
        description="Crea formularios dinámicos en una pantalla dedicada.",
        content=constructor_content,
        hide_floating_actions=False,
        floating_actions_screen="plantillas",
    )


@app.get("/api/plantillas")
def listar_plantillas():
    return JSONResponse({"success": True, "data": _load_plantillas_store()})


@app.post("/api/plantillas")
def guardar_plantilla(data: dict = Body(...)):
    nombre = (data.get("nombre") or "").strip()
    html_code = data.get("html") or ""
    css_code = data.get("css") or ""
    template_id = (data.get("id") or "").strip()
    if not nombre:
        return JSONResponse({"success": False, "error": "El nombre es obligatorio"}, status_code=400)
    if not html_code.strip() and not css_code.strip():
        return JSONResponse({"success": False, "error": "Debes agregar HTML o CSS"}, status_code=400)

    templates = _load_plantillas_store()
    now_iso = datetime.utcnow().isoformat()
    if template_id:
        updated = False
        for template in templates:
            if template.get("id") == template_id:
                template["nombre"] = nombre
                template["html"] = html_code
                template["css"] = css_code
                template["updated_at"] = now_iso
                updated = True
                break
        if not updated:
            return JSONResponse({"success": False, "error": "Plantilla no encontrada"}, status_code=404)
    else:
        template_id = secrets.token_hex(8)
        templates.insert(
            0,
            {
                "id": template_id,
                "nombre": nombre,
                "html": html_code,
                "css": css_code,
                "created_at": now_iso,
                "updated_at": now_iso,
            },
        )
    _save_plantillas_store(templates)
    return JSONResponse({"success": True, "data": {"id": template_id}})


def _forms_scope_query_by_tenant(query, request: Request):
    tenant_id = _normalize_tenant_id(get_current_tenant(request))
    if is_superadmin(request):
        header_tenant = request.headers.get("x-tenant-id")
        if header_tenant and _normalize_tenant_id(header_tenant) != "all":
            return query.filter(func.lower(FormDefinition.tenant_id) == _normalize_tenant_id(header_tenant).lower())
        if not header_tenant:
            return query.filter(func.lower(FormDefinition.tenant_id) == tenant_id.lower())
        return query
    return query.filter(func.lower(FormDefinition.tenant_id) == tenant_id.lower())


def _resolve_form_tenant_for_write(request: Request, requested_tenant: Optional[str]) -> str:
    if is_superadmin(request):
        if requested_tenant:
            return _normalize_tenant_id(requested_tenant)
        header_tenant = request.headers.get("x-tenant-id")
        if header_tenant and _normalize_tenant_id(header_tenant) != "all":
            return _normalize_tenant_id(header_tenant)
    return _normalize_tenant_id(get_current_tenant(request))


def _normalize_form_allowed_roles(request: Request, db, raw_roles: Any) -> List[str]:
    allowed = {
        normalize_role_name(role.nombre)
        for role in db.query(Rol).all()
        if (role.nombre or "").strip()
    }
    if is_admin(request):
        allowed = {item for item in allowed if item != "superadministrador"}
    normalized_roles: List[str] = []
    if isinstance(raw_roles, list):
        for role in raw_roles:
            role_name = normalize_role_name(str(role))
            if role_name in allowed:
                normalized_roles.append(role_name)
    return sorted(set(normalized_roles))


def _form_role_is_allowed(form: FormDefinition, request: Request) -> bool:
    if is_superadmin(request):
        return True
    allowed_roles_raw = form.allowed_roles if isinstance(form.allowed_roles, list) else []
    if not allowed_roles_raw:
        return True
    allowed_roles = {normalize_role_name(str(item)) for item in allowed_roles_raw if str(item).strip()}
    return get_current_role(request) in allowed_roles


def _get_form_by_id_for_request(db, form_id: int, request: Request) -> FormDefinition:
    query = _forms_scope_query_by_tenant(
        db.query(FormDefinition).filter(FormDefinition.id == form_id),
        request,
    )
    form = query.first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")
    return form


def _get_form_by_slug_for_request(db, slug: str, request: Request, active_only: bool = True) -> FormDefinition:
    query = db.query(FormDefinition).filter(FormDefinition.slug == slug)
    if active_only:
        query = query.filter(FormDefinition.is_active == True)  # noqa: E712
    query = _forms_scope_query_by_tenant(query, request)
    form = query.first()
    if not form:
        raise HTTPException(status_code=404, detail="Formulario no encontrado")
    if not _form_role_is_allowed(form, request):
        raise HTTPException(status_code=403, detail="Tu rol no tiene acceso a este formulario")
    return form


@app.post("/api/admin/forms", response_model=FormDefinitionResponseSchema)
def create_form_definition(
    request: Request,
    form_data: FormDefinitionCreateSchema,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    slug = slugify_value(form_data.slug or form_data.name)
    existing = db.query(FormDefinition).filter(FormDefinition.slug == slug).first()
    if existing:
        raise HTTPException(status_code=400, detail="El slug ya existe")

    tenant_id = _resolve_form_tenant_for_write(request, form_data.tenant_id)
    allowed_roles = _normalize_form_allowed_roles(request, db, form_data.allowed_roles)
    form = FormDefinition(
        name=form_data.name,
        slug=slug,
        tenant_id=tenant_id,
        description=form_data.description,
        config=form_data.config or {},
        allowed_roles=allowed_roles,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow(),
        is_active=form_data.is_active,
    )
    db.add(form)
    db.flush()

    for field_data in form_data.fields:
        db.add(
            FormField(
                form_id=form.id,
                field_type=field_data.field_type,
                label=field_data.label,
                name=field_data.name,
                placeholder=field_data.placeholder,
                help_text=field_data.help_text,
                default_value=field_data.default_value,
                is_required=field_data.is_required,
                validation_rules=field_data.validation_rules,
                options=field_data.options,
                order=field_data.order,
                conditional_logic=field_data.conditional_logic,
            )
        )

    db.commit()
    db.refresh(form)
    return form


@app.get("/api/admin/forms", response_model=List[FormDefinitionResponseSchema])
def list_form_definitions(
    request: Request,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    query = _forms_scope_query_by_tenant(db.query(FormDefinition), request)
    return query.order_by(FormDefinition.id.desc()).all()


@app.get("/api/admin/forms/{form_id}", response_model=FormDefinitionResponseSchema)
def get_form_definition(
    form_id: int,
    request: Request,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    return _get_form_by_id_for_request(db, form_id, request)


@app.put("/api/admin/forms/{form_id}")
def update_form_definition(
    form_id: int,
    request: Request,
    form_data: FormDefinitionCreateSchema,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    form = _get_form_by_id_for_request(db, form_id, request)

    new_slug = slugify_value(form_data.slug or form_data.name)
    duplicate = (
        db.query(FormDefinition)
        .filter(FormDefinition.slug == new_slug, FormDefinition.id != form_id)
        .first()
    )
    if duplicate:
        raise HTTPException(status_code=400, detail="El slug ya existe")

    form.name = form_data.name
    form.slug = new_slug
    form.tenant_id = _resolve_form_tenant_for_write(request, form_data.tenant_id or form.tenant_id)
    form.description = form_data.description
    form.config = form_data.config or {}
    form.allowed_roles = _normalize_form_allowed_roles(request, db, form_data.allowed_roles)
    form.is_active = form_data.is_active
    form.updated_at = datetime.utcnow()
    db.add(form)

    db.query(FormField).filter(FormField.form_id == form_id).delete()
    for field_data in form_data.fields:
        db.add(
            FormField(
                form_id=form_id,
                field_type=field_data.field_type,
                label=field_data.label,
                name=field_data.name,
                placeholder=field_data.placeholder,
                help_text=field_data.help_text,
                default_value=field_data.default_value,
                is_required=field_data.is_required,
                validation_rules=field_data.validation_rules,
                options=field_data.options,
                order=field_data.order,
                conditional_logic=field_data.conditional_logic,
            )
        )

    db.commit()
    return {"success": True, "message": "Formulario actualizado"}


@app.delete("/api/admin/forms/{form_id}")
def delete_form_definition(
    form_id: int,
    request: Request,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    form = _get_form_by_id_for_request(db, form_id, request)
    db.delete(form)
    db.commit()
    return {"success": True, "message": "Formulario eliminado"}


@app.get("/api/admin/forms/{form_id}/submissions/export/{formato}")
def export_form_submissions(
    form_id: int,
    formato: str,
    request: Request,
    db=Depends(get_db),
):
    require_admin_or_superadmin(request)
    form = _get_form_by_id_for_request(db, form_id, request)

    submissions = (
        db.query(FormSubmission)
        .filter(FormSubmission.form_id == form_id)
        .order_by(FormSubmission.submitted_at.asc(), FormSubmission.id.asc())
        .all()
    )
    columns = _build_submission_export_columns(form, submissions)
    base_headers = ["submission_id", "submitted_at", "ip_address", "user_agent"]
    headers = [*base_headers, *columns]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_slug = slugify_value(form.slug or form.name)

    if formato.lower() == "csv":
        buffer = StringIO()
        writer = csv.DictWriter(buffer, fieldnames=headers)
        writer.writeheader()
        for submission in submissions:
            data = submission.data if isinstance(submission.data, dict) else {}
            row = {
                "submission_id": submission.id,
                "submitted_at": submission.submitted_at.isoformat() if submission.submitted_at else "",
                "ip_address": submission.ip_address or "",
                "user_agent": submission.user_agent or "",
            }
            for field_name in columns:
                row[field_name] = _normalize_submission_value(data.get(field_name))
            writer.writerow(row)
        filename = f"{safe_slug}_submissions_{timestamp}.csv"
        return Response(
            content=buffer.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    if formato.lower() in {"excel", "xlsx"}:
        wb = Workbook()
        ws = wb.active
        ws.title = "Submissions"
        ws.append(headers)
        for submission in submissions:
            data = submission.data if isinstance(submission.data, dict) else {}
            row_values = [
                submission.id,
                submission.submitted_at.isoformat() if submission.submitted_at else "",
                submission.ip_address or "",
                submission.user_agent or "",
            ]
            row_values.extend(_normalize_submission_value(data.get(field_name)) for field_name in columns)
            ws.append(row_values)
        output = BytesIO()
        wb.save(output)
        filename = f"{safe_slug}_submissions_{timestamp}.xlsx"
        return Response(
            content=output.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    raise HTTPException(status_code=400, detail="Formato no soportado. Usa csv o excel")


@app.get("/api/forms/{slug}")
def get_public_form_definition(
    slug: str,
    request: Request,
    db=Depends(get_db),
):
    form = _get_form_by_slug_for_request(db, slug, request, active_only=True)
    return {"success": True, "data": FormRenderer.render_to_json(form)}


@app.post("/api/forms/{slug}/submit")
def submit_public_form(
    slug: str,
    request: Request,
    payload: Dict[str, Any] = Body(default_factory=dict),
    db=Depends(get_db),
):
    form = _get_form_by_slug_for_request(db, slug, request, active_only=True)

    normalized_payload = _normalize_form_submission_payload(form, payload)
    visible_fields = FormRenderer.visible_field_names(form, normalized_payload)
    dynamic_model = FormRenderer.generate_pydantic_model(form, visible_field_names=visible_fields)
    try:
        validated = dynamic_model.model_validate(normalized_payload)
        validated_data = validated.model_dump()
        FormRenderer._apply_custom_rules(form, validated_data, visible_field_names=visible_fields)
    except ValidationError as exc:
        return JSONResponse(
            {"success": False, "error": "Datos inválidos", "details": exc.errors()},
            status_code=422,
        )
    except ValueError as exc:
        return JSONResponse(
            {"success": False, "error": str(exc)},
            status_code=422,
        )
    validated_data = {key: value for key, value in validated_data.items() if key in visible_fields}

    submission = FormSubmission(
        form_id=form.id,
        data=validated_data,
        submitted_at=datetime.utcnow(),
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)
    email_status = send_form_submission_email_notification(form, submission)
    webhook_status = send_form_submission_webhooks(form, submission)

    return {
        "success": True,
        "message": "Formulario enviado correctamente",
        "submission_id": submission.id,
        "notification": {
            "email": email_status,
            "webhooks": webhook_status,
        },
    }


@app.get("/forms/{slug}", response_class=HTMLResponse)
def render_public_form(
    slug: str,
    request: Request,
    db=Depends(get_db),
):
    form = _get_form_by_slug_for_request(db, slug, request, active_only=True)
    login_identity = _get_login_identity_context()
    return request.app.state.templates.TemplateResponse(
        "form.html",
        {
            "request": request,
            "title": form.name,
            "form": form,
            "form_json": json.dumps(FormRenderer.render_to_json(form), ensure_ascii=False),
            "app_favicon_url": login_identity.get("login_favicon_url"),
        },
    )


@app.get("/forms/api/{slug}")
def get_public_form_definition_v2(
    slug: str,
    request: Request,
    db=Depends(get_db),
):
    form = _get_form_by_slug_for_request(db, slug, request, active_only=True)
    return {"success": True, "data": FormRenderer.render_to_json(form)}


@app.post("/forms/api/{slug}/submit")
async def submit_public_form_v2(
    slug: str,
    request: Request,
    db=Depends(get_db),
):
    form = _get_form_by_slug_for_request(db, slug, request, active_only=True)

    form_data = await request.form()
    payload: Dict[str, Any] = {}
    for key in form_data.keys():
        values = form_data.getlist(key)
        payload[key] = values if len(values) > 1 else values[0]
    normalized_payload = _normalize_form_submission_payload(form, payload)
    visible_fields = FormRenderer.visible_field_names(form, normalized_payload)
    dynamic_model = FormRenderer.generate_pydantic_model(form, visible_field_names=visible_fields)
    try:
        validated = dynamic_model.model_validate(normalized_payload)
        validated_data = validated.model_dump()
        FormRenderer._apply_custom_rules(form, validated_data, visible_field_names=visible_fields)
    except ValidationError as exc:
        return JSONResponse(
            {"success": False, "error": "Datos inválidos", "details": exc.errors()},
            status_code=422,
        )
    except ValueError as exc:
        return JSONResponse(
            {"success": False, "error": str(exc)},
            status_code=422,
        )
    validated_data = {key: value for key, value in validated_data.items() if key in visible_fields}

    submission = FormSubmission(
        form_id=form.id,
        data=validated_data,
        submitted_at=datetime.utcnow(),
        ip_address=request.client.host if request.client else None,
        user_agent=request.headers.get("user-agent"),
    )
    db.add(submission)
    db.commit()
    db.refresh(submission)
    email_status = send_form_submission_email_notification(form, submission)
    webhook_status = send_form_submission_webhooks(form, submission)
    return {
        "success": True,
        "message": "Formulario enviado correctamente",
        "submission_id": submission.id,
        "notification": {
            "email": email_status,
            "webhooks": webhook_status,
        },
    }


@app.delete("/api/plantillas/{template_id}")
def eliminar_plantilla(template_id: str):
    template_id = (template_id or "").strip()
    if not template_id:
        return JSONResponse({"success": False, "error": "ID de plantilla invalido"}, status_code=400)
    if template_id == SYSTEM_REPORT_HEADER_TEMPLATE_ID:
        return JSONResponse(
            {"success": False, "error": "La plantilla Encabezado es obligatoria para reportes"},
            status_code=400,
        )

    templates = _load_plantillas_store()
    remaining = [tpl for tpl in templates if str(tpl.get("id", "")).strip() != template_id]
    if len(remaining) == len(templates):
        return JSONResponse({"success": False, "error": "Plantilla no encontrada"}, status_code=404)
    _save_plantillas_store(remaining)
    return JSONResponse({"success": True})


@app.post("/api/usuarios/registro-seguro")
def crear_usuario_seguro(request: Request, data: dict = Body(...)):
    require_admin_or_superadmin(request)
    nombre = (data.get("nombre") or "").strip()
    usuario_login = (data.get("usuario") or "").strip()
    correo = (data.get("correo") or "").strip()
    password = data.get("contrasena") or ""
    rol_nombre = normalize_role_name(data.get("rol"))

    if not nombre or not usuario_login or not correo or not password:
        return JSONResponse(
            {"success": False, "error": "nombre, usuario, correo y contrasena son obligatorios"},
            status_code=400,
        )
    if len(password) < 8:
        return JSONResponse(
            {"success": False, "error": "La contraseña debe tener al menos 8 caracteres"},
            status_code=400,
        )

    db = SessionLocal()
    try:
        login_hash = _sensitive_lookup_hash(usuario_login)
        email_hash = _sensitive_lookup_hash(correo)
        exists_login = db.query(Usuario).filter(Usuario.usuario_hash == login_hash).first()
        if not exists_login:
            exists_login = db.query(Usuario).filter(Usuario.usuario == usuario_login).first()
        if exists_login:
            return JSONResponse({"success": False, "error": "El usuario ya existe"}, status_code=409)
        exists_email = db.query(Usuario).filter(Usuario.correo_hash == email_hash).first()
        if not exists_email:
            exists_email = db.query(Usuario).filter(Usuario.correo == correo).first()
        if exists_email:
            return JSONResponse({"success": False, "error": "El correo ya existe"}, status_code=409)

        rol_id = None
        if not rol_nombre:
            rol_nombre = "usuario"
        if not can_assign_role(request, rol_nombre):
            rol_nombre = "usuario"
        if rol_nombre:
            rol = db.query(Rol).filter(Rol.nombre == rol_nombre).first()
            if not rol:
                return JSONResponse({"success": False, "error": "Rol no encontrado"}, status_code=404)
            rol_id = rol.id

        nuevo = Usuario(
            nombre=nombre,
            usuario=_encrypt_sensitive(usuario_login),
            usuario_hash=login_hash,
            correo=_encrypt_sensitive(correo),
            correo_hash=email_hash,
            contrasena=hash_password(password),
            rol_id=rol_id,
            role=rol_nombre,
            is_active=True,
        )
        db.add(nuevo)
        db.commit()
        db.refresh(nuevo)
        return JSONResponse(
            {
                "success": True,
                "data": {
                    "id": nuevo.id,
                    "nombre": nuevo.nombre,
                    "correo": correo,
                    "rol_id": nuevo.rol_id,
                },
            }
        )
    finally:
        db.close()


@app.get("/api/roles-disponibles")
def listar_roles_disponibles(request: Request):
    require_admin_or_superadmin(request)
    allowed = set(get_visible_role_names(request))
    db = SessionLocal()
    try:
        roles = (
            db.query(Rol)
            .filter(Rol.nombre.in_(allowed))
            .order_by(Rol.id.asc())
            .all()
        )
        data = [
            {
                "id": role.id,
                "nombre": role.nombre,
                "descripcion": role.descripcion,
                "label": role.nombre.replace("_", " ").capitalize(),
            }
            for role in roles
        ]
        return JSONResponse({"success": True, "data": data})
    finally:
        db.close()


@app.get("/api/usuarios")
def listar_usuarios_sanitizados(request: Request):
    require_admin_or_superadmin(request)
    db = SessionLocal()
    try:
        roles = {r.id: r.nombre for r in db.query(Rol).all()}
        usuarios = db.query(Usuario).all()

        session_username = (getattr(request.state, "user_name", None) or "").strip()
        session_lookup_hash = _sensitive_lookup_hash(session_username) if session_username else ""
        session_user = None
        if session_username:
            session_user = (
                db.query(Usuario)
                .filter(
                    (Usuario.usuario_hash == session_lookup_hash)
                    | (func.lower(Usuario.usuario) == session_username.lower())
                )
                .first()
            )

        def resolved_role(u: Usuario) -> str:
            if u.rol_id and roles.get(u.rol_id):
                return normalize_role_name(roles.get(u.rol_id))
            return normalize_role_name(u.role)

        session_role_from_db = resolved_role(session_user) if session_user else ""
        session_is_superadmin = is_superadmin(request) or session_role_from_db == "superadministrador"

        data = [
            {
                "id": u.id,
                "nombre": u.nombre,
                "correo": _decrypt_sensitive(u.correo),
                "rol": resolved_role(u),
                "imagen": u.imagen,
            }
            for u in usuarios
            if not is_hidden_user(request, _decrypt_sensitive(u.usuario))
            and (session_is_superadmin or resolved_role(u) != "superadministrador")
        ]
        return JSONResponse({"success": True, "data": data})
    finally:
        db.close()


def _render_areas_page(
    request: Request,
    title: str = "Áreas organizacionales",
    description: str = "Administra la estructura de áreas de la organización",
) -> HTMLResponse:
    areas_content = """
        <section id="area-panel" class="usuario-panel">
            <div id="area-view"></div>
        </section>
    """
    return render_backend_page(
        request,
        title=title,
        description=description,
        content=areas_content,
        hide_floating_actions=False,
        view_buttons=[
            {"label": "Form", "icon": "/templates/icon/formulario.svg", "view": "form", "active": True},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
            {"label": "Organigrama", "view": "organigrama"},
        ],
        floating_actions_screen="areas",
    )


@app.get("/areas-organizacionales", response_class=HTMLResponse)
def areas_page(request: Request):
    return _render_areas_page(request)


@app.get("/inicio", response_class=HTMLResponse)
def inicio_page(request: Request):
    return render_backend_page(
        request,
        title="Inicio",
        description="Tablero de control estratégico y táctico",
        content=INICIO_BSC_HTML,
        hide_floating_actions=True,
        show_page_header=True,
        view_buttons=[
            {"label": "Formulario", "icon": "/templates/icon/formulario.svg", "view": "form"},
            {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list"},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
            {"label": "Dashboard", "icon": "/templates/icon/tablero.svg", "view": "dashboard", "active": True},
        ],
    )


@app.get("/inicio/departamentos", response_class=HTMLResponse)
def inicio_departamentos_page(request: Request):
    return _render_areas_page(
        request,
        title="Departamentos",
        description="Administra la estructura de departamentos de la organización",
    )


@app.get("/inicio/regiones", response_class=HTMLResponse)
def inicio_regiones_page(request: Request):
    return _render_regiones_page(request)


@app.get("/inicio/sucursales", response_class=HTMLResponse)
def inicio_sucursales_page(request: Request):
    return _render_sucursales_page(request)


@app.get("/inicio/colaboradores", response_class=HTMLResponse)
def inicio_colaboradores_page(request: Request):
    return _render_usuarios_page(
        request,
        title="Colaboradores",
        description="Gestiona colaboradores, roles y permisos desde la misma pantalla",
    )


@app.get("/proyectando", response_class=HTMLResponse)
def proyectando_page(request: Request):
    return render_backend_page(
        request,
        title="Proyectando",
        description="Herramienta de proyección financiera",
        content=PROYECTANDO_HTML,
        hide_floating_actions=True,
        show_page_header=True,
        view_buttons=[
            {"label": "Formulario", "icon": "/templates/icon/formulario.svg", "view": "form"},
            {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list"},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
            {"label": "Dashboard", "icon": "/templates/icon/tablero.svg", "view": "dashboard", "active": True},
        ],
    )


@app.post("/api/proyectando/datos-preliminares/datos-generales")
async def guardar_datos_preliminares_generales(data: dict = Body(...)):
    current = _load_datos_preliminares_store()
    updated = dict(current)
    for key in DEFAULT_DATOS_GENERALES.keys():
        if key in data:
            updated[key] = str(data.get(key) or "").strip()
    _save_datos_preliminares_store(updated)
    return {"success": True, "data": updated}


@app.get("/api/inicio/regiones")
def listar_regiones():
    return {"success": True, "data": _load_regiones_store()}


@app.post("/api/inicio/regiones")
async def guardar_regiones(data: dict = Body(...)):
    incoming = data.get("data", [])
    if not isinstance(incoming, list):
        raise HTTPException(status_code=400, detail="Formato inválido")
    _save_regiones_store(incoming)
    return {"success": True, "data": _load_regiones_store()}


@app.get("/api/inicio/sucursales")
def listar_sucursales():
    return {"success": True, "data": _load_sucursales_store()}


@app.post("/api/inicio/sucursales")
async def guardar_sucursales(data: dict = Body(...)):
    incoming = data.get("data", [])
    if not isinstance(incoming, list):
        raise HTTPException(status_code=400, detail="Formato inválido")
    _save_sucursales_store(incoming)
    return {"success": True, "data": _load_sucursales_store()}


@app.get("/proyectando/datos-preliminares", response_class=HTMLResponse)
def proyectando_datos_preliminares_page(request: Request):
    login_identity = _get_login_identity_context()
    logo_url = escape((login_identity.get("login_logo_url") or "").strip())
    datos_generales = _load_datos_preliminares_store()
    responsable_actual = (datos_generales.get("responsable_general") or "").strip()
    primer_anio_actual = (datos_generales.get("primer_anio_proyeccion") or "").strip()
    anios_proyeccion_actual = (datos_generales.get("anios_proyeccion") or "3").strip() or "3"
    db = SessionLocal()
    try:
        users = db.query(Usuario).all()
    finally:
        db.close()
    responsables_options: List[str] = []
    for user in users:
        if user.is_active is False:
            continue
        username_plain = (_decrypt_sensitive(user.usuario or "") or "").strip()
        full_name_plain = (user.nombre or "").strip()
        if not username_plain and not full_name_plain:
            continue
        if full_name_plain and username_plain:
            label = f"{full_name_plain} ({username_plain})"
        else:
            label = full_name_plain or username_plain
        selected_attr = " selected" if username_plain == responsable_actual else ""
        responsables_options.append(
            f'<option value="{escape(username_plain)}"{selected_attr}>{escape(label)}</option>'
        )
    if not responsables_options:
        responsables_options.append('<option value="">Sin usuarios disponibles</option>')
    anio_actual = datetime.now().year
    anios_options = "".join(
        f'<option value="{year}"{" selected" if str(year) == primer_anio_actual else ""}>{year}</option>'
        for year in range(anio_actual, anio_actual + 3)
    )
    try:
        primer_anio_num = int(primer_anio_actual)
    except (TypeError, ValueError):
        primer_anio_num = anio_actual
    anios_proyeccion_options = "".join(
        f'<option value="{years}"{" selected" if str(years) == anios_proyeccion_actual else ""}>{years}</option>'
        for years in range(1, 6)
    )
    macro_rows = []
    for row_idx, offset in enumerate((-2, -1, 0, 1, 2, 3)):
        year_value = anio_actual + offset
        macro_rows.append(
            f"""
            <tr data-macro-row="{row_idx}">
                <td class="year" style="padding:10px; border-bottom:1px solid #e2e8f0; color:#0f172a; font-weight:600;">{year_value}</td>
                <td class="num" style="padding:10px; border-bottom:1px solid #e2e8f0;">
                    <input class="macro-input" data-row="{row_idx}" data-col="0" type="text" name="inflacion_{year_value}" inputmode="decimal" autocomplete="off" placeholder="Ej. 4.50%" style="width:100%; height:36px; border:1px solid #cbd5e1; border-radius:8px; padding:0 10px;">
                </td>
                <td class="num" style="padding:10px; border-bottom:1px solid #e2e8f0;">
                    <input class="macro-input" data-row="{row_idx}" data-col="1" type="text" name="udi_{year_value}" inputmode="decimal" autocomplete="off" placeholder="Ej. 8.15" style="width:100%; height:36px; border:1px solid #cbd5e1; border-radius:8px; padding:0 10px;">
                </td>
            </tr>
            """
        )
    macro_table_rows = "".join(macro_rows)
    activos_fijos_defaults = [
        ("Terrenos", "0"),
        ("Construcciones", "20"),
        ("Construcciones en proceso", "5"),
        ("Equipo de transporte", "4"),
        ("Equipo de cómputo", "3"),
        ("Mobiliario", "3"),
        ("Otras propiedades, mobiliario y equipo", "2"),
    ]
    activos_fijos_rows = []
    for idx, (rubro, depreciacion) in enumerate(activos_fijos_defaults, start=1):
        activos_fijos_rows.append(
            f"""
            <tr>
                <td>
                    <input class="table-input" type="text" name="activo_fijo_rubro_{idx}" value="{escape(rubro)}">
                </td>
                <td>
                    <input class="table-input num" type="number" min="0" step="1" name="activo_fijo_anios_{idx}" value="{escape(depreciacion)}">
                </td>
                <td class="table-actions-cell">
                    <button type="button" class="delete-activo-fijo-row table-delete-btn">−</button>
                </td>
            </tr>
            """
        )
    activos_fijos_table_rows = "".join(activos_fijos_rows)
    initial_sucursal_row = """
        <tr>
            <td><input class="table-input" type="text" name="sucursal_nombre_1" placeholder="Nombre de la sucursal"></td>
            <td><input class="table-input num" type="number" min="0" step="1" name="sucursal_socios_1" placeholder="0"></td>
            <td><input class="table-input num" type="number" min="0" step="1" name="sucursal_menores_ahorradores_1" placeholder="0"></td>
            <td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_ahorro_menor_1" placeholder="0.00"></td>
            <td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_captacion_vista_1" placeholder="0.00"></td>
            <td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_inversiones_1" placeholder="0.00"></td>
            <td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_colocacion_1" placeholder="0.00"></td>
            <td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_cartera_vencida_1" placeholder="0.00"></td>
            <td class="table-actions-cell"><button type="button" class="add-sucursal-row table-add-btn">+</button></td>
        </tr>
    """
    logo_html = (
        f'<img src="{logo_url}" alt="Logo de la empresa" '
        'style="width:min(88px, 14vw); max-width:100%; height:auto; object-fit:contain;">'
        if logo_url
        else ""
    )
    preliminares_content = dedent(f"""
        <section class="sipet-ui-template" style="padding: 12px 4px 8px;">
            <style>
                {SIPET_PREMIUM_UI_TEMPLATE_CSS}
                .avan-premium {{
                    --bg: #f6f8fc;
                    --card: rgba(255,255,255,.86);
                    --border: rgba(15, 23, 42, .10);
                    --border-strong: rgba(15, 23, 42, .16);
                    --text: var(--body-text, #0f172a);
                    --muted: #475569;
                    --field: rgba(255,255,255,.92);
                    --shadow: 0 18px 60px rgba(2, 6, 23, .12);
                    --ring: 0 0 0 4px rgba(37, 99, 235, .16);
                    --radius: 22px;
                    --radius-sm: 14px;
                    --gap: 18px;
                    --pad: 20px;
                    color: var(--text);
                    background:
                        radial-gradient(1200px 700px at 10% 0%, rgba(37,99,235,.10), transparent 55%),
                        radial-gradient(900px 520px at 95% 10%, rgba(14,165,233,.10), transparent 55%),
                        var(--bg);
                    border-radius: calc(var(--radius) + 8px);
                }}
                .avan-premium .form-grid {{
                    display: grid;
                    grid-template-columns: repeat(3, minmax(0, 1fr));
                    gap: var(--gap);
                }}
                .avan-premium label {{
                    font-size: 13px;
                    font-weight: 700;
                    color: var(--text);
                }}
                .avan-premium input[type="text"],
                .avan-premium select {{
                    width: 100%;
                    background: var(--field) !important;
                    border: 1px solid var(--border) !important;
                    border-radius: var(--radius-sm) !important;
                    padding: 12px 14px !important;
                    font-size: 15px;
                    color: var(--text);
                    transition: border-color .18s ease, box-shadow .18s ease, background .18s ease;
                }}
                .avan-premium input:hover,
                .avan-premium select:hover {{
                    border-color: var(--border-strong) !important;
                }}
                .avan-premium input:focus,
                .avan-premium select:focus {{
                    border-color: rgba(37,99,235,.55) !important;
                    box-shadow: var(--ring) !important;
                    background: rgba(255,255,255,.98) !important;
                    outline: none;
                }}
                .tab-status {{
                    display:inline-flex;
                    align-items:center;
                    justify-content:center;
                    width:20px;
                    height:20px;
                    border-radius:999px;
                    text-align:center;
                    font-size: 0.8rem;
                    font-weight: 900;
                    color:#ffffff;
                    background:#dc2626;
                    box-shadow: inset 0 -1px 0 rgba(0,0,0,0.15);
                }}
                .tab-status.is-complete {{
                    background:#16a34a;
                }}
                [data-param-panel] input[type="number"],
                [data-param-panel] .macro-input {{
                    text-align: right !important;
                    font-variant-numeric: tabular-nums;
                }}
                @media (max-width: 1100px) {{
                    .avan-premium .form-grid {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
                }}
                @media (max-width: 720px) {{
                    .avan-premium .form-grid {{ grid-template-columns: 1fr; }}
                }}
                .macro-premium {{
                    --card: rgba(255,255,255,.86);
                    --border: rgba(15, 23, 42, .10);
                    --border-strong: rgba(15, 23, 42, .16);
                    --text: #0f172a;
                    --field: rgba(255,255,255,.92);
                    --shadow: 0 18px 60px rgba(2, 6, 23, .12);
                    --ring: 0 0 0 4px rgba(37, 99, 235, .16);
                    --radius: 22px;
                }}
                .macro-premium.table-card {{
                    background: var(--card) !important;
                    border: 1px solid var(--border) !important;
                    border-radius: var(--radius) !important;
                    box-shadow: var(--shadow) !important;
                    backdrop-filter: blur(10px);
                    -webkit-backdrop-filter: blur(10px);
                    overflow: hidden;
                }}
                .macro-premium .section-title {{
                    font-size: 28px !important;
                    line-height: 1.15;
                    font-weight: 850 !important;
                    letter-spacing: -0.02em;
                    margin: 0 0 14px 0 !important;
                }}
                .macro-premium .table-wrap {{
                    overflow: auto;
                    border: 1px solid var(--border);
                    border-radius: calc(var(--radius) - 6px);
                    background: rgba(255,255,255,.65);
                }}
                .macro-premium #macroeconomia-table {{
                    width: 100% !important;
                    border-collapse: collapse !important;
                    border-spacing: 0 !important;
                    min-width: 720px !important;
                    border: 0 !important;
                    background: transparent !important;
                    table-layout: fixed;
                }}
                .macro-premium #macroeconomia-table thead th {{
                    position: sticky;
                    top: 0;
                    z-index: 2;
                    text-align: left !important;
                    font-size: 13px !important;
                    letter-spacing: .08em;
                    text-transform: uppercase;
                    color: rgba(15,23,42,.75) !important;
                    background: linear-gradient(180deg, rgba(255,255,255,.92), rgba(255,255,255,.74)) !important;
                    border-bottom: 1px solid var(--border) !important;
                    border-right: 1px solid var(--border) !important;
                    padding: 14px 16px !important;
                }}
                .macro-premium #macroeconomia-table thead th:last-child {{
                    border-right: 0 !important;
                }}
                .macro-premium #macroeconomia-table thead th:first-child {{
                    border-top-left-radius: calc(var(--radius) - 8px);
                }}
                .macro-premium #macroeconomia-table thead tr th:last-child {{
                    border-top-right-radius: calc(var(--radius) - 8px);
                }}
                .macro-premium #macroeconomia-table tbody td {{
                    padding: 0 !important;
                    border-bottom: 1px solid rgba(15,23,42,.08) !important;
                    border-right: 1px solid rgba(15,23,42,.10) !important;
                    vertical-align: middle;
                    background: #ffffff !important;
                }}
                .macro-premium #macroeconomia-table tbody tr:nth-child(even) td {{
                    background: #ecfdf3 !important;
                }}
                .macro-premium #macroeconomia-table tbody td:last-child {{
                    border-right: 0 !important;
                }}
                .macro-premium #macroeconomia-table tbody td:first-child {{
                    font-weight: 850 !important;
                    font-size: 16px !important;
                    letter-spacing: 0;
                    color: var(--text) !important;
                    width: 160px;
                    padding: 8px 12px !important;
                    background: rgba(248,250,252,.85) !important;
                }}
                .macro-premium #macroeconomia-table tbody tr:hover td {{
                    background: #dcfce7 !important;
                }}
                .macro-premium #macroeconomia-table tbody tr:last-child td {{
                    border-bottom: 0 !important;
                }}
                .macro-premium #macroeconomia-table tbody tr:last-child td:first-child {{
                    border-bottom-left-radius: calc(var(--radius) - 8px);
                }}
                .macro-premium #macroeconomia-table tbody tr:last-child td:last-child {{
                    border-bottom-right-radius: calc(var(--radius) - 8px);
                }}
                .macro-premium #macroeconomia-table .macro-input {{
                    width: 100% !important;
                    max-width: none;
                    background: transparent !important;
                    border: 0 !important;
                    border-radius: 0 !important;
                    padding: 8px 10px !important;
                    font-size: 15px !important;
                    color: var(--text) !important;
                    outline: none;
                    box-shadow: none !important;
                    transition: box-shadow .15s ease, background .15s ease;
                    min-height: 34px;
                    font-variant-numeric: tabular-nums;
                }}
                .macro-premium #macroeconomia-table .macro-input:focus {{
                    box-shadow: inset 0 0 0 2px rgba(37,99,235,.45) !important;
                    background: rgba(239,246,255,.9) !important;
                }}
                .macro-premium #macroeconomia-table tbody td.num .macro-input {{
                    text-align: right;
                    font-variant-numeric: tabular-nums;
                }}
                @media (max-width: 860px) {{
                    .macro-premium #macroeconomia-table {{ min-width: 640px !important; }}
                    .macro-premium #macroeconomia-table tbody td:first-child {{ font-size: 14px !important; width: 120px; }}
                    .macro-premium #macroeconomia-table .macro-input {{ padding: 7px 9px !important; }}
                }}
            </style>
            <div style="display:flex; flex-wrap:wrap; align-items:flex-start; justify-content:space-between; gap:16px; margin-bottom: 12px;">
                <div style="flex:1 1 480px; min-width:280px; display:flex; align-items:flex-start; gap:16px;">
                    <div style="flex:0 0 auto;">
                        {logo_html}
                    </div>
                    <div style="flex:1 1 auto; padding-top: 2px;">
                        <div style="display:flex; align-items:center; justify-content:space-between; gap:12px; flex-wrap:wrap;">
                            <h2 style="margin:0; font-size: clamp(1.4rem, 2.5vw, 2rem); color:#0f172a;">Fase 1: Parametrización</h2>
                            <div style="display:flex; align-items:baseline; gap:8px; color:#334155; font-weight:700;">
                                <span style="font-size:0.95rem;">Avance parcial</span>
                                <span id="param-avance-parcial" style="font-size:1.55rem; line-height:1; color:#0f172a;">0%</span>
                            </div>
                        </div>
                        <p style="margin:0; font-size: clamp(1rem, 1.5vw, 1.1rem); color:#334155;">Capture los datos requeridos</p>
                    </div>
                </div>
            </div>
            <div id="param-tabs" role="tablist" aria-label="Secciones de parametrización" style="display:flex; align-items:center; justify-content:flex-start; gap:8px; flex-wrap:wrap; overflow:visible; white-space:normal; width:100%; margin: 10px 0 4px; padding:0 0 8px; border-bottom:1px solid #cbd5e1;">
                <button type="button" data-param-tab="datos-generales" aria-selected="true" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid var(--sidebar-bottom, #0f172a); border-bottom:4px solid var(--sidebar-bottom, #0f172a); border-radius:10px; background:var(--sidebar-text, #ffffff); padding:12px 14px; color:var(--sidebar-bottom, #0f172a); font-weight:700; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-datos-generales" class="tab-status">✕</span>
                    <img src="/templates/icon/datos_generales.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Datos generales</span>
                </button>
                <button type="button" data-param-tab="macroeconomia" aria-selected="false" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid transparent; border-bottom:4px solid transparent; border-radius:10px; background:transparent; padding:12px 14px; color:var(--body-text, #0f172a); font-weight:600; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-macroeconomia" class="tab-status">✕</span>
                    <img src="/templates/icon/macroeconomia.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Macroeconomia</span>
                </button>
                <button type="button" data-param-tab="sucursales" aria-selected="false" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid transparent; border-bottom:4px solid transparent; border-radius:10px; background:transparent; padding:12px 14px; color:var(--body-text, #0f172a); font-weight:600; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-sucursales" class="tab-status">✕</span>
                    <img src="/templates/icon/sucursales.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Sucursales</span>
                </button>
                <button type="button" data-param-tab="info-financiera" aria-selected="false" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid transparent; border-bottom:4px solid transparent; border-radius:10px; background:transparent; padding:12px 14px; color:var(--body-text, #0f172a); font-weight:600; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-info-financiera" class="tab-status">✕</span>
                    <img src="/templates/icon/informacion_financiera.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Información financiera</span>
                </button>
                <button type="button" data-param-tab="activo-fijo" aria-selected="false" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid transparent; border-bottom:4px solid transparent; border-radius:10px; background:transparent; padding:12px 14px; color:var(--body-text, #0f172a); font-weight:600; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-activo-fijo" class="tab-status">✕</span>
                    <img src="/templates/icon/activo_fijo.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Activo fijo</span>
                </button>
                <button type="button" data-param-tab="gastos" aria-selected="false" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid transparent; border-bottom:4px solid transparent; border-radius:10px; background:transparent; padding:12px 14px; color:var(--body-text, #0f172a); font-weight:600; cursor:pointer; outline:none; box-shadow:none;">
                    <span id="tab-status-gastos" class="tab-status">✕</span>
                    <img src="/templates/icon/gastos.svg" alt="" style="width:22px; height:22px; object-fit:contain;">
                    <span>Gastos</span>
                </button>
            </div>
            <div data-param-panel="datos-generales" class="avan-premium" style="display:block; margin-top: 22px; border: 1px solid #cbd5e1; border-radius: 14px; padding: 16px; background:#f8fafc;">
                <h3 style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Datos generales</h3>
                <div class="form-grid dg-grid" style="display:grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 14px;">
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Responsable general
                        <select id="dg-responsable-general" name="responsable_general" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                            <option value="">Seleccione un usuario</option>
                            {"".join(responsables_options)}
                        </select>
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Primer año de proyección
                        <select id="dg-primer-anio-proyeccion" name="primer_anio_proyeccion" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                            {anios_options}
                        </select>
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Años de proyección
                        <select id="anios-proyeccion" name="anios_proyeccion" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                            {anios_proyeccion_options}
                        </select>
                    </label>
                </div>
                <div class="form-grid dg-grid" style="display:grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 14px; margin-top: 14px;">
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Sociedad
                        <input id="dg-sociedad" type="text" value="{escape(datos_generales.get('sociedad', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Figura jurídica
                        <input id="dg-figura-juridica" type="text" value="{escape(datos_generales.get('figura_juridica', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Calle
                        <input id="dg-calle" type="text" value="{escape(datos_generales.get('calle', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Número exterior
                        <input id="dg-numero-exterior" type="text" value="{escape(datos_generales.get('numero_exterior', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Número interior
                        <input id="dg-numero-interior" type="text" value="{escape(datos_generales.get('numero_interior', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Colonia
                        <input id="dg-colonia" type="text" value="{escape(datos_generales.get('colonia', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Ciudad
                        <input id="dg-ciudad" type="text" value="{escape(datos_generales.get('ciudad', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Municipio
                        <input id="dg-municipio" type="text" value="{escape(datos_generales.get('municipio', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        Estado
                        <input id="dg-estado" type="text" value="{escape(datos_generales.get('estado', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        C.P.
                        <input id="dg-cp" type="text" value="{escape(datos_generales.get('cp', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                    <label style="display:flex; flex-direction:column; gap:6px; font-weight:600; color:#1e293b;">
                        País
                        <input id="dg-pais" type="text" value="{escape(datos_generales.get('pais', ''))}" style="height:40px; border:1px solid #cbd5e1; border-radius:10px; padding:0 10px; background:#ffffff; color:#0f172a;">
                    </label>
                </div>
                <div style="display:flex; align-items:center; gap:10px; margin-top:14px;">
                    <button type="button" id="dg-save-btn" style="height:38px; padding:0 14px; border-radius:10px; border:1px solid #0f172a; background:#0f172a; color:#fff; font-weight:600; cursor:pointer;">Guardar datos generales</button>
                    <span id="dg-save-status" style="font-size:0.9rem; color:#334155;"></span>
                </div>
            </div>
            <div data-param-panel="macroeconomia" class="avan-premium macro-premium table-card" style="display:none; margin-top: 18px; border: 1px solid #cbd5e1; border-radius: 14px; padding: 16px; background:#f8fafc;">
                <h3 class="section-title" style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Macroeconomía</h3>
                <div class="table-wrap" style="overflow-x:auto;">
                    <table id="macroeconomia-table" class="table-excel table-excel--compact" style="width:100%; border-collapse:collapse; min-width:680px; background:#ffffff; border:1px solid #e2e8f0; border-radius:10px;">
                        <thead>
                            <tr>
                                <th style="text-align:left; padding:10px; border-bottom:1px solid #cbd5e1; color:#1e293b; font-size:0.95rem;">Año</th>
                                <th style="text-align:left; padding:10px; border-bottom:1px solid #cbd5e1; color:#1e293b; font-size:0.95rem;">Inflación</th>
                                <th style="text-align:left; padding:10px; border-bottom:1px solid #cbd5e1; color:#1e293b; font-size:0.95rem;">Valor de la UDI</th>
                            </tr>
                        </thead>
                        <tbody id="macroeconomia-rows">
                            {macro_table_rows}
                        </tbody>
                    </table>
                </div>
            </div>
            <div data-param-panel="sucursales" style="display:none; margin-top:18px;">
                <h3 style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Participación de las sucursales al año (<span id="sucursales-anio-titulo">{escape(primer_anio_actual or str(anio_actual))}</span>)</h3>
                <div class="table-wrap" style="overflow-x:auto;">
                    <table id="sucursales-table" class="table-excel" style="width:100%; min-width:1200px;">
                        <thead>
                            <tr>
                                <th>Nombre de la sucursal</th>
                                <th>Socios</th>
                                <th>Menores ahorradores</th>
                                <th>Ahorro menor</th>
                                <th>Captación a la vista (sin ahorro menor)</th>
                                <th>Inversiones</th>
                                <th>Colocación</th>
                                <th>Cartera vencida</th>
                                <th class="table-actions-head">+</th>
                            </tr>
                        </thead>
                        <tbody id="sucursales-rows">{initial_sucursal_row}</tbody>
                    </table>
                </div>
            </div>
            <div data-param-panel="info-financiera" style="display:none; margin-top:18px;">
                <h3 style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Información financiera</h3>
                <div class="table-wrap" style="overflow-x:auto;">
                    <table id="info-financiera-basica-table" class="table-excel" style="width:100%; min-width:760px;">
                        <thead>
                            <tr>
                                <th>Rubro</th>
                                <th><span id="ifb-year-m3">{primer_anio_num - 3}</span></th>
                                <th><span id="ifb-year-m2">{primer_anio_num - 2}</span></th>
                                <th><span id="ifb-year-m1">{primer_anio_num - 1}</span></th>
                            </tr>
                        </thead>
                        <tbody id="info-financiera-basica-rows">
                            <tr>
                                <td class="ifb-row-label">Activos</td>
                                <td><input id="ifb-activos-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_activos_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-activos-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_activos_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-activos-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_activos_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Pasivos</td>
                                <td><input id="ifb-pasivos-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_pasivos_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-pasivos-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_pasivos_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-pasivos-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_pasivos_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Capital contable</td>
                                <td><input id="ifb-capital-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_capital_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-capital-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_capital_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-capital-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_capital_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Ingresos</td>
                                <td><input id="ifb-ingresos-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_ingresos_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-ingresos-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_ingresos_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-ingresos-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_ingresos_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Egresos</td>
                                <td><input id="ifb-egresos-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_egresos_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-egresos-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_egresos_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-egresos-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_egresos_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Resultado</td>
                                <td><input id="ifb-resultado-m3" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_resultado_m3', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-resultado-m2" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_resultado_m2', ''))}" placeholder="0.00"></td>
                                <td><input id="ifb-resultado-m1" class="table-input num" type="number" step="0.01" value="{escape(datos_generales.get('ifb_resultado_m1', ''))}" placeholder="0.00"></td>
                            </tr>
                            <tr>
                                <td class="ifb-row-label">Validación</td>
                                <td class="ifb-validation-cell">
                                    <div id="ifb-val-m3" class="ifb-validation-output"></div>
                                </td>
                                <td class="ifb-validation-cell">
                                    <div id="ifb-val-m2" class="ifb-validation-output"></div>
                                </td>
                                <td class="ifb-validation-cell">
                                    <div id="ifb-val-m1" class="ifb-validation-output"></div>
                                </td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
            <div data-param-panel="activo-fijo" style="display:none; margin-top:18px;">
                <h3 style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Activo fijo</h3>
                <div class="table-wrap" style="overflow-x:auto;">
                    <table id="activo-fijo-table" class="table-excel" style="width:100%; min-width:620px;">
                        <thead>
                            <tr>
                                <th>Rubro</th>
                                <th>Años de depreciación</th>
                                <th class="table-actions-head">Acción</th>
                            </tr>
                        </thead>
                        <tbody id="activo-fijo-rows">
                            {activos_fijos_table_rows}
                        </tbody>
                    </table>
                </div>
                <div style="margin-top:12px;">
                    <button type="button" id="add-activo-fijo-row" class="table-primary-btn">+ Agregar rubro</button>
                </div>
            </div>
            <div data-param-panel="gastos" style="display:none; margin-top:18px;">
                <h3 style="margin:0 0 14px; font-size: 1.1rem; color:#0f172a;">Gastos</h3>
                <div class="table-wrap" style="overflow-x:auto;">
                    <table id="gastos-table" class="table-excel" style="width:100%; min-width:620px;">
                        <thead>
                            <tr>
                                <th>Rubro</th>
                                <th><span id="gastos-year-m2">{primer_anio_num - 2}</span></th>
                                <th><span id="gastos-year-m1">{primer_anio_num - 1}</span></th>
                                <th><span id="gastos-year-0">{primer_anio_num}</span></th>
                            </tr>
                        </thead>
                        <tbody>
                            <tr>
                                <td class="ifb-row-label">Gastos</td>
                                <td><input id="ifb-gastos-m2" class="table-input num" type="number" step="0.01" placeholder="0.00"></td>
                                <td><input id="ifb-gastos-m1" class="table-input num" type="number" step="0.01" placeholder="0.00"></td>
                                <td><input id="ifb-gastos-0" class="table-input num" type="number" step="0.01" placeholder="0.00"></td>
                            </tr>
                        </tbody>
                    </table>
                </div>
            </div>
            <script>
                (function() {{
                    const currentYear = {anio_actual};
                    const projectionYearsSelect = document.getElementById("anios-proyeccion");
                    const macroRowsContainer = document.getElementById("macroeconomia-rows");
                    const primerAnioProyeccionSelect = document.getElementById("dg-primer-anio-proyeccion");
                    const sucursalesAnioTitulo = document.getElementById("sucursales-anio-titulo");
                    const ifbYearM3 = document.getElementById("ifb-year-m3");
                    const ifbYearM2 = document.getElementById("ifb-year-m2");
                    const ifbYearM1 = document.getElementById("ifb-year-m1");
                    const gastosYearM2 = document.getElementById("gastos-year-m2");
                    const gastosYearM1 = document.getElementById("gastos-year-m1");
                    const gastosYear0 = document.getElementById("gastos-year-0");
                    const sucursalesRows = document.getElementById("sucursales-rows");
                    const activoFijoRows = document.getElementById("activo-fijo-rows");
                    const addActivoFijoRowBtn = document.getElementById("add-activo-fijo-row");
                    const infoFinancieraBasicaTable = document.getElementById("info-financiera-basica-table");
                    const saveDatosGeneralesBtn = document.getElementById("dg-save-btn");
                    const saveDatosGeneralesStatus = document.getElementById("dg-save-status");
                    const avanceParcialElem = document.getElementById("param-avance-parcial");
                    const datosGeneralesTabStatus = document.getElementById("tab-status-datos-generales");
                    const macroeconomiaTabStatus = document.getElementById("tab-status-macroeconomia");
                    const sucursalesTabStatus = document.getElementById("tab-status-sucursales");
                    const infoFinancieraTabStatus = document.getElementById("tab-status-info-financiera");
                    const activoFijoTabStatus = document.getElementById("tab-status-activo-fijo");
                    const gastosTabStatus = document.getElementById("tab-status-gastos");
                    const paramTabButtons = Array.from(document.querySelectorAll("[data-param-tab]"));
                    const paramPanels = Array.from(document.querySelectorAll("[data-param-panel]"));
                    const totalSeccionesParam = 6;
                    const datosGeneralesRequiredIds = [
                        "dg-responsable-general",
                        "dg-primer-anio-proyeccion",
                        "anios-proyeccion",
                        "dg-sociedad",
                        "dg-figura-juridica",
                        "dg-calle",
                        "dg-numero-exterior",
                        "dg-numero-interior",
                        "dg-colonia",
                        "dg-ciudad",
                        "dg-municipio",
                        "dg-estado",
                        "dg-cp",
                        "dg-pais",
                    ];

                    function setActiveParamTab(tabName) {{
                        paramTabButtons.forEach((button) => {{
                            const isActive = button.getAttribute("data-param-tab") === tabName;
                            button.setAttribute("aria-selected", isActive ? "true" : "false");
                            button.style.borderColor = isActive ? "var(--sidebar-bottom, #0f172a)" : "transparent";
                            button.style.borderBottomColor = isActive ? "var(--sidebar-bottom, #0f172a)" : "transparent";
                            button.style.background = isActive ? "var(--sidebar-text, #ffffff)" : "transparent";
                            button.style.color = isActive ? "var(--sidebar-bottom, #0f172a)" : "var(--body-text, #0f172a)";
                            button.style.fontWeight = isActive ? "700" : "600";
                            button.style.outline = "none";
                            button.style.boxShadow = "none";
                        }});
                        paramPanels.forEach((panel) => {{
                            panel.style.display = panel.getAttribute("data-param-panel") === tabName ? "block" : "none";
                        }});
                    }}

                    function isFieldComplete(id) {{
                        const elem = document.getElementById(id);
                        if (!elem) return false;
                        return String(elem.value || "").trim().length > 0;
                    }}

                    function isDatosGeneralesComplete() {{
                        return datosGeneralesRequiredIds.every(isFieldComplete);
                    }}

                    function areAllFilled(selector) {{
                        const elems = Array.from(document.querySelectorAll(selector));
                        if (!elems.length) return false;
                        return elems.every((elem) => String(elem.value || "").trim().length > 0);
                    }}

                    function setTabStatus(tabElem, isComplete) {{
                        if (!tabElem) return;
                        tabElem.textContent = isComplete ? "✓" : "✕";
                        tabElem.classList.toggle("is-complete", isComplete);
                    }}

                    function updateParamProgress() {{
                        const datosGeneralesComplete = isDatosGeneralesComplete();
                        const macroeconomiaComplete = areAllFilled("#macroeconomia-table tbody input");
                        const sucursalesComplete = areAllFilled("#sucursales-table tbody input");
                        const infoFinancieraComplete = areAllFilled("#info-financiera-basica-table tbody input");
                        const activoFijoComplete = areAllFilled("#activo-fijo-table tbody input");
                        const gastosComplete = areAllFilled("#ifb-gastos-m2, #ifb-gastos-m1, #ifb-gastos-0");

                        setTabStatus(datosGeneralesTabStatus, datosGeneralesComplete);
                        setTabStatus(macroeconomiaTabStatus, macroeconomiaComplete);
                        setTabStatus(sucursalesTabStatus, sucursalesComplete);
                        setTabStatus(infoFinancieraTabStatus, infoFinancieraComplete);
                        setTabStatus(activoFijoTabStatus, activoFijoComplete);
                        setTabStatus(gastosTabStatus, gastosComplete);

                        if (!avanceParcialElem) return;
                        const completed = [
                            datosGeneralesComplete,
                            macroeconomiaComplete,
                            sucursalesComplete,
                            infoFinancieraComplete,
                            activoFijoComplete,
                            gastosComplete,
                        ].filter(Boolean).length;
                        const percent = Math.round((completed / totalSeccionesParam) * 100);
                        avanceParcialElem.textContent = percent + "%";
                    }}

                    datosGeneralesRequiredIds.forEach((id) => {{
                        const elem = document.getElementById(id);
                        if (!elem) return;
                        elem.addEventListener("input", updateParamProgress);
                        elem.addEventListener("change", updateParamProgress);
                    }});
                    document.getElementById("macroeconomia-table")?.addEventListener("input", updateParamProgress);
                    document.getElementById("macroeconomia-table")?.addEventListener("change", updateParamProgress);
                    document.getElementById("sucursales-table")?.addEventListener("input", updateParamProgress);
                    document.getElementById("sucursales-table")?.addEventListener("change", updateParamProgress);
                    document.getElementById("activo-fijo-table")?.addEventListener("input", updateParamProgress);
                    document.getElementById("activo-fijo-table")?.addEventListener("change", updateParamProgress);
                    document.getElementById("info-financiera-basica-table")?.addEventListener("input", updateParamProgress);
                    document.getElementById("info-financiera-basica-table")?.addEventListener("change", updateParamProgress);
                    document.getElementById("ifb-gastos-m2")?.addEventListener("input", updateParamProgress);
                    document.getElementById("ifb-gastos-m1")?.addEventListener("input", updateParamProgress);
                    document.getElementById("ifb-gastos-0")?.addEventListener("input", updateParamProgress);

                    if (paramTabButtons.length && paramPanels.length) {{
                        paramTabButtons.forEach((button) => {{
                            button.addEventListener("click", () => {{
                                const tabName = button.getAttribute("data-param-tab") || "";
                                if (!tabName) return;
                                setActiveParamTab(tabName);
                            }});
                        }});
                        setActiveParamTab("datos-generales");
                    }}
                    updateParamProgress();

                    if (!projectionYearsSelect || !macroRowsContainer) return;

                    function buildMacroRows(futureYears) {{
                        const safeFutureYears = Math.min(5, Math.max(1, Number(futureYears) || 3));
                        const offsets = [-2, -1, 0];
                        for (let i = 1; i <= safeFutureYears; i += 1) offsets.push(i);

                        const rowsHtml = offsets.map((offset, rowIndex) => {{
                            const yearValue = currentYear + offset;
                            return (
                                '<tr data-macro-row="' + rowIndex + '">'
                                + '<td class="year" style="padding:10px; border-bottom:1px solid #e2e8f0; color:#0f172a; font-weight:600;">'
                                + yearValue + '</td>'
                                + '<td class="num" style="padding:10px; border-bottom:1px solid #e2e8f0;">'
                                + '<input class="macro-input" data-row="' + rowIndex + '" data-col="0" type="text" name="inflacion_' + yearValue + '" inputmode="decimal" autocomplete="off" placeholder="Ej. 4.50%" style="width:100%; height:36px; border:1px solid #cbd5e1; border-radius:8px; padding:0 10px;">'
                                + '</td>'
                                + '<td class="num" style="padding:10px; border-bottom:1px solid #e2e8f0;">'
                                + '<input class="macro-input" data-row="' + rowIndex + '" data-col="1" type="text" name="udi_' + yearValue + '" inputmode="decimal" autocomplete="off" placeholder="Ej. 8.15" style="width:100%; height:36px; border:1px solid #cbd5e1; border-radius:8px; padding:0 10px;">'
                                + '</td>'
                                + '</tr>'
                            );
                        }}).join("");
                        macroRowsContainer.innerHTML = rowsHtml;
                        updateParamProgress();
                    }}

                    function getMacroMatrix() {{
                        const rows = Array.from(macroRowsContainer.querySelectorAll("tr"));
                        return rows.map((row, rowIndex) => {{
                            const cells = Array.from(row.querySelectorAll(".macro-input"));
                            cells.forEach((input, colIndex) => {{
                                input.dataset.row = String(rowIndex);
                                input.dataset.col = String(colIndex);
                            }});
                            return cells;
                        }});
                    }}

                    function focusMacroCell(rowIndex, colIndex) {{
                        const matrix = getMacroMatrix();
                        const row = matrix[rowIndex];
                        if (!row) return;
                        const cell = row[colIndex];
                        if (!cell) return;
                        cell.focus();
                        cell.select();
                    }}

                    macroRowsContainer.addEventListener("keydown", (event) => {{
                        const target = event.target;
                        if (!(target instanceof HTMLInputElement)) return;
                        if (!target.classList.contains("macro-input")) return;
                        const row = Number(target.dataset.row || "0");
                        const col = Number(target.dataset.col || "0");
                        if (event.key === "ArrowUp") {{
                            event.preventDefault();
                            focusMacroCell(Math.max(0, row - 1), col);
                        }} else if (event.key === "ArrowDown" || event.key === "Enter") {{
                            event.preventDefault();
                            focusMacroCell(row + 1, col);
                        }} else if (event.key === "ArrowLeft") {{
                            if (target.selectionStart === 0 && target.selectionEnd === 0) {{
                                event.preventDefault();
                                focusMacroCell(row, Math.max(0, col - 1));
                            }}
                        }} else if (event.key === "ArrowRight") {{
                            const atEnd = target.selectionStart === target.value.length && target.selectionEnd === target.value.length;
                            if (atEnd) {{
                                event.preventDefault();
                                focusMacroCell(row, col + 1);
                            }}
                        }}
                    }});

                    macroRowsContainer.addEventListener("paste", (event) => {{
                        const target = event.target;
                        if (!(target instanceof HTMLInputElement)) return;
                        if (!target.classList.contains("macro-input")) return;
                        const clipboard = event.clipboardData?.getData("text");
                        if (!clipboard) return;
                        const rows = clipboard
                            .split(/\\r?\\n/)
                            .map((line) => line.trim())
                            .filter((line) => line.length > 0)
                            .map((line) => line.split("\t"));
                        if (!rows.length) return;
                        event.preventDefault();
                        const startRow = Number(target.dataset.row || "0");
                        const startCol = Number(target.dataset.col || "0");
                        const matrix = getMacroMatrix();
                        rows.forEach((rowValues, rowOffset) => {{
                            rowValues.forEach((value, colOffset) => {{
                                const r = startRow + rowOffset;
                                const c = startCol + colOffset;
                                const cell = matrix[r]?.[c];
                                if (!cell) return;
                                cell.value = value.trim();
                            }});
                        }});
                        updateParamProgress();
                    }});

                    projectionYearsSelect.addEventListener("change", () => {{
                        buildMacroRows(projectionYearsSelect.value);
                    }});
                    if (primerAnioProyeccionSelect && sucursalesAnioTitulo) {{
                        const syncSucursalYearTitle = () => {{
                            sucursalesAnioTitulo.textContent = String(primerAnioProyeccionSelect.value || currentYear);
                        }};
                        primerAnioProyeccionSelect.addEventListener("change", syncSucursalYearTitle);
                        syncSucursalYearTitle();
                    }}
                    if (
                        primerAnioProyeccionSelect &&
                        ifbYearM3 && ifbYearM2 && ifbYearM1 &&
                        gastosYearM2 && gastosYearM1 && gastosYear0
                    ) {{
                        const syncFinancialYearHeaders = () => {{
                            const baseYear = Number(primerAnioProyeccionSelect.value) || currentYear;
                            ifbYearM3.textContent = String(baseYear - 3);
                            ifbYearM2.textContent = String(baseYear - 2);
                            ifbYearM1.textContent = String(baseYear - 1);
                            gastosYearM2.textContent = String(baseYear - 2);
                            gastosYearM1.textContent = String(baseYear - 1);
                            gastosYear0.textContent = String(baseYear);
                        }};
                        primerAnioProyeccionSelect.addEventListener("change", syncFinancialYearHeaders);
                        syncFinancialYearHeaders();
                    }}

                    function createSucursalRow(rowIndex) {{
                        const tr = document.createElement("tr");
                        tr.innerHTML =
                            '<td><input class="table-input" type="text" name="sucursal_nombre_' + rowIndex + '" placeholder="Nombre de la sucursal"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="1" name="sucursal_socios_' + rowIndex + '" placeholder="0"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="1" name="sucursal_menores_ahorradores_' + rowIndex + '" placeholder="0"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_ahorro_menor_' + rowIndex + '" placeholder="0.00"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_captacion_vista_' + rowIndex + '" placeholder="0.00"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_inversiones_' + rowIndex + '" placeholder="0.00"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_colocacion_' + rowIndex + '" placeholder="0.00"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="0.01" name="sucursal_cartera_vencida_' + rowIndex + '" placeholder="0.00"></td>'
                            + '<td class="table-actions-cell"><button type="button" class="add-sucursal-row table-add-btn">+</button></td>';
                        return tr;
                    }}

                    if (sucursalesRows) {{
                        let sucursalIndex = Math.max(1, sucursalesRows.querySelectorAll("tr").length);
                        sucursalesRows.addEventListener("click", (event) => {{
                            const target = event.target;
                            if (!(target instanceof HTMLElement)) return;
                            if (!target.classList.contains("add-sucursal-row")) return;
                            sucursalIndex += 1;
                            const currentRow = target.closest("tr");
                            const newRow = createSucursalRow(sucursalIndex);
                            if (currentRow && currentRow.parentNode) {{
                                currentRow.parentNode.insertBefore(newRow, currentRow.nextSibling);
                            }} else {{
                                sucursalesRows.appendChild(newRow);
                            }}
                        }});
                    }}

                    function createActivoFijoRow(rowIndex, rubroValue = "", aniosValue = "") {{
                        const tr = document.createElement("tr");
                        tr.innerHTML =
                            '<td><input class="table-input" type="text" name="activo_fijo_rubro_' + rowIndex + '" value="' + rubroValue.replace(/"/g, '&quot;') + '"></td>'
                            + '<td><input class="table-input num" type="number" min="0" step="1" name="activo_fijo_anios_' + rowIndex + '" value="' + aniosValue.replace(/"/g, '&quot;') + '"></td>'
                            + '<td class="table-actions-cell"><button type="button" class="delete-activo-fijo-row table-delete-btn">−</button></td>';
                        return tr;
                    }}

                    if (activoFijoRows) {{
                        let activoFijoIndex = Math.max(1, activoFijoRows.querySelectorAll("tr").length);
                        addActivoFijoRowBtn && addActivoFijoRowBtn.addEventListener("click", () => {{
                            activoFijoIndex += 1;
                            activoFijoRows.appendChild(createActivoFijoRow(activoFijoIndex));
                        }});
                        activoFijoRows.addEventListener("click", (event) => {{
                            const target = event.target;
                            if (!(target instanceof HTMLElement)) return;
                            if (!target.classList.contains("delete-activo-fijo-row")) return;
                            const row = target.closest("tr");
                            if (!row) return;
                            if (activoFijoRows.querySelectorAll("tr").length <= 1) {{
                                const rubroInput = row.querySelector('input[name^="activo_fijo_rubro_"]');
                                const aniosInput = row.querySelector('input[name^="activo_fijo_anios_"]');
                                if (rubroInput) rubroInput.value = "";
                                if (aniosInput) aniosInput.value = "";
                                return;
                            }}
                            row.remove();
                        }});
                    }}

                    function parseNumber(value) {{
                        const parsed = Number(value);
                        return Number.isFinite(parsed) ? parsed : 0;
                    }}

                    function renderValidationCell(targetId, activos, pasivos, capital) {{
                        const target = document.getElementById(targetId);
                        if (!target) return;
                        const validBalance = Math.abs(activos - (pasivos + capital)) < 0.0001;
                        const okStyle = "display:block; text-align:center; font-style:italic; color:#0f172a; padding:6px 4px;";
                        const errStyle = "display:block; text-align:center; color:#991b1b; background:#fee2e2; border:1px solid #fca5a5; border-radius:6px; padding:6px 8px;";
                        target.innerHTML = validBalance
                            ? '<span style="' + okStyle + '">Ok</span>'
                            : '<span style="' + errStyle + '">Revisar activos, pasivos y capital</span>';
                    }}

                    function validateInformacionFinanciera() {{
                        const periods = ['m3', 'm2', 'm1'];
                        periods.forEach((period) => {{
                            const activos = parseNumber(document.getElementById('ifb-activos-' + period)?.value);
                            const pasivos = parseNumber(document.getElementById('ifb-pasivos-' + period)?.value);
                            const capital = parseNumber(document.getElementById('ifb-capital-' + period)?.value);
                            renderValidationCell('ifb-val-' + period, activos, pasivos, capital);
                        }});
                    }}
                    function isInformacionFinancieraCorrecta() {{
                        const periods = ['m3', 'm2', 'm1'];
                        return periods.every((period) => {{
                            const activos = parseNumber(document.getElementById('ifb-activos-' + period)?.value);
                            const pasivos = parseNumber(document.getElementById('ifb-pasivos-' + period)?.value);
                            const capital = parseNumber(document.getElementById('ifb-capital-' + period)?.value);
                            return Math.abs(activos - (pasivos + capital)) < 0.0001;
                        }});
                    }}

                    if (infoFinancieraBasicaTable) {{
                        infoFinancieraBasicaTable.addEventListener('input', (event) => {{
                            const target = event.target;
                            if (!(target instanceof HTMLElement)) return;
                            if (target.tagName !== 'INPUT') return;
                            validateInformacionFinanciera();
                        }});
                        validateInformacionFinanciera();
                    }}

                    function getValue(id) {{
                        const elem = document.getElementById(id);
                        if (!elem) return "";
                        return (elem.value || "").trim();
                    }}

                    async function saveDatosGenerales() {{
                        if (!saveDatosGeneralesBtn) return;
                        saveDatosGeneralesBtn.disabled = true;
                        if (saveDatosGeneralesStatus) saveDatosGeneralesStatus.textContent = "Guardando...";
                        const payload = {{
                            responsable_general: getValue("dg-responsable-general"),
                            primer_anio_proyeccion: getValue("dg-primer-anio-proyeccion"),
                            anios_proyeccion: getValue("anios-proyeccion"),
                            sociedad: getValue("dg-sociedad"),
                            figura_juridica: getValue("dg-figura-juridica"),
                            calle: getValue("dg-calle"),
                            numero_exterior: getValue("dg-numero-exterior"),
                            numero_interior: getValue("dg-numero-interior"),
                            colonia: getValue("dg-colonia"),
                            ciudad: getValue("dg-ciudad"),
                            municipio: getValue("dg-municipio"),
                            estado: getValue("dg-estado"),
                            cp: getValue("dg-cp"),
                            pais: getValue("dg-pais"),
                            ifb_activos_m3: getValue("ifb-activos-m3"),
                            ifb_activos_m2: getValue("ifb-activos-m2"),
                            ifb_activos_m1: getValue("ifb-activos-m1"),
                            ifb_pasivos_m3: getValue("ifb-pasivos-m3"),
                            ifb_pasivos_m2: getValue("ifb-pasivos-m2"),
                            ifb_pasivos_m1: getValue("ifb-pasivos-m1"),
                            ifb_capital_m3: getValue("ifb-capital-m3"),
                            ifb_capital_m2: getValue("ifb-capital-m2"),
                            ifb_capital_m1: getValue("ifb-capital-m1"),
                            ifb_ingresos_m3: getValue("ifb-ingresos-m3"),
                            ifb_ingresos_m2: getValue("ifb-ingresos-m2"),
                            ifb_ingresos_m1: getValue("ifb-ingresos-m1"),
                            ifb_egresos_m3: getValue("ifb-egresos-m3"),
                            ifb_egresos_m2: getValue("ifb-egresos-m2"),
                            ifb_egresos_m1: getValue("ifb-egresos-m1"),
                            ifb_resultado_m3: getValue("ifb-resultado-m3"),
                            ifb_resultado_m2: getValue("ifb-resultado-m2"),
                            ifb_resultado_m1: getValue("ifb-resultado-m1"),
                        }};
                        try {{
                            const res = await fetch("/api/proyectando/datos-preliminares/datos-generales", {{
                                method: "POST",
                                headers: {{ "Content-Type": "application/json" }},
                                body: JSON.stringify(payload),
                            }});
                            const json = await res.json();
                            if (!res.ok || !json?.success) throw new Error(json?.error || "No se pudo guardar");
                            if (saveDatosGeneralesStatus) {{
                                saveDatosGeneralesStatus.textContent = isInformacionFinancieraCorrecta()
                                    ? "Información financiera correcta. Guardado correctamente."
                                    : "Información financiera incorrecta. Guardado con observaciones.";
                            }}
                        }} catch (error) {{
                            if (saveDatosGeneralesStatus) saveDatosGeneralesStatus.textContent = "Error al guardar.";
                        }} finally {{
                            saveDatosGeneralesBtn.disabled = false;
                        }}
                    }}

                    saveDatosGeneralesBtn && saveDatosGeneralesBtn.addEventListener("click", saveDatosGenerales);

                    buildMacroRows(projectionYearsSelect.value);
                }})();
            </script>
        </section>
    """)
    return render_backend_page(
        request,
        title="Datos preliminares",
        description="",
        content=preliminares_content,
        hide_floating_actions=True,
        show_page_header=False,
    )


@app.get("/proyectando/crecimiento-general", response_class=HTMLResponse)
def proyectando_crecimiento_general_page(request: Request):
    preliminares = _load_datos_preliminares_store()
    current_year = datetime.now().year
    try:
        base_year = int((preliminares.get("primer_anio_proyeccion") or "").strip() or current_year)
    except (TypeError, ValueError):
        base_year = current_year
    try:
        projection_years = int((preliminares.get("anios_proyeccion") or "").strip() or 3)
    except (TypeError, ValueError):
        projection_years = 3
    projection_years = max(1, min(projection_years, 10))

    def _to_float(raw: Any) -> float:
        text = str(raw or "").strip().replace(",", "")
        if not text:
            return 0.0
        try:
            return float(text)
        except ValueError:
            return 0.0

    activos_m3 = _to_float(preliminares.get("ifb_activos_m3"))
    activos_m2 = _to_float(preliminares.get("ifb_activos_m2"))
    activos_m1 = _to_float(preliminares.get("ifb_activos_m1"))
    activos_y0_raw = str(preliminares.get("ifb_activos_0") or "").strip()
    activos_y0 = _to_float(activos_y0_raw) if activos_y0_raw else activos_m1
    historical_activos = {
        "m3": activos_m3,
        "m2": activos_m2,
        "m1": activos_m1,
        "y0": activos_y0,
    }
    historical_activos_json = json.dumps(historical_activos)

    login_identity = _get_login_identity_context()
    logo_url = escape((login_identity.get("login_logo_url") or "").strip())
    logo_html = (
        f'<img src="{logo_url}" alt="Logo de la empresa" '
        'style="width:min(88px, 14vw); max-width:100%; height:auto; object-fit:contain;">'
        if logo_url
        else ""
    )
    content = dedent(f"""
        <section class="sipet-ui-template" style="padding: 12px 4px 8px;">
            <style>
                {SIPET_PREMIUM_UI_TEMPLATE_CSS}
            </style>
            <div style="display:flex; flex-wrap:wrap; align-items:flex-start; justify-content:space-between; gap:16px; margin-bottom: 12px;">
                <div style="flex:1 1 480px; min-width:280px; display:flex; align-items:flex-start; gap:16px;">
                    <div style="flex:0 0 auto;">
                        {logo_html}
                    </div>
                    <div style="flex:1 1 auto; padding-top: 2px;">
                        <div style="display:flex; align-items:center; justify-content:space-between; gap:12px; flex-wrap:wrap;">
                            <h2 style="margin:0; font-size: clamp(1.4rem, 2.5vw, 2rem); color:#0f172a;">Fase 2: Crecimiento General</h2>
                            <div style="display:flex; align-items:baseline; gap:10px; color:#334155; font-weight:700;">
                                <span style="font-size: clamp(1.2rem, 2vw, 2rem);">Avance parcial</span>
                                <span id="crecimiento-avance-parcial" style="font-size: clamp(1.8rem, 3.2vw, 3.2rem); line-height:1; color:#0f172a;">0%</span>
                            </div>
                        </div>
                        <p style="margin:0; font-size: clamp(1rem, 1.5vw, 1.1rem); color:#334155;">Capture los datos requeridos</p>
                    </div>
                </div>
            </div>
            <div id="cg-tabs" role="tablist" aria-label="Secciones de crecimiento general" style="display:flex; align-items:center; justify-content:flex-start; gap:8px; flex-wrap:wrap; overflow:visible; white-space:normal; width:100%; margin: 10px 0 4px; padding:0 0 8px; border-bottom:1px solid #cbd5e1;">
                <button type="button" data-cg-tab="activo-total" aria-selected="true" style="display:inline-flex; align-items:center; gap:10px; flex:0 0 auto; border:1px solid var(--sidebar-bottom, #0f172a); border-bottom:4px solid var(--sidebar-bottom, #0f172a); border-radius:10px; background:var(--sidebar-text, #ffffff); padding:12px 14px; color:var(--sidebar-bottom, #0f172a); font-weight:700; cursor:pointer; outline:none; box-shadow:none;">
                    <img src="/templates/icon/crecimiento.svg" alt="" style="width:20px; height:20px; object-fit:contain;">
                    <span>Crecimiento del activo total</span>
                </button>
            </div>
            <div data-cg-panel="activo-total" style="display:block; margin-top:18px;">
                <div class="table-wrap" style="overflow-x:auto;">
                    <table class="table-excel" style="width:100%; min-width:980px;">
                        <thead>
                            <tr>
                                <th style="width:26%;">Periodo</th>
                                <th>Saldo</th>
                                <th>Crecimiento</th>
                                <th>%</th>
                            </tr>
                        </thead>
                        <tbody id="cg-activo-total-rows"></tbody>
                    </table>
                </div>
            </div>
            <script>
                (function() {{
                    const tabButtons = Array.from(document.querySelectorAll("[data-cg-tab]"));
                    const panels = Array.from(document.querySelectorAll("[data-cg-panel]"));
                    if (!tabButtons.length) return;
                    function setActiveTab(tabName) {{
                        tabButtons.forEach((button) => {{
                            const isActive = button.getAttribute("data-cg-tab") === tabName;
                            button.setAttribute("aria-selected", isActive ? "true" : "false");
                            button.style.borderColor = isActive ? "var(--sidebar-bottom, #0f172a)" : "transparent";
                            button.style.borderBottomColor = isActive ? "var(--sidebar-bottom, #0f172a)" : "transparent";
                            button.style.background = isActive ? "var(--sidebar-text, #ffffff)" : "transparent";
                            button.style.color = isActive ? "var(--sidebar-bottom, #0f172a)" : "var(--body-text, #0f172a)";
                            button.style.fontWeight = isActive ? "700" : "600";
                            button.style.outline = "none";
                            button.style.boxShadow = "none";
                        }});
                        panels.forEach((panel) => {{
                            panel.style.display = panel.getAttribute("data-cg-panel") === tabName ? "block" : "none";
                        }});
                    }}
                    tabButtons.forEach((button) => {{
                        button.addEventListener("click", () => {{
                            const tabName = button.getAttribute("data-cg-tab") || "";
                            if (!tabName) return;
                            setActiveTab(tabName);
                        }});
                    }});

                    const cgRows = document.getElementById("cg-activo-total-rows");
                    const baseYear = {base_year};
                    const projectionYears = {projection_years};
                    const historicalActivos = {historical_activos_json};
                    const offsets = [-3, -2, -1, 0].concat(Array.from({{ length: projectionYears }}, (_, idx) => idx + 1));
                    const futureGrowthPct = {{}};
                    offsets.forEach((offset) => {{
                        if (offset > 0) futureGrowthPct[offset] = 0;
                    }});

                    const fmtNumber = (value) => Number(value || 0).toLocaleString("en-US", {{
                        minimumFractionDigits: 2,
                        maximumFractionDigits: 2,
                    }});
                    const fmtPercent = (value) => Number(value || 0).toLocaleString("en-US", {{
                        minimumFractionDigits: 2,
                        maximumFractionDigits: 2,
                    }}) + "%";
                    const periodLabel = (offset) => {{
                        if (offset === 0) return `Año actual (${{baseYear}})`;
                        if (offset < 0) return `${{offset}} (${{baseYear + offset}})`;
                        return `+${{offset}} (${{baseYear + offset}})`;
                    }};
                    const computeRows = () => {{
                        const rows = [];
                        let prevSaldo = null;
                        offsets.forEach((offset) => {{
                            let saldo = 0;
                            let crecimiento = null;
                            let pct = null;
                            if (offset === -3) {{
                                saldo = Number(historicalActivos.m3 || 0);
                            }} else if (offset === -2) {{
                                saldo = Number(historicalActivos.m2 || 0);
                            }} else if (offset === -1) {{
                                saldo = Number(historicalActivos.m1 || 0);
                            }} else if (offset === 0) {{
                                saldo = Number(historicalActivos.y0 || 0);
                            }} else {{
                                const userPct = Number(futureGrowthPct[offset] || 0);
                                crecimiento = Number((Number(prevSaldo || 0) * (userPct / 100)).toFixed(2));
                                saldo = Number((Number(prevSaldo || 0) + crecimiento).toFixed(2));
                                pct = userPct;
                            }}
                            if (offset <= 0 && prevSaldo !== null) {{
                                crecimiento = Number((saldo - prevSaldo).toFixed(2));
                                pct = prevSaldo !== 0 ? Number(((crecimiento / prevSaldo) * 100).toFixed(4)) : 0;
                            }}
                            rows.push({{ offset, saldo, crecimiento, pct }});
                            prevSaldo = saldo;
                        }});
                        return rows;
                    }};
                    const renderRows = () => {{
                        if (!cgRows) return;
                        const rows = computeRows();
                        cgRows.innerHTML = rows.map((row) => {{
                            const crecimientoValue = row.crecimiento === null ? "-" : fmtNumber(row.crecimiento);
                            const pctReadonly = row.pct === null ? "-" : fmtPercent(row.pct);
                            const pctCell = row.offset > 0
                                ? `<input class="table-input num cg-growth-input" data-offset="${{row.offset}}" type="number" step="0.01" min="-100" value="${{Number(futureGrowthPct[row.offset] || 0).toFixed(2)}}" style="background:#dbeafe; border-color:#93c5fd; font-weight:700;">`
                                : `<input class="table-input num" type="text" value="${{pctReadonly}}" readonly>`;
                            return `
                                <tr>
                                    <td class="year">${{periodLabel(row.offset)}}</td>
                                    <td><input class="table-input num" type="text" value="${{fmtNumber(row.saldo)}}" readonly></td>
                                    <td><input class="table-input num" type="text" value="${{crecimientoValue}}" readonly></td>
                                    <td>${{pctCell}}</td>
                                </tr>
                            `;
                        }}).join("");
                    }};
                    document.addEventListener("input", (event) => {{
                        const target = event.target;
                        if (!(target instanceof HTMLInputElement)) return;
                        if (!target.classList.contains("cg-growth-input")) return;
                        const offset = Number(target.dataset.offset || "0");
                        if (offset <= 0) return;
                        futureGrowthPct[offset] = Number(target.value || 0);
                        renderRows();
                    }});

                    setActiveTab("activo-total");
                    renderRows();
                }})();
            </script>
        </section>
    """)
    return render_backend_page(
        request,
        title="Crecimiento general",
        description="",
        content=content,
        hide_floating_actions=True,
        show_page_header=False,
    )


@app.get("/proyectando/crecimiento-general/activo-total", response_class=HTMLResponse)
def proyectando_crecimiento_activo_total_page(request: Request):
    return RedirectResponse(url="/proyectando/crecimiento-general", status_code=307)


@app.get("/planes", response_class=HTMLResponse)
def plan_estrategico_page(request: Request):
    create_plan_button = ""
    if is_admin_or_superadmin(request):
        create_plan_button = (
            "<div style=\"display:flex; justify-content:flex-end; margin: 4px 0 12px;\">"
            "<button class=\"pe-btn pe-btn--primary\" type=\"button\" "
            "onclick=\"window.location.href='/ejes-estrategicos'\">Crear plan estratégico</button>"
            "</div>"
        )
    plan_content = PLAN_ESTRATEGICO_HTML.replace("__CREATE_PLAN_BUTTON__", create_plan_button)
    return render_backend_page(
        request,
        title="Plan estratégico",
        description="Consolidación de planificación estratégica institucional.",
        content=plan_content,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/poa", response_class=HTMLResponse)
def poa_page(request: Request):
    create_poa_button = (
        "<div style=\"display:flex; justify-content:flex-end; margin: 4px 0 12px;\">"
        "<button class=\"pe-btn pe-btn--primary\" type=\"button\" "
        "onclick=\"window.location.href='/poa/crear'\">Crear POA</button>"
        "</div>"
    )
    poa_content = POA_HTML.replace("__CREATE_POA_BUTTON__", create_poa_button)
    return render_backend_page(
        request,
        title="POA",
        description="Programación operativa anual alineada al plan estratégico.",
        content=poa_content,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/poa/crear", response_class=HTMLResponse)
def poa_create_page(request: Request):
    return render_backend_page(
        request,
        title="Crear POA",
        description="Tablero de creación y delegación de actividades POA.",
        content=POA_CREAR_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )


def _render_blank_management_screen(request: Request, title: str) -> HTMLResponse:
    return render_backend_page(
        request,
        title=title,
        description="",
        content="",
        hide_floating_actions=True,
        show_page_header=False,
    )


def _render_regiones_page(request: Request) -> HTMLResponse:
    regiones_content = dedent("""
        <section id="regiones-module" style="display:grid; gap:14px;">
            <style>
                .reg-card {
                    background:#ffffff;
                    border:1px solid #dbe3ef;
                    border-radius:14px;
                    padding:14px;
                }
                .reg-title {
                    margin:0 0 10px;
                    font-size:1.02rem;
                    color:#0f172a;
                }
                .reg-grid {
                    display:grid;
                    grid-template-columns:repeat(3, minmax(0, 1fr));
                    gap:12px;
                }
                .reg-field {
                    display:flex;
                    flex-direction:column;
                    gap:6px;
                }
                .reg-field label {
                    font-size:0.85rem;
                    font-weight:700;
                    color:#334155;
                }
                .reg-field input,
                .reg-field textarea {
                    width:100%;
                    border:1px solid #cbd5e1;
                    border-radius:10px;
                    padding:10px;
                    color:#0f172a;
                    background:#ffffff;
                    font-size:0.95rem;
                }
                .reg-field textarea {
                    min-height:82px;
                    resize:vertical;
                }
                .reg-actions {
                    margin-top:12px;
                    display:flex;
                    gap:10px;
                    align-items:center;
                }
                .reg-actions button {
                    height:36px;
                    padding:0 14px;
                    border:1px solid #0f172a;
                    background:#0f172a;
                    color:#ffffff;
                    border-radius:10px;
                    font-weight:600;
                    cursor:pointer;
                }
                .reg-msg {
                    font-size:0.88rem;
                    color:#334155;
                }
                .reg-table {
                    width:100%;
                    border-collapse:collapse;
                }
                .reg-table th,
                .reg-table td {
                    border-bottom:1px solid #e2e8f0;
                    padding:10px;
                    text-align:left;
                    vertical-align:top;
                }
                .reg-table th {
                    color:#334155;
                    font-size:0.85rem;
                    text-transform:uppercase;
                    letter-spacing:.04em;
                }
                .reg-kanban {
                    display:grid;
                    grid-template-columns:repeat(3, minmax(0, 1fr));
                    gap:12px;
                }
                .reg-col {
                    border:1px solid #dbe3ef;
                    border-radius:12px;
                    background:#f8fafc;
                    padding:10px;
                }
                .reg-col h4 {
                    margin:0 0 10px;
                    font-size:0.9rem;
                    color:#0f172a;
                }
                .reg-item {
                    border:1px solid #dbe3ef;
                    border-radius:10px;
                    background:#ffffff;
                    padding:10px;
                    margin-bottom:8px;
                }
                .reg-item strong {
                    color:#0f172a;
                    display:block;
                    margin-bottom:4px;
                }
                .reg-item p {
                    margin:0;
                    color:#475569;
                    font-size:0.9rem;
                }
                @media (max-width: 980px) {
                    .reg-grid { grid-template-columns:1fr; }
                    .reg-kanban { grid-template-columns:1fr; }
                }
            </style>
            <div id="regiones-view"></div>
        </section>
        <script>
            (() => {
                const mount = document.getElementById('regiones-view');
                if (!mount) return;
                const data = [];
                let currentView = 'form';
                let editingIndex = -1;

                const escapeHtml = (value) => String(value || '').replace(/[&<>"']/g, (char) => (
                    { '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;' }[char] || char
                ));
                const normalizeRegion = (row) => {
                    const item = row && typeof row === 'object' ? row : {};
                    return {
                        nombre: String(item.nombre || '').trim(),
                        codigo: String(item.codigo || '').trim(),
                        descripcion: String(item.descripcion || '').trim(),
                    };
                };
                const replaceData = (rows) => {
                    const normalized = Array.isArray(rows) ? rows.map(normalizeRegion) : [];
                    data.splice(0, data.length, ...normalized);
                };
                const loadRegiones = async () => {
                    try {
                        const res = await fetch('/api/inicio/regiones');
                        const json = await res.json().catch(() => ({}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron cargar regiones');
                        replaceData(json?.data || []);
                    } catch (_error) {
                        replaceData([]);
                    }
                };
                const persistRegiones = async () => {
                    try {
                        const res = await fetch('/api/inicio/regiones', {
                            method: 'POST',
                            headers: { 'Content-Type': 'application/json' },
                            body: JSON.stringify({ data }),
                        });
                        const json = await res.json().catch(() => ({}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron guardar regiones');
                        replaceData(json?.data || []);
                        return true;
                    } catch (_error) {
                        return false;
                    }
                };

                const renderForm = () => {
                    const current = editingIndex >= 0 ? data[editingIndex] : { nombre: '', codigo: '', descripcion: '' };
                    mount.innerHTML = `
                        <article class="reg-card">
                            <h3 class="reg-title">Formulario de regiones</h3>
                            <form id="regiones-form">
                                <div class="reg-grid">
                                    <div class="reg-field">
                                        <label for="region-nombre">Nombre</label>
                                        <input id="region-nombre" type="text" value="${escapeHtml(current.nombre)}" required>
                                    </div>
                                    <div class="reg-field">
                                        <label for="region-codigo">Código</label>
                                        <input id="region-codigo" type="text" value="${escapeHtml(current.codigo)}" required>
                                    </div>
                                    <div class="reg-field">
                                        <label for="region-descripcion">Descripción</label>
                                        <textarea id="region-descripcion">${escapeHtml(current.descripcion)}</textarea>
                                    </div>
                                </div>
                                <div class="reg-actions">
                                    <button type="submit">${editingIndex >= 0 ? 'Actualizar' : 'Guardar'} región</button>
                                    <span class="reg-msg">${data.length} registro(s)</span>
                                </div>
                            </form>
                        </article>
                    `;
                    const form = document.getElementById('regiones-form');
                    form && form.addEventListener('submit', async (event) => {
                        event.preventDefault();
                        const nombre = (document.getElementById('region-nombre')?.value || '').trim();
                        const codigo = (document.getElementById('region-codigo')?.value || '').trim();
                        const descripcion = (document.getElementById('region-descripcion')?.value || '').trim();
                        if (!nombre || !codigo) return;
                        const payload = { nombre, codigo, descripcion };
                        if (editingIndex >= 0) {
                            data[editingIndex] = payload;
                        } else {
                            data.push(payload);
                        }
                        editingIndex = -1;
                        await persistRegiones();
                        renderForm();
                    });
                };

                const renderList = () => {
                    mount.innerHTML = `
                        <article class="reg-card">
                            <h3 class="reg-title">Lista de regiones</h3>
                            <table class="reg-table">
                                <thead>
                                    <tr>
                                        <th>Nombre</th>
                                        <th>Código</th>
                                        <th>Descripción</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    ${data.length ? data.map((row) => `
                                        <tr>
                                            <td>${escapeHtml(row.nombre)}</td>
                                            <td>${escapeHtml(row.codigo)}</td>
                                            <td>${escapeHtml(row.descripcion)}</td>
                                        </tr>
                                    `).join('') : `
                                        <tr><td colspan="3" style="color:#64748b;">Sin registros.</td></tr>
                                    `}
                                </tbody>
                            </table>
                        </article>
                    `;
                };

                const laneName = (nombre) => {
                    const first = (nombre || '').trim().charAt(0).toUpperCase();
                    if (!first) return 'Sin clasificar';
                    if (first <= 'H') return 'A - H';
                    if (first <= 'P') return 'I - P';
                    return 'Q - Z';
                };

                const renderKanban = () => {
                    const lanes = { 'A - H': [], 'I - P': [], 'Q - Z': [], 'Sin clasificar': [] };
                    data.forEach((row, index) => { lanes[laneName(row.nombre)].push({ ...row, index }); });
                    mount.innerHTML = `
                        <article class="reg-card">
                            <h3 class="reg-title">Kanban de regiones</h3>
                            <div class="reg-kanban">
                                ${Object.keys(lanes).map((key) => `
                                    <div class="reg-col">
                                        <h4>${key}</h4>
                                        ${lanes[key].length ? lanes[key].map((row) => `
                                            <article class="reg-item">
                                                <strong>${escapeHtml(row.nombre)}</strong>
                                                <p>${escapeHtml(row.codigo)}</p>
                                                <p>${escapeHtml(row.descripcion)}</p>
                                            </article>
                                        `).join('') : '<p class="reg-msg">Sin registros.</p>'}
                                    </div>
                                `).join('')}
                            </div>
                        </article>
                    `;
                };

                const render = (view) => {
                    currentView = ['form', 'list', 'kanban'].includes(view) ? view : 'form';
                    if (currentView === 'list') return renderList();
                    if (currentView === 'kanban') return renderKanban();
                    return renderForm();
                };

                document.addEventListener('backend-view-change', (event) => {
                    const view = event.detail?.view;
                    if (!view) return;
                    render(view);
                });

                (async () => {
                    await loadRegiones();
                    render('form');
                })();
            })();
        </script>
    """)
    return render_backend_page(
        request,
        title="Regiones",
        description="Registro y visualización de regiones.",
        content=regiones_content,
        hide_floating_actions=True,
        show_page_header=True,
        view_buttons=[
            {"label": "Form", "icon": "/templates/icon/formulario.svg", "view": "form", "active": True},
            {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list"},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
        ],
    )


def _render_sucursales_page(request: Request) -> HTMLResponse:
    preliminares = _load_datos_preliminares_store()
    current_year = datetime.now().year
    try:
        base_year = int((preliminares.get("primer_anio_proyeccion") or "").strip() or current_year)
    except (TypeError, ValueError):
        base_year = current_year
    try:
        projection_years = int((preliminares.get("anios_proyeccion") or "").strip() or 3)
    except (TypeError, ValueError):
        projection_years = 3
    projection_years = max(1, min(projection_years, 10))
    column_offsets = [-4, -3, -2, -1, 0] + list(range(1, projection_years))

    def _header_label(offset: int) -> str:
        if offset < 0:
            return f"{offset} ({base_year + offset})"
        if offset == 0:
            return f"Año actual ({base_year})"
        return f"+{offset} ({base_year + offset})"

    header_cells = "".join(f"<th>{escape(_header_label(offset))}</th>" for offset in column_offsets)
    rubros = [
        "Socios",
        "Menores Ahorradores",
        "Ahorro menores",
        "Captación a la vista",
        "Inversión",
        "Cartera de préstamos",
        "Cartera vencida",
    ]
    results_rows = []
    for row_idx, rubro in enumerate(rubros, start=1):
        inputs = "".join(
            (
                f'<td><input class="suc-result-input" type="number" step="0.01" min="0" '
                f'name="suc_result_{row_idx}_{offset}" placeholder="0.00"></td>'
            )
            for offset in column_offsets
        )
        results_rows.append(
            f"""
            <tr>
                <td class="suc-result-rubro">{escape(rubro)}</td>
                {inputs}
            </tr>
            """
        )
    resultados_rows_html = "".join(results_rows)
    activo_fijo_catalog = [
        {"rubro": "Terrenos", "years": 0},
        {"rubro": "Construcciones", "years": 20},
        {"rubro": "Construcciones en proceso", "years": 5},
        {"rubro": "Equipo de transporte", "years": 4},
        {"rubro": "Equipo de cómputo", "years": 3},
        {"rubro": "Mobiliario", "years": 3},
        {"rubro": "Otras propiedades, mobiliario y equipo", "years": 2},
    ]
    activo_fijo_catalog_json = json.dumps(activo_fijo_catalog, ensure_ascii=False)

    sucursales_content = dedent(f"""
        <section id="sucursales-module" style="display:grid; gap:14px;">
            <style>
                .suc-tabs {{
                    display:flex;
                    flex-wrap:wrap;
                    gap:10px;
                    border-bottom:1px solid #cbd5e1;
                    padding-bottom:8px;
                }}
                .suc-tab-btn {{
                    display:inline-flex;
                    align-items:center;
                    gap:8px;
                    border:1px solid transparent;
                    border-bottom:3px solid transparent;
                    border-radius:10px;
                    background:transparent;
                    padding:10px 12px;
                    color:#334155;
                    font-weight:700;
                    cursor:pointer;
                }}
                .suc-tab-btn img {{
                    width:18px;
                    height:18px;
                    object-fit:contain;
                }}
                .suc-tab-btn.active {{
                    border-color: var(--sidebar-bottom, #0f172a);
                    border-bottom-color: var(--sidebar-bottom, #0f172a);
                    background: var(--sidebar-text, #ffffff);
                    color: var(--sidebar-bottom, #0f172a);
                }}
                .suc-tab-panel {{
                    display:none;
                }}
                .suc-tab-panel.active {{
                    display:block;
                }}
                .suc-card {{
                    background:#ffffff;
                    border:1px solid #dbe3ef;
                    border-radius:14px;
                    padding:14px;
                }}
                .suc-title {{
                    margin:0 0 10px;
                    font-size:1.02rem;
                    color:#0f172a;
                }}
                .suc-grid {{
                    display:grid;
                    grid-template-columns:repeat(4, minmax(0, 1fr));
                    gap:12px;
                }}
                .suc-field {{
                    display:flex;
                    flex-direction:column;
                    gap:6px;
                }}
                .suc-field label {{
                    font-size:0.85rem;
                    font-weight:700;
                    color:#334155;
                }}
                .suc-field input,
                .suc-field select,
                .suc-field textarea {{
                    width:100%;
                    border:1px solid #cbd5e1;
                    border-radius:10px;
                    padding:10px;
                    color:#0f172a;
                    background:#ffffff;
                    font-size:0.95rem;
                }}
                .suc-field textarea {{
                    min-height:82px;
                    resize:vertical;
                }}
                .suc-actions {{
                    margin-top:12px;
                    display:flex;
                    gap:10px;
                    align-items:center;
                }}
                .suc-actions button {{
                    height:36px;
                    padding:0 14px;
                    border:1px solid #0f172a;
                    background:#0f172a;
                    color:#ffffff;
                    border-radius:10px;
                    font-weight:600;
                    cursor:pointer;
                }}
                .suc-actions .suc-btn-alt {{
                    border:1px solid #cbd5e1;
                    background:#ffffff;
                    color:#0f172a;
                }}
                .suc-msg {{
                    font-size:0.88rem;
                    color:#334155;
                }}
                .suc-table {{
                    width:100%;
                    border-collapse:collapse;
                }}
                .suc-table th,
                .suc-table td {{
                    border-bottom:1px solid #e2e8f0;
                    padding:10px;
                    text-align:left;
                    vertical-align:top;
                }}
                .suc-table th {{
                    color:#334155;
                    font-size:0.85rem;
                    text-transform:uppercase;
                    letter-spacing:.04em;
                }}
                .suc-kanban {{
                    display:grid;
                    grid-template-columns:repeat(3, minmax(0, 1fr));
                    gap:12px;
                }}
                .suc-col {{
                    border:1px solid #dbe3ef;
                    border-radius:12px;
                    background:#f8fafc;
                    padding:10px;
                }}
                .suc-col h4 {{
                    margin:0 0 10px;
                    font-size:0.9rem;
                    color:#0f172a;
                }}
                .suc-item {{
                    border:1px solid #dbe3ef;
                    border-radius:10px;
                    background:#ffffff;
                    padding:10px;
                    margin-bottom:8px;
                    cursor:pointer;
                }}
                .suc-item strong {{
                    color:#0f172a;
                    display:block;
                    margin-bottom:4px;
                }}
                .suc-item p {{
                    margin:0;
                    color:#475569;
                    font-size:0.9rem;
                }}
                .suc-results-wrap {{
                    overflow-x:auto;
                }}
                .suc-results-table {{
                    width:100%;
                    min-width:980px;
                    border-collapse:collapse;
                    border-spacing:0;
                }}
                .suc-results-table thead th {{
                    text-align:left;
                    font-size:13px;
                    letter-spacing:.08em;
                    text-transform:uppercase;
                    color:rgba(15,23,42,.75);
                    background:linear-gradient(180deg, rgba(255,255,255,.92), rgba(255,255,255,.74));
                    border-bottom:1px solid rgba(15,23,42,.10);
                    border-right:1px solid rgba(15,23,42,.10);
                    padding:14px 12px;
                    white-space:nowrap;
                }}
                .suc-results-table thead th:last-child {{
                    border-right:0;
                }}
                .suc-results-table tbody td {{
                    border-bottom:1px solid rgba(15,23,42,.08);
                    border-right:1px solid rgba(15,23,42,.10);
                    background:#ffffff;
                    padding:10px 12px;
                    vertical-align:middle;
                }}
                .suc-results-table tbody td:last-child {{
                    border-right:0;
                }}
                .suc-results-table tbody tr:nth-child(even) td {{
                    background:#ecfdf3;
                }}
                .suc-results-table tbody tr:hover td {{
                    background:#dcfce7;
                }}
                .suc-result-rubro {{
                    font-weight:700;
                    color:#0f172a;
                    white-space:nowrap;
                }}
                .suc-result-input {{
                    width:100%;
                    height:34px;
                    border:1px solid #cbd5e1;
                    border-radius:8px;
                    padding:0 10px;
                    background:#ffffff;
                    color:#0f172a;
                    text-align:right;
                    font-variant-numeric:tabular-nums;
                }}
                .suc-af-toolbar {{
                    display:flex;
                    align-items:center;
                    gap:10px;
                    margin-bottom:10px;
                    flex-wrap:wrap;
                }}
                .suc-af-btn {{
                    height:36px;
                    padding:0 12px;
                    border:1px solid #0f172a;
                    border-radius:8px;
                    background:#0f172a;
                    color:#ffffff;
                    font-weight:600;
                    cursor:pointer;
                }}
                .suc-af-note {{
                    font-size:0.82rem;
                    color:#64748b;
                }}
                .suc-af-input,
                .suc-af-select {{
                    width:100%;
                    height:34px;
                    border:1px solid #cbd5e1;
                    border-radius:8px;
                    padding:0 8px;
                    background:#ffffff;
                    color:#0f172a;
                }}
                .suc-af-input.num {{
                    text-align:right;
                    font-variant-numeric:tabular-nums;
                }}
                @media (max-width: 1100px) {{
                    .suc-grid {{ grid-template-columns:repeat(2, minmax(0, 1fr)); }}
                }}
                @media (max-width: 980px) {{
                    .suc-grid {{ grid-template-columns:1fr; }}
                    .suc-kanban {{ grid-template-columns:1fr; }}
                }}
            </style>
            <div id="sucursales-view"></div>
        </section>
        <script>
            (() => {{
                const mount = document.getElementById('sucursales-view');
                if (!mount) return;
                const data = [];
                const activoFijoCatalog = {activo_fijo_catalog_json};
                const projectionYears = {projection_years};
                const baseYear = {base_year};
                const purchaseYearOptions = Array.from({{ length: projectionYears }}, (_, idx) => baseYear + idx);
                const monthOptions = [
                    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
                ];
                const statusOptions = ["Solicitado", "Autorizado", "Comprado"];
                const activoFijoRowsData = [];
                const regionCatalog = [];
                const regionRowsData = [];
                let currentView = 'list';
                let editingIndex = -1;
                let formTab = 'captura';

                const escapeHtml = (value) => String(value || '').replace(/[&<>"']/g, (char) => (
                    {{ '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;' }}[char] || char
                ));
                const normalizeSucursal = (row) => {{
                    const item = row && typeof row === 'object' ? row : {{}};
                    return {{
                        nombre: String(item.nombre || '').trim(),
                        region: String(item.region || '').trim(),
                        codigo: String(item.codigo || '').trim(),
                        descripcion: String(item.descripcion || '').trim(),
                    }};
                }};
                const replaceData = (rows) => {{
                    const normalized = Array.isArray(rows) ? rows.map(normalizeSucursal) : [];
                    data.splice(0, data.length, ...normalized);
                }};
                const normalizeRegionName = (value) => String(value || '').trim();
                const replaceRegionCatalog = (rows) => {{
                    const source = Array.isArray(rows) ? rows : [];
                    const cleanedRows = source
                        .filter((row) => row && typeof row === 'object')
                        .map((row) => ({{
                            nombre: normalizeRegionName(row.nombre),
                            codigo: String(row.codigo || '').trim(),
                            descripcion: String(row.descripcion || '').trim(),
                        }}))
                        .filter((row) => Boolean(row.nombre));
                    regionRowsData.splice(0, regionRowsData.length, ...cleanedRows);
                    const values = cleanedRows.map((row) => row.nombre);
                    const unique = Array.from(new Set(values)).sort((a, b) => a.localeCompare(b, 'es', {{ sensitivity: 'base' }}));
                    regionCatalog.splice(0, regionCatalog.length, ...unique);
                }};
                const loadSucursales = async () => {{
                    try {{
                        const res = await fetch('/api/inicio/sucursales');
                        const json = await res.json().catch(() => ({{}}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron cargar sucursales');
                        replaceData(json?.data || []);
                    }} catch (_error) {{
                        replaceData([]);
                    }}
                }};
                const persistSucursales = async () => {{
                    try {{
                        const res = await fetch('/api/inicio/sucursales', {{
                            method: 'POST',
                            headers: {{ 'Content-Type': 'application/json' }},
                            body: JSON.stringify({{ data }}),
                        }});
                        const json = await res.json().catch(() => ({{}}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron guardar sucursales');
                        replaceData(json?.data || []);
                        return true;
                    }} catch (_error) {{
                        return false;
                    }}
                }};
                const loadRegionesCatalog = async () => {{
                    try {{
                        const res = await fetch('/api/inicio/regiones');
                        const json = await res.json().catch(() => ({{}}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron cargar regiones');
                        replaceRegionCatalog(json?.data || []);
                    }} catch (_error) {{
                        replaceRegionCatalog([]);
                    }}
                }};
                const persistRegionesCatalog = async () => {{
                    try {{
                        const res = await fetch('/api/inicio/regiones', {{
                            method: 'POST',
                            headers: {{ 'Content-Type': 'application/json' }},
                            body: JSON.stringify({{ data: regionRowsData }}),
                        }});
                        const json = await res.json().catch(() => ({{}}));
                        if (!res.ok || json?.success === false) throw new Error('No se pudieron guardar regiones');
                        replaceRegionCatalog(json?.data || []);
                        return true;
                    }} catch (_error) {{
                        return false;
                    }}
                }};

                const renderForm = () => {{
                    const current = editingIndex >= 0
                        ? data[editingIndex]
                        : {{ nombre: '', region: '', codigo: '', descripcion: '' }};
                    const currentRegion = normalizeRegionName(current.region);
                    const regionValues = Array.from(new Set([
                        ...regionCatalog,
                        ...(currentRegion ? [currentRegion] : []),
                    ]));
                    const regionOptionsHtml = regionValues
                        .sort((a, b) => a.localeCompare(b, 'es', {{ sensitivity: 'base' }}))
                        .map((name) => `<option value="${{escapeHtml(name)}}" ${{name === currentRegion ? 'selected' : ''}}>${{escapeHtml(name)}}</option>`)
                        .join('');
                    const isCaptura = formTab === 'captura';
                    const isResultados = formTab === 'resultados';
                    const isActivoFijo = formTab === 'activo-fijo';
                    const isReparaciones = formTab === 'reparaciones';
                    mount.innerHTML = `
                        <article class="suc-card">
                            <h3 class="suc-title">Formulario de sucursales</h3>
                            <div class="suc-tabs" role="tablist" aria-label="Control por sucursal">
                                <button type="button" class="suc-tab-btn ${{isCaptura ? 'active' : ''}}" data-suc-form-tab="captura" aria-selected="${{isCaptura ? 'true' : 'false'}}">Captura</button>
                                <button type="button" class="suc-tab-btn ${{isResultados ? 'active' : ''}}" data-suc-form-tab="resultados" aria-selected="${{isResultados ? 'true' : 'false'}}">
                                    <img src="/templates/icon/resultados.svg" alt="">
                                    Resultados
                                </button>
                                <button type="button" class="suc-tab-btn ${{isActivoFijo ? 'active' : ''}}" data-suc-form-tab="activo-fijo" aria-selected="${{isActivoFijo ? 'true' : 'false'}}">
                                    <img src="/templates/icon/activo_fijo.svg" alt="">
                                    Compras de activo fijo
                                </button>
                                <button type="button" class="suc-tab-btn ${{isReparaciones ? 'active' : ''}}" data-suc-form-tab="reparaciones" aria-selected="${{isReparaciones ? 'true' : 'false'}}">
                                    <img src="/templates/icon/reparaciones.svg" alt="">
                                    Reparaciones
                                </button>
                            </div>
                            <div class="suc-tab-panel ${{isCaptura ? 'active' : ''}}" data-suc-form-panel="captura">
                            <form id="sucursales-form">
                                <div class="suc-grid">
                                    <div class="suc-field">
                                        <label for="sucursal-nombre">Nombre</label>
                                        <input id="sucursal-nombre" type="text" value="${{escapeHtml(current.nombre)}}" required>
                                    </div>
                                    <div class="suc-field">
                                        <label for="sucursal-region-select">Región</label>
                                        <select id="sucursal-region-select" required>
                                            <option value="">Seleccione región</option>
                                            ${{regionOptionsHtml}}
                                            <option value="__new__">+ Agregar región</option>
                                        </select>
                                        <input id="sucursal-region-new" type="text" placeholder="Nueva región" style="display:none; margin-top:6px;">
                                    </div>
                                    <div class="suc-field">
                                        <label for="sucursal-codigo">Código</label>
                                        <input id="sucursal-codigo" type="text" value="${{escapeHtml(current.codigo)}}" required>
                                    </div>
                                    <div class="suc-field">
                                        <label for="sucursal-descripcion">Descripción</label>
                                        <textarea id="sucursal-descripcion">${{escapeHtml(current.descripcion)}}</textarea>
                                    </div>
                                </div>
                                <div class="suc-actions">
                                    <button type="button" class="suc-btn-alt" id="suc-btn-new">Nuevo</button>
                                    <button type="button" class="suc-btn-alt" id="suc-btn-edit">Editar</button>
                                    <button type="submit" id="suc-btn-save">Guardar</button>
                                    <button type="button" class="suc-btn-alt" id="suc-btn-delete">Eliminar</button>
                                    <span class="suc-msg" id="suc-form-msg">${{data.length}} registro(s)</span>
                                </div>
                            </form>
                            </div>
                            <div class="suc-tab-panel ${{isResultados ? 'active' : ''}}" data-suc-form-panel="resultados">
                                <h3 class="suc-title">Resultados</h3>
                                <div class="suc-results-wrap">
                                    <table class="suc-results-table">
                                        <thead>
                                            <tr>
                                                <th>Rubro</th>
                                                {header_cells}
                                            </tr>
                                        </thead>
                                        <tbody>
                                            {resultados_rows_html}
                                        </tbody>
                                    </table>
                                </div>
                            </div>
                            <div class="suc-tab-panel ${{isActivoFijo ? 'active' : ''}}" data-suc-form-panel="activo-fijo">
                                <h3 class="suc-title">Activo fijo</h3>
                                <div class="suc-af-toolbar">
                                    <button type="button" class="suc-af-btn" id="suc-af-add-btn">Compra de activo fijo</button>
                                    <span class="suc-af-note">Procedimiento de autorización pendiente.</span>
                                </div>
                                <div class="suc-results-wrap">
                                    <table class="suc-results-table">
                                        <thead>
                                            <tr>
                                                <th>Código</th>
                                                <th>Artículo</th>
                                                <th>Rubro</th>
                                                <th>Precio</th>
                                                <th>Año de compra</th>
                                                <th>Mes de compra</th>
                                                <th>Status</th>
                                            </tr>
                                        </thead>
                                        <tbody id="suc-af-rows"></tbody>
                                    </table>
                                </div>
                            </div>
                            <div class="suc-tab-panel ${{isReparaciones ? 'active' : ''}}" data-suc-form-panel="reparaciones">
                                <h3 class="suc-title">Reparaciones</h3>
                                <p class="suc-msg">Aquí se llevará el registro de las solicitudes de reparación de la sucursal.</p>
                                <p class="suc-msg">Lógica y código pendientes.</p>
                            </div>
                        </article>
                    `;
                    const tabButtons = Array.from(mount.querySelectorAll('[data-suc-form-tab]'));
                    tabButtons.forEach((button) => {{
                        button.addEventListener('click', () => {{
                            formTab = button.getAttribute('data-suc-form-tab') || 'captura';
                            renderForm();
                        }});
                    }});
                    const form = document.getElementById('sucursales-form');
                    const formMsg = document.getElementById('suc-form-msg');
                    const newBtn = document.getElementById('suc-btn-new');
                    const editBtn = document.getElementById('suc-btn-edit');
                    const deleteBtn = document.getElementById('suc-btn-delete');
                    const regionSelect = document.getElementById('sucursal-region-select');
                    const regionNewInput = document.getElementById('sucursal-region-new');
                    const upsertRegionOption = (name) => {{
                        if (!regionSelect) return;
                        const normalized = normalizeRegionName(name);
                        if (!normalized) return;
                        const options = Array.from(regionSelect.querySelectorAll('option'));
                        const exists = options.some((opt) => normalizeRegionName(opt.value) === normalized);
                        if (!exists) {{
                            const newOption = document.createElement('option');
                            newOption.value = normalized;
                            newOption.textContent = normalized;
                            const newMarker = regionSelect.querySelector('option[value="__new__"]');
                            if (newMarker) {{
                                regionSelect.insertBefore(newOption, newMarker);
                            }} else {{
                                regionSelect.appendChild(newOption);
                            }}
                        }}
                        regionSelect.value = normalized;
                    }};
                    const syncRegionNewVisibility = () => {{
                        if (!regionSelect || !regionNewInput) return;
                        const isNew = regionSelect.value === '__new__';
                        regionNewInput.style.display = isNew ? 'block' : 'none';
                        regionNewInput.required = isNew;
                    }};
                    regionSelect && regionSelect.addEventListener('change', syncRegionNewVisibility);
                    syncRegionNewVisibility();
                    const readValues = () => {{
                        const nombre = (document.getElementById('sucursal-nombre')?.value || '').trim();
                        const selectedRegion = (regionSelect?.value || '').trim();
                        const newRegion = (regionNewInput?.value || '').trim();
                        const region = selectedRegion === '__new__' ? newRegion : selectedRegion;
                        const codigo = (document.getElementById('sucursal-codigo')?.value || '').trim();
                        const descripcion = (document.getElementById('sucursal-descripcion')?.value || '').trim();
                        return {{ nombre, region, codigo, descripcion }};
                    }};
                    const setFormMsg = (text) => {{
                        if (!formMsg) return;
                        formMsg.textContent = text || `${{data.length}} registro(s)`;
                    }};
                    const resetCapturaInputs = () => {{
                        const nombreInput = document.getElementById('sucursal-nombre');
                        const codigoInput = document.getElementById('sucursal-codigo');
                        const descripcionInput = document.getElementById('sucursal-descripcion');
                        if (nombreInput) nombreInput.value = '';
                        if (codigoInput) codigoInput.value = '';
                        if (descripcionInput) descripcionInput.value = '';
                        if (regionSelect) regionSelect.value = '';
                        if (regionNewInput) regionNewInput.value = '';
                        syncRegionNewVisibility();
                    }};
                    form && form.addEventListener('submit', async (event) => {{
                        event.preventDefault();
                        const {{ nombre, region, codigo, descripcion }} = readValues();
                        if (!nombre || !region || !codigo) return;
                        const normalizedRegion = normalizeRegionName(region);
                        let regionSaved = true;
                        if (normalizedRegion && !regionCatalog.includes(normalizedRegion)) {{
                            regionRowsData.push({{ nombre: normalizedRegion, codigo: '', descripcion: '' }});
                            regionSaved = await persistRegionesCatalog();
                            if (!regionSaved) {{
                                setFormMsg('No se pudo guardar la nueva región.');
                                return;
                            }}
                            upsertRegionOption(normalizedRegion);
                        }}
                        const payload = {{ nombre, region: normalizedRegion, codigo, descripcion }};
                        if (editingIndex >= 0) {{
                            data[editingIndex] = payload;
                        }} else {{
                            data.push(payload);
                        }}
                        editingIndex = -1;
                        const saved = await persistSucursales();
                        if (saved) resetCapturaInputs();
                        setFormMsg(saved ? `Sucursal guardada. Total: ${{data.length}}` : 'No se pudo guardar en la BD/store.');
                    }});
                    newBtn && newBtn.addEventListener('click', () => {{
                        editingIndex = -1;
                        renderForm();
                    }});
                    editBtn && editBtn.addEventListener('click', () => {{
                        const {{ codigo }} = readValues();
                        if (!codigo) {{
                            setFormMsg('Capture el código para editar.');
                            return;
                        }}
                        const idx = data.findIndex((row) => String(row.codigo).trim() === codigo);
                        if (idx < 0) {{
                            setFormMsg('No se encontró sucursal con ese código.');
                            return;
                        }}
                        editingIndex = idx;
                        renderForm();
                        setTimeout(() => {{
                            const msg = document.getElementById('suc-form-msg');
                            if (msg) msg.textContent = `Editando sucursal: ${{codigo}}`;
                        }}, 0);
                    }});
                    deleteBtn && deleteBtn.addEventListener('click', async () => {{
                        const {{ codigo }} = readValues();
                        const idx = editingIndex >= 0
                            ? editingIndex
                            : data.findIndex((row) => String(row.codigo).trim() === codigo);
                        if (idx < 0) {{
                            setFormMsg('No hay sucursal para eliminar.');
                            return;
                        }}
                        data.splice(idx, 1);
                        editingIndex = -1;
                        const saved = await persistSucursales();
                        if (saved) resetCapturaInputs();
                        setFormMsg(saved ? `Sucursal eliminada. Total: ${{data.length}}` : 'No se pudo guardar eliminación en la BD/store.');
                    }});
                    const afAddBtn = document.getElementById('suc-af-add-btn');
                    const afRowsEl = document.getElementById('suc-af-rows');
                    afAddBtn && afAddBtn.addEventListener('click', addActivoFijoRow);
                    afRowsEl && afRowsEl.addEventListener('input', (event) => {{
                        const target = event.target;
                        if (!(target instanceof HTMLElement)) return;
                        const rowElem = target.closest('tr[data-af-row]');
                        if (!rowElem) return;
                        const rowIndex = Number(rowElem.getAttribute('data-af-row'));
                        const row = activoFijoRowsData[rowIndex];
                        if (!row) return;
                        const field = target.getAttribute('data-field');
                        if (!field) return;
                        row[field] = target.value;
                    }});
                    afRowsEl && afRowsEl.addEventListener('change', (event) => {{
                        const target = event.target;
                        if (!(target instanceof HTMLElement)) return;
                        const rowElem = target.closest('tr[data-af-row]');
                        if (!rowElem) return;
                        const rowIndex = Number(rowElem.getAttribute('data-af-row'));
                        const row = activoFijoRowsData[rowIndex];
                        if (!row) return;
                        const field = target.getAttribute('data-field');
                        if (!field) return;
                        row[field] = target.value;
                    }});
                    renderActivoFijoRows();
                }};

                const renderList = () => {{
                    mount.innerHTML = `
                        <article class="suc-card">
                            <h3 class="suc-title">Lista de sucursales</h3>
                            <table class="suc-table">
                                <thead>
                                    <tr>
                                        <th>Nombre</th>
                                        <th>Región</th>
                                        <th>Código</th>
                                        <th>Descripción</th>
                                    </tr>
                                </thead>
                                <tbody>
                                    ${{data.length ? data.map((row) => `
                                        <tr class="suc-open-form-row" data-row-code="${{escapeHtml(row.codigo)}}">
                                            <td>${{escapeHtml(row.nombre)}}</td>
                                            <td>${{escapeHtml(row.region)}}</td>
                                            <td>${{escapeHtml(row.codigo)}}</td>
                                            <td>${{escapeHtml(row.descripcion)}}</td>
                                        </tr>
                                    `).join('') : `
                                        <tr><td colspan="4" style="color:#64748b;">Sin registros.</td></tr>
                                    `}}
                                </tbody>
                            </table>
                        </article>
                    `;
                }};

                const laneName = (region) => {{
                    const key = (region || '').trim();
                    return key || 'Sin región';
                }};

                const renderKanban = () => {{
                    const lanes = {{}};
                    data.forEach((row) => {{
                        const lane = laneName(row.region);
                        if (!lanes[lane]) lanes[lane] = [];
                        lanes[lane].push(row);
                    }});
                    if (!Object.keys(lanes).length) lanes['Sin región'] = [];
                    mount.innerHTML = `
                        <article class="suc-card">
                            <h3 class="suc-title">Kanban de sucursales</h3>
                            <div class="suc-kanban">
                                ${{Object.keys(lanes).map((key) => `
                                    <div class="suc-col">
                                        <h4>${{escapeHtml(key)}}</h4>
                                        ${{lanes[key].length ? lanes[key].map((row) => `
                                            <article class="suc-item" data-row-code="${{escapeHtml(row.codigo)}}">
                                                <strong>${{escapeHtml(row.nombre)}}</strong>
                                                <p>${{escapeHtml(row.codigo)}}</p>
                                                <p>${{escapeHtml(row.descripcion)}}</p>
                                            </article>
                                        `).join('') : '<p class="suc-msg">Sin registros.</p>'}}
                                    </div>
                                `).join('')}}
                            </div>
                        </article>
                    `;
                }};

                const render = (view) => {{
                    currentView = ['form', 'list', 'kanban'].includes(view) ? view : 'form';
                    if (currentView === 'list') return renderList();
                    if (currentView === 'kanban') return renderKanban();
                    return renderForm();
                }};
                const openFormByCode = (codigo) => {{
                    const targetCode = String(codigo || '').trim();
                    if (!targetCode) return;
                    const idx = data.findIndex((row) => String(row.codigo || '').trim() === targetCode);
                    if (idx < 0) return;
                    editingIndex = idx;
                    formTab = 'captura';
                    render('form');
                    const msg = document.getElementById('suc-form-msg');
                    if (msg) msg.textContent = `Editando sucursal: ${{targetCode}}`;
                }};

                const normalizeSucursalCode = () => {{
                    const raw = String(data[0]?.codigo || '').trim();
                    if (!raw) return "001";
                    const digits = raw.replace(/\\D+/g, "");
                    if (digits) return digits;
                    const normalized = raw.toUpperCase().replace(/[^A-Z0-9]+/g, "");
                    return normalized || "001";
                }};

                const generateActivoFijoCode = () => {{
                    const branchCode = normalizeSucursalCode();
                    const sequence = String(activoFijoRowsData.length + 1).padStart(3, "0");
                    return `${{branchCode}}-${{sequence}}`;
                }};

                const rubroSelectOptions = (selected) => activoFijoCatalog.map((item) => {{
                    const isSelected = item.rubro === selected ? "selected" : "";
                    return `<option value="${{escapeHtml(item.rubro)}}" ${{isSelected}}>${{escapeHtml(item.rubro)}}</option>`;
                }}).join("");
                const yearSelectOptions = (selected) => purchaseYearOptions.map((year) => {{
                    const isSelected = Number(selected) === Number(year) ? "selected" : "";
                    return `<option value="${{year}}" ${{isSelected}}>${{year}}</option>`;
                }}).join("");
                const monthSelectOptions = (selected) => monthOptions.map((month, idx) => {{
                    const value = idx + 1;
                    const isSelected = Number(selected) === value ? "selected" : "";
                    return `<option value="${{value}}" ${{isSelected}}>${{month}}</option>`;
                }}).join("");
                const statusSelectOptions = (selected) => statusOptions.map((status) => {{
                    const isSelected = status === selected ? "selected" : "";
                    return `<option value="${{escapeHtml(status)}}" ${{isSelected}}>${{escapeHtml(status)}}</option>`;
                }}).join("");

                const renderActivoFijoRows = () => {{
                    const activoFijoRows = document.getElementById('suc-af-rows');
                    if (!activoFijoRows) return;
                    if (!activoFijoRowsData.length) {{
                        activoFijoRows.innerHTML = '<tr><td colspan="7" style="color:#64748b;">Sin registros de compras.</td></tr>';
                        return;
                    }}
                    activoFijoRows.innerHTML = activoFijoRowsData.map((row, idx) => `
                        <tr data-af-row="${{idx}}">
                            <td><input class="suc-af-input" type="text" data-field="code" value="${{escapeHtml(row.code)}}" readonly></td>
                            <td><input class="suc-af-input" type="text" data-field="article" value="${{escapeHtml(row.article)}}"></td>
                            <td><select class="suc-af-select" data-field="rubro">${{rubroSelectOptions(row.rubro)}}</select></td>
                            <td><input class="suc-af-input num" type="number" min="0" step="0.01" data-field="price" value="${{escapeHtml(row.price)}}"></td>
                            <td><select class="suc-af-select" data-field="year">${{yearSelectOptions(row.year)}}</select></td>
                            <td><select class="suc-af-select" data-field="month">${{monthSelectOptions(row.month)}}</select></td>
                            <td><select class="suc-af-select" data-field="status">${{statusSelectOptions(row.status)}}</select></td>
                        </tr>
                    `).join("");
                }};

                const addActivoFijoRow = () => {{
                    const firstRubro = activoFijoCatalog[0]?.rubro || "";
                    const newRow = {{
                        code: generateActivoFijoCode(),
                        article: "",
                        rubro: firstRubro,
                        price: "",
                        year: purchaseYearOptions[0] || baseYear,
                        month: 1,
                        status: "Solicitado",
                    }};
                    activoFijoRowsData.push(newRow);
                    renderActivoFijoRows();
                }};

                document.addEventListener('backend-view-change', (event) => {{
                    const view = event.detail?.view;
                    if (!view) return;
                    render(view);
                }});
                mount.addEventListener('click', (event) => {{
                    const target = event.target;
                    if (!(target instanceof HTMLElement)) return;
                    const listRow = target.closest('.suc-open-form-row');
                    if (listRow instanceof HTMLElement) {{
                        openFormByCode(listRow.getAttribute('data-row-code') || '');
                        return;
                    }}
                    const kanbanCard = target.closest('.suc-item');
                    if (kanbanCard instanceof HTMLElement) {{
                        const codeText = (kanbanCard.getAttribute('data-row-code') || '').trim();
                        openFormByCode(codeText);
                    }}
                }});

                (async () => {{
                    await loadRegionesCatalog();
                    await loadSucursales();
                    render('list');
                }})();
            }})();
        </script>
    """)
    return render_backend_page(
        request,
        title="Sucursales",
        description="Registro y visualización de sucursales.",
        content=sucursales_content,
        hide_floating_actions=True,
        show_page_header=True,
        view_buttons=[
            {"label": "Form", "icon": "/templates/icon/formulario.svg", "view": "form"},
            {"label": "Lista", "icon": "/templates/icon/list.svg", "view": "list", "active": True},
            {"label": "Kanban", "icon": "/templates/icon/kanban.svg", "view": "kanban"},
        ],
    )


@app.get("/diagnostico", response_class=HTMLResponse)
def diagnostico_page(request: Request):
    return _render_blank_management_screen(request, "Diagnóstico")


@app.get("/diagnostico/foda", response_class=HTMLResponse)
def diagnostico_foda_page(request: Request):
    return render_backend_page(
        request,
        title="FODA",
        description="Identifica factores internos y externos para el diagnóstico estratégico.",
        content=FODA_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/diagnostico/pestel", response_class=HTMLResponse)
def diagnostico_pestel_page(request: Request):
    return render_backend_page(
        request,
        title="PESTEL",
        description="Analiza factores externos que impactan la estrategia institucional.",
        content=PESTEL_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/diagnostico/porter", response_class=HTMLResponse)
def diagnostico_porter_page(request: Request):
    return render_backend_page(
        request,
        title="PORTER",
        description="Evalúa las cinco fuerzas competitivas para priorizar decisiones estratégicas.",
        content=PORTER_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/diagnostico/percepcion-cliente", response_class=HTMLResponse)
def diagnostico_percepcion_cliente_page(request: Request):
    return render_backend_page(
        request,
        title="Percepción del cliente",
        description="Monitorea feedback para detectar fortalezas y brechas del servicio.",
        content=PERCEPCION_CLIENTE_HTML,
        hide_floating_actions=True,
        show_page_header=True,
    )


BACKEND_ROLES_PERMISSIONS_URL = os.environ.get(
    "BACKEND_ROLES_PERMISSIONS_URL",
    "http://localhost:8000/api/v1/personalizacion/roles-permisos"
)


async def _fetch_roles_permissions_view() -> str:
    async with httpx.AsyncClient() as client:
        try:
            resp = await client.get(BACKEND_ROLES_PERMISSIONS_URL, timeout=10.0)
            resp.raise_for_status()
        except httpx.HTTPError as exc:
            raise HTTPException(status_code=502, detail=f"Error al cargar el módulo de roles: {exc}") from exc
    return resp.text.replace("http://localhost:8000", "http://localhost:8005")


def _render_local_roles_permissions_page(
    request: Request,
    error_detail: Optional[str] = None,
) -> HTMLResponse:
    db = SessionLocal()
    try:
        roles = db.query(Rol).order_by(Rol.id.asc()).all()
    finally:
        db.close()
    error_html = (
        f"<p style='margin:0 0 12px;color:#991b1b;font-weight:600;'>"
        f"Servicio externo no disponible. Se muestra vista local. Detalle: {escape(error_detail or '')}"
        f"</p>"
        if error_detail
        else ""
    )
    rows_html = "".join(
        [
            (
                "<tr>"
                f"<td>{role.id}</td>"
                f"<td>{escape(role.nombre or '')}</td>"
                f"<td>{escape(role.descripcion or '')}</td>"
                "</tr>"
            )
            for role in roles
        ]
    )
    content = f"""
        <section class="form-section">
            {error_html}
            <div class="section-title">
                <h2>Roles y permisos (vista local)</h2>
            </div>
            <p class="plantillas-hint">El módulo remoto no respondió, por lo que se cargó una vista local de contingencia.</p>
            <div style="overflow:auto;">
                <table style="width:100%; border-collapse:collapse; background:#fff;">
                    <thead>
                        <tr>
                            <th style="border:1px solid #cbd5e1; text-align:left; padding:10px;">ID</th>
                            <th style="border:1px solid #cbd5e1; text-align:left; padding:10px;">Rol</th>
                            <th style="border:1px solid #cbd5e1; text-align:left; padding:10px;">Descripción</th>
                        </tr>
                    </thead>
                    <tbody>
                        {rows_html}
                    </tbody>
                </table>
            </div>
        </section>
    """
    return render_backend_page(
        request,
        title="Roles y permisos",
        description="Gestión de roles y permisos",
        content=content,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/roles-sistema", response_class=HTMLResponse)
@app.get("/roles-permisos", response_class=HTMLResponse)
@app.get("/api/v1/personalizacion/roles-permisos", response_class=HTMLResponse)
@app.get("/personalizacion/roles-permisos", response_class=HTMLResponse)
async def personalizacion_roles_permisos(request: Request):
    require_superadmin(request)
    try:
        content = await _fetch_roles_permissions_view()
        return HTMLResponse(content=content)
    except HTTPException as exc:
        if exc.status_code == 502:
            return _render_local_roles_permissions_page(request, exc.detail)
        raise


@app.get("/", response_class=HTMLResponse)
def root():
    return "<h1>Bienvenido al módulo de planificación estratégica y POA</h1>"

# Área de configuración de imagen (menú)
@app.get("/configura-imagen", response_class=HTMLResponse)
def configura_imagen():
    # Aquí se usará un template en el futuro
    return "<h2>Configuración de imagen (template)</h2>"


def _render_identidad_institucional_page(request: Request) -> HTMLResponse:
    identity = _load_login_identity()
    favicon_url = _build_login_asset_url(identity.get("favicon_filename"), DEFAULT_LOGIN_IDENTITY["favicon_filename"])
    safe_company_short_name = escape(identity.get("company_short_name", DEFAULT_LOGIN_IDENTITY["company_short_name"]))
    safe_login_message = escape(identity.get("login_message", DEFAULT_LOGIN_IDENTITY["login_message"]))
    logo_url = _build_login_asset_url(identity.get("logo_filename"), DEFAULT_LOGIN_IDENTITY["logo_filename"])
    desktop_bg_url = _build_login_asset_url(identity.get("desktop_bg_filename"), DEFAULT_LOGIN_IDENTITY["desktop_bg_filename"])
    mobile_bg_url = _build_login_asset_url(identity.get("mobile_bg_filename"), DEFAULT_LOGIN_IDENTITY["mobile_bg_filename"])
    loaded_assets = sum(1 for value in [favicon_url, logo_url, desktop_bg_url, mobile_bg_url] if (value or "").strip())
    consistency = max(60, min(100, int(round((loaded_assets / 4) * 100)))) if loaded_assets else 60
    saved_flag = request.query_params.get("saved")
    saved_message = "<p class='id-flash'>Identidad institucional actualizada.</p>" if saved_flag == "1" else ""
    content = f"""
        <section class="id-page">
            <style>
                .id-page {{
                    --bg: #f6f8fc;
                    --surface: rgba(255,255,255,.88);
                    --text: #0f172a;
                    --muted: #64748b;
                    --border: rgba(148,163,184,.38);
                    --shadow: 0 18px 40px rgba(15,23,42,.08);
                    --shadow-soft: 0 10px 22px rgba(15,23,42,.06);
                    --radius: 18px;
                    --primary: #0f3d2e;
                    --primary-2: #1f6f52;
                    --ok: #16a34a;
                    --warn: #f59e0b;
                    --crit: #ef4444;
                    width: 100%;
                    font-family: Inter, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif;
                    color: var(--text);
                    background:
                      radial-gradient(1200px 640px at 15% 0%, rgba(15,61,46,.10), transparent 58%),
                      radial-gradient(1000px 540px at 90% 6%, rgba(37,99,235,.10), transparent 55%),
                      var(--bg);
                    border-radius: 18px;
                }}
                .id-wrap {{ width: 100%; margin: 0 auto; padding: 18px 0 34px; }}
                .id-flash {{
                    margin: 0 0 12px;
                    padding: 10px 12px;
                    border-radius: 12px;
                    background: rgba(22,163,74,.10);
                    border: 1px solid rgba(22,163,74,.20);
                    color: #166534;
                    font-weight: 700;
                }}
                .id-btn {{
                    border-radius: 14px;
                    padding: 10px 14px;
                    font-weight: 800;
                    border: 1px solid var(--border);
                    background: rgba(255,255,255,.75);
                    cursor: pointer;
                    box-shadow: var(--shadow-soft);
                    transition: transform .15s ease, box-shadow .15s ease, background .15s ease;
                }}
                .id-btn:hover {{ transform: translateY(-1px); box-shadow: var(--shadow); background: rgba(255,255,255,.95); }}
                .id-btn--primary {{
                    background: linear-gradient(135deg, var(--primary), var(--primary-2));
                    color: #fff;
                    border-color: rgba(15,61,46,.35);
                }}
                .id-btn--primary2 {{
                    background: rgba(37,99,235,.12);
                    color: #1d4ed8;
                    border-color: rgba(37,99,235,.24);
                }}
                .id-btn--soft {{
                    background: rgba(15,61,46,.10);
                    border-color: rgba(15,61,46,.18);
                    color: #0b2a20;
                }}
                .id-btn--ghost2 {{
                    background: rgba(255,255,255,.85);
                    color: var(--text);
                    border-color: var(--border);
                }}
                .id-btn--ghost3 {{
                    background: rgba(255,255,255,.90);
                    color: var(--text);
                    border-color: var(--border);
                }}
                .id-btn--danger {{
                    border-color: rgba(239,68,68,.25);
                    color: #991b1b;
                    background: rgba(239,68,68,.08);
                }}
                .id-stats {{
                    display: grid;
                    grid-template-columns: repeat(4, minmax(0, 1fr));
                    gap: 12px;
                    margin-bottom: 14px;
                }}
                .id-stat {{
                    background: var(--surface);
                    border: 1px solid var(--border);
                    border-radius: var(--radius);
                    box-shadow: var(--shadow-soft);
                    padding: 14px;
                }}
                .id-stat__k {{ color: var(--muted); font-size: 12px; font-weight: 700; }}
                .id-stat__v {{ margin-top: 8px; font-size: 28px; font-weight: 900; letter-spacing: -0.02em; }}
                .id-stat__meta {{ margin-top: 8px; display: flex; gap: 8px; flex-wrap: wrap; align-items: center; }}
                .id-chip {{
                    font-size: 12px;
                    padding: 6px 10px;
                    border-radius: 999px;
                    background: rgba(15,23,42,.05);
                    border: 1px solid rgba(15,23,42,.08);
                    color: rgba(15,23,42,.72);
                }}
                .id-chip--ok {{ background: rgba(22,163,74,.10); border-color: rgba(22,163,74,.20); color: #166534; }}
                .id-bar {{
                    height: 10px;
                    flex: 1 1 auto;
                    min-width: 110px;
                    border-radius: 999px;
                    background: rgba(148,163,184,.25);
                    border: 1px solid rgba(148,163,184,.25);
                    overflow: hidden;
                }}
                .id-bar__fill {{
                    height: 100%;
                    border-radius: 999px;
                    background: linear-gradient(90deg, rgba(15,61,46,1), rgba(31,111,82,1));
                }}
                .id-grid {{ display: grid; grid-template-columns: 1.1fr .9fr; gap: 14px; align-items: start; margin-bottom: 14px; }}
                .id-card {{
                    background: var(--surface);
                    border: 1px solid var(--border);
                    border-radius: 22px;
                    box-shadow: var(--shadow-soft);
                    overflow: hidden;
                }}
                .id-card--pad {{ padding: 16px; }}
                .id-card__head {{
                    display: flex;
                    align-items: flex-start;
                    justify-content: space-between;
                    gap: 12px;
                    margin-bottom: 14px;
                }}
                .id-card__head h2 {{ margin: 0; font-size: 20px; letter-spacing: -0.02em; }}
                .id-card__head p {{ margin: 6px 0 0; color: var(--muted); font-size: 13px; }}
                .id-card__tools {{ display: flex; gap: 10px; align-items: center; }}
                .id-pill {{
                    font-size: 12px;
                    padding: 6px 10px;
                    border-radius: 999px;
                    background: rgba(255,255,255,.70);
                    border: 1px solid var(--border);
                    color: rgba(15,23,42,.72);
                }}
                .id-pill--soft {{
                    background: rgba(15,61,46,.10);
                    border-color: rgba(15,61,46,.18);
                    color: #0b2a20;
                }}
                .id-iconbtn {{
                    width: 38px;
                    height: 38px;
                    border-radius: 14px;
                    border: 1px solid var(--border);
                    background: rgba(255,255,255,.75);
                    box-shadow: var(--shadow-soft);
                    cursor: pointer;
                }}
                .id-form {{ display: flex; flex-direction: column; gap: 12px; }}
                .id-field {{ display: flex; flex-direction: column; gap: 8px; }}
                .id-field label {{ font-size: 13px; font-weight: 700; color: #334155; }}
                .id-field input,
                .id-field textarea {{
                    border: 1px solid var(--border);
                    border-radius: 14px;
                    padding: 12px;
                    background: rgba(255,255,255,.85);
                    box-shadow: 0 10px 20px rgba(15,23,42,.04);
                }}
                .id-field textarea {{ min-height: 90px; resize: vertical; }}
                .id-help {{ color: var(--muted); font-size: 12px; }}
                .id-actions {{ display: flex; gap: 10px; justify-content: flex-end; margin-top: 4px; }}
                .id-tips {{ display: flex; flex-direction: column; gap: 10px; }}
                .id-tip {{
                    display: flex;
                    gap: 10px;
                    align-items: flex-start;
                    background: rgba(255,255,255,.82);
                    border: 1px solid rgba(148,163,184,.30);
                    border-radius: 14px;
                    padding: 10px;
                }}
                .id-tip__icon {{ font-size: 18px; }}
                .id-tip__text strong {{ display: block; font-size: 13px; }}
                .id-tip__text p {{ margin: 4px 0 0; color: var(--muted); font-size: 12px; line-height: 1.35; }}
                .id-divider {{ height: 1px; background: rgba(148,163,184,.25); margin: 4px 0; }}
                .id-cta {{ display: flex; justify-content: space-between; align-items: center; gap: 10px; }}
                .id-cta p {{ margin: 4px 0 0; color: var(--muted); font-size: 12px; }}
                .id-assets__grid {{ display: flex; flex-direction: column; gap: 10px; }}
                .id-asset {{
                    display: grid;
                    grid-template-columns: 240px minmax(0, 1fr) auto;
                    gap: 14px;
                    align-items: center;
                    background: rgba(255,255,255,.86);
                    border: 1px solid rgba(148,163,184,.30);
                    border-radius: 16px;
                    padding: 12px;
                }}
                .id-asset__label {{ display: flex; gap: 8px; align-items: center; }}
                .id-asset__meta {{ margin-top: 4px; font-size: 12px; color: var(--muted); }}
                .id-dot {{ width: 10px; height: 10px; border-radius: 999px; display: inline-block; }}
                .id-dot--ok {{ background: var(--ok); box-shadow: 0 0 0 6px rgba(22,163,74,.10); }}
                .id-asset__preview {{
                    border: 1px solid rgba(148,163,184,.30);
                    background: rgba(248,250,252,.95);
                    border-radius: 12px;
                    overflow: hidden;
                    display: flex;
                    align-items: center;
                    justify-content: center;
                }}
                .id-asset__preview img {{ width: 100%; height: 100%; object-fit: cover; display: block; }}
                .id-asset__preview--square {{ width: 96px; height: 96px; }}
                .id-asset__preview--square img {{ object-fit: contain; padding: 8px; }}
                .id-asset__preview--logo {{ height: 120px; }}
                .id-asset__preview--logo img {{ object-fit: contain; padding: 10px; }}
                .id-asset__preview--wide {{ height: 140px; }}
                .id-asset__actions {{ display: flex; flex-direction: column; gap: 8px; min-width: 116px; }}
                @media (max-width: 1200px) {{
                    .id-stats {{ grid-template-columns: repeat(2, minmax(0, 1fr)); }}
                    .id-grid {{ grid-template-columns: 1fr; }}
                    .id-asset {{ grid-template-columns: 1fr; }}
                    .id-asset__actions {{ flex-direction: row; }}
                }}
                @media (max-width: 640px) {{
                    .id-stats {{ grid-template-columns: 1fr; }}
                    .id-actions {{ justify-content: flex-start; flex-wrap: wrap; }}
                    .id-cta {{ flex-direction: column; align-items: flex-start; }}
                }}
            </style>
            <div class="id-wrap">
                <form id="identity-form" method="post" action="/identidad-institucional" enctype="multipart/form-data">
                    <input type="hidden" id="remove_favicon" name="remove_favicon" value="0">
                    <input type="hidden" id="remove_logo" name="remove_logo" value="0">
                    <input type="hidden" id="remove_desktop" name="remove_desktop" value="0">
                    <input type="hidden" id="remove_mobile" name="remove_mobile" value="0">
                    <div style="display:none;">
                        <input type="file" id="favicon" name="favicon" accept="image/*">
                        <input type="file" id="logo_empresa" name="logo_empresa" accept="image/*">
                        <input type="file" id="fondo_escritorio" name="fondo_escritorio" accept="image/*">
                        <input type="file" id="fondo_movil" name="fondo_movil" accept="image/*">
                    </div>
                    {saved_message}
                    <section class="id-stats">
                        <article class="id-stat">
                            <div class="id-stat__k">Nombre corto</div>
                            <div class="id-stat__v">{safe_company_short_name}</div>
                            <div class="id-stat__meta"><span class="id-chip">Activo</span></div>
                        </article>
                        <article class="id-stat">
                            <div class="id-stat__k">Recursos cargados</div>
                            <div class="id-stat__v">{loaded_assets}</div>
                            <div class="id-stat__meta"><span class="id-chip id-chip--ok">Completo</span></div>
                        </article>
                        <article class="id-stat">
                            <div class="id-stat__k">Formato recomendado</div>
                            <div class="id-stat__v">SVG/PNG</div>
                            <div class="id-stat__meta"><span class="id-chip">Alta nitidez</span></div>
                        </article>
                        <article class="id-stat">
                            <div class="id-stat__k">Consistencia visual</div>
                            <div class="id-stat__v">{consistency}%</div>
                            <div class="id-stat__meta">
                                <div class="id-bar"><div class="id-bar__fill" style="width:{consistency}%"></div></div>
                                <span class="id-chip id-chip--ok">Optimo</span>
                            </div>
                        </article>
                    </section>
                    <section class="id-grid">
                        <section class="id-card id-card--pad">
                            <header class="id-card__head">
                                <div>
                                    <h2>Datos institucionales</h2>
                                    <p>Estos datos se muestran en login y en elementos de marca del sistema.</p>
                                </div>
                                <div class="id-card__tools">
                                    <span class="id-pill id-pill--soft">Configuracion</span>
                                    <button class="id-iconbtn" type="button" title="Ayuda">?</button>
                                </div>
                            </header>
                            <div class="id-form">
                                <div class="id-field">
                                    <label for="company_short_name">Nombre corto de la empresa</label>
                                    <input class="campo campo-personalizado" id="company_short_name" name="company_short_name" type="text" value="{safe_company_short_name}" placeholder="Ej. AVAN">
                                    <div class="id-help">Se usa en titulos, encabezados y login. Recomendado: 3-18 caracteres.</div>
                                </div>
                                <div class="id-field">
                                    <label for="login_message">Mensaje para pantalla de login</label>
                                    <textarea class="campo campo-personalizado" id="login_message" name="login_message" placeholder="Ej. Incrementando el nivel de eficiencia">{safe_login_message}</textarea>
                                    <div class="id-help">Sugerencia: frase corta y orientada a valor.</div>
                                </div>
                                <div class="id-actions">
                                    <button class="id-btn id-btn--soft" type="reset">Restablecer</button>
                                    <button class="id-btn id-btn--primary2" type="submit">Guardar identidad</button>
                                </div>
                            </div>
                        </section>
                        <aside class="id-card id-card--pad">
                            <header class="id-card__head">
                                <div>
                                    <h2>Recomendaciones</h2>
                                    <p>Buenas practicas para mantener calidad visual en web y movil.</p>
                                </div>
                                <div class="id-card__tools"><span class="id-pill">Guia rapida</span></div>
                            </header>
                            <div class="id-tips">
                                <div class="id-tip"><div class="id-tip__icon">L</div><div class="id-tip__text"><strong>Logo</strong><p>Preferible SVG; si usas PNG que sea transparente y amplio.</p></div></div>
                                <div class="id-tip"><div class="id-tip__icon">F</div><div class="id-tip__text"><strong>Favicon</strong><p>Usa 32x32 y 64x64 con buena legibilidad.</p></div></div>
                                <div class="id-tip"><div class="id-tip__icon">M</div><div class="id-tip__text"><strong>Fondo movil</strong><p>Formato vertical con punto focal centrado.</p></div></div>
                                <div class="id-tip"><div class="id-tip__icon">E</div><div class="id-tip__text"><strong>Fondo escritorio</strong><p>Recomendado 1920x1080 y buen contraste.</p></div></div>
                                <div class="id-divider"></div>
                                <div class="id-cta">
                                    <div><strong>Vista previa</strong><p>Valida como se vera el login con los recursos actuales.</p></div>
                                    <button class="id-btn id-btn--ghost2" type="button">Abrir preview</button>
                                </div>
                            </div>
                        </aside>
                    </section>
                    <section class="id-card id-card--pad id-assets">
                        <header class="id-card__head">
                            <div>
                                <h2>Recursos visuales</h2>
                                <p>Administra favicon, logo y fondos para web y movil.</p>
                            </div>
                            <div class="id-card__tools">
                                <span class="id-pill">4 recursos</span>
                                <button class="id-btn id-btn--primary" type="submit">Guardar cambios</button>
                            </div>
                        </header>
                        <div class="id-assets__grid">
                            <article class="id-asset">
                                <div class="id-asset__left">
                                    <div class="id-asset__label"><span class="id-dot id-dot--ok"></span><strong>Favicon</strong></div>
                                    <div class="id-asset__meta">Recomendado: 32x32 / 64x64 - PNG/ICO</div>
                                </div>
                                <div class="id-asset__preview id-asset__preview--square">
                                    <img src="{favicon_url}" alt="Favicon">
                                </div>
                                <div class="id-asset__actions identity-asset-actions">
                                    <button class="id-btn id-btn--ghost3 identity-asset-edit" data-target-input="favicon" type="button">Editar</button>
                                    <button class="id-btn id-btn--danger identity-asset-delete" data-target-remove="remove_favicon" type="button">Eliminar</button>
                                </div>
                            </article>
                            <article class="id-asset">
                                <div class="id-asset__left">
                                    <div class="id-asset__label"><span class="id-dot id-dot--ok"></span><strong>Logo de la empresa</strong></div>
                                    <div class="id-asset__meta">Preferible SVG - alternativa PNG transparente</div>
                                </div>
                                <div class="id-asset__preview id-asset__preview--logo">
                                    <img src="{logo_url}" alt="Logo">
                                </div>
                                <div class="id-asset__actions identity-asset-actions">
                                    <button class="id-btn id-btn--ghost3 identity-asset-edit" data-target-input="logo_empresa" type="button">Editar</button>
                                    <button class="id-btn id-btn--danger identity-asset-delete" data-target-remove="remove_logo" type="button">Eliminar</button>
                                </div>
                            </article>
                            <article class="id-asset">
                                <div class="id-asset__left">
                                    <div class="id-asset__label"><span class="id-dot id-dot--ok"></span><strong>Fondo de escritorio</strong></div>
                                    <div class="id-asset__meta">Recomendado: 1920x1080 - JPG/PNG</div>
                                </div>
                                <div class="id-asset__preview id-asset__preview--wide">
                                    <img src="{desktop_bg_url}" alt="Fondo de escritorio">
                                </div>
                                <div class="id-asset__actions identity-asset-actions">
                                    <button class="id-btn id-btn--ghost3 identity-asset-edit" data-target-input="fondo_escritorio" type="button">Editar</button>
                                    <button class="id-btn id-btn--danger identity-asset-delete" data-target-remove="remove_desktop" type="button">Eliminar</button>
                                </div>
                            </article>
                            <article class="id-asset">
                                <div class="id-asset__left">
                                    <div class="id-asset__label"><span class="id-dot id-dot--ok"></span><strong>Fondo movil</strong></div>
                                    <div class="id-asset__meta">Recomendado: 1080x1920 - JPG/PNG</div>
                                </div>
                                <div class="id-asset__preview id-asset__preview--wide">
                                    <img src="{mobile_bg_url}" alt="Fondo movil">
                                </div>
                                <div class="id-asset__actions identity-asset-actions">
                                    <button class="id-btn id-btn--ghost3 identity-asset-edit" data-target-input="fondo_movil" type="button">Editar</button>
                                    <button class="id-btn id-btn--danger identity-asset-delete" data-target-remove="remove_mobile" type="button">Eliminar</button>
                                </div>
                            </article>
                        </div>
                    </section>
                </form>
            </div>
        </section>
    """
    return render_backend_page(
        request,
        title="Identidad institucional",
        description="Configuración de identidad para la pantalla de login.",
        content=content,
        hide_floating_actions=True,
        show_page_header=True,
    )


@app.get("/identidad-institucional", response_class=HTMLResponse)
def identidad_institucional_page(request: Request):
    require_superadmin(request)
    return _render_identidad_institucional_page(request)


@app.post("/identidad-institucional", response_class=HTMLResponse)
async def identidad_institucional_save(
    request: Request,
    company_short_name: str = Form(""),
    login_message: str = Form(""),
    favicon: Optional[UploadFile] = File(None),
    logo_empresa: Optional[UploadFile] = File(None),
    fondo_escritorio: Optional[UploadFile] = File(None),
    fondo_movil: Optional[UploadFile] = File(None),
    remove_favicon: str = Form("0"),
    remove_logo: str = Form("0"),
    remove_desktop: str = Form("0"),
    remove_mobile: str = Form("0"),
):
    require_superadmin(request)
    current = _load_login_identity()
    current["company_short_name"] = company_short_name.strip() or DEFAULT_LOGIN_IDENTITY["company_short_name"]
    current["login_message"] = login_message.strip() or DEFAULT_LOGIN_IDENTITY["login_message"]

    if str(remove_favicon).strip() == "1":
        _remove_login_image_if_custom(current.get("favicon_filename"))
        current["favicon_filename"] = DEFAULT_LOGIN_IDENTITY["favicon_filename"]
    if str(remove_logo).strip() == "1":
        _remove_login_image_if_custom(current.get("logo_filename"))
        current["logo_filename"] = DEFAULT_LOGIN_IDENTITY["logo_filename"]
    if str(remove_desktop).strip() == "1":
        _remove_login_image_if_custom(current.get("desktop_bg_filename"))
        current["desktop_bg_filename"] = DEFAULT_LOGIN_IDENTITY["desktop_bg_filename"]
    if str(remove_mobile).strip() == "1":
        _remove_login_image_if_custom(current.get("mobile_bg_filename"))
        current["mobile_bg_filename"] = DEFAULT_LOGIN_IDENTITY["mobile_bg_filename"]

    new_favicon = await _store_login_image(favicon, "favicon") if favicon else None
    if new_favicon:
        _remove_login_image_if_custom(current.get("favicon_filename"))
        current["favicon_filename"] = new_favicon

    new_logo = await _store_login_image(logo_empresa, "logo_empresa") if logo_empresa else None
    if new_logo:
        _remove_login_image_if_custom(current.get("logo_filename"))
        current["logo_filename"] = new_logo

    new_desktop = await _store_login_image(fondo_escritorio, "fondo_escritorio") if fondo_escritorio else None
    if new_desktop:
        _remove_login_image_if_custom(current.get("desktop_bg_filename"))
        current["desktop_bg_filename"] = new_desktop

    new_mobile = await _store_login_image(fondo_movil, "fondo_movil") if fondo_movil else None
    if new_mobile:
        _remove_login_image_if_custom(current.get("mobile_bg_filename"))
        current["mobile_bg_filename"] = new_mobile

    _save_login_identity(current)
    return RedirectResponse(url="/identidad-institucional?saved=1", status_code=303)

# Placeholder para templates
# En el futuro, importar y usar templates para todas las respuestas

if __name__ == "__main__":
    import os
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=int(os.environ.get("PORT", "8005")))
