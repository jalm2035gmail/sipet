import io
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable, Image,
)

logger = logging.getLogger(__name__)


# ── PDF ───────────────────────────────────────────────────────────────────────

def generate_pdf(
    title: str,
    sections: list[dict],  # [{"heading": str, "body": str, "table": [[...]] }]
    output_path: str | None = None,
) -> bytes:
    """
    Genera un PDF con título, secciones de texto y tablas opcionales.
    Retorna los bytes del PDF. Si output_path se provee, también lo guarda.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    style_title = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=20,
        textColor=colors.HexColor("#1e293b"),
        spaceAfter=12,
    )
    style_heading = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=colors.HexColor("#334155"),
        spaceBefore=16,
        spaceAfter=6,
    )
    style_body = ParagraphStyle(
        "CustomBody",
        parent=styles["Normal"],
        fontSize=10,
        leading=14,
        textColor=colors.HexColor("#475569"),
    )

    story = []

    # Header
    story.append(Paragraph(title, style_title))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        style_body,
    ))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#e2e8f0")))
    story.append(Spacer(1, 0.4 * cm))

    for section in sections:
        if heading := section.get("heading"):
            story.append(Paragraph(heading, style_heading))

        if body := section.get("body"):
            story.append(Paragraph(body, style_body))
            story.append(Spacer(1, 0.3 * cm))

        if table_data := section.get("table"):
            col_count = len(table_data[0]) if table_data else 1
            col_width = (A4[0] - 4 * cm) / col_count

            t = Table(table_data, colWidths=[col_width] * col_count, repeatRows=1)
            t.setStyle(TableStyle([
                ("BACKGROUND",  (0, 0), (-1, 0),  colors.HexColor("#1e40af")),
                ("TEXTCOLOR",   (0, 0), (-1, 0),  colors.white),
                ("FONTNAME",    (0, 0), (-1, 0),  "Helvetica-Bold"),
                ("FONTSIZE",    (0, 0), (-1, 0),  10),
                ("ALIGN",       (0, 0), (-1, -1), "LEFT"),
                ("VALIGN",      (0, 0), (-1, -1), "MIDDLE"),
                ("FONTSIZE",    (0, 1), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
                ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#cbd5e1")),
                ("TOPPADDING",  (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ]))
            story.append(t)
            story.append(Spacer(1, 0.4 * cm))

    doc.build(story)
    pdf_bytes = buffer.getvalue()

    if output_path:
        Path(output_path).write_bytes(pdf_bytes)
        logger.info("PDF saved: %s", output_path)

    return pdf_bytes


def dataframe_to_pdf(
    df: pd.DataFrame,
    title: str,
    output_path: str | None = None,
) -> bytes:
    """Shortcut: convierte un DataFrame directamente a PDF."""
    headers = list(df.columns)
    rows = df.astype(str).values.tolist()
    return generate_pdf(
        title=title,
        sections=[{"heading": "Data", "table": [headers] + rows}],
        output_path=output_path,
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

def _apply_header_style(ws, row: int, col_count: int):
    header_fill = PatternFill("solid", fgColor="1e40af")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        bottom=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
    )
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = border


def generate_excel(
    sheets: list[dict],  # [{"name": str, "dataframe": pd.DataFrame}]
    output_path: str | None = None,
) -> bytes:
    """
    Genera un Excel con múltiples hojas a partir de DataFrames.
    Retorna bytes. Si output_path se provee también guarda el archivo.
    """
    wb = Workbook()
    wb.remove(wb.active)  # elimina la hoja por defecto

    even_fill = PatternFill("solid", fgColor="F1F5F9")
    border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["name"][:31])
        df: pd.DataFrame = sheet_def["dataframe"]

        # Headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=header)
        _apply_header_style(ws, row=1, col_count=len(headers))
        ws.row_dimensions[1].height = 22

        # Data rows
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if row_idx % 2 == 0:
                    cell.fill = even_fill

        # Auto column width
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        # Freeze header
        ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    excel_bytes = buffer.getvalue()

    if output_path:
        Path(output_path).write_bytes(excel_bytes)
        logger.info("Excel saved: %s", output_path)

    return excel_bytes


def dataframe_to_excel(
    df: pd.DataFrame,
    sheet_name: str = "Sheet1",
    output_path: str | None = None,
) -> bytes:
    """Shortcut: un solo DataFrame a Excel."""
    return generate_excel(
        sheets=[{"name": sheet_name, "dataframe": df}],
        output_path=output_path,
    )


def excel_to_dataframe(file_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    """Lee un Excel subido y lo convierte a DataFrame."""
    return pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name or 0)
