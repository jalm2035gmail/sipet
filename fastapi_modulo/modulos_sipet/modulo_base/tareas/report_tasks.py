from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any

from fastapi_modulo.modulos_sipet.modulo_base.core.task_queue import (
    build_module_task_registry,
    create_module_task_queue,
)
from fastapi_modulo.modulos_sipet.modulo_base.tareas.celery_app import celery_app

logger = logging.getLogger(__name__)

registry = build_module_task_registry("modulo_base")
registry.register("report_export", queue="modulo_base_reports")
task_queue = create_module_task_queue("modulo_base", celery_app=celery_app, registry=registry)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _utcnow_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _require_pandas() -> Any:
    try:
        import pandas as pd
        return pd
    except ImportError as exc:
        raise RuntimeError("pandas no esta disponible. Instala: pip install pandas") from exc


def _require_openpyxl() -> None:
    try:
        import openpyxl  # noqa: F401
    except ImportError as exc:
        raise RuntimeError("openpyxl no esta disponible. Instala: pip install openpyxl") from exc


def _require_reportlab() -> Any:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        return {
            "colors": colors,
            "A4": A4,
            "landscape": landscape,
            "getSampleStyleSheet": getSampleStyleSheet,
            "cm": cm,
            "Paragraph": Paragraph,
            "SimpleDocTemplate": SimpleDocTemplate,
            "Spacer": Spacer,
            "Table": Table,
            "TableStyle": TableStyle,
        }
    except ImportError as exc:
        raise RuntimeError("reportlab no esta disponible. Instala: pip install reportlab") from exc


# ── Generadores de archivo ────────────────────────────────────────────────────

def build_excel_report(
    rows: list[dict[str, Any]],
    *,
    sheet_name: str = "Reporte",
    title: str = "",
) -> bytes:
    """
    Recibe una lista de dicts y devuelve bytes de un archivo .xlsx.
    Usa pandas + openpyxl. La primera fila es el encabezado en negrita.
    """
    pd = _require_pandas()
    _require_openpyxl()

    df = pd.DataFrame(rows)
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        start_row = 0
        if title:
            # Escribir título en la primera fila y dejar una fila vacía
            df_title = pd.DataFrame([[title]])
            df_title.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=0)
            start_row = 2

        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=start_row)

        # Dar formato al encabezado
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        from openpyxl.styles import Alignment, Font, PatternFill

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_row = start_row + 1  # openpyxl es 1-indexed

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = worksheet.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            # Autoajustar ancho de columna
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) else 0)
            worksheet.column_dimensions[cell.column_letter].width = min(max_len + 4, 50)

        if title:
            title_cell = worksheet.cell(row=1, column=1)
            title_cell.font = Font(bold=True, size=14)

    return output.getvalue()


def build_pdf_report(
    rows: list[dict[str, Any]],
    *,
    title: str = "Reporte",
    landscape_mode: bool = False,
    tenant_id: str = "",
) -> bytes:
    """
    Recibe una lista de dicts y devuelve bytes de un archivo .pdf.
    Usa reportlab con tabla automática basada en las claves del primer dict.
    """
    rl = _require_reportlab()

    page_size = rl["landscape"](rl["A4"]) if landscape_mode else rl["A4"]
    output = io.BytesIO()
    doc = rl["SimpleDocTemplate"](
        output,
        pagesize=page_size,
        leftMargin=1.5 * rl["cm"],
        rightMargin=1.5 * rl["cm"],
        topMargin=2 * rl["cm"],
        bottomMargin=2 * rl["cm"],
    )

    styles = rl["getSampleStyleSheet"]()
    elements = []

    # Título
    elements.append(rl["Paragraph"](f"<b>{title}</b>", styles["Title"]))
    if tenant_id:
        elements.append(rl["Paragraph"](
            f"<font size=9>Tenant: {tenant_id} — Generado: {_utcnow_iso()}</font>",
            styles["Normal"],
        ))
    elements.append(rl["Spacer"](1, 0.4 * rl["cm"]))

    if not rows:
        elements.append(rl["Paragraph"]("Sin datos para mostrar.", styles["Normal"]))
    else:
        columns = list(rows[0].keys())
        header = [str(col).upper() for col in columns]
        data = [header] + [[str(row.get(col, "")) for col in columns] for row in rows]

        col_width = (page_size[0] - 3 * rl["cm"]) / len(columns)
        table = rl["Table"](data, colWidths=[col_width] * len(columns), repeatRows=1)
        table.setStyle(rl["TableStyle"]([
            ("BACKGROUND", (0, 0), (-1, 0), rl["colors"].HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl["colors"].white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl["colors"].white, rl["colors"].HexColor("#EBF3FB")]),
            ("GRID", (0, 0), (-1, -1), 0.4, rl["colors"].HexColor("#CCCCCC")),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(table)

    doc.build(elements)
    return output.getvalue()


# ── Tarea Celery ──────────────────────────────────────────────────────────────

if celery_app is not None:
    @celery_app.task(name="modulo_base.report_export", bind=True, max_retries=3)
    def report_export_task(
        self: Any,
        *,
        task_id: str = "",
        tenant_id: str = "default",
        formato: str = "xlsx",
        titulo: str = "Reporte",
        rows: list[dict[str, Any]] | None = None,
        landscape_mode: bool = False,
        **kwargs: Any,
    ) -> dict[str, Any]:
        """
        Genera un reporte en formato xlsx o pdf a partir de una lista de dicts.

        Uso desde un servicio:
            task_queue.queue_task("report_export", kwargs={
                "tenant_id": "acme",
                "formato": "xlsx",
                "titulo": "Registros Q1",
                "rows": [{"nombre": "X", "estado": "activo"}, ...],
            })
        """
        task_queue.report_task_state(
            "report_export", task_id, status="running"
        )
        try:
            data = rows or []
            formato_norm = str(formato or "xlsx").strip().lower()

            if formato_norm == "pdf":
                content = build_pdf_report(
                    data,
                    title=titulo,
                    landscape_mode=landscape_mode,
                    tenant_id=tenant_id,
                )
                extension = "pdf"
                mime = "application/pdf"
            else:
                content = build_excel_report(data, title=titulo)
                extension = "xlsx"
                mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            filename = f"reporte_{tenant_id}_{_utcnow_iso().replace(':', '-')}.{extension}"

            result = {
                "kind": "report",
                "tenant_id": tenant_id,
                "formato": extension,
                "mime": mime,
                "filename": filename,
                "size_bytes": len(content),
                "generated_at": _utcnow_iso(),
                # content en base64 para poder enviarlo por Redis si es pequeño
                "content_b64": __import__("base64").b64encode(content).decode(),
            }

            logger.info(
                "report_export_completed",
                extra={"tenant_id": tenant_id, "formato": extension, "size_bytes": len(content)},
            )

            return task_queue.report_task_state(
                "report_export", task_id, status="completed", result=result
            )

        except Exception as exc:
            logger.error(
                "report_export_failed",
                extra={"tenant_id": tenant_id, "error": str(exc)},
                exc_info=True,
            )
            task_queue.report_task_state(
                "report_export", task_id, status="failed", error=str(exc)
            )
            raise self.retry(exc=exc, countdown=10)


__all__ = [
    "build_excel_report",
    "build_pdf_report",
    "registry",
    "task_queue",
]
