from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from fastapi_modulo.modulos_sipet.aplicaciones.repositorios.persistence_repository import (
    list_registry_state,
)
from fastapi_modulo.modulos_sipet.aplicaciones.servicios.catalog_service import (
    decorate_modules_payload,
)

# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="EAF0FB")
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def _write_header(ws, columns: list[str]) -> None:
    for col_idx, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
    ws.row_dimensions[1].height = 22


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _bool_label(value: Any) -> str:
    return "Sí" if value else "No"


# ---------------------------------------------------------------------------
# sheet 1 – catálogo de módulos
# ---------------------------------------------------------------------------

_CATALOG_COLS = [
    "Clave",
    "Nombre",
    "Versión instalada",
    "Habilitado",
    "Protocolo OK",
    "Tiene __init__",
    "Tiene __manifest__",
    "Archivos faltantes",
    "Subida paquete",
    "Nombre archivo subido",
    "Fecha subida",
    "Directorio módulo",
]


def _fill_catalog_sheet(ws, modules: list[dict[str, Any]]) -> None:
    _write_header(ws, _CATALOG_COLS)
    for row_idx, mod in enumerate(modules, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        values = [
            mod.get("key", ""),
            mod.get("name", ""),
            mod.get("installed_version", ""),
            _bool_label(mod.get("enabled")),
            _bool_label(mod.get("protocol_ok")),
            _bool_label(mod.get("protocol_has_init")),
            _bool_label(mod.get("protocol_has_manifest")),
            ", ".join(mod.get("protocol_missing") or []),
            _bool_label(mod.get("package_upload_enabled")),
            mod.get("uploaded_filename", ""),
            mod.get("uploaded_at", ""),
            mod.get("module_dir", ""),
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _LEFT
            if fill:
                cell.fill = fill
    _auto_width(ws)


# ---------------------------------------------------------------------------
# sheet 2 – historial de auditoría
# ---------------------------------------------------------------------------

_AUDIT_COLS = [
    "ID",
    "Clave módulo",
    "Acción",
    "Resultado",
    "Usuario",
    "IP",
    "Fecha (UTC)",
    "Payload",
]


def _list_all_registry_audits() -> list[Any]:
    """Query all AppRegistryAudit rows ordered by created_at desc."""
    from sqlalchemy import desc
    from sqlalchemy.exc import OperationalError, ProgrammingError

    from fastapi_modulo.core import db as core_db
    from fastapi_modulo.modulos_sipet.aplicaciones.modelos.db_models import AppRegistryAudit

    try:
        session_factory = core_db.get_session_factory_for_host("")
        with session_factory() as db:
            return (
                db.query(AppRegistryAudit)
                .order_by(desc(AppRegistryAudit.created_at), desc(AppRegistryAudit.id))
                .all()
            )
    except (OperationalError, ProgrammingError):
        return []


def _fill_audit_sheet(ws) -> None:
    _write_header(ws, _AUDIT_COLS)
    rows = _list_all_registry_audits()
    for row_idx, audit in enumerate(rows, start=2):
        fill = _ALT_FILL if row_idx % 2 == 0 else None
        created = audit.created_at
        if isinstance(created, datetime):
            created_str = created.strftime("%Y-%m-%d %H:%M:%S")
        else:
            created_str = str(created or "")
        values = [
            audit.id,
            audit.module_key,
            audit.action,
            audit.result,
            audit.user_id or "",
            audit.ip or "",
            created_str,
            audit.payload_json or "{}",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _LEFT
            if fill:
                cell.fill = fill
    _auto_width(ws)


# ---------------------------------------------------------------------------
# public API
# ---------------------------------------------------------------------------

def generate_catalog_excel(tenant_key: str | None = None) -> bytes:
    """Return Excel bytes with two sheets: catalog + audit history."""
    modules = decorate_modules_payload(tenant_key=tenant_key)

    wb = openpyxl.Workbook()

    ws_catalog = wb.active
    ws_catalog.title = "Catálogo de Módulos"
    ws_catalog.freeze_panes = "A2"
    _fill_catalog_sheet(ws_catalog, modules)

    ws_audit = wb.create_sheet(title="Historial de Auditoría")
    ws_audit.freeze_panes = "A2"
    _fill_audit_sheet(ws_audit)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def get_export_filename(prefix: str = "catalogo_modulos") -> str:
    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{ts}.xlsx"


def export_catalog_xlsx(tenant_key: str | None = None) -> io.BytesIO:
    """Compatibility wrapper that returns a BytesIO stream."""
    data = generate_catalog_excel(tenant_key=tenant_key)
    buf = io.BytesIO(data)
    buf.seek(0)
    return buf


__all__ = ["export_catalog_xlsx", "generate_catalog_excel", "get_export_filename"]
